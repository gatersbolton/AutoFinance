from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List, Tuple

import yaml
from openpyxl import load_workbook

from ..models import AliasRecord, RelationRecord, TemplateSubject
from ..normalize.text import normalize_label_for_matching


SUBJECT_RE = re.compile(r"^(?P<code>[A-Z]{2}_[0-9]{3})\s+(?P<name>.+)$")


def load_template_subjects(template_path: Path) -> Tuple[List[TemplateSubject], str, int]:
    workbook = load_workbook(template_path)
    best_sheet = None
    best_subjects: List[TemplateSubject] = []
    best_header_row = 1

    for worksheet in workbook.worksheets:
        subjects: List[TemplateSubject] = []
        header_row = 1
        for row_idx in range(1, worksheet.max_row + 1):
            value = worksheet.cell(row=row_idx, column=1).value
            if value is None:
                continue
            text = str(value).strip()
            if text == "科目名称":
                header_row = row_idx
            match = SUBJECT_RE.match(text)
            if not match:
                continue
            subjects.append(
                TemplateSubject(
                    code=match.group("code"),
                    canonical_name=match.group("name").strip(),
                    row_index=row_idx,
                    sheet_name=worksheet.title,
                    source_value=text,
                )
            )
        if len(subjects) > len(best_subjects):
            best_sheet = worksheet.title
            best_subjects = subjects
            best_header_row = header_row

    if not best_subjects or not best_sheet:
        raise ValueError(f"No standard subject master found in template: {template_path}")
    return best_subjects, best_sheet, best_header_row


def build_subject_index(subjects: List[TemplateSubject]) -> Dict[str, TemplateSubject]:
    return {
        normalize_label_for_matching(subject.canonical_name): subject
        for subject in subjects
    }


def load_alias_records(config_path: Path, subjects: List[TemplateSubject]) -> List[AliasRecord]:
    if not config_path.exists():
        return []
    payload = yaml.safe_load(config_path.read_text(encoding="utf-8")) or {}
    subject_by_code = {subject.code: subject for subject in subjects}
    subject_by_name = {normalize_label_for_matching(subject.canonical_name): subject for subject in subjects}
    aliases = payload.get("aliases", [])
    records: List[AliasRecord] = []

    if isinstance(aliases, dict):
        for canonical_name, alias_values in aliases.items():
            subject = subject_by_name.get(normalize_label_for_matching(canonical_name))
            if not subject:
                continue
            candidates = alias_values if isinstance(alias_values, list) else [alias_values]
            for alias in candidates:
                records.append(
                    AliasRecord(
                        canonical_code=subject.code,
                        canonical_name=subject.canonical_name,
                        alias=str(alias).strip(),
                        alias_type="exact_alias",
                        enabled=True,
                        note="migrated_from_stage2_dict",
                    )
                )
        return records

    if not isinstance(aliases, list):
        return records

    for item in aliases:
        if not isinstance(item, dict):
            continue
        subject = None
        canonical_code = str(item.get("canonical_code", "")).strip()
        canonical_name = str(item.get("canonical_name", "")).strip()
        if canonical_code and canonical_code in subject_by_code:
            subject = subject_by_code[canonical_code]
        elif canonical_name:
            subject = subject_by_name.get(normalize_label_for_matching(canonical_name))
        if not subject:
            continue
        records.append(
            AliasRecord(
                canonical_code=subject.code,
                canonical_name=subject.canonical_name,
                alias=str(item.get("alias", "")).strip(),
                alias_type=str(item.get("alias_type", "exact_alias")).strip() or "exact_alias",
                enabled=bool(item.get("enabled", True)),
                note=str(item.get("note", "")).strip(),
            )
        )
    return records


def load_subject_relations(config_path: Path, subjects: List[TemplateSubject]) -> List[RelationRecord]:
    if not config_path.exists():
        return []
    payload = yaml.safe_load(config_path.read_text(encoding="utf-8")) or {}
    relations = payload.get("relations", [])
    subject_by_code = {subject.code: subject for subject in subjects}
    subject_by_name = {normalize_label_for_matching(subject.canonical_name): subject for subject in subjects}
    results: List[RelationRecord] = []

    for item in relations:
        if not isinstance(item, dict):
            continue
        subject = None
        canonical_code = str(item.get("canonical_code", "")).strip()
        canonical_name = str(item.get("canonical_name", "")).strip()
        if canonical_code and canonical_code in subject_by_code:
            subject = subject_by_code[canonical_code]
        elif canonical_name:
            subject = subject_by_name.get(normalize_label_for_matching(canonical_name))
        if not subject:
            continue
        relation_type = str(item.get("relation_type", "")).strip()
        related_codes = [str(value).strip() for value in item.get("related_codes", []) if str(value).strip()]
        related_names = [str(value).strip() for value in item.get("related_names", []) if str(value).strip()]
        results.append(
            RelationRecord(
                canonical_code=subject.code,
                canonical_name=subject.canonical_name,
                relation_type=relation_type,
                related_codes=related_codes,
                related_names=related_names,
                enabled=bool(item.get("enabled", True)),
                review_required=bool(item.get("review_required", True)),
                note=str(item.get("note", "")).strip(),
            )
        )
    return results
