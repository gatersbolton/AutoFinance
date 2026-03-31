from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook


SUBJECT_RE = re.compile(r"^(?P<code>[A-Z]{2}_[0-9]{3})\s*(?P<name>.*)$")


def load_workbook_main_sheet(path: Path) -> Dict[str, Any]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header_row = detect_header_row(worksheet)
    headers = [str(worksheet.cell(row=header_row, column=index).value or "").strip() for index in range(1, worksheet.max_column + 1)]
    rows: List[Dict[str, Any]] = []
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        subject_value = str(worksheet.cell(row=row_idx, column=1).value or "").strip()
        if not subject_value.startswith("ZT_"):
            continue
        match = SUBJECT_RE.match(subject_value)
        if not match:
            continue
        values = {
            headers[col_idx - 1]: worksheet.cell(row=row_idx, column=col_idx).value
            for col_idx in range(2, worksheet.max_column + 1)
            if headers[col_idx - 1]
        }
        rows.append(
            {
                "mapping_code": match.group("code"),
                "mapping_name": match.group("name").strip(),
                "row_index": row_idx,
                "values": values,
            }
        )
    return {
        "path": str(path),
        "sheet_name": worksheet.title,
        "header_row": header_row,
        "headers": headers,
        "rows": rows,
    }


def detect_header_row(worksheet) -> int:
    for row_idx in range(1, min(worksheet.max_row, 20) + 1):
        value = str(worksheet.cell(row=row_idx, column=1).value or "").strip()
        if value == "科目名称":
            return row_idx
    return 3
