from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


PROMOTION_HEADERS = [
    "promotion_id",
    "source_type",
    "candidate_alias",
    "canonical_code",
    "canonical_name",
    "statement_type",
    "evidence_count",
    "amount_coverage_gain",
    "benchmark_support",
    "rule_id",
    "formula_payload_json",
    "target_scope_rule_json",
    "mapping_code",
    "mapping_name",
    "period_key",
    "benchmark_value",
    "gap_cause",
    "action_type",
    "action_value",
    "reviewer_name",
    "reviewer_note",
    "review_status",
    "applied_status",
    "apply_message",
]

PROMOTION_GUIDE = [
    {"action_type": "promote_alias", "action_value_format": "leave blank or override alias text", "effect": "Append a curated exact alias entry."},
    {"action_type": "promote_legacy_alias", "action_value_format": "leave blank or override alias text", "effect": "Append a curated legacy alias entry."},
    {"action_type": "promote_formula_rule", "action_value_format": "leave blank to use formula_payload_json", "effect": "Append a curated formula rule entry."},
    {"action_type": "promote_target_rule", "action_value_format": "leave blank to use target_scope_rule_json", "effect": "Append a curated target scope rule entry."},
    {"action_type": "reject", "action_value_format": "optional reason text", "effect": "Record the promotion as explicitly rejected."},
    {"action_type": "defer", "action_value_format": "optional note", "effect": "Keep the promotion for a later batch without applying it."},
]


def export_promotion_actions_template(
    output_dir: Path,
    alias_candidates: Sequence[Dict[str, Any]],
    formula_candidates: Sequence[Dict[str, Any]],
    benchmark_gap_rows: Sequence[Dict[str, Any]],
    benchmark_missing_true_rows: Sequence[Dict[str, Any]],
    unmapped_value_bearing_rows: Sequence[Dict[str, Any]],
    target_gap_backlog_rows: Sequence[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for index, row in enumerate(alias_candidates, start=1):
        rows.append(
            base_row(
                promotion_id=f"PROM_ALIAS_{index:04d}",
                source_type="alias_candidate",
                candidate_alias=row.get("candidate_alias", ""),
                canonical_code=row.get("canonical_code", ""),
                canonical_name=row.get("canonical_name", ""),
                statement_type=row.get("statement_type", ""),
                evidence_count=row.get("evidence_count", 0),
                amount_coverage_gain=row.get("amount_coverage_gain", 0.0),
                benchmark_support=row.get("benchmark_support", 0),
                formula_payload_json="",
                target_scope_rule_json="",
                mapping_code=row.get("canonical_code", ""),
                mapping_name=row.get("canonical_name", ""),
            )
        )
    for index, row in enumerate(formula_candidates, start=1):
        payload = row.get("formula_payload_json", "")
        rows.append(
            base_row(
                promotion_id=f"PROM_FORMULA_{index:04d}",
                source_type="formula_candidate",
                candidate_alias="",
                canonical_code=row.get("mapping_code", ""),
                canonical_name=row.get("mapping_name", ""),
                statement_type=row.get("statement_type", ""),
                evidence_count=1,
                amount_coverage_gain=abs(float(row.get("value_num") or 0.0)),
                benchmark_support=0,
                rule_id=row.get("rule_id", ""),
                formula_payload_json=payload,
                target_scope_rule_json="",
                mapping_code=row.get("mapping_code", ""),
                mapping_name=row.get("mapping_name", ""),
                period_key=row.get("period_key", ""),
                benchmark_value=row.get("value_num", ""),
            )
        )
    for index, row in enumerate(target_gap_backlog_rows, start=1):
        rows.append(
            base_row(
                promotion_id=f"PROM_TARGET_{index:04d}",
                source_type="target_gap",
                candidate_alias="",
                canonical_code=row.get("mapping_code", ""),
                canonical_name=row.get("mapping_name", ""),
                statement_type="",
                evidence_count=0,
                amount_coverage_gain=row.get("priority_score", 0.0),
                benchmark_support=0,
                formula_payload_json="",
                target_scope_rule_json=json.dumps(
                    {"mapping_code": row.get("mapping_code", ""), "target_scope": "main_export_target"},
                    ensure_ascii=False,
                    separators=(",", ":"),
                ),
                mapping_code=row.get("mapping_code", ""),
                mapping_name=row.get("mapping_name", ""),
                period_key=row.get("aligned_period_key", ""),
                benchmark_value=row.get("benchmark_value", ""),
                gap_cause=row.get("gap_cause", ""),
            )
        )
    for index, row in enumerate(benchmark_gap_rows, start=1):
        rows.append(
            base_row(
                promotion_id=f"PROM_GAP_{index:04d}",
                source_type="benchmark_gap",
                candidate_alias="",
                canonical_code=row.get("mapping_code", ""),
                canonical_name=row.get("mapping_name", ""),
                statement_type="",
                evidence_count=0,
                amount_coverage_gain=abs(float(row.get("benchmark_value") or 0.0)) if _is_numeric(row.get("benchmark_value")) else 0.0,
                benchmark_support=1,
                formula_payload_json="",
                target_scope_rule_json="",
                mapping_code=row.get("mapping_code", ""),
                mapping_name=row.get("mapping_name", ""),
                period_key=row.get("aligned_period_key", ""),
                benchmark_value=row.get("benchmark_value", ""),
                gap_cause=row.get("gap_cause", ""),
            )
        )
    for index, row in enumerate(unmapped_value_bearing_rows, start=1):
        rows.append(
            base_row(
                promotion_id=f"PROM_UNMAPPED_{index:04d}",
                source_type="unmapped_value_bearing",
                candidate_alias=row.get("row_label_canonical_candidate") or row.get("row_label_norm") or row.get("row_label_std", ""),
                canonical_code="",
                canonical_name="",
                statement_type=row.get("statement_type", ""),
                evidence_count=1,
                amount_coverage_gain=abs(float(row.get("value_num") or 0.0)) if _is_numeric(row.get("value_num")) else 0.0,
                benchmark_support=0,
                formula_payload_json="",
                target_scope_rule_json="",
                mapping_code="",
                mapping_name="",
                period_key=row.get("period_key", ""),
                benchmark_value=row.get("value_num", ""),
            )
        )
    csv_path = output_dir / "promotion_actions_template.csv"
    xlsx_path = output_dir / "promotion_actions_template.xlsx"
    write_csv(csv_path, rows)
    write_workbook(xlsx_path, rows)
    return rows


def base_row(**values: Any) -> Dict[str, Any]:
    row = {header: "" for header in PROMOTION_HEADERS}
    row.update(values)
    return row


def write_csv(path: Path, rows: Sequence[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=PROMOTION_HEADERS)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: serialize(row.get(header)) for header in PROMOTION_HEADERS})


def write_workbook(path: Path, rows: Sequence[Dict[str, Any]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Promotions"
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(PROMOTION_HEADERS))}1"
    worksheet.append(PROMOTION_HEADERS)
    for row in rows:
        worksheet.append([serialize(row.get(header)) for header in PROMOTION_HEADERS])
    guide = workbook.create_sheet("ActionGuide")
    guide.append(["action_type", "action_value_format", "effect"])
    for row in PROMOTION_GUIDE:
        guide.append([row["action_type"], row["action_value_format"], row["effect"]])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def serialize(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"), sort_keys=True)
    return "" if value is None else value


def _is_numeric(value: Any) -> bool:
    try:
        float(value)
        return True
    except (TypeError, ValueError):
        return False
