from __future__ import annotations

import csv
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook


def parse_promotion_actions_file(path: Path) -> List[Dict[str, str]]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))
    workbook = load_workbook(path, data_only=True)
    sheet_name = "Promotions" if "Promotions" in workbook.sheetnames else workbook.sheetnames[0]
    worksheet = workbook[sheet_name]
    headers = [str(cell.value or "").strip() for cell in worksheet[1]]
    rows: List[Dict[str, str]] = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        row = {headers[index]: ("" if value is None else str(value)) for index, value in enumerate(values) if index < len(headers)}
        if any(str(value).strip() for value in row.values()):
            rows.append(row)
    return rows
