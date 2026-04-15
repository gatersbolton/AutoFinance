from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from ..models import DiscoveredSource, ProviderCell, ProviderPage


def load_xlsx_fallback_page(source: DiscoveredSource) -> ProviderPage:
    if not source.artifact_file:
        raise ValueError(f"Fallback source for page {source.page_no} is missing artifact_file")

    workbook_path = Path(source.artifact_file)
    table_cells, context_lines, worksheet_title = load_table_cells_from_xlsx(
        workbook_path,
        table_id_override="",
        extra_meta={"source_kind": "xlsx_fallback"},
    )

    page_text = str(source.result_page_meta.get("text", "") or "\n".join(context_lines))
    return ProviderPage(
        doc_id=source.doc_id,
        page_no=source.page_no,
        provider=source.provider,
        source_file=source.artifact_file,
        source_kind="xlsx_fallback",
        page_text=page_text,
        tables={worksheet_title: table_cells},
        context_lines=context_lines,
        meta={
            "notes": list(source.notes) + ["missing_bbox", "missing_confidence"],
            "worksheet_title": worksheet_title,
        },
    )


def load_table_cells_from_xlsx(
    workbook_path: Path,
    *,
    table_id_override: str = "",
    extra_meta: Optional[Dict[str, Any]] = None,
) -> Tuple[List[ProviderCell], List[str], str]:
    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook.active

    start_row, end_row, start_col, end_col, context_lines = detect_table_window(worksheet)
    merged_lookup = build_merged_lookup(worksheet, start_row, end_row, start_col, end_col)

    table_id = table_id_override or worksheet.title
    table_cells: List[ProviderCell] = []
    seen_anchors = set()
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(start_col, end_col + 1):
            merge_range = merged_lookup.get((row_idx, col_idx))
            if merge_range:
                anchor = (merge_range[0], merge_range[1])
                if (row_idx, col_idx) != anchor or anchor in seen_anchors:
                    continue
                seen_anchors.add(anchor)
                cell = worksheet.cell(row=anchor[0], column=anchor[1])
                table_cells.append(
                    ProviderCell(
                        table_id=table_id,
                        row_start=anchor[0] - start_row,
                        row_end=merge_range[2] - start_row,
                        col_start=anchor[1] - start_col,
                        col_end=merge_range[3] - start_col,
                        text=str(cell.value or ""),
                        bbox=None,
                        confidence=None,
                        cell_type="body",
                        meta=dict(extra_meta or {}),
                    )
                )
                continue

            cell = worksheet.cell(row=row_idx, column=col_idx)
            table_cells.append(
                ProviderCell(
                    table_id=table_id,
                    row_start=row_idx - start_row,
                    row_end=row_idx - start_row,
                    col_start=col_idx - start_col,
                    col_end=col_idx - start_col,
                    text=str(cell.value or ""),
                    bbox=None,
                    confidence=None,
                    cell_type="body",
                    meta=dict(extra_meta or {}),
                )
            )
    return table_cells, context_lines, worksheet.title


def detect_table_window(worksheet) -> Tuple[int, int, int, int, List[str]]:
    row_details: List[Tuple[int, int, List[str]]] = []
    for row_idx in range(1, worksheet.max_row + 1):
        values = []
        for col_idx in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row=row_idx, column=col_idx).value
            if value not in (None, ""):
                values.append(str(value).strip())
        row_details.append((row_idx, len(values), values))

    start_row = 1
    for row_idx, count, values in row_details:
        joined = " ".join(values)
        if count >= 3 or "行次" in joined or "期初数" in joined or "期末数" in joined:
            start_row = row_idx
            break

    end_row = start_row
    for row_idx, count, values in row_details:
        if row_idx < start_row:
            continue
        joined = " ".join(values)
        if count >= 2 or "行次" in joined or "期初数" in joined or "期末数" in joined:
            end_row = row_idx

    used_cols = []
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(1, worksheet.max_column + 1):
            if worksheet.cell(row=row_idx, column=col_idx).value not in (None, ""):
                used_cols.append(col_idx)

    start_col = min(used_cols) if used_cols else 1
    end_col = max(used_cols) if used_cols else worksheet.max_column

    context_lines: List[str] = []
    for row_idx, count, values in row_details:
        if row_idx < start_row or row_idx > end_row:
            if count:
                context_lines.extend(values)

    return start_row, end_row, start_col, end_col, context_lines


def build_merged_lookup(
    worksheet,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
) -> Dict[Tuple[int, int], Tuple[int, int, int, int]]:
    lookup: Dict[Tuple[int, int], Tuple[int, int, int, int]] = {}
    for merge in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merge.bounds
        if max_row < start_row or min_row > end_row or max_col < start_col or min_col > end_col:
            continue
        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                lookup[(row_idx, col_idx)] = (min_row, min_col, max_row, max_col)
    return lookup
