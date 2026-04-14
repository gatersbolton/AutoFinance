from __future__ import annotations

import json
import csv
from pathlib import Path
from typing import Any, Dict, List, Sequence

from openpyxl import load_workbook

from ..metadata import scan_summary_run_ids


def run_full_run_contract(
    output_dir: Path,
    workbook_path: Path,
    run_id: str,
    feature_flags: Dict[str, Any],
    export_stats: Dict[str, Any],
    required_helper_sheets: Sequence[str],
) -> Dict[str, Any]:
    workbook = load_workbook(workbook_path)
    actual_sheets = list(workbook.sheetnames)
    checks: List[Dict[str, Any]] = []
    produced = {path.name for path in output_dir.iterdir() if path.is_file()}
    required_summary_files = build_required_summary_files(feature_flags)

    add_check(
        checks,
        "helper_sheets_present",
        all(sheet in actual_sheets for sheet in required_helper_sheets),
        {"required_helper_sheets": list(required_helper_sheets), "actual_sheets": actual_sheets},
    )
    if feature_flags.get("emit_benchmark_report"):
        required = {
            "benchmark_summary.json",
            "benchmark_summary.csv",
            "benchmark_missing_in_auto.csv",
            "benchmark_missing_true.csv",
            "benchmark_alignment_audit.csv",
            "benchmark_alignment_summary.json",
            "benchmark_value_diff.csv",
            "benchmark_gap_explanations.csv",
            "benchmark_gap_summary.json",
        }
        add_check(checks, "benchmark_outputs_present", required.issubset(produced), {"required": sorted(required), "produced": sorted(produced)})
    if feature_flags.get("enable_benchmark_alignment_repair"):
        required = {
            "benchmark_alignment_audit.csv",
            "benchmark_alignment_summary.json",
            "benchmark_missing_true.csv",
            "benchmark_alignment_only.csv",
        }
        add_check(checks, "benchmark_alignment_outputs_present", required.issubset(produced), {"required": sorted(required), "produced": sorted(produced)})
    if feature_flags.get("enable_derived_facts"):
        required = {
            "derived_facts.csv",
            "derived_formula_audit.csv",
            "derived_formula_summary.json",
            "derived_conflicts.csv",
        }
        add_check(checks, "derived_outputs_present", required.issubset(produced), {"required": sorted(required), "produced": sorted(produced)})
    if feature_flags.get("enable_export_target_scoping"):
        required = {
            "export_target_scope.csv",
            "export_target_kpi_summary.json",
            "main_target_review_queue.csv",
            "note_detail_review_queue.csv",
            "target_gap_backlog.csv",
            "target_gap_summary.json",
        }
        add_check(checks, "target_scope_outputs_present", required.issubset(produced), {"required": sorted(required), "produced": sorted(produced)})
    if feature_flags.get("emit_promotion_template"):
        required = {"promotion_actions_template.xlsx", "promotion_actions_template.csv"}
        add_check(checks, "promotion_template_outputs_present", required.issubset(produced), {"required": sorted(required), "produced": sorted(produced)})
    if feature_flags.get("apply_promotions"):
        required = {
            "applied_promotions.csv",
            "rejected_promotions.csv",
            "promotion_audit.csv",
            "promotion_delta.json",
            "promotion_delta.csv",
            "promoted_aliases.csv",
            "promoted_formula_rules.csv",
        }
        add_check(checks, "promotion_outputs_present", required.issubset(produced), {"required": sorted(required), "produced": sorted(produced)})
    if feature_flags.get("emit_run_manifest"):
        required = {"run_manifest.json", "artifact_manifest_core.csv", "artifact_manifest.csv"}
        if str(feature_flags.get("artifact_manifest_mode", "core")).strip().lower() == "full":
            required.add("artifact_manifest_full.csv")
        add_check(checks, "manifest_outputs_present", required.issubset(produced), {"required": sorted(required), "produced": sorted(produced)})
        manifest = read_json(output_dir / "run_manifest.json")
        add_check(
            checks,
            "manifest_run_id_matches",
            str(manifest.get("run_id", "")) == run_id,
            {"expected_run_id": run_id, "manifest_run_id": manifest.get("run_id", "")},
        )
        artifact_rows = read_csv_rows(output_dir / "artifact_manifest_core.csv")
        artifact_run_ids = sorted({row.get("run_id", "") for row in artifact_rows if row.get("run_id")})
        add_check(
            checks,
            "artifact_manifest_run_id_matches",
            artifact_run_ids == [run_id] if artifact_run_ids else False,
            {"expected_run_id": run_id, "artifact_run_ids": artifact_run_ids},
        )
    add_check(
        checks,
        "export_source_facts_is_deduped",
        str(export_stats.get("source_facts", "")) == "facts_deduped",
        {"source_facts": export_stats.get("source_facts", "")},
    )
    metadata_contract = scan_summary_run_ids(
        output_dir=output_dir,
        expected_run_id=run_id,
        required_summary_files=required_summary_files,
    )
    add_check(
        checks,
        "summary_run_ids_present_and_match",
        bool(metadata_contract.get("pass", False)),
        {
            "required_summary_files": required_summary_files,
            "checked_summary_files": metadata_contract.get("checked_summary_files", []),
            "missing_run_id_files": metadata_contract.get("missing_run_id_files", []),
            "mismatched_run_id_files": metadata_contract.get("mismatched_run_id_files", []),
        },
    )
    return {
        "run_id": run_id,
        "feature_flags": feature_flags,
        "required_artifacts": build_required_artifacts(feature_flags),
        "required_summary_files": required_summary_files,
        "produced_artifacts": sorted(produced),
        "workbook_helper_sheets": actual_sheets,
        "checks_total": len(checks),
        "contract_fail_total": sum(1 for item in checks if item["status"] == "fail"),
        "metadata_contract": metadata_contract,
        "checks": checks,
    }


def build_required_artifacts(feature_flags: Dict[str, Any]) -> List[str]:
    required = ["run_summary.json", "summary.json", "会计报表_填充结果.xlsx"]
    required.extend(
        [
            "pipeline_stage_timings.json",
            "pipeline_stage_status.json",
            "pipeline_completion_summary.json",
            "pages_skipped_metric_audit.json",
            "metadata_contract_summary.json",
            "run_id_propagation_audit.json",
            "hardening_summary.json",
        ]
    )
    if feature_flags.get("emit_benchmark_report"):
        required.extend(["benchmark_summary.json", "benchmark_gap_summary.json"])
    if feature_flags.get("enable_benchmark_alignment_repair"):
        required.extend(["benchmark_alignment_summary.json", "benchmark_missing_true.csv"])
    if feature_flags.get("enable_derived_facts"):
        required.extend(["derived_formula_summary.json", "derived_facts.csv"])
    if feature_flags.get("enable_export_target_scoping"):
        required.extend(["export_target_kpi_summary.json", "target_gap_backlog.csv"])
        required.extend(["source_backed_gap_closure.csv", "source_backed_gap_closure_summary.json"])
    if feature_flags.get("emit_reocr_tasks"):
        required.extend(["reocr_task_pruned_deduped.csv", "reocr_dedupe_audit.json"])
    if feature_flags.get("emit_promotion_template"):
        required.extend(["promotion_actions_template.xlsx", "promotion_actions_template.csv"])
    if feature_flags.get("apply_promotions"):
        required.extend(["applied_promotions.csv", "promotion_delta.json"])
    if feature_flags.get("emit_run_manifest"):
        required.extend(["run_manifest.json", "artifact_manifest_core.csv", "artifact_manifest.csv"])
        if str(feature_flags.get("artifact_manifest_mode", "core")).strip().lower() == "full":
            required.append("artifact_manifest_full.csv")
    return required


def build_required_summary_files(feature_flags: Dict[str, Any]) -> List[str]:
    required = [
        "alias_acceptance_summary.json",
        "coverage_opportunity_summary.json",
        "curated_alias_pack_summary.json",
        "hardening_summary.json",
        "label_normalization_summary.json",
        "mapping_lift_summary.json",
        "metadata_contract_summary.json",
        "pipeline_completion_summary.json",
        "period_role_resolution_summary.json",
        "review_actionable_summary.json",
        "review_summary.json",
        "run_summary.json",
        "statement_classification_summary.json",
        "summary.json",
        "validation_summary.json",
    ]
    if feature_flags.get("emit_reocr_tasks"):
        required.extend(
            [
                "reocr_task_pruned_summary.json",
                "reocr_task_summary.json",
            ]
        )
    if feature_flags.get("emit_benchmark_report"):
        required.extend(
            [
                "benchmark_gap_summary.json",
                "benchmark_summary.json",
            ]
        )
    if feature_flags.get("enable_benchmark_alignment_repair"):
        required.extend(
            [
                "benchmark_alignment_summary.json",
                "benchmark_missing_true_summary.json",
            ]
        )
    if feature_flags.get("enable_derived_facts"):
        required.append("derived_formula_summary.json")
    if feature_flags.get("enable_export_target_scoping"):
        required.extend(
            [
                "export_target_kpi_summary.json",
                "source_backed_gap_closure_summary.json",
                "target_gap_summary.json",
            ]
        )
        if feature_flags.get("target_gap_enabled", True):
            required.extend(
                [
                    "no_source_gap_summary.json",
                    "target_backfill_summary.json",
                ]
            )
    if feature_flags.get("emit_delta_report"):
        required.append("export_delta_summary.json")
    if feature_flags.get("emit_stage6_kpis"):
        required.append("stage6_kpi_summary.json")
    if feature_flags.get("emit_stage7_kpis"):
        required.append("stage7_kpi_summary.json")
    if feature_flags.get("apply_promotions"):
        pass
    return sorted(set(required))


def add_check(checks: List[Dict[str, Any]], name: str, ok: bool, meta: Dict[str, Any]) -> None:
    checks.append({"check_name": name, "status": "pass" if ok else "fail", "meta": meta})


def read_json(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def read_csv_rows(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))
