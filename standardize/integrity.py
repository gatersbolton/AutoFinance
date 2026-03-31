from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, Dict, List, Sequence

from openpyxl import load_workbook

from .models import ArtifactIntegrityRecord, compact_json


def run_artifact_integrity(
    output_dir: Path,
    workbook_path: Path,
    run_summary: Dict[str, Any],
    export_stats: Dict[str, Any],
    export_rules: Dict[str, Any],
) -> Dict[str, Any]:
    records: List[ArtifactIntegrityRecord] = []
    required_sheets = export_rules.get(
        "required_helper_sheets",
        ["_meta_summary", "_issues", "_validation", "_duplicates", "_conflicts", "_unplaced_facts", "_review_queue"],
    )
    workbook = load_workbook(workbook_path)
    main_sheet = workbook[workbook.sheetnames[0]]
    header_row = int(export_stats.get("header_row", 3))
    main_headers = [str(main_sheet.cell(row=header_row, column=index).value or "").strip() for index in range(1, main_sheet.max_column + 1)]
    period_columns = [header for header in main_headers if "__" in header]
    unplaced_rows = read_sheet_rows(workbook["_unplaced_facts"]) if "_unplaced_facts" in workbook.sheetnames else []
    unplaced_by_fact = {row.get("fact_id", ""): row for row in unplaced_rows if row.get("fact_id")}

    add_record(records, "required_helper_sheets", "error", all(sheet in workbook.sheetnames for sheet in required_sheets), "Workbook helper sheet contract check.", {"required_sheets": required_sheets, "actual_sheets": workbook.sheetnames})
    add_record(records, "duplicate_main_period_columns", "error", len(period_columns) == len(set(period_columns)), "Main sheet must not contain duplicate period columns.", {"period_columns": period_columns})
    meta_rows = read_sheet_rows(workbook["_meta_summary"]) if "_meta_summary" in workbook.sheetnames else []
    meta_map = {row.get("key", ""): row.get("value", "") for row in meta_rows}
    if run_summary.get("run_id"):
        add_record(
            records,
            "meta_summary_run_id_matches",
            "error",
            str(meta_map.get("run_id", "")) == str(run_summary.get("run_id", "")),
            "Workbook _meta_summary run_id must match run_summary run_id.",
            {"workbook_run_id": meta_map.get("run_id", ""), "summary_run_id": run_summary.get("run_id", "")},
        )
        artifact_run_id_mismatches = []
        for artifact_name in ("benchmark_summary.json", "derived_formula_summary.json", "run_manifest.json"):
            artifact_path = output_dir / artifact_name
            if not artifact_path.exists():
                continue
            payload = parse_json(artifact_path.read_text(encoding="utf-8"))
            if payload.get("run_id", "") != run_summary.get("run_id", ""):
                artifact_run_id_mismatches.append({"artifact": artifact_name, "artifact_run_id": payload.get("run_id", "")})
        add_record(
            records,
            "artifact_run_ids_match_summary",
            "error",
            not artifact_run_id_mismatches,
            "All major artifacts with run_id must match the current run_summary run_id.",
            {"mismatches": artifact_run_id_mismatches},
        )

    if int(run_summary.get("unknown_date_total", 0)) == 0:
        add_record(
            records,
            "no_unknown_date_columns_when_summary_zero",
            "error",
            not any(column.startswith("unknown_date__") for column in period_columns),
            "run_summary says unknown_date_total=0, so main sheet must not expose unknown_date columns.",
            {"period_columns": period_columns},
        )

    add_record(
        records,
        "no_unknown_role_columns_on_main_sheet",
        "error",
        not any(column.endswith("__unknown") for column in period_columns),
        "Main sheet must not expose unresolved period-role columns.",
        {"period_columns": period_columns},
    )

    facts_rows = read_csv_rows(output_dir / "facts_deduped.csv")
    exportable_rows = [row for row in facts_rows if is_exportable_fact_row(row)]
    expected_periods = sorted({row.get("period_key", "") for row in exportable_rows if row.get("period_key")})
    add_record(
        records,
        "main_sheet_periods_match_exportable_facts",
        "error",
        set(expected_periods) == set(period_columns),
        "Main sheet period columns must match exportable resolved deduped facts.",
        {"expected_periods": expected_periods, "actual_periods": period_columns},
    )

    source_facts = str(export_stats.get("source_facts", ""))
    add_record(
        records,
        "export_uses_deduped_facts",
        "error",
        source_facts == "facts_deduped",
        "Export must explicitly use deduped facts as its source of truth.",
        {"source_facts": source_facts},
    )

    accepted_conflicts = read_csv_rows(output_dir / "conflicts_enriched.csv")
    main_values = extract_main_sheet_values(main_sheet, header_row)
    accepted_missing = []
    for row in accepted_conflicts:
        if row.get("decision") not in {"accepted", "accepted_with_rule_support", "accepted_with_validation_support"}:
            continue
        accepted_fact_id = row.get("accepted_fact_id", "")
        provider_values = parse_json(row.get("provider_values_json", ""))
        accepted_fact = find_fact_payload(provider_values, accepted_fact_id)
        if not accepted_fact:
            continue
        mapping_code = lookup_fact_field(facts_rows, accepted_fact_id, "mapping_code")
        period_key = lookup_fact_field(facts_rows, accepted_fact_id, "period_key")
        if mapping_code and period_key and (mapping_code, period_key) in main_values:
            continue
        if accepted_fact_id in unplaced_by_fact:
            continue
        accepted_missing.append({"conflict_id": row.get("conflict_id"), "accepted_fact_id": accepted_fact_id})
    add_record(
        records,
        "accepted_conflicts_are_explained",
        "error",
        not accepted_missing,
        "Accepted conflict outcomes must either land on the main sheet or be present in _unplaced_facts with a reason.",
        {"missing": accepted_missing},
    )

    duplicate_keys = duplicate_exportable_keys(exportable_rows)
    add_record(
        records,
        "no_duplicate_exportable_fact_keys",
        "error",
        not duplicate_keys,
        "Exportable deduped facts must not contain duplicate mapping_code + period_key keys.",
        {"duplicate_keys": duplicate_keys},
    )

    status_counts = {
        "pass": sum(1 for record in records if record.status == "pass"),
        "fail": sum(1 for record in records if record.status == "fail"),
        "review": sum(1 for record in records if record.status == "review"),
    }
    summary = {
        "checks_total": len(records),
        "integrity_pass_total": status_counts["pass"],
        "integrity_fail_total": status_counts["fail"],
        "integrity_review_total": status_counts["review"],
        "records": [record.__dict__ for record in records],
    }
    return {"records": records, "summary": summary}


def add_record(records: List[ArtifactIntegrityRecord], check_name: str, severity: str, ok: bool, message: str, meta: Dict[str, Any]) -> None:
    records.append(
        ArtifactIntegrityRecord(
            check_id=f"INT{len(records)+1:03d}",
            check_name=check_name,
            severity=severity,
            status="pass" if ok else "fail",
            message=message,
            meta_json=compact_json(meta),
        )
    )


def read_csv_rows(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def read_sheet_rows(worksheet) -> List[Dict[str, str]]:
    headers = [str(cell.value or "").strip() for cell in worksheet[1]]
    rows: List[Dict[str, str]] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        rows.append({headers[index]: ("" if value is None else str(value)) for index, value in enumerate(row) if index < len(headers)})
    return rows


def is_exportable_fact_row(row: Dict[str, str]) -> bool:
    status = row.get("status", "")
    conflict_decision = row.get("conflict_decision", "")
    report_date_norm = row.get("report_date_norm", "")
    period_role_norm = row.get("period_role_norm", "")
    mapping_code = row.get("mapping_code", "")
    value_num = row.get("value_num", "")
    unplaced_reason = row.get("unplaced_reason", "")
    mapping_review_required = str(row.get("mapping_review_required", "")).lower() == "true"
    if not mapping_code or not value_num:
        return False
    if unplaced_reason:
        return False
    if mapping_review_required:
        return False
    if report_date_norm in {"", "unknown_date"}:
        return False
    if period_role_norm in {"", "unknown"}:
        return False
    if row.get("period_key", "").startswith("unknown_date__") or row.get("period_key", "").endswith("__unknown"):
        return False
    if status not in {"observed", "repaired"}:
        return False
    if conflict_decision in {"review_required", "unresolved"}:
        return False
    return True


def extract_main_sheet_values(worksheet, header_row: int) -> set[tuple[str, str]]:
    headers = [str(worksheet.cell(row=header_row, column=index).value or "").strip() for index in range(1, worksheet.max_column + 1)]
    values = set()
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        subject_value = str(worksheet.cell(row=row_idx, column=1).value or "").strip()
        if not subject_value.startswith("ZT_"):
            continue
        mapping_code = subject_value.split(" ", 1)[0]
        for col_idx, header in enumerate(headers, start=1):
            if "__" not in header:
                continue
            value = worksheet.cell(row=row_idx, column=col_idx).value
            if value is not None:
                values.add((mapping_code, header))
    return values


def duplicate_exportable_keys(rows: Sequence[Dict[str, str]]) -> List[str]:
    counter: Dict[str, int] = {}
    duplicates: List[str] = []
    for row in rows:
        key = f"{row.get('mapping_code','')}::{row.get('period_key','')}"
        counter[key] = counter.get(key, 0) + 1
        if counter[key] == 2:
            duplicates.append(key)
    return duplicates


def parse_json(value: str) -> Dict[str, Any]:
    if not value:
        return {}
    try:
        return json.loads(value)
    except json.JSONDecodeError:
        return {}


def find_fact_payload(provider_values: Dict[str, Any], fact_id: str) -> Dict[str, Any]:
    for items in provider_values.values():
        if not isinstance(items, list):
            continue
        for item in items:
            if isinstance(item, dict) and item.get("fact_id") == fact_id:
                return item
    return {}


def lookup_fact_field(rows: Sequence[Dict[str, str]], fact_id: str, field_name: str) -> str:
    for row in rows:
        if row.get("fact_id") == fact_id:
            return row.get(field_name, "")
    return ""
