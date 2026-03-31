from __future__ import annotations

import csv
import json
from collections import defaultdict
from pathlib import Path
from typing import Dict, Iterable, List, Sequence

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .actions import ACTION_GUIDE


TEMPLATE_HEADERS = [
    "review_id",
    "priority_score",
    "source_type",
    "fact_id",
    "source_cell_ref",
    "related_conflict_ids",
    "doc_id",
    "page_no",
    "statement_type",
    "row_label_raw",
    "row_label_std",
    "period_key",
    "value_raw",
    "value_num",
    "provider",
    "candidate_mapping_code",
    "candidate_mapping_name",
    "candidate_conflict_fact_id",
    "candidate_period_override",
    "suggested_reocr_task_id",
    "action_type",
    "action_value",
    "reviewer_note",
    "reviewer_name",
    "review_status",
    "applied_status",
    "apply_message",
]


def export_review_actions_template(
    output_dir: Path,
    review_items: Sequence,
    mapping_candidates: Sequence,
    conflicts: Sequence,
    validations: Sequence,
    unmapped_summary: Sequence,
    reocr_tasks: Sequence,
) -> List[Dict[str, object]]:
    mapping_lookup = build_mapping_lookup(mapping_candidates)
    conflict_lookup = build_conflict_lookup(conflicts)
    reocr_lookup = {task.source_review_id: task.task_id for task in reocr_tasks}
    validation_lookup = build_validation_lookup(validations)
    rows: List[Dict[str, object]] = []

    for item in review_items:
        best_mapping = mapping_lookup.get(item.row_label_std or item.row_label_raw, {})
        conflict_info = first_conflict_info(item.related_conflict_ids, conflict_lookup)
        source_type = infer_source_type(item.reason_codes)
        rows.append(
            {
                "review_id": item.review_id,
                "priority_score": item.priority_score,
                "source_type": source_type,
                "fact_id": item.related_fact_ids[0] if item.related_fact_ids else "",
                "source_cell_ref": parse_json_payload(item.meta_json).get("source_cell_ref", ""),
                "related_conflict_ids": ",".join(item.related_conflict_ids),
                "doc_id": item.doc_id,
                "page_no": item.page_no,
                "statement_type": item.statement_type,
                "row_label_raw": item.row_label_raw,
                "row_label_std": item.row_label_std,
                "period_key": item.period_key,
                "value_raw": item.value_raw,
                "value_num": item.value_num,
                "provider": item.provider,
                "candidate_mapping_code": best_mapping.get("candidate_code", ""),
                "candidate_mapping_name": best_mapping.get("candidate_name", ""),
                "candidate_conflict_fact_id": conflict_info.get("candidate_fact_id", ""),
                "candidate_period_override": infer_candidate_period_override(item, validation_lookup),
                "suggested_reocr_task_id": reocr_lookup.get(item.review_id, ""),
                "action_type": "",
                "action_value": "",
                "reviewer_note": "",
                "reviewer_name": "",
                "review_status": "",
                "applied_status": "",
                "apply_message": "",
            }
        )

    csv_path = output_dir / "review_actions_template.csv"
    xlsx_path = output_dir / "review_actions_template.xlsx"
    write_template_csv(csv_path, rows)
    write_template_workbook(xlsx_path, rows)
    return rows


def infer_source_type(reason_codes: Iterable[str]) -> str:
    reasons = list(reason_codes)
    if any(str(reason).startswith("conflict:") for reason in reasons):
        return "conflict"
    if any(str(reason).startswith("mapping:") for reason in reasons):
        return "unmapped"
    if any(str(reason).startswith("validation:") for reason in reasons):
        return "validation"
    if any("suspicious" in str(reason) for reason in reasons):
        return "suspicious"
    if any(str(reason).startswith("unplaced:") for reason in reasons):
        return "unplaced"
    if any(str(reason).startswith("source:") for reason in reasons):
        return "fallback"
    return "review"


def build_mapping_lookup(mapping_candidates: Sequence) -> Dict[str, Dict[str, str]]:
    lookup: Dict[str, Dict[str, str]] = {}
    for candidate in sorted(mapping_candidates, key=lambda item: (item.row_label_std or item.row_label_raw, item.candidate_rank, -item.candidate_score)):
        key = candidate.row_label_std or candidate.row_label_raw
        if key in lookup:
            continue
        lookup[key] = {"candidate_code": candidate.candidate_code, "candidate_name": candidate.candidate_name}
    return lookup


def build_conflict_lookup(conflicts: Sequence) -> Dict[str, Dict[str, str]]:
    result: Dict[str, Dict[str, str]] = {}
    for conflict in conflicts:
        payload = parse_json_payload(getattr(conflict, "provider_values_json", ""))
        candidate_fact_id = ""
        for items in payload.values():
            if items:
                candidate_fact_id = str(items[0].get("fact_id", "")).strip()
                if candidate_fact_id:
                    break
        result[conflict.conflict_id] = {"candidate_fact_id": candidate_fact_id}
    return result


def first_conflict_info(conflict_ids: Sequence[str], lookup: Dict[str, Dict[str, str]]) -> Dict[str, str]:
    for conflict_id in conflict_ids:
        if conflict_id in lookup:
            return lookup[conflict_id]
    return {}


def build_validation_lookup(validations: Sequence) -> Dict[str, List[str]]:
    lookup: Dict[str, List[str]] = defaultdict(list)
    for validation in validations:
        for ref in validation.evidence_fact_refs:
            lookup[ref].append(validation.rule_name)
    return lookup


def infer_candidate_period_override(item, validation_lookup: Dict[str, List[str]]) -> str:
    if item.period_key and not item.period_key.startswith("unknown_date__") and not item.period_key.endswith("__unknown"):
        return item.period_key
    meta = parse_json_payload(item.meta_json)
    source_ref = meta.get("source_cell_ref", "")
    if source_ref and source_ref in validation_lookup:
        return item.period_key
    return ""


def write_template_csv(path: Path, rows: List[Dict[str, object]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=TEMPLATE_HEADERS)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: serialize_value(row.get(header)) for header in TEMPLATE_HEADERS})


def write_template_workbook(path: Path, rows: List[Dict[str, object]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Actions"
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(TEMPLATE_HEADERS))}1"
    worksheet.append(TEMPLATE_HEADERS)
    for row in rows:
        worksheet.append([serialize_value(row.get(header)) for header in TEMPLATE_HEADERS])

    guide = workbook.create_sheet("ActionGuide")
    guide.append(["action_type", "action_value_format", "effect"])
    for row in ACTION_GUIDE:
        guide.append([row["action_type"], row["action_value_format"], row["effect"]])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def parse_json_payload(payload: str) -> Dict[str, object]:
    if not payload:
        return {}
    try:
        return json.loads(payload)
    except json.JSONDecodeError:
        return {}


def serialize_value(value):
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"), sort_keys=True)
    return value if value is not None else ""
