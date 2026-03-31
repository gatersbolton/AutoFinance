from __future__ import annotations

import csv
import json
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence, Tuple

from openpyxl import load_workbook


def load_artifact_snapshot(output_dir: Path) -> Dict[str, Any]:
    snapshot = {
        "run_summary": load_json(output_dir / "run_summary.json"),
        "review_rows": load_csv(output_dir / "review_queue.csv"),
        "unmapped_rows": load_csv(output_dir / "unmapped_labels_summary.csv"),
        "reocr_rows": load_csv(output_dir / "reocr_tasks.csv"),
        "conflict_rows": load_csv(output_dir / "conflicts_enriched.csv"),
        "facts_rows": load_csv(output_dir / "facts_deduped.csv"),
        "unplaced_rows": load_csv(output_dir / "unplaced_facts.csv"),
    }
    if not snapshot["unplaced_rows"]:
        snapshot["unplaced_rows"] = load_workbook_sheet_rows(output_dir / "会计报表_填充结果.xlsx", "_unplaced_facts")
    return snapshot


def build_delta_reports(before: Dict[str, Any], after: Dict[str, Any]) -> Dict[str, Any]:
    before_summary = before.get("run_summary", {})
    after_summary = after.get("run_summary", {})
    delta_keys = [
        "mapped_facts_ratio",
        "amount_coverage_ratio",
        "review_total",
        "validation_fail_total",
        "provider_conflict_pairs",
    ]
    unresolved_before = unresolved_conflicts_total(before.get("conflict_rows", []))
    unresolved_after = unresolved_conflicts_total(after.get("conflict_rows", []))
    exportable_before = exportable_facts_total(before.get("facts_rows", []))
    exportable_after = exportable_facts_total(after.get("facts_rows", []))
    reocr_before = len(before.get("reocr_rows", []))
    reocr_after = len(after.get("reocr_rows", []))

    coverage_rows: List[Dict[str, Any]] = []
    for key in delta_keys:
        coverage_rows.append(build_metric_delta_row(key, before_summary.get(key, 0), after_summary.get(key, 0)))
    coverage_rows.extend(
        [
            build_metric_delta_row("unresolved_conflicts_total", unresolved_before, unresolved_after),
            build_metric_delta_row("exportable_facts_total", exportable_before, exportable_after),
            build_metric_delta_row("unplaced_facts_total", len(before.get("unplaced_rows", [])), len(after.get("unplaced_rows", []))),
            build_metric_delta_row("reocr_tasks_total", reocr_before, reocr_after),
        ]
    )

    before_review = {row.get("review_id", ""): row for row in before.get("review_rows", []) if row.get("review_id")}
    after_review = {row.get("review_id", ""): row for row in after.get("review_rows", []) if row.get("review_id")}
    review_delta_rows: List[Dict[str, Any]] = []
    for review_id in sorted(set(before_review) | set(after_review)):
        status = "unchanged"
        if review_id in before_review and review_id not in after_review:
            status = "resolved"
        elif review_id not in before_review and review_id in after_review:
            status = "new"
        review_delta_rows.append(
            {
                "review_id": review_id,
                "status": status,
                "before_priority_score": before_review.get(review_id, {}).get("priority_score", ""),
                "after_priority_score": after_review.get(review_id, {}).get("priority_score", ""),
                "row_label_std": after_review.get(review_id, before_review.get(review_id, {})).get("row_label_std", ""),
                "period_key": after_review.get(review_id, before_review.get(review_id, {})).get("period_key", ""),
            }
        )

    top_resolved_items = sorted(
        [row for row in review_delta_rows if row["status"] == "resolved"],
        key=lambda row: float(row["before_priority_score"] or 0.0),
        reverse=True,
    )[:50]
    top_remaining_unmapped = sorted(
        after.get("unmapped_rows", []),
        key=lambda row: (float(row.get("amount_abs_total", 0) or 0.0), int(row.get("occurrences", 0) or 0)),
        reverse=True,
    )[:50]
    top_remaining_review_items = sorted(
        after.get("review_rows", []),
        key=lambda row: float(row.get("priority_score", 0) or 0.0),
        reverse=True,
    )[:50]

    export_delta_summary = {
        "exportable_facts_total_before": exportable_before,
        "exportable_facts_total_after": exportable_after,
        "delta": exportable_after - exportable_before,
    }
    coverage_delta = {
        "metrics": coverage_rows,
        "before": before_summary,
        "after": after_summary,
    }
    return {
        "coverage_delta": coverage_delta,
        "coverage_rows": coverage_rows,
        "review_delta_rows": review_delta_rows,
        "export_delta_summary": export_delta_summary,
        "top_resolved_items": top_resolved_items,
        "top_remaining_unmapped": top_remaining_unmapped,
        "top_remaining_review_items": top_remaining_review_items,
    }


def build_priority_backlog(
    review_rows: Sequence[Dict[str, Any]],
    unmapped_rows: Sequence[Dict[str, Any]],
    reocr_rows: Sequence[Dict[str, Any]],
    priority_rules: Dict[str, Any],
) -> Tuple[List[Dict[str, Any]], Dict[str, Any], List[Dict[str, Any]]]:
    weights = priority_rules.get("weights", {}) if isinstance(priority_rules, dict) else {}
    reocr_review_ids = {row.get("source_review_id", "") for row in reocr_rows if row.get("source_review_id")}
    occurrence_by_label = Counter(row.get("row_label_std", "") for row in review_rows if row.get("row_label_std"))
    mapping_by_label = {row.get("row_label_std", ""): row for row in unmapped_rows if row.get("row_label_std")}
    backlog_rows: List[Dict[str, Any]] = []

    for row in review_rows:
        amount = abs(float(row.get("value_num", 0) or 0.0))
        label = row.get("row_label_std", "")
        statement_type = row.get("statement_type", "")
        statement_score = 1.0 if statement_type in {"balance_sheet", "income_statement", "cash_flow"} else 0.3
        occurrence_score = float(occurrence_by_label.get(label, 1))
        has_candidate = 1.0 if mapping_by_label.get(label, {}).get("top_candidate_code") else 0.0
        has_reocr = 1.0 if row.get("review_id") in reocr_review_ids else 0.0
        priority_score = (
            amount * float(weights.get("amount_scale", 0.000001))
            + occurrence_score * float(weights.get("occurrence", 1.0))
            + statement_score * float(weights.get("statement_criticality", 3.0))
            + has_candidate * float(weights.get("strong_mapping_candidate", 2.0))
            + has_reocr * float(weights.get("reocr_exists", 1.0))
            + float(row.get("priority_score", 0) or 0.0) * float(weights.get("base_review_priority", 1.0))
        )
        backlog_rows.append(
            {
                "review_id": row.get("review_id", ""),
                "row_label_std": label,
                "statement_type": statement_type,
                "period_key": row.get("period_key", ""),
                "value_num": row.get("value_num", ""),
                "reason_codes": row.get("reason_codes", ""),
                "priority_score": round(priority_score, 6),
                "has_mapping_candidate": bool(has_candidate),
                "has_reocr_task": bool(has_reocr),
                "occurrences_for_label": occurrence_by_label.get(label, 1),
            }
        )

    backlog_rows.sort(key=lambda row: float(row["priority_score"]), reverse=True)
    mapping_opportunities = sorted(
        [
            {
                "row_label_std": row.get("row_label_std", ""),
                "occurrences": row.get("occurrences", ""),
                "amount_abs_total": row.get("amount_abs_total", ""),
                "top_candidate_code": row.get("top_candidate_code", ""),
                "top_candidate_name": row.get("top_candidate_name", ""),
                "top_candidate_score": row.get("top_candidate_score", ""),
            }
            for row in unmapped_rows
        ],
        key=lambda row: (float(row.get("amount_abs_total", 0) or 0.0), int(row.get("occurrences", 0) or 0)),
        reverse=True,
    )
    summary = {
        "review_backlog_total": len(backlog_rows),
        "high_priority_total": sum(1 for row in backlog_rows if float(row["priority_score"]) >= float(priority_rules.get("high_priority_threshold", 10.0))),
        "mapping_opportunities_total": len(mapping_opportunities),
    }
    return backlog_rows, summary, mapping_opportunities


def build_metric_delta_row(metric_name: str, before_value: Any, after_value: Any) -> Dict[str, Any]:
    before_num = float(before_value or 0.0)
    after_num = float(after_value or 0.0)
    return {"metric": metric_name, "before": before_value, "after": after_value, "delta": round(after_num - before_num, 6)}


def unresolved_conflicts_total(rows: Sequence[Dict[str, Any]]) -> int:
    return sum(1 for row in rows if row.get("decision") in {"review_required", "unresolved"})


def exportable_facts_total(rows: Sequence[Dict[str, Any]]) -> int:
    return sum(
        1
        for row in rows
        if row.get("mapping_code")
        and row.get("value_num")
        and row.get("report_date_norm") not in {"", "unknown_date"}
        and row.get("period_role_norm") not in {"", "unknown"}
        and row.get("status") in {"observed", "repaired"}
        and row.get("conflict_decision") not in {"review_required", "unresolved"}
        and row.get("unplaced_reason", "") == ""
    )


def load_json(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def load_csv(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def load_sheet_like_csv(path: Path, default: List[Dict[str, str]]) -> List[Dict[str, str]]:
    return load_csv(path) if path.exists() else default


def load_workbook_sheet_rows(path: Path, sheet_name: str) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    workbook = load_workbook(path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        return []
    worksheet = workbook[sheet_name]
    headers = [str(cell.value or "").strip() for cell in worksheet[1]]
    rows: List[Dict[str, str]] = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        rows.append({headers[index]: ("" if value is None else str(value)) for index, value in enumerate(values) if index < len(headers)})
    return rows
