from __future__ import annotations

import csv
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook


def parse_review_actions_file(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return [normalize_row(row) for row in csv.DictReader(handle)]

    workbook = load_workbook(path)
    worksheet = workbook["Actions"] if "Actions" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    headers = [str(cell.value or "").strip() for cell in worksheet[1]]
    rows: List[Dict[str, str]] = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        row = {headers[index]: ("" if value is None else str(value).strip()) for index, value in enumerate(values) if index < len(headers)}
        rows.append(normalize_row(row))
    return rows


def normalize_row(row: Dict[str, str]) -> Dict[str, str]:
    return {str(key).strip(): ("" if value is None else str(value).strip()) for key, value in row.items()}
