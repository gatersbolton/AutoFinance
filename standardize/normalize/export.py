from __future__ import annotations

from collections import defaultdict
from dataclasses import fields, is_dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import load_workbook

from ..models import ConflictRecord, DuplicateRecord, FactRecord, IssueRecord, ReviewQueueRecord, ValidationResultRecord
from ..statement.export_filters import classify_export_blocker
from .mapping import load_template_subjects


UNPLACED_HEADERS = [
    "fact_id",
    "doc_id",
    "page_no",
    "provider",
    "statement_type",
    "mapping_code",
    "mapping_name",
    "row_label_raw",
    "row_label_std",
    "period_key",
    "report_date_norm",
    "period_role_norm",
    "value_raw",
    "value_num",
    "status",
    "conflict_id",
    "conflict_decision",
    "unplaced_reason",
    "source_cell_ref",
]


def export_template(
    template_path: Path,
    output_path: Path,
    facts: List[FactRecord],
    derived_facts: List[FactRecord] | None = None,
    run_summary: Dict[str, Any] | None = None,
    issues: List[IssueRecord] | None = None,
    validations: List[ValidationResultRecord] | None = None,
    duplicates: List[DuplicateRecord] | None = None,
    conflicts: List[ConflictRecord] | None = None,
    review_queue: List[ReviewQueueRecord] | None = None,
    applied_actions: List[Dict[str, Any]] | None = None,
    classification_audit: List[Dict[str, Any]] | None = None,
    period_role_audit: List[Dict[str, Any]] | None = None,
    export_rules: Dict[str, Any] | None = None,
) -> Dict[str, Any]:
    export_rules = export_rules or {}
    subjects, sheet_name, header_row = load_template_subjects(template_path)
    row_by_code = {subject.code: subject.row_index for subject in subjects}

    workbook = load_workbook(template_path)
    worksheet = workbook[sheet_name]

    derived_facts = derived_facts or []
    export_facts = list(facts) + list(derived_facts)
    grouped: Dict[Tuple[str, str], List[FactRecord]] = defaultdict(list)
    unplaced_rows: List[Dict[str, Any]] = []
    for fact in export_facts:
        reason = determine_unplaced_reason(fact, export_rules)
        if reason:
            fact.unplaced_reason = fact.unplaced_reason or reason
            unplaced_rows.append(unplaced_row(fact))
            continue
        grouped[(fact.mapping_code, fact.period_key)].append(fact)

    resolved_groups: Dict[Tuple[str, str], FactRecord] = {}
    for group_key, group_items in grouped.items():
        selected, dropped_rows = select_export_fact(group_items)
        if selected is None:
            for item in group_items:
                item.unplaced_reason = item.unplaced_reason or "multiple_export_candidates"
                unplaced_rows.append(unplaced_row(item))
            continue
        resolved_groups[group_key] = selected
        for row in dropped_rows:
            unplaced_rows.append(row)

    period_keys = sorted({period_key for (_, period_key) in resolved_groups})
    period_columns = ensure_period_columns(worksheet, header_row, period_keys)

    written = 0
    for (mapping_code, period_key), fact in sorted(resolved_groups.items()):
        row_idx = row_by_code.get(mapping_code)
        col_idx = period_columns.get(period_key)
        if row_idx is None or col_idx is None:
            fact.unplaced_reason = fact.unplaced_reason or "template_row_or_column_missing"
            unplaced_rows.append(unplaced_row(fact))
            continue
        adjusted_value = float(fact.value_num) * float(fact.unit_multiplier or 1.0)
        worksheet.cell(row=row_idx, column=col_idx, value=adjusted_value)
        written += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    replace_sheet_with_key_values(workbook, "_meta_summary", run_summary or {})
    replace_sheet_with_rows(workbook, "_issues", issues or [], model_cls=IssueRecord)
    replace_sheet_with_rows(workbook, "_validation", validations or [], model_cls=ValidationResultRecord)
    replace_sheet_with_rows(workbook, "_duplicates", duplicates or [], model_cls=DuplicateRecord)
    replace_sheet_with_rows(workbook, "_conflicts", conflicts or [], model_cls=ConflictRecord)
    replace_sheet_with_rows(workbook, "_review_queue", review_queue or [], model_cls=ReviewQueueRecord)
    replace_sheet_with_rows(workbook, "_unplaced_facts", unplaced_rows, headers=UNPLACED_HEADERS)
    replace_sheet_with_rows(workbook, "_applied_actions", applied_actions or [])
    replace_sheet_with_rows(workbook, "_derived_facts", derived_facts or [])
    replace_sheet_with_rows(workbook, "_classification_audit", classification_audit or [])
    replace_sheet_with_rows(workbook, "_period_role_audit", period_role_audit or [])
    replace_sheet_with_key_values(workbook, "_benchmark_summary", {})
    replace_sheet_with_rows(workbook, "_gap_explanations", [])
    workbook.save(output_path)
    return {
        "written_cells": written,
        "conflicted_cells": 0,
        "period_columns": len(period_columns),
        "period_keys": period_keys,
        "sheet_name": sheet_name,
        "header_row": header_row,
        "helper_sheets": [
            "_meta_summary",
            "_issues",
            "_validation",
            "_duplicates",
            "_conflicts",
            "_unplaced_facts",
            "_review_queue",
            "_applied_actions",
            "_derived_facts",
            "_classification_audit",
            "_period_role_audit",
            "_benchmark_summary",
            "_gap_explanations",
        ],
        "source_facts": "facts_deduped",
        "unplaced_count": len(unplaced_rows),
        "unplaced_rows": unplaced_rows,
    }


def determine_unplaced_reason(fact: FactRecord, export_rules: Dict[str, Any]) -> str:
    return classify_export_blocker(fact, export_rules)


def select_export_fact(group_items: List[FactRecord]) -> Tuple[FactRecord | None, List[Dict[str, Any]]]:
    if len(group_items) == 1:
        return group_items[0], []

    unique_values = {round(float(item.value_num or 0.0), 8) for item in group_items}
    dropped_rows: List[Dict[str, Any]] = []
    if len(unique_values) == 1:
        ordered = sorted(group_items, key=export_fact_score, reverse=True)
        selected = ordered[0]
        for item in ordered[1:]:
            item.unplaced_reason = "duplicate_same_value_not_exported"
            dropped_rows.append(unplaced_row(item))
        return selected, dropped_rows

    accepted = [item for item in group_items if item.conflict_decision in {"accepted", "accepted_with_rule_support", "accepted_with_validation_support"}]
    if len(accepted) == 1:
        selected = accepted[0]
        for item in group_items:
            if item.fact_id == selected.fact_id:
                continue
            item.unplaced_reason = "replaced_by_resolved_candidate"
            dropped_rows.append(unplaced_row(item))
        return selected, dropped_rows

    return None, []


def export_fact_score(fact: FactRecord) -> Tuple[int, ...]:
    return (
        1 if fact.source_kind != "derived_formula" else 0,
        1 if fact.conflict_decision == "accepted_with_validation_support" else 0,
        1 if fact.conflict_decision == "accepted_with_rule_support" else 0,
        1 if fact.comparison_status == "accepted" else 0,
        1 if fact.status == "observed" else 0,
        1 if fact.mapping_code else 0,
        1 if fact.period_role_norm and fact.period_role_norm != "unknown" else 0,
        len(fact.col_header_path or []),
        1 if fact.source_kind != "xlsx_fallback" else 0,
    )


def ensure_period_columns(worksheet, header_row: int, period_keys: List[str]) -> Dict[str, int]:
    existing_columns: Dict[str, int] = {}
    for col_idx in range(1, worksheet.max_column + 1):
        header = worksheet.cell(row=header_row, column=col_idx).value
        if header is not None and str(header).strip():
            existing_columns[str(header).strip()] = col_idx

    period_columns: Dict[str, int] = {}
    next_col = worksheet.max_column + 1
    for period_key in period_keys:
        if period_key in existing_columns:
            period_columns[period_key] = existing_columns[period_key]
            continue
        period_columns[period_key] = next_col
        worksheet.cell(row=header_row, column=next_col, value=period_key)
        next_col += 1
    return period_columns


def unplaced_row(fact: FactRecord) -> Dict[str, Any]:
    return {
        "fact_id": fact.fact_id,
        "doc_id": fact.doc_id,
        "page_no": fact.page_no,
        "provider": fact.provider,
        "statement_type": fact.statement_type,
        "mapping_code": fact.mapping_code,
        "mapping_name": fact.mapping_name,
        "row_label_raw": fact.row_label_raw,
        "row_label_std": fact.row_label_std,
        "period_key": fact.period_key,
        "report_date_norm": fact.report_date_norm,
        "period_role_norm": fact.period_role_norm,
        "value_raw": fact.value_raw,
        "value_num": fact.value_num,
        "status": fact.status,
        "conflict_id": fact.conflict_id,
        "conflict_decision": fact.conflict_decision,
        "unplaced_reason": fact.unplaced_reason,
        "source_cell_ref": fact.source_cell_ref,
    }


def replace_sheet_with_key_values(workbook, sheet_name: str, payload: Dict[str, Any]) -> None:
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.cell(row=1, column=1, value="key")
    worksheet.cell(row=1, column=2, value="value")
    for row_index, (key, value) in enumerate(sorted(payload.items()), start=2):
        worksheet.cell(row=row_index, column=1, value=key)
        worksheet.cell(row=row_index, column=2, value=serialize_sheet_value(value))


def replace_sheet_with_rows(
    workbook,
    sheet_name: str,
    rows: Iterable[Any],
    model_cls=None,
    headers: List[str] | None = None,
) -> None:
    rows = list(rows)
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    worksheet = workbook.create_sheet(sheet_name)

    if headers is None and model_cls is not None:
        headers = [field.name for field in fields(model_cls)]
    elif headers is None and rows and is_dataclass(rows[0]):
        headers = [field.name for field in fields(rows[0])]
    elif headers is None and rows and isinstance(rows[0], dict):
        headers = list(rows[0].keys())
    elif headers is None:
        headers = ["value"]

    worksheet.append(headers)
    if not rows:
        return

    if model_cls is not None or is_dataclass(rows[0]):
        for row in rows:
            worksheet.append([serialize_sheet_value(getattr(row, header, "")) for header in headers])
        return

    if isinstance(rows[0], dict):
        for row in rows:
            worksheet.append([serialize_sheet_value(row.get(header)) for header in headers])
        return

    for row in rows:
        worksheet.append([serialize_sheet_value(row)])


def serialize_sheet_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return str(value)
    return value


def rewrite_meta_summary(output_path: Path, run_summary: Dict[str, Any]) -> None:
    workbook = load_workbook(output_path)
    replace_sheet_with_key_values(workbook, "_meta_summary", run_summary or {})
    workbook.save(output_path)


def rewrite_stage5_helper_sheets(
    output_path: Path,
    benchmark_summary: Dict[str, Any] | None = None,
    gap_rows: List[Dict[str, Any]] | None = None,
    derived_facts: List[FactRecord] | None = None,
    classification_audit: List[Dict[str, Any]] | None = None,
    period_role_audit: List[Dict[str, Any]] | None = None,
) -> None:
    workbook = load_workbook(output_path)
    replace_sheet_with_key_values(workbook, "_benchmark_summary", benchmark_summary or {})
    replace_sheet_with_rows(workbook, "_gap_explanations", gap_rows or [])
    replace_sheet_with_rows(workbook, "_derived_facts", derived_facts or [])
    replace_sheet_with_rows(workbook, "_classification_audit", classification_audit or [])
    replace_sheet_with_rows(workbook, "_period_role_audit", period_role_audit or [])
    workbook.save(output_path)
