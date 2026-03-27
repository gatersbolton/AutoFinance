from __future__ import annotations

import difflib
import re
from pathlib import Path
from typing import Any, Dict, List, Tuple

import yaml
from openpyxl import load_workbook

from ..models import FactRecord, MappingReviewRecord, TemplateSubject
from .text import normalize_label_for_matching


SUBJECT_RE = re.compile(r"^(?P<code>[A-Z]{2}_[0-9]{3})\s+(?P<name>.+)$")


def load_template_subjects(template_path: Path) -> Tuple[List[TemplateSubject], str, int]:
    """Load the standard subject master from the template workbook."""

    workbook = load_workbook(template_path)
    best_sheet = None
    best_subjects: List[TemplateSubject] = []
    best_header_row = 1

    for worksheet in workbook.worksheets:
        subjects: List[TemplateSubject] = []
        header_row = 1
        for row_idx in range(1, worksheet.max_row + 1):
            value = worksheet.cell(row=row_idx, column=1).value
            if value is None:
                continue
            text = str(value).strip()
            if text == "科目名称":
                header_row = row_idx
            match = SUBJECT_RE.match(text)
            if match:
                subjects.append(
                    TemplateSubject(
                        code=match.group("code"),
                        canonical_name=match.group("name").strip(),
                        row_index=row_idx,
                        sheet_name=worksheet.title,
                        source_value=text,
                    )
                )
        if len(subjects) > len(best_subjects):
            best_sheet = worksheet.title
            best_subjects = subjects
            best_header_row = header_row

    if not best_subjects or not best_sheet:
        raise ValueError(f"No standard subject master found in template: {template_path}")

    return best_subjects, best_sheet, best_header_row


def load_alias_mapping(config_path: Path, subjects: List[TemplateSubject]) -> Dict[str, TemplateSubject]:
    if not config_path.exists():
        return {}

    payload = yaml.safe_load(config_path.read_text(encoding="utf-8")) or {}
    aliases = payload.get("aliases", {}) or {}
    normalized_subjects = {
        normalize_label_for_matching(subject.canonical_name): subject
        for subject in subjects
    }
    alias_mapping: Dict[str, TemplateSubject] = {}

    for key, values in aliases.items():
        source_key = normalize_label_for_matching(key)
        candidates = values if isinstance(values, list) else [values]

        if source_key in normalized_subjects:
            canonical = normalized_subjects[source_key]
            for candidate in candidates:
                alias_mapping[normalize_label_for_matching(candidate)] = canonical
            continue

        for candidate in candidates:
            candidate_norm = normalize_label_for_matching(candidate)
            if candidate_norm in normalized_subjects:
                alias_mapping[source_key] = normalized_subjects[candidate_norm]
                break

    return alias_mapping


def apply_subject_mapping(
    facts: List[FactRecord],
    subjects: List[TemplateSubject],
    alias_mapping: Dict[str, TemplateSubject],
) -> Tuple[List[FactRecord], List[MappingReviewRecord]]:
    normalized_subjects = {
        normalize_label_for_matching(subject.canonical_name): subject
        for subject in subjects
    }
    mapping_review: List[MappingReviewRecord] = []

    for fact in facts:
        label = fact.row_label_std
        if not label:
            continue

        exact = normalized_subjects.get(label)
        if exact:
            fact.mapping_code = exact.code
            fact.mapping_name = exact.canonical_name
            fact.mapping_method = "exact"
            fact.mapping_confidence = 1.0
            continue

        alias_match = alias_mapping.get(label)
        if alias_match:
            fact.mapping_code = alias_match.code
            fact.mapping_name = alias_match.canonical_name
            fact.mapping_method = "alias"
            fact.mapping_confidence = 0.95
            continue

        choices = difflib.get_close_matches(label, normalized_subjects.keys(), n=3, cutoff=0.65)
        candidate_subjects = [normalized_subjects[choice] for choice in choices]
        if candidate_subjects:
            best = candidate_subjects[0]
            mapping_review.append(
                MappingReviewRecord(
                    doc_id=fact.doc_id,
                    page_no=fact.page_no,
                    provider=fact.provider,
                    statement_type=fact.statement_type,
                    row_label_raw=fact.row_label_raw,
                    row_label_std=fact.row_label_std,
                    best_code=best.code,
                    best_name=best.canonical_name,
                    candidate_codes_json=";".join(f"{item.code} {item.canonical_name}" for item in candidate_subjects),
                    reason="fuzzy_candidate_only",
                    source_cell_ref=fact.source_cell_ref,
                    status="review",
                )
            )
            fact.mapping_method = "fuzzy_candidate"
            fact.mapping_confidence = 0.5
            continue

        mapping_review.append(
            MappingReviewRecord(
                doc_id=fact.doc_id,
                page_no=fact.page_no,
                provider=fact.provider,
                statement_type=fact.statement_type,
                row_label_raw=fact.row_label_raw,
                row_label_std=fact.row_label_std,
                best_code="",
                best_name="",
                candidate_codes_json="",
                reason="no_match",
                source_cell_ref=fact.source_cell_ref,
                status="review",
            )
        )

    return facts, mapping_review
