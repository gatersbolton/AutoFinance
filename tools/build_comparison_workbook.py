from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import sys
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from project_paths import COMPARISONS_ROOT, CORPUS_ROOT, DEFAULT_TEMPLATE_PATH, LEGACY_ROOT, STANDARDIZE_ARCHIVE_ROOT


TEMPLATE_PATH = DEFAULT_TEMPLATE_PATH
OUTPUT_PATH = COMPARISONS_ROOT / "会计报表_项目与大模型对比汇总_20260405.xlsx"


@dataclass(frozen=True)
class ComparisonSpec:
    company: str
    project_path: Path
    llm_path: Path
    project_note: str = ""
    llm_note: str = ""


@dataclass(frozen=True)
class UnmatchedSpec:
    status: str
    company: str
    project_file: str
    llm_file: str
    note: str


COMPARISONS = [
    ComparisonSpec(
        company="泰兴市泰泽实业有限公司",
        project_path=STANDARDIZE_ARCHIVE_ROOT / "RUN_20260402T161926Z_79bd9692" / "会计报表_填充结果.xlsx",
        llm_path=CORPUS_ROOT / "D01" / "benchmarks" / "会计报表_泰兴市泰泽实业有限公司2022年审计报告_gpt5.4填写.xlsx",
        project_note="基线项目结果（债务人审计报告-2022年）",
        llm_note="采用新命名的泰泽 LLM 文件",
    ),
    ComparisonSpec(
        company="泰兴市长虹现代农业开发有限公司",
        project_path=LEGACY_ROOT / "20260404_table_only_new_docs" / "results" / "1.4.2 债务人——三年一期的财务报表（审计报告）.xlsx",
        llm_path=CORPUS_ROOT / "D02" / "benchmarks" / "会计报表_泰兴市长虹现代农业开发有限公司2022年审计报告_gpt5.4填写.xlsx",
        project_note="项目批次 job 01",
    ),
    ComparisonSpec(
        company="盐城平坦住房租赁有限公司",
        project_path=LEGACY_ROOT / "20260404_table_only_new_docs" / "results" / "1.4.2.3盐城平坦住房租赁有限公司2022年审计报告.xlsx",
        llm_path=CORPUS_ROOT / "D03" / "benchmarks" / "会计报表_盐城平坦住房租赁有限公司2022年审计报告_gpt5.4填写.xlsx",
        project_note="项目批次 job 02",
    ),
    ComparisonSpec(
        company="泰州新滨江科技发展有限公司",
        project_path=LEGACY_ROOT / "20260404_table_only_new_docs" / "results" / "2021年新滨江科技审计报告.xlsx",
        llm_path=CORPUS_ROOT / "D05" / "benchmarks" / "会计报表_泰州新滨江科技发展有限公司2021年审计报告_gpt5.4填写.xlsx",
        project_note="项目批次 job 04",
    ),
    ComparisonSpec(
        company="金湖县润金现代农业发展有限公司",
        project_path=LEGACY_ROOT / "20260404_table_only_new_docs" / "results" / "润金2019-2021年度审计报告.xlsx",
        llm_path=CORPUS_ROOT / "D06" / "benchmarks" / "会计报表_金湖县润金现代农业发展有限公司2021年审计报告_gpt5.4填写.xlsx",
        project_note="项目批次 job 05",
    ),
    ComparisonSpec(
        company="淮安市清浦现代农村建设开发有限公司",
        project_path=LEGACY_ROOT / "20260404_table_only_new_docs" / "results" / "现代农村2021年审计报告.xlsx",
        llm_path=CORPUS_ROOT / "D07" / "benchmarks" / "会计报表_淮安市清浦现代农村建设开发有限公司2021年审计报告_gpt5.4填写.xlsx",
        project_note="项目批次 job 06",
    ),
    ComparisonSpec(
        company="阜宁县鑫泽源污水处理有限公司",
        project_path=LEGACY_ROOT / "20260404_table_only_new_docs" / "results" / "鑫泽源2022年审计报告.xlsx",
        llm_path=CORPUS_ROOT / "D08" / "benchmarks" / "会计报表_阜宁县鑫泽源污水处理有限公司2022年审计报告_gpt5.4填写.xlsx",
        project_note="项目批次 job 07",
    ),
]


UNMATCHED = [
    UnmatchedSpec(
        status="项目仅有",
        company="泰兴市中汇贸易有限公司",
        project_file=str(LEGACY_ROOT / "20260404_table_only_new_docs" / "results" / "1.4.3债务人2022年审计报告.xlsx"),
        llm_file="",
        note="OCR 首屏和正文均识别为泰兴市中汇贸易有限公司；data 中未找到同名 LLM 填写工作簿。",
    ),
    UnmatchedSpec(
        status="LLM仅有/重复",
        company="泰泽实业（旧命名文件）",
        project_file="",
        llm_file=str(LEGACY_ROOT / "会计报表_债务人审计报告-2022年-gpt5.4填写.xlsx"),
        note="与 data/corpus/D01/benchmarks 下的新命名文件指向同一公司，主表已采用新命名文件。",
    ),
]


def assert_files_exist(paths: Iterable[Path]) -> None:
    missing = [str(path) for path in paths if not path.exists()]
    if missing:
        raise FileNotFoundError("Missing required files:\n" + "\n".join(missing))


def load_template_labels(template_path: Path) -> list[str]:
    workbook = load_workbook(template_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    labels: list[str] = []
    for row_index in range(4, worksheet.max_row + 1):
        value = worksheet.cell(row_index, 1).value
        if value is None:
            continue
        labels.append(str(value).strip())
    workbook.close()
    return labels


def load_main_sheet(path: Path) -> tuple[list[object], dict[str, list[object]]]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    headers = [worksheet.cell(3, column_index).value for column_index in range(1, worksheet.max_column + 1)]
    rows: dict[str, list[object]] = {}
    for row_index in range(4, worksheet.max_row + 1):
        label = worksheet.cell(row_index, 1).value
        if label is None:
            continue
        row_values = [worksheet.cell(row_index, column_index).value for column_index in range(1, worksheet.max_column + 1)]
        rows[str(label).strip()] = row_values
    workbook.close()
    return headers, rows


def extract_project_pair(headers: list[object], row_values: list[object]) -> tuple[object | None, object | None]:
    dynamic_values: list[object] = []
    for column_index, header in enumerate(headers, start=1):
        if column_index < 5 or header is None:
            continue
        if column_index - 1 >= len(row_values):
            continue
        value = row_values[column_index - 1]
        if value in (None, ""):
            continue
        dynamic_values.append(value)

    if not dynamic_values:
        fallback = []
        for column_index in (3, 4):
            if column_index - 1 >= len(row_values):
                continue
            value = row_values[column_index - 1]
            if value not in (None, ""):
                fallback.append(value)
        dynamic_values = fallback

    if len(dynamic_values) >= 2:
        return dynamic_values[-2], dynamic_values[-1]
    if len(dynamic_values) == 1:
        return None, dynamic_values[0]
    return None, None


def extract_llm_pair(row_values: list[object]) -> tuple[object | None, object | None]:
    left_value = row_values[2] if len(row_values) >= 3 else None
    right_value = row_values[3] if len(row_values) >= 4 else None
    return left_value, right_value


def build_workbook() -> Path:
    required = [TEMPLATE_PATH]
    for item in COMPARISONS:
        required.extend([item.project_path, item.llm_path])
    assert_files_exist(required)

    labels = load_template_labels(TEMPLATE_PATH)
    loaded = []
    for item in COMPARISONS:
        project_headers, project_rows = load_main_sheet(item.project_path)
        llm_headers, llm_rows = load_main_sheet(item.llm_path)
        loaded.append(
            {
                "spec": item,
                "project_headers": project_headers,
                "project_rows": project_rows,
                "llm_headers": llm_headers,
                "llm_rows": llm_rows,
            }
        )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "对比汇总"

    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill = PatternFill("solid", fgColor="DCE6F1")
    project_fill = PatternFill("solid", fgColor="EAF3FF")
    llm_fill = PatternFill("solid", fgColor="FFF0E0")
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    total_columns = 1 + len(loaded) * 4
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = worksheet.cell(1, 1, "会计报表项目与大模型对比汇总")
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.merge_cells("A2:A3")
    header_cell = worksheet["A2"]
    header_cell.value = "科目名称"
    header_cell.font = Font(bold=True)
    header_cell.alignment = Alignment(horizontal="center", vertical="center")
    header_cell.fill = title_fill

    start_column = 2
    for item in loaded:
        worksheet.merge_cells(start_row=2, start_column=start_column, end_row=2, end_column=start_column + 3)
        company_cell = worksheet.cell(2, start_column, item["spec"].company)
        company_cell.font = Font(bold=True)
        company_cell.alignment = Alignment(horizontal="center", vertical="center")
        company_cell.fill = title_fill

        subheaders = [
            "本项目-期初/上期",
            "本项目-期末/本期",
            "大模型-期初/上期",
            "大模型-期末/本期",
        ]
        for offset, text in enumerate(subheaders):
            cell = worksheet.cell(3, start_column + offset, text)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = project_fill if offset < 2 else llm_fill
        start_column += 4

    for output_row, label in enumerate(labels, start=4):
        label_cell = worksheet.cell(output_row, 1, label)
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        column_index = 2
        for item in loaded:
            project_row = item["project_rows"].get(label, [])
            llm_row = item["llm_rows"].get(label, [])
            project_left, project_right = extract_project_pair(item["project_headers"], project_row) if project_row else (None, None)
            llm_left, llm_right = extract_llm_pair(llm_row) if llm_row else (None, None)
            values = [project_left, project_right, llm_left, llm_right]
            for offset, value in enumerate(values):
                cell = worksheet.cell(output_row, column_index + offset, value)
                cell.fill = project_fill if offset < 2 else llm_fill
                if isinstance(value, (int, float)):
                    cell.number_format = "#,##0.00"
            column_index += 4

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=total_columns):
        for cell in row:
            cell.border = border

    worksheet.freeze_panes = "B4"
    worksheet.auto_filter.ref = f"A3:{get_column_letter(total_columns)}{worksheet.max_row}"
    worksheet.row_dimensions[1].height = 24
    worksheet.row_dimensions[2].height = 24
    worksheet.row_dimensions[3].height = 32
    worksheet.column_dimensions["A"].width = 32
    for column_index in range(2, total_columns + 1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = 16

    meta_sheet = workbook.create_sheet("匹配说明")
    meta_sheet.merge_cells("A1:E1")
    note = meta_sheet["A1"]
    note.value = "项目侧列值按主表每行最后两个非空数据列提取；通常对应期初/上期与期末/本期。若该行只有一列有效值，则左列留空。"
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    note.fill = header_fill
    note.font = Font(bold=True)

    headers = ["状态", "公司", "项目文件", "LLM文件", "备注"]
    for column_index, text in enumerate(headers, start=1):
        cell = meta_sheet.cell(3, column_index, text)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    meta_row = 4
    for item in loaded:
        spec = item["spec"]
        row_values = [
            "已纳入主表",
            spec.company,
            str(spec.project_path),
            str(spec.llm_path),
            "；".join(part for part in [spec.project_note, spec.llm_note] if part),
        ]
        for column_index, value in enumerate(row_values, start=1):
            cell = meta_sheet.cell(meta_row, column_index, value)
            cell.alignment = Alignment(horizontal="left" if column_index >= 3 else "center", vertical="center", wrap_text=True)
            cell.border = border
        meta_row += 1

    for item in UNMATCHED:
        row_values = [item.status, item.company, item.project_file, item.llm_file, item.note]
        for column_index, value in enumerate(row_values, start=1):
            cell = meta_sheet.cell(meta_row, column_index, value)
            cell.alignment = Alignment(horizontal="left" if column_index >= 3 else "center", vertical="center", wrap_text=True)
            cell.border = border
        meta_row += 1

    meta_sheet.freeze_panes = "A4"
    meta_sheet.column_dimensions["A"].width = 16
    meta_sheet.column_dimensions["B"].width = 28
    meta_sheet.column_dimensions["C"].width = 85
    meta_sheet.column_dimensions["D"].width = 85
    meta_sheet.column_dimensions["E"].width = 90

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)
    return OUTPUT_PATH


if __name__ == "__main__":
    print(build_workbook())
