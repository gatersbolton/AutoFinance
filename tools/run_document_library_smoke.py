from __future__ import annotations

import json
import shutil
import sys
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from project_paths import (
    RAW_METRICS_GENERATED_ROOT,
    STANDARD_METRICS_GENERATED_ROOT,
    WEB_COMBINED_DOWNLOAD_SUMMARY_PATH,
    WEB_DOCUMENT_LIBRARY_SUMMARY_PATH,
)
from webapp.config import load_settings
from webapp.document_library import document_to_job, execute_delete, list_documents, load_document, update_document_status
from webapp.document_models import STATUS_COMPLETED
from webapp.main import create_app
from webapp.simple_flow import load_simple_flow_state


def main() -> int:
    settings = load_settings()
    settings.enable_local_worker = False
    settings.auth_required = False
    settings.admin_password = ""
    settings.ensure_directories()

    summary: dict[str, object] = {"pass": False}
    try:
        with TestClient(create_app(settings)) as client:
            home = client.get("/")
            home_pass = home.status_code == 200 and "财务报表数据提取" in home.text
            cleanup_stale_smoke_docs(settings)
            before_ids = {document.doc_id for document in list_documents(settings)}

            sample_pdf = build_sample_pdf_bytes()
            upload = client.post(
                "/documents/upload",
                files=[("uploaded_files", ("stage14_smoke_sample.pdf", sample_pdf, "application/pdf"))],
                follow_redirects=False,
            )
            if upload.status_code != 303:
                raise RuntimeError(f"Upload failed: {upload.status_code}")
            after_documents = [document for document in list_documents(settings) if document.doc_id not in before_ids]
            if not after_documents:
                raise RuntimeError("Uploaded document was not found in library.")
            document = after_documents[0]
            uploaded_doc_id = document.doc_id
            uploaded_pdf_path = document.pdf_path

            initial_home = client.get("/")
            initial_buttons = detect_buttons_near(initial_home.text, document.original_filename, ["开始识别", "重新OCR", "继续处理", "删除"])
            if "开始识别" not in initial_buttons or "继续处理" in initial_buttons:
                raise RuntimeError(f"Unexpected initial buttons: {initial_buttons}")

            write_tiny_aliyun_ocr(settings, uploaded_doc_id)
            update_document_status(settings, uploaded_doc_id, ocr_status=STATUS_COMPLETED)

            ocr_home = client.get("/")
            buttons_after_ocr = detect_buttons_near(ocr_home.text, document.original_filename, ["开始识别", "重新OCR", "继续处理", "删除"])
            if not {"重新OCR", "继续处理", "删除"}.issubset(set(buttons_after_ocr)):
                raise RuntimeError(f"Unexpected buttons after OCR: {buttons_after_ocr}")

            continue_page = client.get(f"/documents/{uploaded_doc_id}/continue")
            if continue_page.status_code != 200 or "一键提取原始数据" not in continue_page.text:
                raise RuntimeError("Continue page did not show Step 1.")

            step1_response = client.post(f"/documents/{uploaded_doc_id}/raw-metrics/run", follow_redirects=False)
            if step1_response.status_code != 303:
                raise RuntimeError(f"Step 1 failed: {step1_response.status_code}")
            document = load_document(settings, uploaded_doc_id)
            job = document_to_job(settings, document)
            state = load_simple_flow_state(job)
            raw_metrics_output = str(state.get("raw_metrics_csv", "") or "")
            step1_status = "completed" if state.get("raw_ready") else "failed"

            step2_response = client.post(f"/documents/{uploaded_doc_id}/standard-metrics/run", follow_redirects=False)
            if step2_response.status_code != 303:
                raise RuntimeError(f"Step 2 failed: {step2_response.status_code}")
            document = load_document(settings, uploaded_doc_id)
            job = document_to_job(settings, document)
            state = load_simple_flow_state(job)
            standardized_metrics_output = str(state.get("standardized_metrics_csv", "") or "")
            combined_workbook_output = str(state.get("combined_metrics_xlsx", "") or "")
            step2_status = "completed" if state.get("standard_ready") else "failed"

            downloads_page = client.get(f"/documents/{uploaded_doc_id}/continue")
            primary_download_button_pass = "下载数据表" in downloads_page.text
            advanced_downloads_available = "高级下载" in downloads_page.text and "原始数据 CSV" in downloads_page.text and "标准化数据 CSV" in downloads_page.text
            legacy_primary_downloads_hidden = "下载原始数据</a>" not in downloads_page.text and "下载标准化数据</a>" not in downloads_page.text
            if not primary_download_button_pass:
                raise RuntimeError("Primary combined download link missing after Step 2.")

            workbook = load_workbook(combined_workbook_output)
            sheets_created = workbook.sheetnames
            total_sheet = workbook["数据总表"]
            headers = [cell.value for cell in total_sheet[1]]
            value_col = headers.index("指标数值") + 1
            fill_date_col = headers.index("填表日期") + 1
            numeric_formatting_pass = total_sheet.cell(row=2, column=value_col).number_format == "#,##0.00"
            date_formatting_pass = total_sheet.cell(row=2, column=fill_date_col).number_format == "yyyy-mm-dd"

            combined_summary = {}
            if WEB_COMBINED_DOWNLOAD_SUMMARY_PATH.exists():
                combined_summary = json.loads(WEB_COMBINED_DOWNLOAD_SUMMARY_PATH.read_text(encoding="utf-8"))
            combined_summary.update(
                {
                    "pass": bool(combined_summary.get("pass") and primary_download_button_pass and advanced_downloads_available and legacy_primary_downloads_hidden and numeric_formatting_pass and date_formatting_pass),
                    "doc_id": uploaded_doc_id,
                    "job_id": uploaded_doc_id,
                    "workbook_path": combined_workbook_output,
                    "sheets_created": sheets_created,
                    "raw_rows_total": combined_summary.get("raw_rows_total", 0),
                    "standardized_rows_total": combined_summary.get("standardized_rows_total", 0),
                    "numeric_formatting_pass": numeric_formatting_pass,
                    "date_formatting_pass": date_formatting_pass,
                    "primary_download_button_pass": primary_download_button_pass,
                    "advanced_downloads_available": advanced_downloads_available,
                    "legacy_primary_downloads_hidden": legacy_primary_downloads_hidden,
                    "path_hygiene_pass": combined_summary.get("path_hygiene_pass", False),
                }
            )
            WEB_COMBINED_DOWNLOAD_SUMMARY_PATH.parent.mkdir(parents=True, exist_ok=True)
            WEB_COMBINED_DOWNLOAD_SUMMARY_PATH.write_text(json.dumps(combined_summary, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")

            raw_review = client.get(f"/documents/{uploaded_doc_id}/raw-review")
            mapping_review = client.get(f"/documents/{uploaded_doc_id}/mapping-review")
            raw_review_page_pass = raw_review.status_code == 200 and "原始数据校对" in raw_review.text
            mapping_review_page_pass = mapping_review.status_code == 200 and "术语映射校对" in mapping_review.text

            delete_doc_id = upload_delete_sample(client, settings, before_ids | {uploaded_doc_id})
            delete_confirm = client.get(f"/documents/{delete_doc_id}/delete-confirm")
            delete_confirm_pass = delete_confirm.status_code == 200 and "确认删除" in delete_confirm.text and "OCR 输出" in delete_confirm.text
            delete_response = client.post(f"/documents/{delete_doc_id}/delete", follow_redirects=False)
            if delete_response.status_code != 303:
                raise RuntimeError(f"Delete failed: {delete_response.status_code}")
            delete_summaries = sorted(settings.deletions_root.glob(f"{delete_doc_id}_*_delete_summary.json"))
            delete_summary_path = str(delete_summaries[-1]) if delete_summaries else ""
            path_hygiene_pass = all(
                is_under_allowed_outputs(Path(path))
                for path in [uploaded_pdf_path, raw_metrics_output, standardized_metrics_output, combined_workbook_output, delete_summary_path]
                if path
            )

            summary.update(
                {
                    "pass": bool(
                        home_pass
                        and step1_status == "completed"
                        and step2_status == "completed"
                        and raw_review_page_pass
                        and mapping_review_page_pass
                        and delete_confirm_pass
                        and delete_summary_path
                        and primary_download_button_pass
                        and advanced_downloads_available
                        and legacy_primary_downloads_hidden
                        and numeric_formatting_pass
                        and date_formatting_pass
                        and path_hygiene_pass
                    ),
                    "uploaded_doc_id": uploaded_doc_id,
                    "uploaded_pdf_path": uploaded_pdf_path,
                    "initial_buttons": initial_buttons,
                    "buttons_after_ocr": buttons_after_ocr,
                    "step1_status": step1_status,
                    "step2_status": step2_status,
                    "raw_metrics_output": raw_metrics_output,
                    "standardized_metrics_output": standardized_metrics_output,
                    "combined_workbook_output": combined_workbook_output,
                    "sheets_created": sheets_created,
                    "numeric_formatting_pass": numeric_formatting_pass,
                    "date_formatting_pass": date_formatting_pass,
                    "primary_download_button_pass": primary_download_button_pass,
                    "advanced_downloads_available": advanced_downloads_available,
                    "legacy_primary_downloads_hidden": legacy_primary_downloads_hidden,
                    "raw_review_page_pass": raw_review_page_pass,
                    "mapping_review_page_pass": mapping_review_page_pass,
                    "delete_confirm_pass": delete_confirm_pass,
                    "delete_summary_path": delete_summary_path,
                    "path_hygiene_pass": path_hygiene_pass,
                }
            )
    except Exception as exc:
        summary["error"] = str(exc)
    finally:
        WEB_DOCUMENT_LIBRARY_SUMMARY_PATH.parent.mkdir(parents=True, exist_ok=True)
        WEB_DOCUMENT_LIBRARY_SUMMARY_PATH.write_text(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True))
    return 0 if summary.get("pass") else 1


def build_sample_pdf_bytes() -> bytes:
    try:
        import fitz

        document = fitz.open()
        page = document.new_page()
        page.insert_text((72, 72), "Stage 14 smoke sample")
        payload = document.write()
        document.close()
        return payload
    except Exception:
        return b"%PDF-1.4\n1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n2 0 obj<</Type/Pages/Count 0>>endobj\ntrailer<</Root 1 0 R>>\n%%EOF\n"


def cleanup_stale_smoke_docs(settings) -> None:
    for document in list_documents(settings):
        if document.original_filename in {"stage14_smoke_sample.pdf", "stage14_delete_sample.pdf"}:
            try:
                execute_delete(settings, document.doc_id)
            except Exception:
                pass


def upload_delete_sample(client: TestClient, settings, existing_ids: set[str]) -> str:
    response = client.post(
        "/documents/upload",
        files=[("uploaded_files", ("stage14_delete_sample.pdf", build_sample_pdf_bytes(), "application/pdf"))],
        follow_redirects=False,
    )
    if response.status_code != 303:
        raise RuntimeError(f"Delete sample upload failed: {response.status_code}")
    for document in list_documents(settings):
        if document.doc_id not in existing_ids and document.original_filename == "stage14_delete_sample.pdf":
            return document.doc_id
    raise RuntimeError("Delete sample document was not found.")


def detect_buttons(html: str, labels: list[str]) -> list[str]:
    return [label for label in labels if label in html]


def detect_buttons_near(html: str, needle: str, labels: list[str]) -> list[str]:
    index = html.find(needle)
    if index < 0:
        return detect_buttons(html, labels)
    fragment = html[index : index + 2500]
    return detect_buttons(fragment, labels)


def write_tiny_aliyun_ocr(settings, doc_id: str) -> None:
    document = load_document(settings, doc_id, refresh=False)
    provider_doc_dir = Path(document.ocr_output_dir) / "aliyun_table" / "demo_doc"
    if provider_doc_dir.exists():
        shutil.rmtree(provider_doc_dir)
    raw_dir = provider_doc_dir / "raw"
    raw_dir.mkdir(parents=True, exist_ok=True)
    cells = [
        (1, 0, 0, 0, 0, "项目"),
        (2, 0, 0, 1, 1, "行次"),
        (3, 0, 0, 2, 2, "期初数"),
        (4, 0, 0, 3, 3, "期末数"),
        (5, 1, 1, 0, 0, "货币资金"),
        (6, 1, 1, 1, 1, "1"),
        (7, 1, 1, 2, 2, "100"),
        (8, 1, 1, 3, 3, "200"),
    ]
    raw_payload = {
        "Data": {
            "content": "资产负债表\n编制单位：AAA有限公司\n2022年12月31日\n单位：元",
            "tableHeadTail": [{"head": ["资产负债表", "编制单位：AAA有限公司", "2022年12月31日", "单位：元"], "tail": []}],
            "prism_tablesInfo": [
                {
                    "tableId": "1",
                    "xCellSize": 4,
                    "yCellSize": 2,
                    "cellInfos": [
                        {
                            "tableCellId": cell_id,
                            "ysc": row_start,
                            "yec": row_end,
                            "xsc": col_start,
                            "xec": col_end,
                            "word": text,
                            "pos": [
                                {"x": 10 + col_start * 80, "y": 20 + row_start * 30},
                                {"x": 70 + col_end * 80, "y": 20 + row_start * 30},
                                {"x": 70 + col_end * 80, "y": 45 + row_end * 30},
                                {"x": 10 + col_start * 80, "y": 45 + row_end * 30},
                            ],
                        }
                        for cell_id, row_start, row_end, col_start, col_end, text in cells
                    ],
                }
            ],
        }
    }
    (raw_dir / "page_0001.json").write_text(json.dumps(raw_payload, ensure_ascii=False, indent=2), encoding="utf-8")
    (provider_doc_dir / "result.json").write_text(
        json.dumps(
            {
                "provider": "aliyun_table",
                "pages": [
                    {
                        "page_number": 1,
                        "text": "资产负债表\n编制单位：AAA有限公司\n2022年12月31日\n单位：元",
                        "raw_file": "raw/page_0001.json",
                    }
                ],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def is_under_allowed_outputs(path: Path) -> bool:
    resolved = path.resolve()
    roots = [
        Path("data/corpus/library").resolve(),
        RAW_METRICS_GENERATED_ROOT.resolve(),
        STANDARD_METRICS_GENERATED_ROOT.resolve(),
        Path("data/generated/web").resolve(),
    ]
    return any(is_relative_to(resolved, root) for root in roots)


def is_relative_to(path: Path, root: Path) -> bool:
    try:
        path.relative_to(root)
        return True
    except ValueError:
        return False


if __name__ == "__main__":
    raise SystemExit(main())
