from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook


TERM_PATTERN = re.compile(r"^\s*(ZT_\d+)\s+(.+?)\s*$")
METADATA_FIELDS = (
    "category",
    "aliases",
    "statement_scope",
    "metric_type",
    "period_type",
    "legacy_aliases",
    "description",
    "enabled",
    "notes",
)


def build_registry(template_path: Path, existing_registry_path: Path) -> dict[str, Any]:
    existing_payload = _load_yaml(existing_registry_path)
    existing_terms = [item for item in existing_payload.get("terms", []) if isinstance(item, dict)]
    metadata_by_name = {
        _normalize_name(item.get("name", "")): item
        for item in existing_terms
        if str(item.get("name", "") or "").strip()
    }
    outside_template_terms = [
        _clean_existing_term(item)
        for item in existing_terms
        if str(item.get("code", "") or "").strip() and not str(item.get("code", "") or "").startswith("ZT_")
    ]

    template_terms = _read_template_terms(template_path)
    terms: list[dict[str, Any]] = outside_template_terms
    for code, name in template_terms:
        number = int(code.split("_", 1)[1])
        preserved = metadata_by_name.get(_normalize_name(name), {})
        term: dict[str, Any] = {
            "code": code,
            "name": name,
            "category": _category(number),
            "statement_scope": "balance_sheet" if number <= 136 else "income_statement",
            "metric_type": "per_share" if number in {193, 194} else "amount",
            "period_type": "point" if number <= 136 else "flow",
            "enabled": True,
        }
        for field in METADATA_FIELDS:
            if field in {"category", "statement_scope", "metric_type", "period_type", "enabled"}:
                continue
            value = preserved.get(field)
            if value not in (None, "", [], ()):
                term[field] = value
        terms.append(term)

    codes = [str(item["code"]) for item in terms]
    if len(template_terms) != 194:
        raise ValueError(f"Expected 194 template terms, found {len(template_terms)}")
    if len(codes) != len(set(codes)):
        raise ValueError("Generated registry contains duplicate codes")
    return {"terms": terms}


def render_registry(payload: dict[str, Any]) -> str:
    header = "# Generated from data/templates/会计报表.xlsx by tools/sync_standard_terms_from_template.py.\n"
    return header + yaml.safe_dump(payload, allow_unicode=True, sort_keys=False, width=120)


def _read_template_terms(path: Path) -> list[tuple[str, str]]:
    workbook = load_workbook(path, data_only=False, read_only=True)
    try:
        terms: list[tuple[str, str]] = []
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(min_col=1, max_col=1, values_only=True):
                value = row[0]
                if not isinstance(value, str):
                    continue
                match = TERM_PATTERN.match(value)
                if match:
                    terms.append((match.group(1), match.group(2)))
        return terms
    finally:
        workbook.close()


def _load_yaml(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    payload = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    return payload if isinstance(payload, dict) else {}


def _clean_existing_term(item: dict[str, Any]) -> dict[str, Any]:
    result = {"code": str(item.get("code", "") or "").strip(), "name": str(item.get("name", "") or "").strip()}
    for field in METADATA_FIELDS:
        value = item.get(field)
        if value not in (None, "", [], ()):
            result[field] = value
    return result


def _normalize_name(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).replace("（", "(").replace("）", ")")


def _category(number: int) -> str:
    if number <= 67:
        return "资产类"
    if number <= 119:
        return "负债类"
    if number <= 136:
        return "所有者权益类"
    return "损益类"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Synchronize the standard-term registry with the accounting template.")
    parser.add_argument("--template", default="data/templates/会计报表.xlsx")
    parser.add_argument("--registry", default="config/standard_terms.yml")
    parser.add_argument("--check", action="store_true", help="Exit non-zero when the registry is not synchronized.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    template_path = Path(args.template).resolve()
    registry_path = Path(args.registry).resolve()
    rendered = render_registry(build_registry(template_path, registry_path))
    if args.check:
        current = registry_path.read_text(encoding="utf-8") if registry_path.exists() else ""
        if current != rendered:
            print(f"Registry is not synchronized: {registry_path}")
            return 1
        print(f"Registry is synchronized: {registry_path}")
        return 0
    registry_path.write_text(rendered, encoding="utf-8")
    print(f"Wrote synchronized registry: {registry_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
