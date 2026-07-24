from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from project_paths import RAW_METRICS_GENERATED_ROOT

from .models import RawMetricRow


RAW_REQUIRED_COLUMNS = ["填表日期", "当前条目日期", "公司名", "指标名", "指标数值"]

PERIOD_ROLE_LABELS_ZH = {
    "beginning": "期初数",
    "ending": "期末数",
    "previous_ending": "上期期末",
    "current_point": "当前时点",
    "current_period": "本期",
    "current_year": "本年",
    "previous_period": "上期",
    "previous_year": "上年",
    "amount": "金额",
    "explicit_date": "明确日期",
}


def validate_input_path(
    input_path: Path,
    *,
    raw_metrics_root: Path = RAW_METRICS_GENERATED_ROOT,
) -> None:
    raw_root = raw_metrics_root.resolve()
    try:
        input_path.resolve().relative_to(raw_root)
    except ValueError as exc:
        raise ValueError(f"Stage 13 input must be under {raw_root}; got {input_path}") from exc


def load_raw_metrics(
    input_path: Path,
    *,
    company_name_override: str = "",
    raw_metrics_root: Path = RAW_METRICS_GENERATED_ROOT,
) -> list[RawMetricRow]:
    input_path = input_path.resolve()
    validate_input_path(input_path, raw_metrics_root=raw_metrics_root)
    if not input_path.exists() or not input_path.is_file():
        raise ValueError(f"Raw metrics input does not exist: {input_path}")
    if input_path.suffix.lower() == ".csv":
        rows = _read_csv(input_path)
    elif input_path.suffix.lower() == ".xlsx":
        rows = _read_xlsx(input_path)
    else:
        raise ValueError(f"Unsupported raw metrics input type: {input_path.suffix}")
    _validate_raw_columns(rows, input_path)
    detailed_rows = _load_detailed_sidecar(input_path)
    results: list[RawMetricRow] = []
    for index, row in enumerate(rows, start=1):
        detailed = detailed_rows[index - 1] if index - 1 < len(detailed_rows) else {}
        raw_metric_id = str(detailed.get("source_cell_ref", "") or f"raw_{index:06d}").strip()
        provenance = {
            "source_page_no": detailed.get("page_no", ""),
            "source_bbox_json": detailed.get("bbox_json", ""),
            "source_pdf_path": detailed.get("evidence_path", ""),
            "source_file": detailed.get("source_file", ""),
            "provider": detailed.get("provider", ""),
            "doc_id": detailed.get("doc_id", ""),
            "source_cell_ref": detailed.get("source_cell_ref", ""),
            "statement_type": detailed.get("statement_type", ""),
            "header_path": detailed.get("header_path", ""),
            "period_role_raw": detailed.get("period_role_raw", ""),
            "period_role_norm": detailed.get("period_role_norm", ""),
            "row_context_path": detailed.get("row_context_path", ""),
            "value_type": detailed.get("value_type", ""),
        }
        results.append(
            RawMetricRow(
                row_number=index,
                review_item_id=f"maprev_{index:06d}",
                raw_metric_id=raw_metric_id,
                fill_date=row.get("填表日期", ""),
                item_date=row.get("当前条目日期", ""),
                company_name=str(company_name_override or row.get("公司名", "") or ""),
                metric_name=str(row.get("指标名", "") or ""),
                metric_value=row.get("指标数值", ""),
                period_role=_period_role_from_row(row, detailed),
                raw_row=dict(row),
                provenance=provenance,
            )
        )
    return results


def infer_doc_id(
    input_path: Path,
    override: str = "",
    *,
    raw_metrics_root: Path = RAW_METRICS_GENERATED_ROOT,
) -> str:
    if override.strip():
        return override.strip()
    try:
        relative = input_path.resolve().relative_to(raw_metrics_root.resolve())
    except ValueError:
        return ""
    return relative.parts[0] if len(relative.parts) >= 3 else ""


def _read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def _read_xlsx(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["raw_metrics"] if "raw_metrics" in workbook.sheetnames else workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = ["" if value is None else str(value).strip() for value in rows[0]]
    result: list[dict[str, Any]] = []
    for values in rows[1:]:
        result.append({headers[index]: values[index] if index < len(values) else "" for index in range(len(headers)) if headers[index]})
    return result


def _validate_raw_columns(rows: list[dict[str, Any]], input_path: Path) -> None:
    if not rows:
        return
    missing = [column for column in RAW_REQUIRED_COLUMNS if column not in rows[0]]
    if missing:
        raise ValueError(f"Raw metrics input missing required columns {missing}: {input_path}")


def _load_detailed_sidecar(input_path: Path) -> list[dict[str, Any]]:
    sidecar = input_path.parent / "raw_metrics_detailed.csv"
    if not sidecar.exists():
        return []
    with sidecar.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def _period_role_from_row(row: dict[str, Any], detailed: dict[str, Any]) -> str:
    explicit = str(row.get("期间类型", "") or row.get("period_role", "") or "").strip()
    if explicit:
        return explicit
    norm = str(detailed.get("period_role_norm", "") or "").strip()
    raw = str(detailed.get("period_role_raw", "") or "").strip()
    if norm and norm != "unknown":
        return PERIOD_ROLE_LABELS_ZH.get(norm, norm)
    if raw and raw != "unknown":
        return raw
    return ""
