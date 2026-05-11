from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook

from .models import (
    CANDIDATE_OUTPUT_COLUMNS,
    DETAILED_OUTPUT_COLUMNS,
    ISSUE_OUTPUT_COLUMNS,
    REVIEW_ITEM_COLUMNS,
    STANDARD_OUTPUT_COLUMNS,
    MappingResult,
    serialize_value,
)


OUTPUT_FILENAMES = [
    "standardized_metrics.csv",
    "standardized_metrics.xlsx",
    "standardized_metrics.jsonl",
    "standardized_metrics_detailed.csv",
    "mapping_candidates.csv",
    "mapping_issues.csv",
    "standard_mapping_summary.json",
    "mapping_review_items.csv",
    "standard_mapping_run_manifest.json",
]


def export_standard_mapping_run(
    *,
    output_dir: Path,
    rows: list[MappingResult],
    summary: dict[str, Any],
    manifest: dict[str, Any],
) -> list[str]:
    output_dir.mkdir(parents=True, exist_ok=True)
    main_rows = [row.main_row() for row in rows]
    detailed_rows = [row.detailed_row() for row in rows]
    candidate_rows = [candidate.as_output_row() for row in rows for candidate in row.candidates]
    issue_rows = []
    for index, row in enumerate(rows, start=1):
        issue_row = row.issue_row(index)
        if issue_row is not None:
            issue_rows.append(issue_row)
    review_rows = [row.review_item_row() for row in rows]

    write_dict_csv(output_dir / "standardized_metrics.csv", main_rows, STANDARD_OUTPUT_COLUMNS)
    write_jsonl(output_dir / "standardized_metrics.jsonl", main_rows)
    write_dict_csv(output_dir / "standardized_metrics_detailed.csv", detailed_rows, DETAILED_OUTPUT_COLUMNS)
    write_dict_csv(output_dir / "mapping_candidates.csv", candidate_rows, CANDIDATE_OUTPUT_COLUMNS)
    write_dict_csv(output_dir / "mapping_issues.csv", issue_rows, ISSUE_OUTPUT_COLUMNS)
    write_dict_csv(output_dir / "mapping_review_items.csv", review_rows, REVIEW_ITEM_COLUMNS)
    write_xlsx(output_dir / "standardized_metrics.xlsx", main_rows, detailed_rows, candidate_rows, issue_rows, review_rows)
    write_json(output_dir / "standard_mapping_summary.json", summary)
    write_json(output_dir / "standard_mapping_run_manifest.json", manifest)
    return sorted(str(path) for path in output_dir.iterdir() if path.is_file())


def write_dict_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: serialize_value(row.get(field)) for field in fieldnames})


def write_jsonl(path: Path, rows: Iterable[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            payload = {key: serialize_value(value) for key, value in row.items()}
            handle.write(json.dumps(payload, ensure_ascii=False, separators=(",", ":"), sort_keys=True))
            handle.write("\n")


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")


def write_xlsx(
    path: Path,
    main_rows: list[dict[str, Any]],
    detailed_rows: list[dict[str, Any]],
    candidate_rows: list[dict[str, Any]],
    issue_rows: list[dict[str, Any]],
    review_rows: list[dict[str, Any]],
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "standardized_metrics"
    append_rows(sheet, STANDARD_OUTPUT_COLUMNS, main_rows)
    detailed = workbook.create_sheet("standardized_metrics_detailed")
    append_rows(detailed, DETAILED_OUTPUT_COLUMNS, detailed_rows)
    candidates = workbook.create_sheet("mapping_candidates")
    append_rows(candidates, CANDIDATE_OUTPUT_COLUMNS, candidate_rows)
    issues = workbook.create_sheet("mapping_issues")
    append_rows(issues, ISSUE_OUTPUT_COLUMNS, issue_rows)
    review = workbook.create_sheet("mapping_review_items")
    append_rows(review, REVIEW_ITEM_COLUMNS, review_rows)
    workbook.save(path)


def append_rows(sheet, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    sheet.append(fieldnames)
    for row in rows:
        sheet.append([serialize_value(row.get(field)) for field in fieldnames])
