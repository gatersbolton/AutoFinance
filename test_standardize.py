import csv
import json
import shutil
import tempfile
import unittest
from pathlib import Path

import yaml
from PIL import Image, ImageDraw
from openpyxl import load_workbook

from standardize import cli
from standardize.benchmark import compare_benchmark_workbook, explain_benchmark_gaps
from standardize.curation import build_alias_acceptance_candidates, build_formula_rule_impact, load_curated_alias_records, load_curated_formula_rules, load_legacy_alias_records, prune_reocr_tasks, split_unmapped_facts
from standardize.dedupe import assign_fact_ids, dedupe_facts
from standardize.derive import derive_formula_facts
from standardize.feedback import apply_review_actions, build_delta_reports, export_review_actions_template, parse_review_actions_file
from standardize.integrity import run_artifact_integrity
from standardize.manifest import generate_run_id, write_run_manifest
from standardize.mapping.review import apply_subject_mapping
from standardize.models import AliasRecord, CellRecord, ConflictRecord, FactRecord, ProviderCell, ProviderPage, ReOCRTaskRecord, RelationRecord, ReviewQueueRecord, StatementMeta, TemplateSubject, ValidationResultRecord
from standardize.normalize.conflicts import enrich_conflicts, resolve_conflicts
from standardize.normalize.export import export_template
from standardize.normalize.labels import apply_label_canonicalization
from standardize.normalize.mapping import load_template_subjects
from standardize.normalize.numbers import analyze_numeric_text
from standardize.normalize.periods import apply_period_normalization
from standardize.normalize.tables import standardize_page
from standardize.overrides.periods import apply_period_overrides
from standardize.overrides.storage import ensure_override_store
from standardize.overrides.suppression import apply_suppression_overrides
from standardize.promotion import apply_promotions, build_promotion_delta, export_promotion_actions_template, parse_promotion_actions_file
from standardize.providers.aliyun import extract_aliyun_data
from standardize.review import build_review_queue
from standardize.routing.page_selector import build_page_selection
from standardize.routing.secondary_ocr import materialize_reocr_inputs
from standardize.statement import resolve_single_period_annual_roles, run_full_run_contract, specialize_statement_types
from standardize.target import build_stage7_kpis, build_target_kpis, build_target_review_backlogs, investigate_no_source_gaps, repair_benchmark_alignment, scope_facts_to_targets
from standardize.validation import run_validation


class StandardizeTests(unittest.TestCase):
    def test_extract_aliyun_outer_data_string(self):
        raw = {
            "Data": json.dumps(
                {
                    "content": "资产负债表",
                    "prism_tablesInfo": [{"cellInfos": []}],
                },
                ensure_ascii=False,
            )
        }

        data = extract_aliyun_data(raw)

        self.assertEqual(data["content"], "资产负债表")
        self.assertIn("prism_tablesInfo", data)

    def test_dense_grid_preserves_blank_cells(self):
        page = ProviderPage(
            doc_id="demo",
            page_no=1,
            provider="aliyun_table",
            source_file="demo.json",
            source_kind="json",
            page_text="资产负债表",
            tables={
                "1": [
                    ProviderCell(table_id="1", row_start=0, row_end=0, col_start=0, col_end=0, text="项目"),
                    ProviderCell(table_id="1", row_start=0, row_end=0, col_start=2, col_end=2, text="金额"),
                    ProviderCell(table_id="1", row_start=1, row_end=1, col_start=0, col_end=0, text="货币资金"),
                    ProviderCell(table_id="1", row_start=1, row_end=1, col_start=2, col_end=2, text="100"),
                ]
            },
            context_lines=["资产负债表", "2022年12月31日", "单位:元"],
        )
        statement_meta = StatementMeta(
            statement_type="balance_sheet",
            statement_name_raw="资产负债表",
            report_date_raw="2022年12月31日",
            report_date_norm="2022-12-31",
            unit_raw="元",
            unit_multiplier=1.0,
        )

        cells, subtables, issues = standardize_page(
            page=page,
            statement_meta=statement_meta,
            keyword_config={},
        )

        self.assertEqual(len(cells), 6)
        self.assertEqual(sum(1 for cell in cells if cell.is_empty), 2)
        self.assertEqual(len(subtables), 1)
        self.assertEqual(len(issues), 0)

    def test_number_cleaning(self):
        self.assertEqual(analyze_numeric_text("396，149，420.62", expected_numeric=True)["value_num"], 396149420.62)
        self.assertEqual(analyze_numeric_text("(1,234.56)", expected_numeric=True)["value_num"], -1234.56)
        self.assertEqual(analyze_numeric_text("98.26%", expected_numeric=True)["value_num"], 0.9826)

    def test_suspicious_values(self):
        for value in ["务屏20,000,000.00", "t", "章"]:
            result = analyze_numeric_text(value, expected_numeric=True)
            self.assertTrue(result["is_suspicious"])
            self.assertTrue(result["suspicious_reason"])

    def test_template_subject_parsing(self):
        template_path = Path(__file__).resolve().parent.parent / "会计报表.xlsx"

        subjects, sheet_name, header_row = load_template_subjects(template_path)

        self.assertEqual(subjects[0].code, "ZT_001")
        self.assertEqual(subjects[0].canonical_name, "货币资金")
        self.assertTrue(sheet_name)
        self.assertGreaterEqual(header_row, 1)

    def test_period_normalization_parses_and_inherits(self):
        facts = [
            self.make_fact(
                doc_id="demo-2022年",
                page_no=1,
                statement_type="balance_sheet",
                statement_name_raw="资产负债表",
                col_header_raw="截至2022年12月31日 / 期末数",
                col_header_path=["截至2022年12月31日", "期末数"],
                period_role_raw="期末数",
                period_key="unknown_date__期末数",
                row_label_std="货币资金",
                report_date_raw="",
                report_date_norm="unknown_date",
            ),
            self.make_fact(
                doc_id="demo-2022年",
                page_no=2,
                statement_type="note",
                statement_name_raw="2022年度财务报表附注",
                col_header_raw="本期发生额",
                col_header_path=["本期发生额"],
                period_role_raw="本期",
                period_key="unknown_date__本期",
                row_label_std="管理费用",
                report_date_raw="",
                report_date_norm="unknown_date",
            ),
            self.make_fact(
                doc_id="demo-2022年",
                page_no=3,
                statement_type="note",
                statement_name_raw="财务报表附注",
                col_header_raw="期末数",
                col_header_path=["期末数"],
                period_role_raw="期末数",
                period_key="unknown_date__期末数",
                row_label_std="固定资产净额",
                report_date_raw="",
                report_date_norm="unknown_date",
            ),
        ]
        pages = [
            ProviderPage(
                doc_id="demo-2022年",
                page_no=1,
                provider="aliyun_table",
                source_file="page1.json",
                source_kind="json",
                page_text="资产负债表 2022年12月31日",
                tables={},
                context_lines=["资产负债表", "2022年12月31日"],
            ),
            ProviderPage(
                doc_id="demo-2022年",
                page_no=2,
                provider="aliyun_table",
                source_file="page2.json",
                source_kind="json",
                page_text="2022年度财务报表附注",
                tables={},
                context_lines=["2022年度财务报表附注"],
            ),
            ProviderPage(
                doc_id="demo-2022年",
                page_no=3,
                provider="aliyun_table",
                source_file="page3.json",
                source_kind="json",
                page_text="财务报表附注",
                tables={},
                context_lines=["财务报表附注"],
            ),
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            input_dir = Path(tmpdir) / "outputs"
            input_dir.mkdir(parents=True, exist_ok=True)
            normalized = apply_period_normalization(
                facts=facts,
                provider_pages=pages,
                input_dir=input_dir,
                keyword_config=self.statement_config(),
                period_config={"source_weights": {"header": 120, "page_context": 95, "page_text": 70, "doc_id": 18}, "repeat_bonus": 2, "tie_threshold": 5},
                enabled=True,
            )

        by_page = {fact.page_no: fact for fact in normalized}
        self.assertEqual(by_page[1].report_date_norm, "2022-12-31")
        self.assertEqual(by_page[1].period_key, "2022-12-31__期末数")
        self.assertEqual(by_page[2].report_date_norm, "2022年度")
        self.assertEqual(by_page[2].period_key, "2022年度__本期")
        self.assertEqual(by_page[3].report_date_norm, "2022-12-31")
        self.assertEqual(by_page[3].period_source_level, "statement")

    def test_dedupe_prefers_explicit_date_over_unknown(self):
        facts = assign_fact_ids(
            [
                self.make_fact(
                    fact_id="",
                    report_date_norm="2022-12-31",
                    period_key="2022-12-31__期末数",
                    period_role_norm="期末数",
                    row_label_std="货币资金",
                    mapping_code="ZT_001",
                    mapping_name="货币资金",
                    value_num=100.0,
                    value_raw="100",
                ),
                self.make_fact(
                    fact_id="",
                    report_date_norm="unknown_date",
                    period_key="unknown_date__期末数",
                    period_role_norm="期末数",
                    row_label_std="货币资金",
                    mapping_code="ZT_001",
                    mapping_name="货币资金",
                    value_num=100.0,
                    value_raw="100",
                ),
            ]
        )

        deduped, duplicates = dedupe_facts(facts, ["aliyun_table", "tencent_table_v3"])

        self.assertEqual(len(deduped), 1)
        self.assertEqual(deduped[0].period_key, "2022-12-31__期末数")
        self.assertEqual(len(duplicates), 1)
        self.assertEqual(duplicates[0].dedupe_reason, "explicit_date_preferred_over_unknown_date")

    def test_provider_compare_metrics(self):
        equal_fact_a = self.make_fact(
            doc_id="demo",
            page_no=4,
            provider="aliyun_table",
            table_semantic_key="balance_sheet|h:项目,期末数",
            row_label_std="货币资金",
            row_label_raw="货币资金",
            column_semantic_key="期末数",
            period_role_raw="期末数",
            period_role_norm="期末数",
            period_key="2022-12-31__期末数",
            value_num=100.0,
            value_raw="100",
        )
        equal_fact_b = self.make_fact(
            doc_id="demo",
            page_no=4,
            provider="tencent_table_v3",
            table_semantic_key="balance_sheet|h:项目,期末数",
            row_label_std="货币资金",
            row_label_raw="货币资金",
            column_semantic_key="期末数",
            period_role_raw="期末数",
            period_role_norm="期末数",
            period_key="2022-12-31__期末数",
            value_num=100.0,
            value_raw="100",
        )
        conflict_fact_a = self.make_fact(
            doc_id="demo",
            page_no=5,
            provider="aliyun_table",
            table_semantic_key="note|h:项目,本期发生额",
            row_label_std="管理费用",
            row_label_raw="管理费用",
            column_semantic_key="本期发生额",
            period_role_raw="本期",
            period_role_norm="本期",
            period_key="2022年度__本期",
            value_num=10.0,
            value_raw="10",
        )
        conflict_fact_b = self.make_fact(
            doc_id="demo",
            page_no=5,
            provider="tencent_table_v3",
            table_semantic_key="note|h:项目,本期发生额",
            row_label_std="管理费用",
            row_label_raw="管理费用",
            column_semantic_key="本期发生额",
            period_role_raw="本期",
            period_role_norm="本期",
            period_key="2022年度__本期",
            value_num=12.0,
            value_raw="12",
        )
        facts = assign_fact_ids([equal_fact_a, equal_fact_b, conflict_fact_a, conflict_fact_b])

        resolved_facts, conflicts, comparisons = resolve_conflicts(
            facts=facts,
            provider_priority=["aliyun_table", "tencent_table_v3"],
            enabled=True,
        )

        self.assertEqual(len(conflicts), 1)
        comparison_by_page = {record.page_no: record for record in comparisons}
        self.assertGreater(comparison_by_page[4].compared_pairs, 0)
        self.assertEqual(comparison_by_page[4].equal_pairs, 1)
        self.assertEqual(comparison_by_page[5].conflict_pairs, 1)
        self.assertTrue(any(fact.comparison_status == "equal" for fact in resolved_facts if fact.page_no == 4))

    def test_conflict_grouping_uses_table_semantics(self):
        facts = assign_fact_ids(
            [
                self.make_fact(
                    doc_id="demo",
                    page_no=13,
                    provider="aliyun_table",
                    table_semantic_key="note|h:单位名称,期末余额,账龄,比例",
                    logical_subtable_id="4_sub1",
                    row_label_raw="合计",
                    row_label_std="合计",
                    column_semantic_key="期末数",
                    period_role_raw="期末数",
                    period_role_norm="期末数",
                    period_key="unknown_date__期末数",
                    value_raw="735，449，914.02",
                    value_num=735449914.02,
                ),
                self.make_fact(
                    doc_id="demo",
                    page_no=13,
                    provider="tencent_table_v3",
                    table_semantic_key="note|h:项目,期初数,期末数",
                    logical_subtable_id="4_sub1",
                    row_label_raw="合计",
                    row_label_std="合计",
                    column_semantic_key="期末数",
                    period_role_raw="期末数",
                    period_role_norm="期末数",
                    period_key="unknown_date__期末数",
                    value_raw="401,700,701.94",
                    value_num=401700701.94,
                ),
            ]
        )

        resolved_facts, conflicts, comparisons = resolve_conflicts(
            facts=facts,
            provider_priority=["aliyun_table", "tencent_table_v3"],
            enabled=True,
        )

        self.assertEqual(len(conflicts), 0)
        self.assertEqual(comparisons[0].compared_pairs, 0)
        self.assertEqual(comparisons[0].reason, "no_aligned_pairs_found")
        self.assertTrue(all(fact.status == "observed" for fact in resolved_facts))

    def test_validation_rules(self):
        facts = assign_fact_ids(
            [
                self.make_fact(mapping_code="ZT_A", mapping_name="资产总计", row_label_std="资产总计", statement_type="balance_sheet", period_key="2022-12-31__期末数", value_num=300.0, value_raw="300"),
                self.make_fact(mapping_code="ZT_B", mapping_name="负债合计", row_label_std="负债合计", statement_type="balance_sheet", period_key="2022-12-31__期末数", value_num=100.0, value_raw="100"),
                self.make_fact(mapping_code="ZT_C", mapping_name="所有者权益合计", row_label_std="所有者权益合计", statement_type="balance_sheet", period_key="2022-12-31__期末数", value_num=200.0, value_raw="200"),
                self.make_fact(logical_subtable_id="sub1", column_semantic_key="金额", row_label_std="明细1", source_row_start=1, value_num=60.0, value_raw="60"),
                self.make_fact(logical_subtable_id="sub1", column_semantic_key="金额", row_label_std="明细2", source_row_start=2, value_num=40.0, value_raw="40"),
                self.make_fact(logical_subtable_id="sub1", column_semantic_key="金额", row_label_std="合计", source_row_start=3, value_num=100.0, value_raw="100"),
                self.make_fact(logical_subtable_id="sub2", column_semantic_key="比例", row_label_std="合计", source_row_start=1, value_num=1.0, value_raw="100%", value_type="ratio"),
                self.make_fact(row_label_std="异常金额", value_raw="务屏20,000,000.00", value_num=None, issue_flags=["numeric_parse_failed", "contains_chinese_noise"], status="review"),
            ]
        )
        validation_config = {
            "balance_equation": {
                "aliases": {
                    "assets_total": ["资产总计"],
                    "liabilities_total": ["负债合计"],
                    "equity_total": ["所有者权益合计", "股东权益合计"],
                },
                "tolerance": {"absolute": 1, "relative": 0.0001},
            },
            "subtotal_rules": {"min_detail_rows": 2, "tolerance": {"absolute": 1, "relative": 0.0001}},
            "ratio_rules": {"expected_total": 1.0, "tolerance": 0.02},
            "amount_legality": {"status_when_noise_detected": "review"},
        }

        results, summary = run_validation(facts, validation_config)

        rule_names = [result.rule_name for result in results]
        self.assertIn("balance_equation", rule_names)
        self.assertIn("subtotal_check", rule_names)
        self.assertIn("ratio_total_check", rule_names)
        self.assertIn("amount_legality", rule_names)
        self.assertGreater(summary["validation_pass_total"], 0)

    def test_mapping_candidate_and_relation_review(self):
        subjects = [
            TemplateSubject(code="ZT_001", canonical_name="货币资金", row_index=4, sheet_name="Sheet1", source_value="ZT_001 货币资金"),
            TemplateSubject(code="ZT_024", canonical_name="应交税费", row_index=5, sheet_name="Sheet1", source_value="ZT_024 应交税费"),
            TemplateSubject(code="ZT_006", canonical_name="应收票据及应收账款", row_index=6, sheet_name="Sheet1", source_value="ZT_006 应收票据及应收账款"),
        ]
        aliases = [
            AliasRecord(canonical_code="ZT_024", canonical_name="应交税费", alias="应交税金", alias_type="legacy_alias", enabled=True),
        ]
        relations = [
            RelationRecord(
                canonical_code="ZT_006",
                canonical_name="应收票据及应收账款",
                relation_type="aggregate_relation",
                related_codes=[],
                related_names=["应收票据和应收账款"],
                enabled=True,
                review_required=True,
            )
        ]
        facts = [
            self.make_fact(row_label_raw="应交税金", row_label_std="应交税金"),
            self.make_fact(row_label_raw="应收票据和应收账款", row_label_std="应收票据和应收账款"),
        ]

        facts, mapping_review, mapping_candidates, unmapped_summary, mapping_stats = apply_subject_mapping(
            facts,
            subjects,
            aliases,
            relations,
            {"max_candidates": 3},
        )

        self.assertEqual(facts[0].mapping_code, "ZT_024")
        self.assertEqual(facts[0].mapping_method, "legacy_alias")
        self.assertEqual(facts[1].mapping_code, "")
        self.assertTrue(any(candidate.relation_type == "aggregate_relation" for candidate in mapping_candidates))
        self.assertTrue(any(row.normalized_label for row in unmapped_summary))
        self.assertEqual(mapping_stats["mapped_by_alias"], 1)

    def test_validation_aware_conflict_resolution(self):
        facts = assign_fact_ids(
            [
                self.make_fact(
                    doc_id="demo",
                    page_no=1,
                    provider="aliyun_table",
                    statement_type="balance_sheet",
                    statement_name_raw="balance_sheet",
                    table_semantic_key="balance_sheet|h:item,end",
                    row_label_raw="assets_total",
                    row_label_std="assets_total",
                    mapping_code="ZT_A",
                    mapping_name="assets_total",
                    column_semantic_key="end",
                    col_header_raw="end",
                    col_header_path=["end"],
                    period_key="2022-12-31__end",
                    period_role_raw="end",
                    period_role_norm="end",
                    report_date_norm="2022-12-31",
                    value_raw="301",
                    value_num=301.0,
                ),
                self.make_fact(
                    doc_id="demo",
                    page_no=1,
                    provider="tencent_table_v3",
                    statement_type="balance_sheet",
                    statement_name_raw="balance_sheet",
                    table_semantic_key="balance_sheet|h:item,end",
                    row_label_raw="assets_total",
                    row_label_std="assets_total",
                    mapping_code="ZT_A",
                    mapping_name="assets_total",
                    column_semantic_key="end",
                    col_header_raw="end",
                    col_header_path=["end"],
                    period_key="2022-12-31__end",
                    period_role_raw="end",
                    period_role_norm="end",
                    report_date_norm="2022-12-31",
                    value_raw="300",
                    value_num=300.0,
                ),
                self.make_fact(
                    doc_id="demo",
                    page_no=1,
                    provider="aliyun_table",
                    statement_type="balance_sheet",
                    statement_name_raw="balance_sheet",
                    table_semantic_key="balance_sheet|h:item,end",
                    row_label_raw="liabilities_total",
                    row_label_std="liabilities_total",
                    mapping_code="ZT_B",
                    mapping_name="liabilities_total",
                    column_semantic_key="end",
                    col_header_raw="end",
                    col_header_path=["end"],
                    period_key="2022-12-31__end",
                    period_role_raw="end",
                    period_role_norm="end",
                    report_date_norm="2022-12-31",
                    value_raw="100",
                    value_num=100.0,
                ),
                self.make_fact(
                    doc_id="demo",
                    page_no=1,
                    provider="aliyun_table",
                    statement_type="balance_sheet",
                    statement_name_raw="balance_sheet",
                    table_semantic_key="balance_sheet|h:item,end",
                    row_label_raw="equity_total",
                    row_label_std="equity_total",
                    mapping_code="ZT_C",
                    mapping_name="equity_total",
                    column_semantic_key="end",
                    col_header_raw="end",
                    col_header_path=["end"],
                    period_key="2022-12-31__end",
                    period_role_raw="end",
                    period_role_norm="end",
                    report_date_norm="2022-12-31",
                    value_raw="200",
                    value_num=200.0,
                ),
            ]
        )
        resolved, conflicts, _ = resolve_conflicts(facts, ["aliyun_table", "tencent_table_v3"], enabled=False)
        resolved, enriched, audits, impacts = enrich_conflicts(
            facts=resolved,
            conflicts=conflicts,
            provider_priority=["aliyun_table", "tencent_table_v3"],
            validation_config={
                "balance_equation": {
                    "aliases": {
                        "assets_total": ["assets_total"],
                        "liabilities_total": ["liabilities_total"],
                        "equity_total": ["equity_total"],
                    },
                    "tolerance": {"absolute": 0.001, "relative": 0.000001},
                },
                "subtotal_rules": {"min_detail_rows": 2, "tolerance": {"absolute": 1, "relative": 0.0001}},
                "ratio_rules": {"expected_total": 1.0, "tolerance": 0.02},
                "amount_legality": {"status_when_noise_detected": "review"},
            },
            conflict_config={"magnitude_ratio_review_threshold": 10, "magnitude_ratio_force_review_threshold": 100},
            merge_enabled=True,
            validation_aware_enabled=True,
        )

        self.assertEqual(enriched[0].decision, "accepted_with_validation_support")
        self.assertEqual(enriched[0].accepted_provider, "tencent_table_v3")
        self.assertTrue(audits)
        self.assertTrue(impacts)

    def test_large_magnitude_conflict_requires_review(self):
        facts = assign_fact_ids(
            [
                self.make_fact(
                    doc_id="demo",
                    page_no=17,
                    provider="aliyun_table",
                    statement_type="note",
                    table_semantic_key="note|h:项目,本期发生额",
                    row_label_raw="财产保险费",
                    row_label_std="财产保险费",
                    column_semantic_key="本期发生额",
                    col_header_raw="本期发生额",
                    col_header_path=["本期发生额"],
                    period_key="2022年度__本期",
                    period_role_raw="本期",
                    period_role_norm="本期",
                    report_date_norm="2022年度",
                    value_raw="85533955",
                    value_num=85533955.0,
                ),
                self.make_fact(
                    doc_id="demo",
                    page_no=17,
                    provider="tencent_table_v3",
                    statement_type="note",
                    table_semantic_key="note|h:项目,本期发生额",
                    row_label_raw="财产保险费",
                    row_label_std="财产保险费",
                    column_semantic_key="本期发生额",
                    col_header_raw="本期发生额",
                    col_header_path=["本期发生额"],
                    period_key="2022年度__本期",
                    period_role_raw="本期",
                    period_role_norm="本期",
                    report_date_norm="2022年度",
                    value_raw="85,334.55",
                    value_num=85334.55,
                ),
            ]
        )
        resolved, conflicts, _ = resolve_conflicts(facts, ["aliyun_table", "tencent_table_v3"], enabled=False)
        resolved, enriched, audits, impacts = enrich_conflicts(
            facts=resolved,
            conflicts=conflicts,
            provider_priority=["aliyun_table", "tencent_table_v3"],
            validation_config={"balance_equation": {}, "subtotal_rules": {}, "ratio_rules": {}, "amount_legality": {"status_when_noise_detected": "review"}},
            conflict_config={"magnitude_ratio_review_threshold": 10, "magnitude_ratio_force_review_threshold": 100},
            merge_enabled=True,
            validation_aware_enabled=True,
        )
        self.assertEqual(enriched[0].decision, "review_required")
        self.assertGreaterEqual(enriched[0].magnitude_ratio, 100)

    def test_export_and_integrity_contract(self):
        template_path = Path(__file__).resolve().parent.parent / "会计报表.xlsx"
        with tempfile.TemporaryDirectory() as tmpdir:
            output_dir = Path(tmpdir)
            facts = assign_fact_ids(
                [
                    self.make_fact(
                        mapping_code="ZT_001",
                        mapping_name="货币资金",
                        row_label_std="货币资金",
                        period_key="2022-12-31__期末数",
                        report_date_norm="2022-12-31",
                        period_role_norm="期末数",
                        value_raw="100",
                        value_num=100.0,
                    ),
                    self.make_fact(
                        mapping_code="ZT_001",
                        mapping_name="货币资金",
                        row_label_std="货币资金",
                        period_key="unknown_date__期末数",
                        report_date_norm="unknown_date",
                        period_role_norm="期末数",
                        value_raw="100",
                        value_num=100.0,
                    ),
                ]
            )
            export_stats = export_template(
                template_path=template_path,
                output_path=output_dir / "会计报表_填充结果.xlsx",
                facts=facts,
                run_summary={"unknown_date_total": 0},
                issues=[],
                validations=[],
                duplicates=[],
                conflicts=[],
                review_queue=[],
                export_rules={"allowed_statuses": ["observed", "repaired"]},
            )
            write_rows = [
                {
                    "fact_id": fact.fact_id,
                    "mapping_code": fact.mapping_code,
                    "period_key": fact.period_key,
                    "status": fact.status,
                    "conflict_decision": fact.conflict_decision,
                    "report_date_norm": fact.report_date_norm,
                    "period_role_norm": fact.period_role_norm,
                    "value_num": fact.value_num,
                    "mapping_review_required": fact.mapping_review_required,
                }
                for fact in facts
            ]
            with (output_dir / "facts_deduped.csv").open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=list(write_rows[0].keys()))
                writer.writeheader()
                writer.writerows(write_rows)
            with (output_dir / "conflicts_enriched.csv").open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=["conflict_id", "decision", "accepted_fact_id", "provider_values_json"])
                writer.writeheader()

            integrity = run_artifact_integrity(
                output_dir=output_dir,
                workbook_path=output_dir / "会计报表_填充结果.xlsx",
                run_summary={"unknown_date_total": 0},
                export_stats=export_stats,
                export_rules={"required_helper_sheets": ["_meta_summary", "_issues", "_validation", "_duplicates", "_conflicts", "_unplaced_facts", "_review_queue"]},
            )

            workbook = load_workbook(output_dir / "会计报表_填充结果.xlsx")
            headers = [workbook[workbook.sheetnames[0]].cell(row=3, column=idx).value for idx in range(1, workbook[workbook.sheetnames[0]].max_column + 1)]
            self.assertIn("2022-12-31__期末数", headers)
            self.assertNotIn("unknown_date__期末数", headers)
            for sheet_name in ["_meta_summary", "_validation", "_duplicates", "_conflicts", "_unplaced_facts", "_review_queue"]:
                self.assertIn(sheet_name, workbook.sheetnames)
            self.assertEqual(integrity["summary"]["integrity_fail_total"], 0)

    def test_review_queue_generates_crops(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            source_dir = base / "data"
            source_dir.mkdir(parents=True, exist_ok=True)
            image = Image.new("RGB", (300, 300), "white")
            draw = ImageDraw.Draw(image)
            draw.rectangle((50, 50, 200, 120), outline="black", width=3)
            pdf_path = source_dir / "demo.pdf"
            image.save(pdf_path, "PDF")

            bbox = json.dumps([{"x": 50, "y": 50}, {"x": 200, "y": 50}, {"x": 200, "y": 120}, {"x": 50, "y": 120}], ensure_ascii=False)
            cell = CellRecord(
                doc_id="demo",
                page_no=1,
                provider="aliyun_table",
                source_file="page_0001.json",
                table_id="1",
                logical_subtable_id="1_sub1",
                row_start=1,
                row_end=1,
                col_start=1,
                col_end=1,
                bbox_json=bbox,
                text_raw="务屏20,000,000.00",
                text_clean="务屏20,000,000.00",
                ocr_conf=None,
                is_empty=False,
                is_header=False,
                is_suspicious=True,
                suspicious_reason="contains_chinese_noise",
                repair_status="raw",
                meta_json="",
            )
            fact = assign_fact_ids(
                [
                    self.make_fact(
                        doc_id="demo",
                        page_no=1,
                        provider="aliyun_table",
                        source_cell_ref="demo:1:aliyun_table:1:1-1:1-1",
                        row_label_raw="异常金额",
                        row_label_std="异常金额",
                        mapping_code="",
                        value_raw="务屏20,000,000.00",
                        value_num=None,
                        issue_flags=["numeric_parse_failed", "contains_chinese_noise"],
                        status="review",
                    )
                ]
            )
            validation = [
                self.make_validation(
                    validation_id="VAL000001",
                    doc_id="demo",
                    period_key="2022年度__本期",
                    rule_name="amount_legality",
                    status="review",
                    evidence_fact_refs=["demo:1:aliyun_table:1:1-1:1-1"],
                )
            ]

            review_items, summary = build_review_queue(
                facts=fact,
                cells=[cell],
                issues=[],
                conflicts=[],
                validations=validation,
                mapping_candidates=[],
                source_image_dir=source_dir,
                output_dir=base / "normalized",
                review_config={"crop_padding": 8, "reason_weights": {"validation": 3.0, "mapping": 2.0, "quality": 2.0}},
                generate_evidence=True,
            )

            self.assertTrue(review_items)
            self.assertTrue((base / "normalized" / "review_pack" / "index.csv").exists())
            self.assertTrue(any(item.evidence_cell_path for item in review_items))

    def test_page_selector(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            image_dir = base / "images"
            input_dir = base / "outputs"
            image_dir.mkdir(parents=True, exist_ok=True)
            input_dir.mkdir(parents=True, exist_ok=True)

            table_image = Image.new("RGB", (800, 1000), "white")
            draw = ImageDraw.Draw(table_image)
            for x in range(50, 750, 100):
                draw.line((x, 100, x, 900), fill="black", width=3)
            for y in range(100, 900, 80):
                draw.line((50, y, 750, y), fill="black", width=3)
            table_path = image_dir / "page_0001.png"
            table_image.save(table_path)

            text_image = Image.new("RGB", (800, 1000), "white")
            draw = ImageDraw.Draw(text_image)
            for idx in range(15):
                draw.text((80, 80 + idx * 45), f"plain text paragraph line {idx}", fill="black")
            text_path = image_dir / "page_0002.png"
            text_image.save(text_path)

            records, plan = build_page_selection(
                source_image_dir=image_dir,
                input_dir=input_dir,
                routing_config={
                    "pre_ocr": {
                        "selection_threshold": 0.45,
                        "strong_horizontal_ratio": 0.35,
                        "strong_vertical_ratio": 0.18,
                        "table_likelihood_trigger": 0.45,
                        "line_density_trigger": 0.45,
                        "numeric_density_trigger": 0.35,
                        "hard_keyword_hits": 2,
                        "weights": {
                            "table_likelihood_score": 0.4,
                            "numeric_density_score": 0.2,
                            "line_density_score": 0.3,
                            "keyword_score": 0.1,
                        },
                        "keywords": [],
                    }
                },
            )

            by_page = {record.page_no: record for record in records}
            self.assertTrue(by_page[1].is_candidate_table_page)
            self.assertFalse(by_page[2].is_candidate_table_page)
            self.assertEqual(plan["pages_total"], 2)

    def test_review_action_parsing_and_application(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            config_dir = base / "config"
            ensure_override_store(config_dir)
            action_csv = base / "review_actions.csv"
            rows = [
                {
                    "review_id": "REV_demo_1",
                    "action_type": "accept_mapping_alias",
                    "candidate_mapping_code": "ZT_001",
                    "candidate_mapping_name": "货币资金",
                    "row_label_std": "货币资金",
                    "row_label_raw": "货币资金",
                    "action_value": "",
                    "reviewer_note": "accepted",
                    "reviewer_name": "tester",
                },
                {
                    "review_id": "REV_demo_1",
                    "action_type": "unsupported_action",
                    "candidate_mapping_code": "",
                    "candidate_mapping_name": "",
                    "row_label_std": "货币资金",
                    "row_label_raw": "货币资金",
                    "action_value": "",
                    "reviewer_note": "",
                    "reviewer_name": "",
                },
            ]
            with action_csv.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
                writer.writeheader()
                writer.writerows(rows)

            parsed = parse_review_actions_file(action_csv)
            applied, rejected, audit_rows, summary = apply_review_actions(parsed, ["REV_demo_1"], config_dir)

            self.assertEqual(len(applied), 1)
            self.assertEqual(len(rejected), 1)
            self.assertEqual(summary["applied_total"], 1)
            self.assertTrue((config_dir / "manual_overrides" / "mapping_overrides.yml").exists())
            self.assertEqual(audit_rows[0]["action_type"], "accept_mapping_alias")

    def test_period_override_and_suppression_affect_fact_state(self):
        facts = assign_fact_ids(
            [
                self.make_fact(
                    period_key="2022年度__本期",
                    report_date_norm="2022年度",
                    period_role_norm="本期",
                    value_raw="100",
                    value_num=100.0,
                )
            ]
        )
        fact_id = facts[0].fact_id
        facts = apply_period_overrides(
            facts,
            [{"fact_id": fact_id, "period_key": "2022-12-31__期末数", "report_date_norm": "2022-12-31", "period_role_norm": "期末数", "note": "manual"}],
        )
        facts = apply_suppression_overrides(
            facts,
            [{"fact_id": fact_id, "action_type": "suppress_false_positive", "note": "false_positive"}],
        )
        self.assertEqual(facts[0].period_key, "2022-12-31__期末数")
        self.assertEqual(facts[0].report_date_norm, "2022-12-31")
        self.assertEqual(facts[0].status, "suppressed")

    def test_delta_reporting(self):
        before = {
            "run_summary": {"mapped_facts_ratio": 0.1, "amount_coverage_ratio": 0.2, "review_total": 10, "validation_fail_total": 3, "provider_conflict_pairs": 2},
            "review_rows": [{"review_id": "REV1", "priority_score": "5", "row_label_std": "货币资金", "period_key": "2022-12-31__期末数"}],
            "unmapped_rows": [{"row_label_std": "未知科目", "occurrences": "3", "amount_abs_total": "100"}],
            "reocr_rows": [{"task_id": "REOCR1"}],
            "conflict_rows": [{"decision": "review_required"}],
            "facts_rows": [{"mapping_code": "ZT_001", "value_num": "1", "report_date_norm": "2022-12-31", "period_role_norm": "期末数", "status": "observed", "conflict_decision": "", "unplaced_reason": ""}],
            "unplaced_rows": [{"fact_id": "F1"}],
        }
        after = {
            "run_summary": {"mapped_facts_ratio": 0.2, "amount_coverage_ratio": 0.25, "review_total": 4, "validation_fail_total": 1, "provider_conflict_pairs": 1},
            "review_rows": [],
            "unmapped_rows": [{"row_label_std": "未知科目", "occurrences": "1", "amount_abs_total": "10"}],
            "reocr_rows": [],
            "conflict_rows": [],
            "facts_rows": [{"mapping_code": "ZT_001", "value_num": "1", "report_date_norm": "2022-12-31", "period_role_norm": "期末数", "status": "observed", "conflict_decision": "", "unplaced_reason": ""}],
            "unplaced_rows": [],
        }
        payload = build_delta_reports(before, after)
        metric_map = {row["metric"]: row for row in payload["coverage_rows"]}
        self.assertEqual(metric_map["mapped_facts_ratio"]["delta"], 0.1)
        self.assertEqual(metric_map["review_total"]["delta"], -6.0)
        self.assertTrue(any(row["status"] == "resolved" for row in payload["review_delta_rows"]))

    def test_materialize_reocr_inputs(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            source_dir = base / "data"
            output_dir = base / "normalized"
            source_dir.mkdir(parents=True, exist_ok=True)
            image = Image.new("RGB", (300, 300), "white")
            draw = ImageDraw.Draw(image)
            draw.rectangle((50, 50, 200, 120), outline="black", width=3)
            (source_dir / "demo.pdf").parent.mkdir(parents=True, exist_ok=True)
            image.save(source_dir / "demo.pdf", "PDF")
            task = ReOCRTaskRecord(
                task_id="REOCR_demo",
                granularity="cell",
                doc_id="demo",
                page_no=1,
                table_id="1",
                logical_subtable_id="1_sub1",
                bbox=json.dumps([50, 50, 200, 120], ensure_ascii=False),
                reason_codes=["quality:suspicious_numeric"],
                suggested_provider="tencent_table_v3",
                priority_score=5.0,
                expected_benefit="improve_numeric_legibility",
                source_review_id="REV_demo",
                meta_json="",
            )

            manifest_rows, summary = materialize_reocr_inputs([task], source_dir, output_dir)

            self.assertEqual(summary["materialized_total"], 1)
            self.assertTrue(Path(manifest_rows[0]["crop_path"]).exists())

    def test_export_writes_applied_actions_sheet(self):
        repo_root = Path(__file__).resolve().parent
        template_path = repo_root.parent / "会计报表.xlsx"
        with tempfile.TemporaryDirectory() as tmpdir:
            output_dir = Path(tmpdir)
            facts = assign_fact_ids(
                [
                    self.make_fact(
                        mapping_code="ZT_001",
                        mapping_name="货币资金",
                        row_label_std="货币资金",
                        period_key="2022-12-31__期末数",
                        report_date_norm="2022-12-31",
                        period_role_norm="期末数",
                        value_raw="100",
                        value_num=100.0,
                    )
                ]
            )
            export_template(
                template_path=template_path,
                output_path=output_dir / "会计报表_填充结果.xlsx",
                facts=facts,
                run_summary={"unknown_date_total": 0},
                issues=[],
                validations=[],
                duplicates=[],
                conflicts=[],
                review_queue=[],
                applied_actions=[{"action_id": "ACT_1", "review_id": "REV_1", "action_type": "accept_mapping_alias"}],
                export_rules={"allowed_statuses": ["observed", "repaired"], "required_helper_sheets": ["_applied_actions"]},
            )
            workbook = load_workbook(output_dir / "会计报表_填充结果.xlsx")
            self.assertIn("_applied_actions", workbook.sheetnames)
            self.assertEqual(workbook["_applied_actions"].cell(row=1, column=1).value, "action_id")

    def test_label_canonicalization(self):
        facts = [
            self.make_fact(statement_type="income_statement", row_label_raw="一、主营业务收入", row_label_std="一、主营业务收入"),
            self.make_fact(statement_type="income_statement", row_label_raw="减:主营业务成本", row_label_std="减:主营业务成本"),
            self.make_fact(statement_type="note", row_label_raw="其中:应付利息", row_label_std="其中:应付利息"),
        ]
        facts, audit_rows, summary = apply_label_canonicalization(
            facts,
            {
                "general_prefixes": ["加:", "减:", "其中:"],
                "statement_rules": {
                    "income_statement": {"prefixes": ["加:", "减:", "其中:"], "synonyms": {"主营业务收入": "营业收入", "主营业务成本": "营业成本"}},
                    "note": {"prefixes": ["其中:"], "synonyms": {}},
                },
            },
            enabled=True,
        )
        self.assertEqual(facts[0].row_label_norm, "主营业务收入")
        self.assertEqual(facts[0].row_label_canonical_candidate, "营业收入")
        self.assertEqual(facts[1].row_label_norm, "主营业务成本")
        self.assertEqual(facts[2].row_label_norm, "应付利息")
        self.assertTrue(facts[0].normalization_rule_ids)
        self.assertEqual(summary["rows_changed"], 3)
        self.assertEqual(len(audit_rows), 3)

    def test_stage6_statement_specialization(self):
        facts = assign_fact_ids(
            [
                self.make_fact(doc_id="demo", page_no=5, statement_type="unknown", row_label_raw="一、主营业务收入", row_label_std="一、主营业务收入", row_label_norm="主营业务收入", row_label_canonical_candidate="营业收入"),
                self.make_fact(doc_id="demo", page_no=5, statement_type="unknown", row_label_raw="减：主营业务成本", row_label_std="减：主营业务成本", row_label_norm="主营业务成本", row_label_canonical_candidate="营业成本"),
                self.make_fact(doc_id="demo", page_no=5, statement_type="unknown", row_label_raw="四、利润总额", row_label_std="四、利润总额", row_label_norm="利润总额", row_label_canonical_candidate="利润总额"),
            ]
        )
        pages = [
            ProviderPage(
                doc_id="demo",
                page_no=5,
                provider="aliyun_table",
                source_file="page5.json",
                source_kind="json",
                page_text="利润及利润分配表 2022年度",
                tables={},
                context_lines=["利润及利润分配表", "2022年度"],
            )
        ]
        facts, audit_rows, summary = specialize_statement_types(
            facts,
            pages,
            {
                "title_keywords": {"income_statement": ["利润及利润分配表"], "note": ["财务报表附注"]},
                "row_patterns": {"income_statement": ["营业收入", "营业成本", "利润总额", "净利润", "主营业务收入", "主营业务成本"]},
                "header_signatures": {},
                "note_titles": ["财务报表附注"],
                "note_detail_markers": [],
                "main_statement_numbering": {"income_statement_prefixes": ["一", "二", "三", "四", "五"], "cash_flow_prefixes": ["一", "二", "三", "四", "五"]},
                "classification_min_score": 12,
                "classification_margin": 4,
            },
            enabled=True,
        )
        self.assertTrue(all(fact.statement_type == "income_statement" for fact in facts))
        self.assertGreater(summary["unknown_statement_type_total_before"], summary["unknown_statement_type_total_after"])
        self.assertTrue(audit_rows)

    def test_single_period_annual_role_inference(self):
        facts = [
            self.make_fact(
                statement_type="income_statement",
                logical_subtable_id="5_sub1",
                mapping_code="ZT_138",
                mapping_name="营业收入",
                report_date_norm="2022年度",
                period_key="2022年度__unknown",
                period_role_norm="unknown",
                period_role_raw="unknown",
                col_header_raw="金额",
                column_semantic_key="金额",
                col_header_path=["金额"],
                value_num=100.0,
            )
        ]
        facts, audit_rows, summary = resolve_single_period_annual_roles(
            facts,
            {"statement_types": ["income_statement", "cash_flow"], "generic_headers": ["金额"], "inferred_role": "本期"},
            enabled=True,
        )
        self.assertEqual(facts[0].period_key, "2022年度__本期")
        self.assertEqual(facts[0].period_role_norm, "本期")
        self.assertTrue(audit_rows)
        self.assertGreater(summary["unknown_period_role_export_blocking_total_before"], summary["unknown_period_role_export_blocking_total_after"])

    def test_stage6_unmapped_split_and_alias_ranking(self):
        facts = [
            self.make_fact(
                fact_id="F1",
                statement_type="income_statement",
                row_label_raw="一、主营业务收入",
                row_label_std="一、主营业务收入",
                row_label_norm="主营业务收入",
                row_label_canonical_candidate="营业收入",
                period_key="2022年度__本期",
                value_num=100.0,
                mapping_code="",
                source_cell_ref="demo:1:aliyun_table:1:1-1:1-1",
            ),
            self.make_fact(
                fact_id="F2",
                statement_type="income_statement",
                row_label_raw="说明",
                row_label_std="说明",
                row_label_norm="说明",
                row_label_canonical_candidate="说明",
                value_num=None,
                mapping_code="",
                source_cell_ref="demo:1:aliyun_table:1:2-2:1-1",
            ),
        ]
        value_rows, blank_rows, summary = split_unmapped_facts(facts)
        self.assertEqual(len(value_rows), 1)
        self.assertEqual(len(blank_rows), 1)
        candidates, candidate_summary = build_alias_acceptance_candidates(
            value_bearing_rows=value_rows,
            facts=facts,
            mapping_candidates=[
                type("Candidate", (), {
                    "source_cell_ref": "demo:1:aliyun_table:1:1-1:1-1",
                    "candidate_rank": 1,
                    "candidate_score": 1.0,
                    "candidate_code": "ZT_138",
                    "candidate_name": "营业收入",
                    "candidate_method": "exact_normalized_match",
                })()
            ],
            benchmark_missing_rows=[
                {
                    "mapping_code": "ZT_138",
                    "aligned_period_key": "2022年度__本期",
                    "benchmark_value": 100.0,
                }
            ],
            alias_rules={"safe_methods": ["exact_normalized_match"], "safe_min_evidence_count": 1},
        )
        self.assertTrue(candidates)
        self.assertTrue(candidates[0]["safe_to_auto_accept"])
        self.assertEqual(candidate_summary["safe_to_auto_accept_total"], 1)

    def test_stage6_formula_impact_and_pruning(self):
        derived_facts = [
            self.make_fact(
                fact_id="DF_1",
                mapping_code="ZT_006",
                mapping_name="应收票据及应收账款",
                statement_type="balance_sheet",
                period_key="2022-12-31__期末数",
                value_num=100.0,
                source_kind="derived_formula",
                source_cell_ref="derived:ZT_006:2022-12-31__期末数:sum_ar_notes_and_receivables",
                status="derived_resolved",
                unplaced_reason="",
            )
        ]
        impact_summary, placements = build_formula_rule_impact(derived_facts, [])
        self.assertEqual(impact_summary["rule_impact"]["sum_ar_notes_and_receivables"]["newly_exportable_facts"], 1)
        review_item = ReviewQueueRecord(
            review_id="REV_1",
            priority_score=5.0,
            reason_codes=["mapping:unmapped"],
            doc_id="demo",
            page_no=1,
            statement_type="income_statement",
            row_label_raw="主营业务收入",
            row_label_std="主营业务收入",
            period_key="2022年度__本期",
            value_raw="100",
            value_num=100.0,
            provider="aliyun_table",
            source_file="page1.json",
            bbox="",
            related_fact_ids=["F1"],
            related_conflict_ids=[],
            related_validation_ids=[],
            mapping_candidates="",
            evidence_cell_path="",
            evidence_row_path="",
            evidence_table_path="",
            meta_json="",
        )
        task = ReOCRTaskRecord(
            task_id="REOCR_1",
            granularity="row",
            doc_id="demo",
            page_no=1,
            table_id="1",
            logical_subtable_id="1_sub1",
            bbox="",
            reason_codes=["mapping:unmapped"],
            suggested_provider="tencent_table_v3",
            priority_score=5.0,
            expected_benefit="improve_mapping_readability",
            source_review_id="REV_1",
            meta_json="",
        )
        pruned_rows, pruned_summary = prune_reocr_tasks([task], [review_item], {})
        self.assertEqual(pruned_rows, [])
        self.assertEqual(pruned_summary["dropped_mapping_only_total"], 1)

    def test_full_run_contract_benchmark_missing_outputs_fails(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            workbook_path = base / "wb.xlsx"
            template_path = Path(__file__).resolve().parent.parent / "会计报表.xlsx"
            shutil.copy2(template_path, workbook_path)
            summary = run_full_run_contract(
                output_dir=base,
                workbook_path=workbook_path,
                run_id="RUN_TEST",
                feature_flags={"emit_benchmark_report": True, "enable_derived_facts": False, "emit_run_manifest": False},
                export_stats={"source_facts": "facts_deduped"},
                required_helper_sheets=["_meta_summary"],
            )
            self.assertGreater(summary["contract_fail_total"], 0)

    def test_derived_formula_and_conflict(self):
        facts = assign_fact_ids(
            [
                self.make_fact(mapping_code="ZT_007", mapping_name="应收票据", row_label_std="应收票据", period_key="2022-12-31__期末数", report_date_norm="2022-12-31", period_role_norm="期末数", value_num=30.0, value_raw="30", statement_type="balance_sheet"),
                self.make_fact(mapping_code="ZT_008", mapping_name="应收账款", row_label_std="应收账款", period_key="2022-12-31__期末数", report_date_norm="2022-12-31", period_role_norm="期末数", value_num=70.0, value_raw="70", statement_type="balance_sheet"),
                self.make_fact(mapping_code="ZT_009", mapping_name="应收票据及应收账款", row_label_std="应收票据及应收账款", period_key="2022-12-31__期末数", report_date_norm="2022-12-31", period_role_norm="期末数", value_num=120.0, value_raw="120", statement_type="balance_sheet"),
            ]
        )
        derived, audit_rows, summary, conflicts = derive_formula_facts(
            facts=facts,
            formula_rules={
                "rules": [
                    {
                        "rule_id": "sum_receivables",
                        "target_code": "ZT_009",
                        "target_name": "应收票据及应收账款",
                        "rule_type": "sum",
                        "children": ["ZT_007", "ZT_008"],
                        "statement_types": ["balance_sheet"],
                        "enabled": True,
                    }
                ]
            },
            relation_records=[],
            enabled=True,
        )
        self.assertEqual(summary["derived_facts_total"], 1)
        self.assertEqual(derived[0].value_num, 100.0)
        self.assertEqual(derived[0].source_kind, "derived_formula")
        self.assertTrue(conflicts)
        self.assertEqual(conflicts[0]["decision"], "prefer_observed")
        self.assertTrue(audit_rows)

    def test_benchmark_compare_and_gap_mining(self):
        repo_root = Path(__file__).resolve().parent
        template_path = repo_root.parent / "会计报表.xlsx"
        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir_path = Path(tmpdir)
            export_path = tmpdir_path / "auto.xlsx"
            benchmark_path = tmpdir_path / "benchmark.xlsx"
            facts = assign_fact_ids(
                [
                    self.make_fact(mapping_code="ZT_001", mapping_name="货币资金", row_label_std="货币资金", period_key="2022-12-31__期末数", report_date_norm="2022-12-31", period_role_norm="期末数", value_num=100.0, value_raw="100", statement_type="balance_sheet"),
                ]
            )
            export_template(
                template_path=template_path,
                output_path=export_path,
                facts=facts,
                derived_facts=[],
                run_summary={"run_id": "RUN_TEST", "unknown_date_total": 0},
                issues=[],
                validations=[],
                duplicates=[],
                conflicts=[],
                review_queue=[],
                applied_actions=[],
                export_rules={"allowed_statuses": ["observed", "repaired", "derived_resolved"]},
            )
            shutil.copy2(template_path, benchmark_path)
            wb = load_workbook(benchmark_path)
            ws = wb[wb.sheetnames[0]]
            ws.cell(row=4, column=4, value=100.0)
            ws.cell(row=8, column=4, value=50.0)
            wb.save(benchmark_path)

            benchmark_payload = compare_benchmark_workbook(benchmark_path, export_path, {"numeric_tolerance": 0.01})
            self.assertGreaterEqual(benchmark_payload["summary"]["matched_cells"], 1)
            self.assertTrue(any(row["status"] == "missing_in_auto" for row in benchmark_payload["cell_rows"]))

            unmapped_fact = self.make_fact(
                mapping_code="",
                mapping_name="",
                row_label_raw="一、主营业务收入",
                row_label_std="一、主营业务收入",
                row_label_norm="主营业务收入",
                row_label_canonical_candidate="营业收入",
                period_key="2022-12-31__期末数",
                report_date_norm="2022-12-31",
                period_role_norm="期末数",
                value_num=50.0,
                value_raw="50",
                statement_type="balance_sheet",
            )
            gap_payload = explain_benchmark_gaps(
                benchmark_missing_rows=benchmark_payload["missing_rows"],
                facts=[unmapped_fact],
                unplaced_rows=[],
                conflicts=[],
                validations=[],
                mapping_candidates=[
                    type("Candidate", (), {"candidate_code": "ZT_005", "source_cell_ref": unmapped_fact.source_cell_ref})()
                ],
                derived_facts=[],
            )
            self.assertTrue(gap_payload["summary"]["gaps_total"] >= 1)

    def test_stage7_benchmark_alignment_repair(self):
        facts = [
            self.make_fact(
                mapping_code="ZT_138",
                mapping_name="营业收入",
                statement_type="income_statement",
                period_key="2022年度__本期",
                report_date_norm="2022年度",
                period_role_norm="本期",
                value_num=100.0,
                value_raw="100",
            )
        ]
        payload = {
            "summary": {"missing_in_auto": 3},
            "cell_rows": [
                {
                    "mapping_code": "ZT_138",
                    "mapping_name": "营业收入",
                    "benchmark_header": "金额",
                    "aligned_period_key": "",
                    "benchmark_value": 100.0,
                    "auto_value": "",
                    "status": "missing_in_auto",
                    "reason": "legacy_header_unsupported",
                },
                {
                    "mapping_code": "ZT_014",
                    "mapping_name": "其他应收款项",
                    "benchmark_header": "期末",
                    "aligned_period_key": "2022-12-31__期末数",
                    "benchmark_value": 200.0,
                    "auto_value": "",
                    "status": "missing_in_auto",
                    "reason": "legacy_role_exact_date_match",
                },
                {
                    "mapping_code": "ZT_175",
                    "mapping_name": "营业外收入",
                    "benchmark_header": "金额",
                    "aligned_period_key": "",
                    "benchmark_value": 50.0,
                    "auto_value": "",
                    "status": "missing_in_auto",
                    "reason": "legacy_header_unsupported",
                },
            ],
            "export_period_headers": ["2022年度__本期", "2023年度__本期", "2022-12-31__期末数"],
            "export_rows_map": {
                "ZT_138": {"mapping_code": "ZT_138", "values": {"2022年度__本期": 100.0}},
                "ZT_014": {"mapping_code": "ZT_014", "values": {"2022-12-31__期末数": ""}},
            },
        }
        repaired = repair_benchmark_alignment(payload, facts, {"legacy_amount_headers": ["金额"], "annual_amount_statement_types": ["income_statement", "cash_flow"]})
        income_row = next(row for row in repaired["cell_rows"] if row["mapping_code"] == "ZT_138")
        self.assertEqual(income_row["aligned_period_key"], "2022年度__本期")
        self.assertEqual(income_row["status"], "match")
        self.assertEqual(repaired["summary"]["missing_in_auto_true"], 1)
        self.assertEqual(repaired["summary"]["alignment_only_gap_total"], 1)
        self.assertEqual(repaired["summary"]["ambiguous_alignment_total"], 1)

    def test_stage7_target_scoping_and_backlog_split(self):
        facts, scope_rows, summary = scope_facts_to_targets(
            facts=[
                self.make_fact(
                    fact_id="F_MAIN",
                    statement_type="income_statement",
                    mapping_code="ZT_138",
                    mapping_name="营业收入",
                    row_label_raw="一、主营业务收入",
                    row_label_std="主营业务收入",
                    row_label_norm="主营业务收入",
                    row_label_canonical_candidate="营业收入",
                ),
                self.make_fact(
                    fact_id="F_NOTE",
                    statement_type="note",
                    mapping_code="",
                    mapping_name="",
                    row_label_raw="1年以内",
                    row_label_std="1年以内",
                    row_label_norm="1年以内",
                    row_label_canonical_candidate="1年以内",
                ),
            ],
            benchmark_payload={"rows": [{"mapping_code": "ZT_138"}]},
            rules={
                "main_statement_types": ["income_statement", "balance_sheet", "cash_flow", "changes_in_equity"],
                "note_detail_patterns": {"aging_bucket": ["1年以内"], "company_like": [], "note_detail_headers": []},
                "note_aggregation_keywords": ["合计"],
                "promoted_rules": [],
            },
        )
        self.assertEqual(facts[0].target_scope, "main_export_target")
        self.assertEqual(facts[1].target_scope, "note_detail")
        self.assertEqual(summary["scope_breakdown"]["main_export_target"], 1)
        review_items = [
            ReviewQueueRecord(
                review_id="REV_MAIN",
                priority_score=5.0,
                reason_codes=["mapping:unmapped"],
                doc_id="demo",
                page_no=1,
                statement_type="income_statement",
                row_label_raw="主营业务收入",
                row_label_std="主营业务收入",
                period_key="2022年度__本期",
                value_raw="100",
                value_num=100.0,
                provider="aliyun_table",
                source_file="page1.json",
                bbox="",
                related_fact_ids=["F_MAIN"],
                related_conflict_ids=[],
                related_validation_ids=[],
                mapping_candidates="",
                evidence_cell_path="",
                evidence_row_path="",
                evidence_table_path="",
                meta_json="",
            ),
            ReviewQueueRecord(
                review_id="REV_NOTE",
                priority_score=1.0,
                reason_codes=["mapping:unmapped"],
                doc_id="demo",
                page_no=2,
                statement_type="note",
                row_label_raw="1年以内",
                row_label_std="1年以内",
                period_key="2022-12-31__期末数",
                value_raw="",
                value_num=None,
                provider="aliyun_table",
                source_file="page2.json",
                bbox="",
                related_fact_ids=["F_NOTE"],
                related_conflict_ids=[],
                related_validation_ids=[],
                mapping_candidates="",
                evidence_cell_path="",
                evidence_row_path="",
                evidence_table_path="",
                meta_json="",
            ),
        ]
        main_rows, note_rows, suppressed_rows, backlog_summary = build_target_review_backlogs(review_items, facts)
        self.assertEqual(len(main_rows), 1)
        self.assertEqual(len(note_rows), 1)
        self.assertEqual(len(suppressed_rows), 1)
        self.assertEqual(backlog_summary["main_target_review_total"], 1)

    def test_stage7_promotion_workflow(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            config_dir = Path(tmpdir)
            (config_dir / "curated_alias_pack.yml").write_text("aliases: []\n", encoding="utf-8")
            (config_dir / "curated_formula_pack.yml").write_text("rules: []\n", encoding="utf-8")
            (config_dir / "target_scope_rules.yml").write_text("promoted_rules: []\n", encoding="utf-8")
            rows = [
                {
                    "promotion_id": "PROM_1",
                    "candidate_alias": "固定资产原价",
                    "canonical_code": "ZT_046",
                    "canonical_name": "固定资产原值",
                    "statement_type": "balance_sheet",
                    "action_type": "promote_alias",
                    "action_value": "",
                },
                {
                    "promotion_id": "PROM_2",
                    "rule_id": "sum_receivables",
                    "formula_payload_json": json.dumps(
                        {
                            "rule_id": "sum_receivables",
                            "rule_type": "sum",
                            "target_code": "ZT_009",
                            "target_name": "应收票据及应收账款",
                            "children": ["ZT_007", "ZT_008"],
                            "statement_types": ["balance_sheet"],
                            "enabled": True,
                        },
                        ensure_ascii=False,
                    ),
                    "action_type": "promote_formula_rule",
                    "action_value": "",
                },
                {
                    "promotion_id": "PROM_3",
                    "action_type": "reject",
                    "action_value": "not safe",
                },
            ]
            applied, rejected, audit_rows, summary = apply_promotions(rows, config_dir, {"supported_action_types": ["promote_alias", "promote_formula_rule", "reject"]})
            self.assertEqual(len(applied), 2)
            self.assertEqual(len(rejected), 1)
            self.assertTrue(audit_rows)
            alias_payload = json.loads(json.dumps(yaml.safe_load((config_dir / "curated_alias_pack.yml").read_text(encoding="utf-8"))))
            self.assertTrue(any(item.get("alias") == "固定资产原价" for item in alias_payload.get("aliases", [])))
            formula_payload = json.loads(json.dumps(yaml.safe_load((config_dir / "curated_formula_pack.yml").read_text(encoding="utf-8"))))
            self.assertTrue(any(item.get("rule_id") == "sum_receivables" for item in formula_payload.get("rules", [])))
            delta = build_promotion_delta(
                before={"target_missing_total": 10, "target_mapped_ratio": 0.1, "target_amount_coverage_ratio": 0.2, "exportable_facts_total": 5, "benchmark_missing_true_total": 8},
                after={"run_id": "RUN_TEST", "target_missing_total": 8, "target_mapped_ratio": 0.2, "target_amount_coverage_ratio": 0.3, "exportable_facts_total": 6, "benchmark_missing_true_total": 6},
            )
            self.assertEqual(delta["rows"][0]["metric"], "target_missing_total")

    def test_stage7_no_source_investigation(self):
        gap_rows = [
            {"mapping_code": "ZT_001", "mapping_name": "货币资金", "aligned_period_key": "2022-12-31__期末数", "benchmark_value": 100.0},
            {"mapping_code": "ZT_002", "mapping_name": "交易性金融资产", "aligned_period_key": "2022-12-31__期末数", "benchmark_value": 50.0},
        ]
        result = investigate_no_source_gaps(
            benchmark_missing_true_rows=gap_rows,
            facts_raw=[],
            facts_deduped=[],
            unplaced_rows=[{"mapping_code": "ZT_001", "period_key": "2022-12-31__期末数", "fact_id": "F1"}],
            derived_facts=[],
            review_items=[],
            issues=[],
        )
        causes = {row["mapping_code"]: row["gap_cause"] for row in result["rows"]}
        self.assertEqual(causes["ZT_001"], "source_exists_but_unplaced")
        self.assertEqual(causes["ZT_002"], "truly_no_source")
        self.assertEqual(result["backfill_summary"]["tasks_total"], 1)

    def test_stage7_export_blocks_note_detail_and_writes_helper_sheets(self):
        repo_root = Path(__file__).resolve().parent
        template_path = repo_root.parent / "会计报表.xlsx"
        with tempfile.TemporaryDirectory() as tmpdir:
            output_dir = Path(tmpdir)
            facts = assign_fact_ids(
                [
                    self.make_fact(
                        mapping_code="ZT_001",
                        mapping_name="货币资金",
                        row_label_std="货币资金",
                        period_key="2022-12-31__期末数",
                        report_date_norm="2022-12-31",
                        period_role_norm="期末数",
                        value_raw="100",
                        value_num=100.0,
                        target_scope="main_export_target",
                    ),
                    self.make_fact(
                        mapping_code="ZT_014",
                        mapping_name="其他应收款项",
                        row_label_std="账龄1年以内",
                        period_key="2022-12-31__期末数",
                        report_date_norm="2022-12-31",
                        period_role_norm="期末数",
                        value_raw="50",
                        value_num=50.0,
                        target_scope="note_detail",
                    ),
                ]
            )
            stats = export_template(
                template_path=template_path,
                output_path=output_dir / "会计报表_填充结果.xlsx",
                facts=facts,
                run_summary={"unknown_date_total": 0},
                issues=[],
                validations=[],
                duplicates=[],
                conflicts=[],
                review_queue=[],
                applied_actions=[],
                export_rules={"allowed_statuses": ["observed", "repaired"], "required_helper_sheets": ["_benchmark_alignment", "_target_gap_backlog", "_promotions"], "blocked_target_scopes": ["note_detail"]},
            )
            workbook = load_workbook(output_dir / "会计报表_填充结果.xlsx")
            self.assertIn("_benchmark_alignment", workbook.sheetnames)
            self.assertIn("_target_gap_backlog", workbook.sheetnames)
            self.assertIn("_promotions", workbook.sheetnames)
            self.assertEqual(stats["unplaced_count"], 1)
            self.assertEqual(stats["unplaced_rows"][0]["unplaced_reason"], "target_scope_blocked:note_detail")

    def test_run_manifest(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output_dir = Path(tmpdir) / "normalized"
            output_dir.mkdir(parents=True, exist_ok=True)
            (output_dir / "run_summary.json").write_text(json.dumps({"run_id": "RUN_TEST"}), encoding="utf-8")
            (output_dir / "artifact_integrity.json").write_text(json.dumps({"run_id": "RUN_TEST"}), encoding="utf-8")
            run_id = generate_run_id(["--template", "foo.xlsx"])
            payload = write_run_manifest(
                run_id=run_id,
                output_dir=output_dir,
                cli_args=["--template", "foo.xlsx"],
                input_dir=Path(tmpdir) / "outputs",
                template_path=Path(tmpdir) / "foo.xlsx",
                source_files=[],
                run_summary={"run_id": run_id},
                manifest_rules={"snapshot_root": "normalized_runs", "core_artifacts": ["run_summary.json", "artifact_integrity.json"]},
            )
            self.assertEqual(payload["manifest"]["run_id"], run_id)
            self.assertTrue((output_dir / "run_manifest.json").exists())
            self.assertTrue((output_dir / "artifact_manifest.csv").exists())
            self.assertTrue(payload["artifact_rows"])

    def test_page_0004_end_to_end(self):
        repo_root = Path(__file__).resolve().parent
        template_path = repo_root.parent / "会计报表.xlsx"
        benchmark_path = repo_root.parent / "会计报表_GPT5.4Pro填写.xlsx"

        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir_path = Path(tmpdir)
            input_dir = tmpdir_path / "outputs"
            output_root = tmpdir_path / "normalized_archive"

            self.copy_sample_page_0004(repo_root, input_dir)

            exit_code = cli.main(
                [
                    "--input-dir",
                    str(input_dir),
                    "--template",
                    str(template_path),
                    "--output-dir",
                    str(output_root),
                    "--source-image-dir",
                    str(repo_root / "data"),
                    "--provider-priority",
                    "aliyun,tencent",
                    "--enable-conflict-merge",
                    "--enable-period-normalization",
                    "--enable-dedupe",
                    "--enable-validation",
                    "--enable-integrity-check",
                    "--enable-mapping-suggestions",
                    "--enable-review-pack",
                    "--enable-validation-aware-conflicts",
                    "--emit-routing-plan",
                    "--emit-reocr-tasks",
                    "--emit-review-actions-template",
                    "--materialize-reocr-inputs",
                    "--emit-delta-report",
                    "--benchmark-workbook",
                    str(benchmark_path),
                    "--emit-benchmark-report",
                    "--enable-label-canonicalization",
                    "--enable-derived-facts",
                    "--emit-run-manifest",
                    "--enable-main-statement-specialization",
                    "--enable-single-period-role-inference",
                    "--emit-stage6-kpis",
                    "--enable-benchmark-alignment-repair",
                    "--enable-export-target-scoping",
                    "--emit-promotion-template",
                    "--emit-stage7-kpis",
                    "--strict-full-run-contract",
                ]
            )

            self.assertEqual(exit_code, 0)
            run_dirs = self.list_cli_run_dirs(output_root)
            self.assertEqual(len(run_dirs), 1)
            output_dir = run_dirs[0]
            for filename in [
                "cells.csv",
                "facts_raw.csv",
                "facts_deduped.csv",
                "facts.csv",
                "duplicates.csv",
                "provider_comparison_summary.csv",
                "validation_results.csv",
                "mapping_candidates.csv",
                "unmapped_labels_summary.csv",
                "conflicts_enriched.csv",
                "conflict_decision_audit.csv",
                "validation_impact_of_conflicts.csv",
                "review_queue.csv",
                "review_summary.json",
                "review_workbook.xlsx",
                "reocr_tasks.csv",
                "reocr_task_summary.json",
                "review_actions_template.xlsx",
                "review_actions_template.csv",
                "reocr_input_manifest.csv",
                "reocr_input_manifest.json",
                "artifact_integrity.json",
                "run_manifest.json",
                "artifact_manifest.csv",
                "benchmark_summary.json",
                "benchmark_missing_in_auto.csv",
                "benchmark_missing_true.csv",
                "benchmark_alignment_audit.csv",
                "benchmark_alignment_summary.json",
                "benchmark_alignment_only.csv",
                "benchmark_gap_explanations.csv",
                "derived_facts.csv",
                "derived_formula_summary.json",
                "statement_classification_summary.json",
                "period_role_resolution_summary.json",
                "unmapped_value_bearing.csv",
                "alias_acceptance_candidates.csv",
                "formula_rule_impact_summary.json",
                "review_actionable.csv",
                "reocr_task_pruned.csv",
                "stage6_kpi_summary.json",
                "export_target_scope.csv",
                "export_target_kpi_summary.json",
                "main_target_review_queue.csv",
                "note_detail_review_queue.csv",
                "target_gap_backlog.csv",
                "target_gap_summary.json",
                "promotion_actions_template.xlsx",
                "promotion_actions_template.csv",
                "no_source_gap_investigation.csv",
                "no_source_gap_summary.json",
                "target_backfill_tasks.csv",
                "target_backfill_summary.json",
                "stage7_kpi_summary.json",
                "curated_alias_pack_summary.json",
                "full_run_contract_summary.json",
                "run_summary.json",
                "会计报表_填充结果.xlsx",
            ]:
                self.assertTrue((output_dir / filename).exists(), filename)

            with (output_dir / "run_summary.json").open("r", encoding="utf-8") as handle:
                summary = json.load(handle)
            self.assertLessEqual(summary["facts_deduped_total"], summary["facts_raw_total"])
            self.assertGreater(summary["provider_compared_pairs"], 0)

            with (output_dir / "facts_deduped.csv").open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))
            self.assertTrue(any(row["row_label_raw"] == "货币资金" for row in rows))

            workbook = load_workbook(output_dir / "会计报表_填充结果.xlsx")
            worksheet = workbook[workbook.sheetnames[0]]
            headers = [worksheet.cell(row=3, column=idx).value for idx in range(1, worksheet.max_column + 1)]
            self.assertIn("2022-12-31__期初数", headers)
            self.assertIn("2022-12-31__期末数", headers)
            self.assertFalse(any(str(header).startswith("unknown_date__") for header in headers if header))
            self.assertFalse(any(str(header).endswith("__unknown") for header in headers if header))
            self.assertIn("_meta_summary", workbook.sheetnames)
            self.assertIn("_validation", workbook.sheetnames)
            self.assertIn("_duplicates", workbook.sheetnames)
            self.assertIn("_conflicts", workbook.sheetnames)
            self.assertIn("_unplaced_facts", workbook.sheetnames)
            self.assertIn("_review_queue", workbook.sheetnames)
            self.assertIn("_derived_facts", workbook.sheetnames)
            self.assertIn("_benchmark_summary", workbook.sheetnames)
            self.assertIn("_gap_explanations", workbook.sheetnames)
            self.assertIn("_benchmark_alignment", workbook.sheetnames)
            self.assertIn("_target_gap_backlog", workbook.sheetnames)
            self.assertIn("_promotions", workbook.sheetnames)
            self.assertIn("_classification_audit", workbook.sheetnames)
            self.assertIn("_period_role_audit", workbook.sheetnames)

            with (output_dir / "review_actions_template.csv").open("r", encoding="utf-8-sig", newline="") as handle:
                action_rows = list(csv.DictReader(handle))
            self.assertTrue(action_rows)
            sample_action_file = output_dir / "review_actions_filled.csv"
            sample_row = action_rows[0]
            sample_row["action_type"] = "ignore"
            sample_row["reviewer_name"] = "tester"
            sample_row["review_status"] = "done"
            with sample_action_file.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=action_rows[0].keys())
                writer.writeheader()
                writer.writerow(sample_row)

            with (output_dir / "promotion_actions_template.csv").open("r", encoding="utf-8-sig", newline="") as handle:
                promotion_rows = list(csv.DictReader(handle))
            self.assertTrue(promotion_rows)
            promotion_row = next((row for row in promotion_rows if row.get("candidate_alias")), promotion_rows[0])
            promotion_row["action_type"] = "promote_alias" if promotion_row.get("candidate_alias") else "reject"
            promotion_row["reviewer_name"] = "tester"
            promotion_row["review_status"] = "done"
            promotion_file = output_dir / "promotion_actions_filled.csv"
            with promotion_file.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=promotion_rows[0].keys())
                writer.writeheader()
                writer.writerow(promotion_row)

            config_dir = repo_root / "standardize" / "config"
            curated_paths = [
                config_dir / "curated_alias_pack.yml",
                config_dir / "curated_formula_pack.yml",
                config_dir / "target_scope_rules.yml",
                config_dir / "manual_overrides" / "placement_overrides.yml",
            ]
            curated_backup = {path: path.read_text(encoding="utf-8") for path in curated_paths}
            try:
                second_exit_code = cli.main(
                    [
                        "--input-dir",
                        str(input_dir),
                        "--template",
                        str(template_path),
                        "--output-dir",
                        str(output_root),
                        "--source-image-dir",
                        str(repo_root / "data"),
                        "--provider-priority",
                        "aliyun,tencent",
                        "--enable-conflict-merge",
                        "--enable-period-normalization",
                        "--enable-dedupe",
                        "--enable-validation",
                        "--enable-integrity-check",
                        "--enable-mapping-suggestions",
                        "--enable-review-pack",
                        "--enable-validation-aware-conflicts",
                        "--emit-routing-plan",
                        "--emit-reocr-tasks",
                        "--emit-review-actions-template",
                        "--review-actions-file",
                        str(sample_action_file),
                        "--apply-review-actions",
                        "--materialize-reocr-inputs",
                        "--emit-delta-report",
                        "--benchmark-workbook",
                        str(benchmark_path),
                        "--emit-benchmark-report",
                        "--enable-label-canonicalization",
                        "--enable-derived-facts",
                        "--emit-run-manifest",
                        "--enable-main-statement-specialization",
                        "--enable-single-period-role-inference",
                        "--emit-stage6-kpis",
                        "--enable-benchmark-alignment-repair",
                        "--enable-export-target-scoping",
                        "--emit-promotion-template",
                        "--promotion-actions-file",
                        str(promotion_file),
                        "--apply-promotions",
                        "--emit-stage7-kpis",
                        "--strict-full-run-contract",
                    ]
                )
            finally:
                for path, content in curated_backup.items():
                    path.write_text(content, encoding="utf-8")
            self.assertEqual(second_exit_code, 0)
            second_run_dirs = self.list_cli_run_dirs(output_root)
            self.assertEqual(len(second_run_dirs), 2)
            second_output_dir = next(path for path in second_run_dirs if path != output_dir)
            for filename in [
                "applied_review_actions.csv",
                "override_audit.csv",
                "review_decision_summary.json",
                "coverage_delta.json",
                "coverage_delta.csv",
                "review_delta.csv",
                "export_delta_summary.json",
                "top_resolved_items.csv",
                "review_priority_backlog.csv",
                "mapping_opportunities.csv",
                "applied_promotions.csv",
                "rejected_promotions.csv",
                "promotion_audit.csv",
                "promotion_delta.json",
                "promotion_delta.csv",
                "promoted_aliases.csv",
                "promoted_formula_rules.csv",
                "stage7_kpi_summary.json",
            ]:
                self.assertTrue((second_output_dir / filename).exists(), filename)
            workbook = load_workbook(second_output_dir / "会计报表_填充结果.xlsx")
            self.assertIn("_applied_actions", workbook.sheetnames)
            self.assertIn("_promotions", workbook.sheetnames)

    def list_cli_run_dirs(self, output_root: Path) -> list[Path]:
        if not output_root.exists():
            return []
        return sorted(
            [
                path
                for path in output_root.iterdir()
                if path.is_dir() and not path.name.startswith("_") and (path / "run_summary.json").exists()
            ],
            key=lambda path: path.stat().st_mtime,
        )

    def copy_sample_page_0004(self, repo_root: Path, input_dir: Path) -> None:
        doc_name = "债务人审计报告-2022年"

        aliyun_raw_dir = input_dir / "aliyun_table" / doc_name / "raw"
        aliyun_raw_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(
            next((repo_root / "outputs" / "aliyun_table").glob(f"*/raw/page_0004.json")),
            aliyun_raw_dir / "page_0004.json",
        )
        aliyun_result = {
            "provider": "aliyun_table",
            "pages": [
                {
                    "page_number": 4,
                    "text": "资产负债表 编制单位：泰兴市泰泽实业有限公司 2022年12月31日 单位：元",
                    "raw_file": "raw/page_0004.json",
                    "artifact_files": [],
                }
            ],
        }
        (input_dir / "aliyun_table" / doc_name / "result.json").write_text(
            json.dumps(aliyun_result, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

        tencent_doc_dir = input_dir / "tencent_table_v3" / doc_name
        tencent_raw_dir = tencent_doc_dir / "raw"
        tencent_artifacts_dir = tencent_doc_dir / "artifacts"
        tencent_raw_dir.mkdir(parents=True, exist_ok=True)
        tencent_artifacts_dir.mkdir(parents=True, exist_ok=True)

        shutil.copy2(
            next((repo_root / "outputs" / "tencent_table_v3").glob("*/raw/page_0004_tencent.json")),
            tencent_raw_dir / "page_0004_tencent.json",
        )
        shutil.copy2(
            next((repo_root / "outputs" / "tencent_table_v3").glob("*/artifacts/page_0004.xlsx")),
            tencent_artifacts_dir / "page_0004.xlsx",
        )
        tencent_result = {
            "provider": "tencent_table_v3",
            "pages": [
                {
                    "page_number": 4,
                    "text": "资产负债表\n编制单位:泰兴市泰泽实业有限公司\n2022年12月31日\n单位:元",
                    "raw_file": "raw/page_0004.json",
                    "artifact_files": ["artifacts/page_0004.xlsx"],
                }
            ],
        }
        (tencent_doc_dir / "result.json").write_text(
            json.dumps(tencent_result, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

        # Copy text-provider result.json so period normalization and routing can use cheap text hints.
        for provider in ["aliyun_text", "tencent_text"]:
            provider_doc_dir = input_dir / provider / doc_name
            provider_doc_dir.mkdir(parents=True, exist_ok=True)
            provider_result = {
                "provider": provider,
                "pages": [
                    {
                        "page_number": 4,
                        "text": "资产负债表\n编制单位：泰兴市泰泽实业有限公司\n2022年12月31日\n单位：元",
                    }
                ],
            }
            (provider_doc_dir / "result.json").write_text(
                json.dumps(provider_result, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )

    def make_fact(self, **overrides):
        base = {
            "doc_id": "demo",
            "page_no": 1,
            "provider": "aliyun_table",
            "statement_type": "note",
            "statement_name_raw": "2022年度财务报表附注",
            "logical_subtable_id": "1_sub1",
            "table_semantic_key": "note|h:项目,上期发生额,本期发生额",
            "row_label_raw": "项目",
            "row_label_std": "项目",
            "row_label_norm": "项目",
            "row_label_canonical_candidate": "项目",
            "col_header_raw": "本期发生额",
            "col_header_path": ["本期发生额"],
            "column_semantic_key": "本期发生额",
            "period_role_raw": "本期",
            "report_date_raw": "",
            "period_key": "2022年度__本期",
            "value_raw": "1",
            "value_num": 1.0,
            "value_type": "amount",
            "unit_raw": "元",
            "unit_multiplier": 1.0,
            "source_cell_ref": "demo:1:aliyun_table:1:1-1:1-1",
            "status": "observed",
            "mapping_code": "",
            "mapping_name": "",
            "mapping_method": "",
            "mapping_confidence": None,
            "issue_flags": [],
            "fact_id": "",
            "report_date_norm": "2022年度",
            "period_role_norm": "本期",
            "period_source_level": "page",
            "period_reason": "",
            "duplicate_group_id": "",
            "kept_fact_id": "",
            "comparison_status": "single_provider",
            "comparison_reason": "single_provider_only",
            "source_kind": "json",
            "statement_group_key": "note|2022年度财务报表附注",
            "source_row_start": 1,
            "source_row_end": 1,
            "source_col_start": 1,
            "source_col_end": 1,
        }
        base.update(overrides)
        if "row_label_norm" not in overrides:
            base["row_label_norm"] = base.get("row_label_std") or base.get("row_label_raw") or ""
        if "row_label_canonical_candidate" not in overrides:
            base["row_label_canonical_candidate"] = base.get("row_label_norm") or ""
        return FactRecord(**base)

    def make_validation(self, **overrides):
        base = {
            "validation_id": "VAL000001",
            "doc_id": "demo",
            "statement_type": "note",
            "period_key": "2022年度__本期",
            "rule_name": "amount_legality",
            "rule_type": "quality",
            "lhs_value": None,
            "rhs_value": None,
            "diff_value": None,
            "tolerance": None,
            "status": "review",
            "evidence_fact_refs": [],
            "message": "review",
            "meta_json": "",
        }
        base.update(overrides)
        return ValidationResultRecord(**base)

    def statement_config(self):
        return {
            "statement_titles": {
                "balance_sheet": ["资产负债表"],
                "income_statement": ["利润表", "损益表"],
                "cash_flow": ["现金流量表"],
                "equity_statement": ["所有者权益变动表", "股东权益变动表"],
                "note": ["附注", "财务报表附注"],
            },
            "period_roles": {
                "期初数": ["期初数"],
                "期末数": ["期末数"],
                "本期": ["本期", "本年累计"],
                "上期": ["上期", "上年同期"],
            },
        }


if __name__ == "__main__":
    unittest.main()
