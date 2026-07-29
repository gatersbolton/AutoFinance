from __future__ import annotations

import csv
import json
import re
from collections import defaultdict
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font

from standard_map.registry import load_standard_registry
from standardize.mapping.masterdata import load_template_subjects
from standardize.normalize.export import sanitize_template_output_sheet, validate_template_output_sheet


ACCOUNTING_WORKBOOK_DOWNLOAD_NAME = "会计报表.xlsx"
ACCOUNTING_EXPORT_SUMMARY_NAME = "会计报表生成摘要.json"
SUPPORTED_STATEMENT_TYPES = {"balance_sheet", "income_statement"}


def build_accounting_workbook(
    *,
    template_path: str | Path,
    current_data_csv: str | Path,
    output_path: str | Path,
    summary_path: str | Path | None = None,
) -> dict[str, Any]:
    template = Path(template_path).resolve()
    current_csv = Path(current_data_csv).resolve()
    output = Path(output_path).resolve()
    if not template.exists():
        raise ValueError("会计报表模板不存在。")
    if not current_csv.exists():
        raise ValueError("请先生成并校对结构化数据。")

    rows = _read_csv(current_csv)
    subjects, sheet_name, header_row = load_template_subjects(template)
    subject_codes = {subject.code for subject in subjects}
    registry = load_standard_registry()
    term_scopes = {
        code: str(term.statement_scope or "")
        for code, term in registry.term_by_code.items()
    }

    selected: dict[tuple[str, str], dict[str, Any]] = {}
    conflicts: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    conflicted_keys: set[tuple[str, str]] = set()
    skipped: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows, start=2):
        candidate, reason = _accounting_candidate(
            row,
            row_number=row_number,
            subject_codes=subject_codes,
            term_scopes=term_scopes,
        )
        if candidate is None:
            skipped.append(_skipped_row(row, row_number, reason))
            continue
        key = (candidate["code"], candidate["period_key"])
        if key in conflicted_keys:
            conflicts[key].append(candidate)
            continue
        previous = selected.get(key)
        if previous is None:
            selected[key] = candidate
            continue
        if previous["value"] == candidate["value"]:
            skipped.append(_skipped_row(row, row_number, "同一科目和期间的重复相同值"))
            continue
        conflicts[key].extend([previous, candidate])
        conflicted_keys.add(key)
        selected.pop(key, None)

    for key, items in conflicts.items():
        code, period_key = key
        for item in items:
            skipped.append(
                {
                    "行号": item["row_number"],
                    "标准指标编码": code,
                    "标准指标名称": item["name"],
                    "期间": period_key,
                    "指标数值": format(item["value"], "f"),
                    "未写入原因": "同一科目和期间存在不同金额，请先校对",
                }
            )

    workbook = load_workbook(template)
    worksheet = workbook[sheet_name]
    cleanup = sanitize_template_output_sheet(worksheet, header_row)
    row_by_code = {subject.code: subject.row_index for subject in subjects}
    period_keys = sorted({period_key for _, period_key in selected})
    period_columns = _ensure_period_columns(worksheet, header_row, period_keys)
    for (code, period_key), candidate in sorted(selected.items()):
        worksheet.cell(
            row=row_by_code[code],
            column=period_columns[period_key],
            value=float(candidate["value"]),
        )

    validation = validate_template_output_sheet(worksheet, header_row, period_keys)
    if not validation["pass"]:
        raise ValueError("会计报表模板校验失败：" + "；".join(validation["errors"]))

    summary = {
        "pass": True,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "template_path": str(template),
        "current_data_csv": str(current_csv),
        "workbook_path": str(output),
        "rows_total": len(rows),
        "written_cells_total": len(selected),
        "skipped_rows_total": len(skipped),
        "conflicted_cells_total": len(conflicts),
        "period_keys": period_keys,
        "template_values_cleared_total": cleanup["values_cleared_total"],
        "template_output_validation_pass": validation["pass"],
    }
    _write_explanation_sheet(workbook, summary, skipped)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)

    resolved_summary_path = (
        Path(summary_path).resolve()
        if summary_path is not None
        else output.with_name(ACCOUNTING_EXPORT_SUMMARY_NAME)
    )
    resolved_summary_path.parent.mkdir(parents=True, exist_ok=True)
    resolved_summary_path.write_text(
        json.dumps({**summary, "skipped_rows": skipped}, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )
    summary["summary_path"] = str(resolved_summary_path)
    return summary


def _accounting_candidate(
    row: dict[str, Any],
    *,
    row_number: int,
    subject_codes: set[str],
    term_scopes: dict[str, str],
) -> tuple[dict[str, Any] | None, str]:
    code = str(row.get("标准指标编码", "") or "").strip().upper()
    name = str(row.get("标准指标名称", "") or "").strip()
    if code not in subject_codes:
        return None, "未映射到当前会计报表模板"
    if _is_yes(row.get("是否需要人工校对", "")):
        return None, "仍需人工校对"
    statement_type = str(row.get("报表类型", "") or "").strip()
    if statement_type not in SUPPORTED_STATEMENT_TYPES:
        statement_type = term_scopes.get(code, "")
    if statement_type not in SUPPORTED_STATEMENT_TYPES:
        return None, "当前只生成资产负债表和利润表"
    value = _parse_decimal(row.get("指标数值", ""))
    if value is None:
        return None, "金额为空或格式有误"
    period_key = _period_key(row, statement_type)
    if not period_key:
        return None, "日期或期间仍不明确"
    return (
        {
            "row_number": row_number,
            "code": code,
            "name": name,
            "period_key": period_key,
            "value": value,
        },
        "",
    )


def _period_key(row: dict[str, Any], statement_type: str) -> str:
    fill_date = _parse_date(row.get("填表日期", ""))
    item_date = _parse_date(row.get("当前条目日期", ""))
    role = re.sub(r"\s+", "", str(row.get("期间类型", "") or ""))
    if fill_date is None:
        return ""
    if statement_type == "balance_sheet":
        if any(token in role for token in ("期初", "年初", "上期期末")):
            return f"{fill_date.isoformat()}__期初数"
        if any(token in role for token in ("期末", "年末", "当前时点", "明确日期")):
            return f"{fill_date.isoformat()}__期末数"
        if item_date is not None:
            suffix = "期末数" if item_date == fill_date else "期初数"
            return f"{fill_date.isoformat()}__{suffix}"
        return ""

    if any(token in role for token in ("上期", "上年")):
        return f"{fill_date.year:04d}年度__上期"
    if any(token in role for token in ("本期", "本年", "累计")):
        return f"{fill_date.year:04d}年度__本期"
    if item_date is not None:
        suffix = "本期" if item_date.year == fill_date.year else "上期"
        return f"{fill_date.year:04d}年度__{suffix}"
    return ""


def _ensure_period_columns(worksheet, header_row: int, period_keys: list[str]) -> dict[str, int]:
    result: dict[str, int] = {}
    for column, period_key in enumerate(period_keys, start=2):
        worksheet.cell(row=header_row, column=column, value=period_key)
        result[period_key] = column
    return result


def _write_explanation_sheet(workbook, summary: dict[str, Any], skipped: list[dict[str, Any]]) -> None:
    if "生成说明" in workbook.sheetnames:
        del workbook["生成说明"]
    sheet = workbook.create_sheet("生成说明")
    sheet.append(["项目", "内容"])
    sheet["A1"].font = Font(bold=True)
    sheet["B1"].font = Font(bold=True)
    labels = (
        ("生成时间", summary["generated_at"]),
        ("写入单元格", summary["written_cells_total"]),
        ("未写入行", summary["skipped_rows_total"]),
        ("冲突单元格", summary["conflicted_cells_total"]),
        ("说明", "仅写入已映射、日期明确且无需校对的资产负债表和利润表数据。"),
    )
    for label, value in labels:
        sheet.append([label, value])
    sheet.append([])
    headers = ["行号", "标准指标编码", "标准指标名称", "期间", "指标数值", "未写入原因"]
    sheet.append(headers)
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)
    for row in skipped:
        sheet.append([row.get(header, "") for header in headers])
    sheet.column_dimensions["A"].width = 16
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 26
    sheet.column_dimensions["D"].width = 24
    sheet.column_dimensions["E"].width = 18
    sheet.column_dimensions["F"].width = 38
    sheet.freeze_panes = "A2"


def _skipped_row(row: dict[str, Any], row_number: int, reason: str) -> dict[str, Any]:
    return {
        "行号": row_number,
        "标准指标编码": str(row.get("标准指标编码", "") or ""),
        "标准指标名称": str(row.get("标准指标名称", "") or row.get("原始指标名", "") or ""),
        "期间": str(row.get("期间类型", "") or ""),
        "指标数值": str(row.get("指标数值", "") or ""),
        "未写入原因": reason,
    }


def _read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def _parse_decimal(value: Any) -> Decimal | None:
    text = str(value or "").strip().replace(",", "").replace("，", "")
    if not text or "%" in text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _parse_date(value: Any) -> date | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text)
    except ValueError:
        return None


def _is_yes(value: Any) -> bool:
    return str(value or "").strip().lower() in {"是", "true", "1", "yes", "y"}
