from __future__ import annotations

import csv
import hashlib
import json
import re
import shutil
import subprocess
import time
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Iterable

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from project_paths import REPO_ROOT
from standardize.feedback.apply import apply_review_actions as backend_apply_review_actions
from standardize.feedback.actions import SUPPORTED_ACTIONS as BACKEND_SUPPORTED_ACTIONS
from standardize.feedback.actions import validate_action_row
from standardize.feedback.audit import build_review_decision_summary
from standardize.feedback.parser import parse_review_actions_file
from standardize.feedback.template import TEMPLATE_HEADERS

from .config import WebAppSettings
from .db import list_review_actions, upsert_review_action, utc_now_iso
from .labels import (
    REVIEW_COMPATIBILITY_LABELS_ZH as LABEL_REVIEW_COMPATIBILITY_LABELS_ZH,
    reason_code_label_zh,
    review_compatibility_label_zh as label_review_compatibility_label_zh,
    review_source_type_label_zh as label_review_source_type_label_zh,
    review_status_label_zh as label_review_status_label_zh,
)
from .jobs import _repo_relative_or_absolute
from .models import (
    REVIEW_COMPATIBILITY_BACKEND_READY,
    REVIEW_COMPATIBILITY_PARTIAL,
    REVIEW_COMPATIBILITY_SUGGESTION_ONLY,
    REVIEW_COMPATIBILITY_UNSUPPORTED,
    REVIEW_ACTION_TYPES,
    REVIEW_STATUS_DEFERRED,
    REVIEW_STATUS_IGNORED,
    REVIEW_STATUS_REOCR_REQUESTED,
    REVIEW_STATUS_RESOLVED,
    REVIEW_STATUS_UNRESOLVED,
    SUCCESS_LIKE_JOB_STATUSES,
    JobRecord,
    ReviewActionRecord,
    ReviewItemRecord,
    ReviewSourceArtifact,
)
from .quality import build_job_quality_summary, describe_job_status, load_json


class ReviewOperationCancelled(RuntimeError):
    """Raised when a queued review operation is cancelled at a safe checkpoint."""


HIGH_PRIORITY_THRESHOLD = 5.0
MEDIUM_PRIORITY_THRESHOLD = 3.0
ACTIONABLE_REVIEW_STATUSES = {
    REVIEW_STATUS_UNRESOLVED,
    REVIEW_STATUS_DEFERRED,
    REVIEW_STATUS_REOCR_REQUESTED,
}
SUPPORTED_BULK_ACTION_TYPES = {
    "defer",
    "ignore",
    "mark_not_financial_fact",
    "request_reocr",
    "suppress_false_positive",
    "accept_mapping_candidate",
}
REVIEW_COMPATIBILITY_LABELS_ZH = {
    **LABEL_REVIEW_COMPATIBILITY_LABELS_ZH,
}
QUICK_FILTER_OPTIONS = (
    {"value": "high_priority", "label_zh": "只看高优先级"},
    {"value": "backend_ready", "label_zh": "只看可自动应用"},
    {"value": "not_backend_ready", "label_zh": "只看不可自动应用"},
    {"value": "mapping_missing", "label_zh": "只看映射缺失"},
    {"value": "ocr_suspicious", "label_zh": "只看 OCR 可疑"},
    {"value": "validation_fail", "label_zh": "只看校验失败"},
    {"value": "evidence_available", "label_zh": "只看有证据图片"},
)

REVIEW_ACTION_UI_OPTIONS = (
    {
        "action_type": "defer",
        "label_zh": "暂缓处理",
        "help_text": "暂时不处理，后续仍会保留在复核积压中。",
    },
    {
        "action_type": "ignore",
        "label_zh": "忽略此项",
        "help_text": "关闭当前条目，不直接改动底层事实。",
    },
    {
        "action_type": "mark_not_financial_fact",
        "label_zh": "标记为非财务事实",
        "help_text": "将该项视为说明性内容，不进入财务事实。",
    },
    {
        "action_type": "request_reocr",
        "label_zh": "请求重新 OCR",
        "help_text": "记录需要局部重跑 OCR，请尽量填写备注说明原因。",
    },
    {
        "action_type": "accept_mapping_candidate",
        "label_zh": "接受科目建议",
        "help_text": "直接采用当前建议的标准科目；未填值时会优先使用候选编码。",
    },
    {
        "action_type": "set_mapping_override",
        "label_zh": "指定标准科目",
        "help_text": "手动指定标准科目编码，适合候选值不够准确时使用。",
    },
    {
        "action_type": "set_conflict_winner",
        "label_zh": "选择冲突赢家",
        "help_text": "在供应商冲突结果中指定保留哪一条事实或 provider。",
    },
    {
        "action_type": "suppress_false_positive",
        "label_zh": "标记为误报",
        "help_text": "将该项标记为 OCR 误报，并从后续处理中过滤。",
    },
)
REVIEW_ACTION_UI_BY_TYPE = {item["action_type"]: item for item in REVIEW_ACTION_UI_OPTIONS}

REVIEW_SOURCE_DEFS = (
    ("review_queue", "复核队列", "review_queue.csv"),
    ("issues", "问题清单", "issues.csv"),
    ("validation_results", "校验结果", "validation_results.csv"),
    ("conflicts_enriched", "冲突明细", "conflicts_enriched.csv"),
    ("conflict_decision_audit", "冲突决策审计", "conflict_decision_audit.csv"),
    ("unplaced_facts", "未落位事实", "unplaced_facts.csv"),
    ("mapping_candidates", "科目映射候选", "mapping_candidates.csv"),
    ("benchmark_gap_explanations", "基准差异说明", "benchmark_gap_explanations.csv"),
    ("source_backed_gap_closure", "来源支撑缺口闭环", "source_backed_gap_closure.csv"),
    ("review_workbook", "复核工作簿", "review_workbook.xlsx"),
)

EXPORT_EXTRA_HEADERS = [
    "review_item_id",
    "backend_review_id",
    "source_ref",
    "apply_target_type",
    "apply_compatibility_status",
    "apply_incompatibility_reason",
    "created_at",
    "original_action_type",
    "compatibility_status",
    "compatibility_note",
]
EXPORT_HEADERS = [*TEMPLATE_HEADERS, *EXPORT_EXTRA_HEADERS]

PATCHED_STANDARDIZE_CLI_SCRIPT = """
from __future__ import annotations

import json
import pathlib
import sys

import standardize.cli as cli


payload = json.loads(sys.argv[1])
cli.CONFIG_DIR = pathlib.Path(payload["config_dir"])
raise SystemExit(cli.main(payload["argv"]))
"""


def get_review_dir(job: JobRecord) -> Path:
    return Path(job.output_dir).resolve().parent / "review"


def get_allowed_review_roots(job: JobRecord) -> tuple[Path, ...]:
    return (
        Path(job.output_dir).resolve().parent,
        Path(job.result_dir).resolve(),
    )


def review_source_artifacts(job: JobRecord) -> list[ReviewSourceArtifact]:
    output_dir = Path(job.output_dir)
    artifacts: list[ReviewSourceArtifact] = []
    for slug, label, filename in REVIEW_SOURCE_DEFS:
        path = output_dir / filename
        row_count = count_csv_rows(path) if path.suffix.lower() == ".csv" and path.exists() else 0
        artifacts.append(
            ReviewSourceArtifact(
                slug=slug,
                label=label,
                path=str(path),
                relative_path=_repo_relative_or_absolute(path),
                exists=path.exists(),
                row_count=row_count,
            )
        )
    return artifacts


def load_review_items(settings: WebAppSettings, job: JobRecord) -> tuple[list[ReviewItemRecord], list[ReviewSourceArtifact]]:
    output_dir = Path(job.output_dir)
    items: list[ReviewItemRecord] = []

    review_queue_path = output_dir / "review_queue.csv"
    if review_queue_path.exists():
        items.extend(_load_review_queue_items(review_queue_path))

    issues_path = output_dir / "issues.csv"
    if issues_path.exists():
        items.extend(_load_issue_items(issues_path))

    validation_path = output_dir / "validation_results.csv"
    if validation_path.exists():
        items.extend(_load_validation_items(validation_path))

    conflict_path = output_dir / "conflicts_enriched.csv"
    if conflict_path.exists():
        items.extend(_load_conflict_items(conflict_path))

    unplaced_path = output_dir / "unplaced_facts.csv"
    if unplaced_path.exists():
        items.extend(_load_unplaced_fact_items(unplaced_path))

    mapping_candidates_path = output_dir / "mapping_candidates.csv"
    if mapping_candidates_path.exists():
        items.extend(_load_mapping_candidate_items(mapping_candidates_path))

    _normalize_review_item_identity(items)
    _apply_saved_actions(items, list_review_actions(settings, job.job_id))
    _annotate_review_items(job, items)
    items.sort(key=_default_sort_key)
    return items, review_source_artifacts(job)


def get_review_action_ui_options() -> list[dict[str, str]]:
    return [dict(item) for item in REVIEW_ACTION_UI_OPTIONS]


def get_bulk_review_action_ui_options() -> list[dict[str, str]]:
    return [dict(item) for item in REVIEW_ACTION_UI_OPTIONS if item["action_type"] in SUPPORTED_BULK_ACTION_TYPES]


def review_status_label_zh(status: str) -> str:
    return label_review_status_label_zh(status)


def review_source_type_label_zh(source_type: str) -> str:
    return label_review_source_type_label_zh(source_type)


def review_compatibility_label_zh(status: str) -> str:
    return label_review_compatibility_label_zh(status)


def _priority_bucket(priority_score: float) -> str:
    if priority_score >= HIGH_PRIORITY_THRESHOLD:
        return "high"
    if priority_score >= MEDIUM_PRIORITY_THRESHOLD:
        return "medium"
    return "low"


def _priority_bucket_label_zh(bucket: str) -> str:
    return {
        "high": "高优先级",
        "medium": "中优先级",
        "low": "低优先级",
    }.get(bucket, bucket or "未分级")


def _compatibility_sort_rank(status: str) -> int:
    return {
        REVIEW_COMPATIBILITY_BACKEND_READY: 0,
        REVIEW_COMPATIBILITY_PARTIAL: 1,
        REVIEW_COMPATIBILITY_SUGGESTION_ONLY: 2,
        REVIEW_COMPATIBILITY_UNSUPPORTED: 3,
    }.get(status, 9)


def _derive_item_apply_target_type(item: ReviewItemRecord) -> str:
    if item.action_type:
        return _derive_apply_target_type(item.action_type)
    if item.source_type == "mapping_candidate":
        return "mapping_alias"
    if item.source_type == "unplaced_fact":
        return "local_fact_mapping"
    if item.source_type == "conflict":
        return "conflict"
    if item.source_type == "validation":
        return "review_item"
    if item.source_type == "issue":
        return "review_item"
    return "review_item"


def _assess_item_apply_compatibility(item: ReviewItemRecord) -> tuple[str, str]:
    if item.review_id:
        return REVIEW_COMPATIBILITY_BACKEND_READY, ""
    if item.source_type in {"mapping_candidate", "unplaced_fact"}:
        return REVIEW_COMPATIBILITY_SUGGESTION_ONLY, "当前条目缺少 backend_review_id，现阶段更适合作为建议导出。"
    if item.source_ref:
        return REVIEW_COMPATIBILITY_PARTIAL, "当前条目有来源引用，但缺少 backend_review_id，应用时可能被后端拒绝。"
    return REVIEW_COMPATIBILITY_UNSUPPORTED, "当前条目缺少 backend_review_id 和稳定来源引用，暂不支持自动应用。"


def _annotate_review_items(job: JobRecord, items: Iterable[ReviewItemRecord]) -> None:
    for item in items:
        item.backend_review_id = item.review_id
        item.apply_target_type = _derive_item_apply_target_type(item)
        compatibility_status, incompatibility_reason = _assess_item_apply_compatibility(item)
        item.apply_compatibility_status = compatibility_status
        item.apply_incompatibility_reason = incompatibility_reason
        item.priority_bucket = _priority_bucket(item.priority_score)
        evidence_cell_available = resolve_evidence_file(job, item, "cell") is not None
        evidence_row_available = resolve_evidence_file(job, item, "row") is not None
        evidence_table_available = resolve_evidence_file(job, item, "table") is not None
        item.evidence_available = evidence_cell_available or evidence_row_available or evidence_table_available
        item.meta["evidence_cell_available"] = evidence_cell_available
        item.meta["evidence_row_available"] = evidence_row_available
        item.meta["evidence_table_available"] = evidence_table_available
        item.meta["evidence_available"] = item.evidence_available
        item.meta["source_type_label_zh"] = review_source_type_label_zh(item.source_type)
        item.meta["current_status_label_zh"] = review_status_label_zh(item.current_status)
        item.meta["apply_compatibility_label_zh"] = review_compatibility_label_zh(item.apply_compatibility_status)
        item.meta["priority_bucket_label_zh"] = _priority_bucket_label_zh(item.priority_bucket)


def filter_review_items(
    items: Iterable[ReviewItemRecord],
    *,
    status: str = "",
    source_type: str = "",
    reason_code: str = "",
    priority_bucket: str = "",
    apply_compatibility: str = "",
    evidence_available: str = "",
    page_no: str = "",
    statement_type: str = "",
    provider: str = "",
    search: str = "",
    quick_filter: str = "",
    only_high_priority: bool = False,
    sort_by: str = "priority_desc",
) -> list[ReviewItemRecord]:
    search_text = search.strip().lower()
    normalized_reason = reason_code.strip()
    normalized_page = page_no.strip()
    normalized_priority_bucket = priority_bucket.strip()
    normalized_apply_compatibility = apply_compatibility.strip()
    normalized_evidence_available = evidence_available.strip().lower()
    normalized_quick_filter = quick_filter.strip()
    filtered: list[ReviewItemRecord] = []
    for item in items:
        if status and item.current_status != status:
            continue
        if source_type and item.source_type != source_type:
            continue
        if normalized_reason and normalized_reason not in item.reason_codes and normalized_reason != item.reason_code:
            continue
        if normalized_priority_bucket and item.priority_bucket != normalized_priority_bucket:
            continue
        if normalized_apply_compatibility and item.apply_compatibility_status != normalized_apply_compatibility:
            continue
        if normalized_evidence_available in {"1", "true", "yes", "on"} and not item.evidence_available:
            continue
        if normalized_evidence_available in {"0", "false", "no", "off"} and item.evidence_available:
            continue
        if normalized_page:
            try:
                if item.page_no != int(normalized_page):
                    continue
            except ValueError:
                continue
        if statement_type and item.statement_type != statement_type:
            continue
        if provider and item.provider != provider:
            continue
        if only_high_priority and item.priority_score < HIGH_PRIORITY_THRESHOLD:
            continue
        if normalized_quick_filter and not _matches_quick_filter(item, normalized_quick_filter):
            continue
        if search_text:
            haystack = " ".join(
                [
                    item.row_label_raw,
                    item.row_label_std,
                    item.reason_label_zh,
                    item.doc_id,
                    item.mapping_name,
                    item.mapping_code,
                    item.source_ref,
                    item.reviewer_note,
                ]
            ).lower()
            if search_text not in haystack:
                continue
        filtered.append(item)
    filtered.sort(key=_sort_key(sort_by))
    return filtered


def build_review_dashboard_summary(
    items: Iterable[ReviewItemRecord],
    source_artifacts: Iterable[ReviewSourceArtifact] | None = None,
) -> dict[str, Any]:
    rows = list(items)
    artifacts = list(source_artifacts or [])
    source_breakdown = Counter(item.source_type for item in rows)
    reason_breakdown = Counter(item.reason_label_zh for item in rows)
    actions_submitted_total = sum(1 for item in rows if item.action_type)
    unresolved_total = sum(1 for item in rows if item.current_status == REVIEW_STATUS_UNRESOLVED)
    processed_total = len(rows) - unresolved_total
    actionable_total = sum(1 for item in rows if item.current_status in ACTIONABLE_REVIEW_STATUSES)
    source_artifact_rows_total = sum(int(artifact.row_count or 0) for artifact in artifacts)
    backend_ready_total = sum(1 for item in rows if item.apply_compatibility_status == REVIEW_COMPATIBILITY_BACKEND_READY)
    backend_partial_total = sum(1 for item in rows if item.apply_compatibility_status == REVIEW_COMPATIBILITY_PARTIAL)
    backend_suggestion_only_total = sum(1 for item in rows if item.apply_compatibility_status == REVIEW_COMPATIBILITY_SUGGESTION_ONLY)
    backend_unsupported_total = sum(1 for item in rows if item.apply_compatibility_status == REVIEW_COMPATIBILITY_UNSUPPORTED)
    evidence_available_total = sum(1 for item in rows if item.evidence_available)
    return {
        "total_review_items": len(rows),
        "unresolved_review_items": unresolved_total,
        "actions_submitted_total": actions_submitted_total,
        "actionable_items_total": actionable_total,
        "source_artifact_rows_total": source_artifact_rows_total,
        "high_priority_items": sum(1 for item in rows if item.priority_score >= HIGH_PRIORITY_THRESHOLD),
        "validation_fail_count": sum(1 for item in rows if item.source_type == "validation"),
        "mapping_unmapped_count": sum(
            1
            for item in rows
            if item.source_type in {"mapping_candidate", "unplaced_fact"}
            or any(code.startswith("mapping:") for code in item.reason_codes)
        ),
        "suspicious_numeric_count": sum(
            1
            for item in rows
            if any(code == "quality:suspicious_numeric" for code in item.reason_codes)
        ),
        "ocr_suspicious_count": sum(
            1
            for item in rows
            if any(code == "issue:suspicious_value" or code.startswith("source:") for code in item.reason_codes)
        ),
        "provider_conflict_count": sum(
            1 for item in rows if item.source_type == "conflict" or any(code.startswith("conflict:") for code in item.reason_codes)
        ),
        "resolved_count": sum(1 for item in rows if item.current_status == REVIEW_STATUS_RESOLVED),
        "resolved_total": processed_total,
        "deferred_count": sum(1 for item in rows if item.current_status == REVIEW_STATUS_DEFERRED),
        "ignored_total": sum(1 for item in rows if item.current_status == REVIEW_STATUS_IGNORED),
        "reocr_requested_total": sum(1 for item in rows if item.current_status == REVIEW_STATUS_REOCR_REQUESTED),
        "backend_ready_total": backend_ready_total,
        "backend_partial_total": backend_partial_total,
        "backend_suggestion_only_total": backend_suggestion_only_total,
        "backend_unsupported_total": backend_unsupported_total,
        "non_backend_ready_total": backend_partial_total + backend_suggestion_only_total + backend_unsupported_total,
        "evidence_available_total": evidence_available_total,
        "evidence_missing_total": max(len(rows) - evidence_available_total, 0),
        "count_semantics": {
            "待复核项目": "去重后的 Web 复核项数量，每个 review_item_id 只计一次。",
            "已提交动作": "已经在 Web 页面保存过动作的条目数。",
            "可处理项目": "仍建议继续处理的条目数 = 未处理 + 暂缓 + 已请求重新 OCR。",
            "原始问题行": "来源 CSV/XLSX 的原始问题行汇总，可能高于去重后的复核项数量。",
            "高优先级": f"priority_score >= {HIGH_PRIORITY_THRESHOLD} 的条目数。",
        },
        "source_type_breakdown": dict(sorted(source_breakdown.items())),
        "reason_label_breakdown": dict(sorted(reason_breakdown.items())),
    }


def build_review_filters(items: Iterable[ReviewItemRecord]) -> dict[str, list[str]]:
    rows = list(items)
    return {
        "statuses": [
            {"value": value, "label_zh": review_status_label_zh(value)}
            for value in sorted({item.current_status for item in rows})
        ],
        "source_types": [
            {"value": value, "label_zh": review_source_type_label_zh(value)}
            for value in sorted({item.source_type for item in rows})
        ],
        "reason_codes": [
            {"value": value, "label_zh": reason_code_label_zh(value)}
            for value in sorted({code for item in rows for code in item.reason_codes})
        ],
        "priority_buckets": [
            {"value": "high", "label_zh": "高优先级"},
            {"value": "medium", "label_zh": "中优先级"},
            {"value": "low", "label_zh": "低优先级"},
        ],
        "apply_compatibility_statuses": [
            {"value": value, "label_zh": review_compatibility_label_zh(value)}
            for value in sorted(
                {item.apply_compatibility_status for item in rows if item.apply_compatibility_status},
                key=_compatibility_sort_rank,
            )
        ],
        "page_numbers": [str(value) for value in sorted({item.page_no for item in rows if item.page_no is not None})],
        "statement_types": sorted({item.statement_type for item in rows if item.statement_type}),
        "providers": sorted({item.provider for item in rows if item.provider}),
        "quick_filters": [dict(item) for item in QUICK_FILTER_OPTIONS],
    }


def build_review_workbench_summary(job: JobRecord, items: Iterable[ReviewItemRecord]) -> dict[str, Any]:
    rows = list(items)
    unresolved_total = sum(1 for item in rows if item.current_status == REVIEW_STATUS_UNRESOLVED)
    resolved_total = len(rows) - unresolved_total
    backend_ready_total = sum(1 for item in rows if item.apply_compatibility_status == REVIEW_COMPATIBILITY_BACKEND_READY)
    backend_partial_total = sum(1 for item in rows if item.apply_compatibility_status == REVIEW_COMPATIBILITY_PARTIAL)
    backend_suggestion_only_total = sum(1 for item in rows if item.apply_compatibility_status == REVIEW_COMPATIBILITY_SUGGESTION_ONLY)
    backend_unsupported_total = sum(1 for item in rows if item.apply_compatibility_status == REVIEW_COMPATIBILITY_UNSUPPORTED)
    evidence_available_total = sum(1 for item in rows if item.evidence_available)
    return {
        "job_id": job.job_id,
        "review_items_total": len(rows),
        "unresolved_total": unresolved_total,
        "resolved_total": resolved_total,
        "deferred_total": sum(1 for item in rows if item.current_status == REVIEW_STATUS_DEFERRED),
        "ignored_total": sum(1 for item in rows if item.current_status == REVIEW_STATUS_IGNORED),
        "reocr_requested_total": sum(1 for item in rows if item.current_status == REVIEW_STATUS_REOCR_REQUESTED),
        "backend_ready_total": backend_ready_total,
        "backend_partial_total": backend_partial_total,
        "backend_suggestion_only_total": backend_suggestion_only_total,
        "backend_unsupported_total": backend_unsupported_total,
        "evidence_available_total": evidence_available_total,
        "high_priority_total": sum(1 for item in rows if item.priority_score >= HIGH_PRIORITY_THRESHOLD),
        "pass": True,
    }


def persist_review_dashboard_artifacts(
    job: JobRecord,
    items: Iterable[ReviewItemRecord],
    source_artifacts: Iterable[ReviewSourceArtifact],
) -> dict[str, Path]:
    rows = list(items)
    artifacts = list(source_artifacts)
    review_dir = get_review_dir(job)
    review_dir.mkdir(parents=True, exist_ok=True)

    counts_summary = build_review_dashboard_summary(rows, artifacts)
    counts_path = review_dir / "review_dashboard_counts_summary.json"
    counts_path.write_text(json.dumps(counts_summary, ensure_ascii=False, indent=2), encoding="utf-8")

    workbench_summary = build_review_workbench_summary(job, rows)
    workbench_summary_path = review_dir / "review_workbench_summary.json"
    workbench_summary_path.write_text(json.dumps(workbench_summary, ensure_ascii=False, indent=2), encoding="utf-8")

    evidence_summary = build_review_evidence_preview_summary(job, rows)
    evidence_path = review_dir / "review_evidence_preview_summary.json"
    evidence_path.write_text(json.dumps(evidence_summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return {
        "counts_summary_path": counts_path,
        "workbench_summary_path": workbench_summary_path,
        "evidence_summary_path": evidence_path,
    }


def build_review_evidence_preview_summary(job: JobRecord, items: Iterable[ReviewItemRecord]) -> dict[str, Any]:
    rows = list(items)
    available_any = 0
    missing_any = 0
    cell_available = 0
    row_available = 0
    table_available = 0
    for item in rows:
        has_cell = resolve_evidence_file(job, item, "cell") is not None
        has_row = resolve_evidence_file(job, item, "row") is not None
        has_table = resolve_evidence_file(job, item, "table") is not None
        if has_cell:
            cell_available += 1
        if has_row:
            row_available += 1
        if has_table:
            table_available += 1
        if has_cell or has_row or has_table:
            available_any += 1
        elif item.evidence_path or item.meta.get("evidence_cell_path") or item.meta.get("evidence_row_path") or item.meta.get("evidence_table_path"):
            missing_any += 1
    return {
        "job_id": job.job_id,
        "review_items_total": len(rows),
        "evidence_preview_available_count": available_any,
        "evidence_preview_missing_count": missing_any,
        "cell_preview_available_count": cell_available,
        "row_preview_available_count": row_available,
        "table_preview_available_count": table_available,
        "pass": available_any > 0,
    }


def _matches_quick_filter(item: ReviewItemRecord, quick_filter: str) -> bool:
    if quick_filter == "high_priority":
        return item.priority_score >= HIGH_PRIORITY_THRESHOLD
    if quick_filter == "backend_ready":
        return item.apply_compatibility_status == REVIEW_COMPATIBILITY_BACKEND_READY
    if quick_filter == "not_backend_ready":
        return item.apply_compatibility_status != REVIEW_COMPATIBILITY_BACKEND_READY
    if quick_filter == "mapping_missing":
        return item.source_type in {"mapping_candidate", "unplaced_fact"} or any(code.startswith("mapping:") for code in item.reason_codes)
    if quick_filter == "ocr_suspicious":
        return any(code == "issue:suspicious_value" or code.startswith("source:") for code in item.reason_codes)
    if quick_filter == "validation_fail":
        return item.source_type == "validation" or any(code.startswith("validation:") for code in item.reason_codes)
    if quick_filter == "evidence_available":
        return item.evidence_available
    return True


def save_review_action(
    settings: WebAppSettings,
    job: JobRecord,
    item: ReviewItemRecord,
    *,
    action_type: str,
    action_value: str,
    reviewer_note: str,
    reviewer_name: str,
) -> ReviewActionRecord:
    if action_type not in REVIEW_ACTION_TYPES:
        raise ValueError(f"不支持的复核动作: {action_type}")
    now = utc_now_iso()
    action = ReviewActionRecord(
        job_id=job.job_id,
        review_item_id=item.review_item_id,
        action_type=action_type.strip(),
        action_value=action_value.strip(),
        reviewer_note=reviewer_note.strip(),
        reviewer_name=reviewer_name.strip(),
        review_status=_derive_review_status(action_type),
        source_type=item.source_type,
        source_ref=item.source_ref or item.source_cell_ref or item.review_item_id,
        created_at=now,
        updated_at=now,
    )
    return upsert_review_action(settings, action)


def bulk_save_review_actions(
    settings: WebAppSettings,
    job: JobRecord,
    *,
    review_item_ids: list[str],
    action_type: str,
    action_value: str,
    reviewer_note: str,
    reviewer_name: str,
) -> dict[str, Any]:
    if action_type not in SUPPORTED_BULK_ACTION_TYPES:
        raise ValueError(f"当前不支持批量动作: {action_type}")

    items, source_artifacts = load_review_items(settings, job)
    persist_review_dashboard_artifacts(job, items, source_artifacts)
    item_by_id = {item.review_item_id: item for item in items}
    requested_ids = [value.strip() for value in review_item_ids if value.strip()]
    if not requested_ids:
        raise ValueError("请至少选择一个复核项。")

    applied_total = 0
    skipped_total = 0
    rejected_total = 0
    rejected_reasons: Counter[str] = Counter()
    created_action_ids: list[str] = []

    for review_item_id in requested_ids:
        item = item_by_id.get(review_item_id)
        if item is None:
            rejected_total += 1
            rejected_reasons["review_item_not_found"] += 1
            continue
        allowed, reason = _validate_bulk_action_for_item(item, action_type)
        if not allowed:
            rejected_total += 1
            rejected_reasons[reason] += 1
            continue
        save_review_action(
            settings,
            job,
            item,
            action_type=action_type,
            action_value=(action_value or item.mapping_code) if action_type == "accept_mapping_candidate" else action_value,
            reviewer_note=reviewer_note,
            reviewer_name=reviewer_name,
        )
        applied_total += 1
        created_action_ids.append(f"{job.job_id}:{item.review_item_id}")

    summary = {
        "job_id": job.job_id,
        "action_type": action_type,
        "requested_total": len(requested_ids),
        "applied_total": applied_total,
        "skipped_total": skipped_total,
        "rejected_total": rejected_total,
        "rejected_reasons": dict(sorted(rejected_reasons.items())),
        "created_action_ids": created_action_ids,
        "pass": applied_total > 0 and rejected_total == 0,
    }
    review_dir = get_review_dir(job)
    review_dir.mkdir(parents=True, exist_ok=True)
    bulk_dir = review_dir / f"bulk_{_timestamp_slug()}"
    bulk_dir.mkdir(parents=True, exist_ok=True)
    summary_path = bulk_dir / "bulk_review_action_summary.json"
    latest_summary_path = review_dir / "bulk_review_action_summary.json"
    _write_json_file(summary_path, summary)
    _write_json_file(latest_summary_path, summary)
    summary["summary_path"] = _repo_relative_or_absolute(summary_path)
    summary["latest_summary_path"] = _repo_relative_or_absolute(latest_summary_path)
    return summary


def _validate_bulk_action_for_item(item: ReviewItemRecord, action_type: str) -> tuple[bool, str]:
    if action_type in {"defer", "ignore"}:
        return True, ""
    if action_type == "request_reocr":
        if item.page_no is not None or item.source_cell_ref or item.source_file:
            return True, ""
        return False, "missing_source_locator"
    if action_type == "mark_not_financial_fact":
        if item.related_fact_ids or item.source_type in {"review_queue", "unplaced_fact"}:
            return True, ""
        return False, "fact_target_missing"
    if action_type == "suppress_false_positive":
        if item.source_type in {"review_queue", "issue", "unplaced_fact"}:
            return True, ""
        return False, "false_positive_not_applicable"
    if action_type == "accept_mapping_candidate":
        if item.apply_compatibility_status != REVIEW_COMPATIBILITY_BACKEND_READY:
            return False, "compatibility_not_backend_ready"
        if not item.mapping_code:
            return False, "mapping_candidate_missing"
        return True, ""
    return False, "unsupported_bulk_action"


def get_review_operations_dir(job: JobRecord) -> Path:
    return get_review_dir(job) / "operations"


def get_latest_review_operation_summary(job: JobRecord) -> dict[str, Any]:
    from .operations import get_latest_review_operation_summary as load_latest_review_operation_summary

    return load_latest_review_operation_summary(job)


def _duration_seconds(started_at: str, finished_at: str) -> float:
    try:
        started = datetime.fromisoformat(started_at.replace("Z", "+00:00"))
        finished = datetime.fromisoformat(finished_at.replace("Z", "+00:00"))
    except ValueError:
        return 0.0
    return max(round((finished - started).total_seconds(), 3), 0.0)


def _write_json_file(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _raise_if_cancel_requested(cancel_requested: Callable[[], bool] | None) -> None:
    if cancel_requested is not None and cancel_requested():
        raise ReviewOperationCancelled("operation_cancelled")


def export_review_actions(settings: WebAppSettings, job: JobRecord) -> dict[str, Any]:
    items, source_artifacts = load_review_items(settings, job)
    persist_review_dashboard_artifacts(job, items, source_artifacts)
    item_by_id = {item.review_item_id: item for item in items}
    actions = list_review_actions(settings, job.job_id)
    export_rows: list[dict[str, Any]] = []
    compatibility_notes: list[str] = []
    backend_ready_total = 0
    backend_partial_total = 0
    backend_suggestion_only_total = 0
    backend_unsupported_total = 0
    missing_review_id_total = 0
    request_reocr_fallback_total = 0
    action_type_mapping_total = 0

    for action in actions:
        item = item_by_id.get(action.review_item_id)
        if item is None:
            continue
        row, note_bundle = _build_export_row(item, action)
        export_rows.append(row)
        if note_bundle["compatibility_status"] == REVIEW_COMPATIBILITY_BACKEND_READY:
            backend_ready_total += 1
        elif note_bundle["compatibility_status"] == REVIEW_COMPATIBILITY_UNSUPPORTED:
            backend_unsupported_total += 1
        elif note_bundle["compatibility_status"] == REVIEW_COMPATIBILITY_SUGGESTION_ONLY:
            backend_suggestion_only_total += 1
        else:
            backend_partial_total += 1
        if note_bundle["compatibility_note"]:
            compatibility_notes.append(note_bundle["compatibility_note"])
        if not row["backend_review_id"]:
            missing_review_id_total += 1
        if note_bundle["used_reocr_fallback"]:
            request_reocr_fallback_total += 1
        if note_bundle["action_type_mapped"]:
            action_type_mapping_total += 1

    review_dir = get_review_dir(job)
    review_dir.mkdir(parents=True, exist_ok=True)
    csv_path = review_dir / "review_actions_filled.csv"
    xlsx_path = review_dir / "review_actions_filled.xlsx"
    summary_path = review_dir / "review_action_export_summary.json"
    compatibility_summary_path = review_dir / "review_action_compatibility_summary.json"

    _write_export_csv(csv_path, export_rows)
    _write_export_workbook(xlsx_path, export_rows)

    summary = {
        "job_id": job.job_id,
        "exported_at": utc_now_iso(),
        "actions_total": len(export_rows),
        "backend_ready_total": backend_ready_total,
        "backend_partial_total": backend_partial_total,
        "backend_suggestion_only_total": backend_suggestion_only_total,
        "backend_unsupported_total": backend_unsupported_total,
        "missing_review_id_total": missing_review_id_total,
        "request_reocr_fallback_total": request_reocr_fallback_total,
        "mapped_action_type_total": action_type_mapping_total,
        "compatibility_notes": sorted({note for note in compatibility_notes if note}),
        "csv_path": _repo_relative_or_absolute(csv_path),
        "xlsx_path": _repo_relative_or_absolute(xlsx_path),
    }
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    compatibility_summary = {
        "job_id": job.job_id,
        "exported_at": utc_now_iso(),
        "actions_total": len(export_rows),
        "backend_ready_total": backend_ready_total,
        "backend_partial_total": backend_partial_total,
        "backend_suggestion_only_total": backend_suggestion_only_total,
        "backend_unsupported_total": backend_unsupported_total,
        "ready_total": backend_ready_total,
        "partial_total": backend_partial_total,
        "suggestion_only_total": backend_suggestion_only_total,
        "unsupported_total": backend_unsupported_total,
        "missing_backend_review_id_total": missing_review_id_total,
        "by_status": dict(
            sorted(
                Counter(str(row.get("apply_compatibility_status", "")).strip() or "unknown" for row in export_rows).items()
            )
        ),
        "incompatibility_reasons": sorted(
            {
                str(row.get("apply_incompatibility_reason", "")).strip()
                for row in export_rows
                if str(row.get("apply_incompatibility_reason", "")).strip()
            }
        ),
        "csv_path": _repo_relative_or_absolute(csv_path),
        "xlsx_path": _repo_relative_or_absolute(xlsx_path),
    }
    compatibility_summary_path.write_text(json.dumps(compatibility_summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return {
        "csv_path": csv_path,
        "xlsx_path": xlsx_path,
        "summary_path": summary_path,
        "compatibility_summary_path": compatibility_summary_path,
        "summary": summary,
        "actions_total": len(export_rows),
    }


def build_review_apply_preview(settings: WebAppSettings, job: JobRecord) -> dict[str, Any]:
    action_file = find_latest_review_actions_file(job)
    if action_file is None:
        raise ValueError("未找到可用的复核动作导出文件，请先生成 review_actions_filled.csv 或 XLSX。")

    items, source_artifacts = load_review_items(settings, job)
    persist_review_dashboard_artifacts(job, items, source_artifacts)
    valid_review_ids = {item.review_id for item in items if item.review_id}
    action_rows = [row for row in parse_review_actions_file(action_file) if str(row.get("action_type", "")).strip()]
    if not action_rows:
        raise ValueError("最近导出的复核动作文件中没有可预览的动作。")

    backend_ready_total = 0
    partial_total = 0
    suggestion_only_total = 0
    unsupported_total = 0
    likely_applied_total = 0
    likely_rejected_total = 0
    warning_messages: set[str] = set()

    for row in action_rows:
        compatibility_status = _normalize_compatibility_status(str(row.get("apply_compatibility_status", "")).strip())
        incompatibility_reason = str(row.get("apply_incompatibility_reason", "")).strip()
        is_valid, reject_reason = validate_action_row(row, valid_review_ids)
        if compatibility_status == REVIEW_COMPATIBILITY_BACKEND_READY:
            backend_ready_total += 1
        elif compatibility_status == REVIEW_COMPATIBILITY_PARTIAL:
            partial_total += 1
        elif compatibility_status == REVIEW_COMPATIBILITY_SUGGESTION_ONLY:
            suggestion_only_total += 1
        else:
            unsupported_total += 1
        if is_valid:
            likely_applied_total += 1
        else:
            likely_rejected_total += 1
            warning_messages.add(reject_reason)
        if incompatibility_reason:
            warning_messages.add(incompatibility_reason)

    summary = {
        "job_id": job.job_id,
        "previewed_at": utc_now_iso(),
        "source_action_file": _repo_relative_or_absolute(action_file),
        "actions_total": len(action_rows),
        "backend_ready_total": backend_ready_total,
        "partial_total": partial_total,
        "suggestion_only_total": suggestion_only_total,
        "unsupported_total": unsupported_total,
        "likely_applied_total": likely_applied_total,
        "likely_rejected_total": likely_rejected_total,
        "warning_messages": sorted(warning_messages),
    }
    preview_path = get_review_dir(job) / "review_apply_preview_summary.json"
    _write_json_file(preview_path, summary)
    summary["summary_path"] = _repo_relative_or_absolute(preview_path)
    return summary


def resolve_evidence_file(job: JobRecord, item: ReviewItemRecord, evidence_kind: str) -> Path | None:
    candidate = ""
    if evidence_kind == "cell":
        candidate = item.meta.get("evidence_cell_path", "") or item.evidence_path
    elif evidence_kind == "row":
        candidate = item.meta.get("evidence_row_path", "")
    elif evidence_kind == "table":
        candidate = item.meta.get("evidence_table_path", "")
    if not candidate:
        return None

    path = Path(candidate)
    if not path.is_absolute():
        path = Path(job.output_dir).resolve().parent / path
    resolved = path.resolve()
    for root in get_allowed_review_roots(job):
        try:
            resolved.relative_to(root)
            return resolved if resolved.exists() and resolved.is_file() else None
        except ValueError:
            continue
    return None


def apply_review_actions_from_web(
    settings: WebAppSettings,
    job: JobRecord,
    *,
    operation_id: str = "",
    progress_callback: Callable[[str, str], None] | None = None,
    log_callback: Callable[[str], None] | None = None,
    cancel_requested: Callable[[], bool] | None = None,
) -> dict[str, Any]:
    progress = progress_callback or (lambda stage, message: None)
    log = log_callback or (lambda message: None)

    _raise_if_cancel_requested(cancel_requested)
    progress("preparing_actions", "正在校验最近一次导出的复核动作。")
    review_dir = get_review_dir(job)
    action_file = find_latest_review_actions_file(job)
    if action_file is None:
        raise ValueError("未找到可用的复核动作导出文件，请先生成 review_actions_filled.csv 或 XLSX。")

    action_rows = [row for row in parse_review_actions_file(action_file) if str(row.get("action_type", "")).strip()]
    if not action_rows:
        raise ValueError("最近导出的复核动作文件中没有可 apply 的动作。")

    apply_id = f"apply_{_timestamp_slug()}"
    apply_dir = review_dir / apply_id
    apply_dir.mkdir(parents=True, exist_ok=True)

    exported_copy_path = apply_dir / action_file.name
    shutil.copy2(action_file, exported_copy_path)
    config_snapshot_dir = _copy_config_snapshot(apply_dir / "config_snapshot")
    log(f"source_action_file={_repo_relative_or_absolute(action_file)}")
    log(f"config_snapshot_dir={_repo_relative_or_absolute(config_snapshot_dir)}")

    _raise_if_cancel_requested(cancel_requested)
    progress("applying_actions", "正在应用复核动作到本次任务配置快照。")
    items, source_artifacts = load_review_items(settings, job)
    persist_review_dashboard_artifacts(job, items, source_artifacts)
    valid_review_ids = [item.review_id for item in items if item.review_id]
    applied_rows, rejected_rows, override_audit_rows, apply_summary = backend_apply_review_actions(
        action_rows=action_rows,
        valid_review_ids=valid_review_ids,
        config_dir=config_snapshot_dir,
    )
    review_decision_summary = build_review_decision_summary(
        applied_rows=applied_rows,
        rejected_rows=rejected_rows,
        touched_files=apply_summary.get("touched_files", []),
    )

    original_rows_by_key: dict[str, dict[str, Any]] = {}
    for row in action_rows:
        for key in _review_action_lookup_keys(row):
            if key and key not in original_rows_by_key:
                original_rows_by_key[key] = row
    applied_export_rows = [_merge_apply_result_row(row, original_rows_by_key, status="applied") for row in applied_rows]
    rejected_export_rows = [_merge_apply_result_row(row, original_rows_by_key, status="rejected") for row in rejected_rows]
    override_audit_export_rows = [_merge_apply_result_row(row, original_rows_by_key, status="audit") for row in override_audit_rows]

    applied_csv_path = apply_dir / "applied_review_actions.csv"
    rejected_csv_path = apply_dir / "rejected_review_actions.csv"
    override_audit_csv_path = apply_dir / "override_audit.csv"
    review_decision_summary_path = apply_dir / "review_decision_summary.json"
    review_apply_summary_path = apply_dir / "review_apply_summary.json"

    _write_generic_csv(applied_csv_path, applied_export_rows)
    _write_generic_csv(rejected_csv_path, rejected_export_rows)
    _write_generic_csv(override_audit_csv_path, override_audit_export_rows)
    _write_json_file(review_decision_summary_path, review_decision_summary)

    apply_status = "applied" if not rejected_rows else "partial" if applied_rows else "rejected"
    ready_total = sum(
        1
        for row in action_rows
        if _normalize_compatibility_status(str(row.get("apply_compatibility_status", "")).strip()) == REVIEW_COMPATIBILITY_BACKEND_READY
    )
    partial_total = sum(
        1
        for row in action_rows
        if _normalize_compatibility_status(str(row.get("apply_compatibility_status", "")).strip()) == REVIEW_COMPATIBILITY_PARTIAL
    )
    suggestion_only_total = sum(
        1
        for row in action_rows
        if _normalize_compatibility_status(str(row.get("apply_compatibility_status", "")).strip()) == REVIEW_COMPATIBILITY_SUGGESTION_ONLY
    )
    unsupported_total = sum(
        1
        for row in action_rows
        if _normalize_compatibility_status(str(row.get("apply_compatibility_status", "")).strip()) == REVIEW_COMPATIBILITY_UNSUPPORTED
    )

    progress("writing_apply_summary", "正在写入复核动作应用摘要。")
    review_apply_summary = {
        "job_id": job.job_id,
        "apply_id": apply_id,
        "operation_id": operation_id,
        "applied_at": utc_now_iso(),
        "source_action_file": _repo_relative_or_absolute(action_file),
        "copied_action_file": _repo_relative_or_absolute(exported_copy_path),
        "apply_dir": _repo_relative_or_absolute(apply_dir),
        "config_snapshot_dir": _repo_relative_or_absolute(config_snapshot_dir),
        "actions_total": len(action_rows),
        "applied_actions_total": len(applied_rows),
        "rejected_actions_total": len(rejected_rows),
        "status": apply_status,
        "status_label": {"applied": "已应用", "partial": "部分应用", "rejected": "全部拒绝"}.get(apply_status, apply_status),
        "touched_files": [_repo_relative_or_absolute(Path(path)) for path in apply_summary.get("touched_files", [])],
        "applied_review_actions_path": _repo_relative_or_absolute(applied_csv_path),
        "rejected_review_actions_path": _repo_relative_or_absolute(rejected_csv_path),
        "override_audit_path": _repo_relative_or_absolute(override_audit_csv_path),
        "review_decision_summary_path": _repo_relative_or_absolute(review_decision_summary_path),
        "backend_ready_total": ready_total,
        "backend_partial_total": partial_total,
        "backend_suggestion_only_total": suggestion_only_total,
        "backend_unsupported_total": unsupported_total,
        "ready_total": ready_total,
        "partial_total": partial_total,
        "suggestion_only_total": suggestion_only_total,
        "unsupported_total": unsupported_total,
        "pass": bool(applied_rows or rejected_rows),
    }
    _write_json_file(review_apply_summary_path, review_apply_summary)
    review_apply_summary["review_apply_summary_path"] = _repo_relative_or_absolute(review_apply_summary_path)
    log(f"applied_actions_total={len(applied_rows)}")
    log(f"rejected_actions_total={len(rejected_rows)}")
    return review_apply_summary


def _resolve_rerun_config_snapshot(
    job: JobRecord,
    rerun_root: Path,
    config_snapshot_dir: Path | None,
) -> Path:
    if config_snapshot_dir is not None:
        resolved = Path(config_snapshot_dir)
        if not resolved.is_absolute():
            resolved = (REPO_ROOT / resolved).resolve()
        return resolved
    latest_apply_summary = get_latest_review_apply_summary(job)
    latest_config_snapshot_dir = Path(str(latest_apply_summary.get("config_snapshot_dir", "") or ""))
    if latest_config_snapshot_dir:
        if not latest_config_snapshot_dir.is_absolute():
            latest_config_snapshot_dir = (REPO_ROOT / latest_config_snapshot_dir).resolve()
        if latest_config_snapshot_dir.exists():
            return latest_config_snapshot_dir
    return _copy_config_snapshot(rerun_root / "config_snapshot")


def rerun_only_from_web(
    settings: WebAppSettings,
    job: JobRecord,
    *,
    operation_id: str = "",
    config_snapshot_dir: Path | None = None,
    progress_callback: Callable[[str, str], None] | None = None,
    log_callback: Callable[[str], None] | None = None,
    cancel_requested: Callable[[], bool] | None = None,
) -> dict[str, Any]:
    progress = progress_callback or (lambda stage, message: None)
    log = log_callback or (lambda message: None)

    _raise_if_cancel_requested(cancel_requested)
    rerun_id = next_rerun_id(job)
    rerun_root = get_rerun_root(job, rerun_id)
    rerun_standardize_dir = rerun_root / "standardize"
    rerun_result_dir = get_rerun_result_root(job, rerun_id)
    rerun_root.mkdir(parents=True, exist_ok=True)
    rerun_standardize_dir.mkdir(parents=True, exist_ok=True)
    rerun_result_dir.mkdir(parents=True, exist_ok=True)

    progress("preparing_rerun", "正在准备重新生成所需的运行目录。")
    resolved_config_snapshot_dir = _resolve_rerun_config_snapshot(job, rerun_root, config_snapshot_dir)
    stdout_path = rerun_root / "standardize_stdout.txt"
    stderr_path = rerun_root / "standardize_stderr.txt"

    _raise_if_cancel_requested(cancel_requested)
    progress("running_standardize", "正在重新生成标准化结果，通常需要几分钟。")
    run_result = _run_patched_standardize_cli(
        settings=settings,
        job=job,
        output_dir=rerun_standardize_dir,
        config_dir=resolved_config_snapshot_dir,
        stdout_path=stdout_path,
        stderr_path=stderr_path,
        cancel_requested=cancel_requested,
        timeout_seconds=settings.operation_timeout_seconds,
    )
    if bool(run_result.get("cancelled")):
        raise ReviewOperationCancelled("operation_cancelled")

    progress("building_rerun_summary", "正在整理重跑结果与前后对比。")
    before_snapshot = build_review_result_snapshot(
        job.job_id,
        Path(job.output_dir),
        Path(job.result_dir),
        job.exit_code or 0,
        write_quality_summary=False,
    )
    after_snapshot = build_review_result_snapshot(
        job.job_id,
        rerun_standardize_dir,
        rerun_result_dir,
        run_result["exit_code"],
        write_quality_summary=True,
    )
    review_rerun_delta = build_review_rerun_delta(before_snapshot, after_snapshot)
    review_rerun_delta_explained = build_review_rerun_delta_explained(
        review_rerun_delta,
        {"applied_actions_total": 0, "rejected_actions_total": 0},
    )

    review_rerun_summary_path = rerun_result_dir / "review_rerun_summary.json"
    review_rerun_delta_path = rerun_result_dir / "review_rerun_delta.json"
    review_rerun_delta_explained_path = rerun_result_dir / "review_rerun_delta_explained.json"
    review_rerun_only_summary_path = rerun_result_dir / "review_rerun_only_summary.json"
    rerun_summary = {
        "job_id": job.job_id,
        "rerun_id": rerun_id,
        "operation_id": operation_id,
        "rerun_at": utc_now_iso(),
        "output_dir": _repo_relative_or_absolute(rerun_standardize_dir),
        "result_dir": _repo_relative_or_absolute(rerun_result_dir),
        "stdout_path": _repo_relative_or_absolute(stdout_path),
        "stderr_path": _repo_relative_or_absolute(stderr_path),
        "config_snapshot_dir": _repo_relative_or_absolute(resolved_config_snapshot_dir),
        "logical_command": run_result["logical_command"],
        "runner_command": run_result["runner_command"],
        **after_snapshot,
    }
    rerun_only_summary = {
        "job_id": job.job_id,
        "rerun_id": rerun_id,
        "operation_id": operation_id,
        "rerun_status": after_snapshot["final_job_status"],
        "rerun_status_label": describe_job_status(after_snapshot["final_job_status"]),
        "original_output_dir": _repo_relative_or_absolute(Path(job.output_dir)),
        "rerun_output_dir": _repo_relative_or_absolute(rerun_standardize_dir),
        "review_rerun_summary_path": _repo_relative_or_absolute(review_rerun_summary_path),
        "review_rerun_delta_path": _repo_relative_or_absolute(review_rerun_delta_path),
        "review_rerun_delta_explained_path": _repo_relative_or_absolute(review_rerun_delta_explained_path),
        "recommended_result_version": rerun_id if after_snapshot["final_job_status"] in SUCCESS_LIKE_JOB_STATUSES else "original",
        "pass": after_snapshot["final_job_status"] in SUCCESS_LIKE_JOB_STATUSES,
    }
    _write_json_file(review_rerun_summary_path, rerun_summary)
    _write_json_file(review_rerun_delta_path, review_rerun_delta)
    _write_json_file(review_rerun_delta_explained_path, review_rerun_delta_explained)
    _write_json_file(review_rerun_only_summary_path, rerun_only_summary)
    log(f"rerun_id={rerun_id}")
    log(f"rerun_status={after_snapshot['final_job_status']}")
    return {
        **rerun_only_summary,
        "review_rerun_only_summary_path": _repo_relative_or_absolute(review_rerun_only_summary_path),
        "stdout_path": _repo_relative_or_absolute(stdout_path),
        "stderr_path": _repo_relative_or_absolute(stderr_path),
        "config_snapshot_dir": _repo_relative_or_absolute(resolved_config_snapshot_dir),
    }


def apply_and_rerun_review_actions_from_web(
    settings: WebAppSettings,
    job: JobRecord,
    *,
    operation_id: str = "",
    progress_callback: Callable[[str, str], None] | None = None,
    log_callback: Callable[[str], None] | None = None,
    cancel_requested: Callable[[], bool] | None = None,
) -> dict[str, Any]:
    progress = progress_callback or (lambda stage, message: None)
    log = log_callback or (lambda message: None)

    apply_summary = apply_review_actions_from_web(
        settings,
        job,
        operation_id=operation_id,
        progress_callback=progress,
        log_callback=log,
        cancel_requested=cancel_requested,
    )
    _raise_if_cancel_requested(cancel_requested)
    rerun_summary = rerun_only_from_web(
        settings,
        job,
        operation_id=operation_id,
        config_snapshot_dir=Path(str(apply_summary["config_snapshot_dir"])),
        progress_callback=progress,
        log_callback=log,
        cancel_requested=cancel_requested,
    )
    rerun_result_dir = get_rerun_result_root(job, str(rerun_summary["rerun_id"]))
    review_apply_and_rerun_summary_path = rerun_result_dir / "review_apply_and_rerun_summary.json"
    combined_summary = {
        "job_id": job.job_id,
        "rerun_id": rerun_summary["rerun_id"],
        "operation_id": operation_id,
        "apply_status": apply_summary["status"],
        "applied_actions_total": apply_summary["applied_actions_total"],
        "rejected_actions_total": apply_summary["rejected_actions_total"],
        "rerun_status": rerun_summary["rerun_status"],
        "rerun_status_label": rerun_summary["rerun_status_label"],
        "original_output_dir": _repo_relative_or_absolute(Path(job.output_dir)),
        "rerun_output_dir": rerun_summary["rerun_output_dir"],
        "review_apply_summary_path": str(apply_summary["review_apply_summary_path"]),
        "review_rerun_summary_path": rerun_summary["review_rerun_summary_path"],
        "review_rerun_delta_path": rerun_summary["review_rerun_delta_path"],
        "review_rerun_delta_explained_path": rerun_summary["review_rerun_delta_explained_path"],
        "recommended_result_version": rerun_summary["recommended_result_version"],
        "pass": rerun_summary["pass"],
    }
    _write_json_file(review_apply_and_rerun_summary_path, combined_summary)
    combined_summary["review_apply_and_rerun_summary_path"] = _repo_relative_or_absolute(review_apply_and_rerun_summary_path)
    return combined_summary


def get_latest_review_apply_summary(job: JobRecord) -> dict[str, Any]:
    review_dir = get_review_dir(job)
    apply_dirs = sorted(
        [path for path in review_dir.glob("apply_*") if path.is_dir()],
        key=lambda path: path.name,
    )
    if not apply_dirs:
        return {}
    summary_path = apply_dirs[-1] / "review_apply_summary.json"
    return load_json(summary_path)


def get_latest_review_apply_preview_summary(job: JobRecord) -> dict[str, Any]:
    return load_json(get_review_dir(job) / "review_apply_preview_summary.json")


def get_latest_review_rerun_summary(job: JobRecord) -> dict[str, Any]:
    rerun_root = get_reruns_root(job)
    rerun_dirs = sorted(
        [path for path in rerun_root.glob("rerun_*") if path.is_dir()],
        key=lambda path: path.name,
    )
    if not rerun_dirs:
        return {}
    summary_path = get_rerun_result_root(job, rerun_dirs[-1].name) / "review_rerun_summary.json"
    return load_json(summary_path)


def get_latest_review_rerun_delta_explained(job: JobRecord) -> dict[str, Any]:
    rerun_root = get_reruns_root(job)
    rerun_dirs = sorted(
        [path for path in rerun_root.glob("rerun_*") if path.is_dir()],
        key=lambda path: path.name,
    )
    if not rerun_dirs:
        return {}
    summary_path = get_rerun_result_root(job, rerun_dirs[-1].name) / "review_rerun_delta_explained.json"
    return load_json(summary_path)


def find_latest_review_actions_file(job: JobRecord) -> Path | None:
    review_dir = get_review_dir(job)
    candidates = [
        review_dir / "review_actions_filled.xlsx",
        review_dir / "review_actions_filled.csv",
    ]
    existing = [path for path in candidates if path.exists() and path.is_file()]
    if not existing:
        return None
    return max(existing, key=lambda path: (path.stat().st_mtime_ns, 1 if path.suffix.lower() == ".xlsx" else 0))


def get_reruns_root(job: JobRecord) -> Path:
    return Path(job.output_dir).resolve().parent / "reruns"


def get_rerun_root(job: JobRecord, rerun_id: str) -> Path:
    return get_reruns_root(job) / rerun_id


def get_rerun_result_root(job: JobRecord, rerun_id: str) -> Path:
    return Path(job.result_dir).resolve() / "reruns" / rerun_id


def next_rerun_id(job: JobRecord) -> str:
    rerun_root = get_reruns_root(job)
    pattern = re.compile(r"rerun_(\d{3})$")
    numbers: list[int] = []
    for path in rerun_root.glob("rerun_*"):
        match = pattern.fullmatch(path.name)
        if match:
            numbers.append(int(match.group(1)))
    return f"rerun_{(max(numbers) + 1 if numbers else 1):03d}"


def build_review_result_snapshot(
    job_id: str,
    output_dir: Path,
    result_dir: Path,
    exit_code: int | None,
    *,
    write_quality_summary: bool,
) -> dict[str, Any]:
    result_dir.mkdir(parents=True, exist_ok=True)
    quality_job = JobRecord(
        job_id=job_id,
        display_name="",
        mode="existing_ocr_outputs",
        provider_mode="cloud_first",
        input_path="",
        source_image_dir="",
        upload_dir="",
        ocr_output_dir="",
        template_path="",
        output_dir=str(output_dir),
        result_dir=str(result_dir),
        log_dir="",
        provider_priority="",
        status="",
        current_stage="",
        progress_summary="",
        created_at="",
        updated_at="",
        started_at="",
        finished_at="",
        error_message="",
        raw_error_message="",
        user_friendly_error="",
        recommended_action="",
        run_id="",
        command_executed="",
        exit_code=exit_code,
        timeout_seconds=0,
    )
    quality_summary = build_job_quality_summary(quality_job, command_exit_code=exit_code)
    if write_quality_summary:
        (result_dir / "job_quality_summary.json").write_text(json.dumps(quality_summary, ensure_ascii=False, indent=2), encoding="utf-8")
    run_summary = load_json(output_dir / "run_summary.json")
    artifact_integrity = load_json(output_dir / "artifact_integrity.json")
    review_summary = load_json(output_dir / "review_summary.json")
    validation_summary = load_json(output_dir / "validation_summary.json")
    return {
        "run_id": str(run_summary.get("run_id", "") or ""),
        "review_total": int(review_summary.get("review_total", run_summary.get("review_total", 0)) or 0),
        "validation_fail_total": int(validation_summary.get("validation_fail_total", run_summary.get("validation_fail_total", 0)) or 0),
        "mapped_facts_ratio": float(run_summary.get("mapped_facts_ratio", 0.0) or 0.0),
        "exportable_facts_total": int(run_summary.get("exportable_facts_total", 0) or 0),
        "workbook_generated": bool(quality_summary.get("workbook_generated", False)),
        "integrity_fail_total": int(artifact_integrity.get("integrity_fail_total", run_summary.get("integrity_fail_total", 0)) or 0),
        "final_job_status": str(quality_summary.get("final_job_status", "") or ""),
        "final_job_status_label": str(quality_summary.get("status_label", "") or ""),
        "quality_summary_path": _repo_relative_or_absolute(result_dir / "job_quality_summary.json"),
    }


def build_review_rerun_delta(before: dict[str, Any], after: dict[str, Any]) -> dict[str, Any]:
    metric_rows: list[dict[str, Any]] = []
    for key in ("review_total", "validation_fail_total", "mapped_facts_ratio", "exportable_facts_total", "integrity_fail_total"):
        before_value = before.get(key, 0)
        after_value = after.get(key, 0)
        metric_rows.append(
            {
                "metric": key,
                "before": before_value,
                "after": after_value,
                "delta": round(float(after_value or 0) - float(before_value or 0), 6),
            }
        )
    return {
        "before": before,
        "after": after,
        "metrics": metric_rows,
        "workbook_generated": {
            "before": bool(before.get("workbook_generated", False)),
            "after": bool(after.get("workbook_generated", False)),
        },
        "run_id": {
            "before": str(before.get("run_id", "") or ""),
            "after": str(after.get("run_id", "") or ""),
        },
        "final_job_status": {
            "before": str(before.get("final_job_status", "") or ""),
            "after": str(after.get("final_job_status", "") or ""),
        },
    }


def build_review_rerun_delta_explained(delta: dict[str, Any], apply_summary: dict[str, Any]) -> dict[str, Any]:
    changed_metrics: list[dict[str, Any]] = []
    unchanged_metrics: list[dict[str, Any]] = []
    metric_labels = {
        "review_total": "待复核项",
        "validation_fail_total": "校验失败",
        "mapped_facts_ratio": "已映射事实比例",
        "exportable_facts_total": "可导出事实数",
        "integrity_fail_total": "完整性失败数",
    }
    for row in delta.get("metrics", []):
        labeled_row = dict(row)
        labeled_row["metric_label_zh"] = metric_labels.get(str(row.get("metric", "")), str(row.get("metric", "")))
        if float(row.get("delta", 0) or 0) == 0:
            unchanged_metrics.append(labeled_row)
        else:
            changed_metrics.append(labeled_row)

    after = dict(delta.get("after", {}))
    why_still_needs_review: list[str] = []
    if str(after.get("final_job_status", "")) == "needs_review":
        if int(after.get("review_total", 0) or 0) > 0:
            why_still_needs_review.append(f"重跑后仍有 {int(after.get('review_total', 0) or 0)} 条待复核项目。")
        if int(after.get("validation_fail_total", 0) or 0) > 0:
            why_still_needs_review.append(f"仍有 {int(after.get('validation_fail_total', 0) or 0)} 条校验失败未消除。")
        if int(after.get("integrity_fail_total", 0) or 0) > 0:
            why_still_needs_review.append(f"仍存在 {int(after.get('integrity_fail_total', 0) or 0)} 条完整性失败。")
        if not why_still_needs_review:
            why_still_needs_review.append("重跑后仍未达到可直接交付状态，需要继续人工判断。")

    if changed_metrics:
        summary_lines = ["本次重跑已更新以下关键指标："]
        for row in changed_metrics:
            summary_lines.append(
                f"{row['metric_label_zh']} 从 {row.get('before')} 变为 {row.get('after')}。"
            )
    else:
        summary_lines = ["本次重跑未改变关键质量指标，结果主要用于保留审计轨迹。"]
    if why_still_needs_review:
        summary_lines.extend(why_still_needs_review)

    recommended_next_action_zh = "可下载最新重跑结果并继续处理剩余复核项。"
    final_status = str(after.get("final_job_status", "") or "")
    if final_status == "succeeded":
        recommended_next_action_zh = "可优先采用最新重跑结果作为推荐交付版本。"
    elif final_status == "succeeded_with_warnings":
        recommended_next_action_zh = "建议先检查警告摘要，再决定是否采用最新重跑结果。"
    elif final_status == "needs_review":
        recommended_next_action_zh = "建议按前后对比和剩余问题原因继续筛选并提交下一轮复核动作。"
    elif final_status == "failed":
        recommended_next_action_zh = "建议先查看 rerun 日志，确认失败阶段后再重试。"

    return {
        "headline_status_before": describe_job_status(str(delta.get("before", {}).get("final_job_status", "") or "")),
        "headline_status_after": describe_job_status(final_status),
        "metrics_changed": changed_metrics,
        "metrics_unchanged": unchanged_metrics,
        "actions_applied_total": int(apply_summary.get("applied_actions_total", 0) or 0),
        "actions_rejected_total": int(apply_summary.get("rejected_actions_total", 0) or 0),
        "why_still_needs_review": why_still_needs_review,
        "user_friendly_summary_zh": " ".join(summary_lines),
        "recommended_next_action_zh": recommended_next_action_zh,
    }


def _normalize_compatibility_status(status: str) -> str:
    normalized = (status or "").strip()
    if normalized == "ready":
        return REVIEW_COMPATIBILITY_BACKEND_READY
    if normalized in {
        REVIEW_COMPATIBILITY_BACKEND_READY,
        REVIEW_COMPATIBILITY_PARTIAL,
        REVIEW_COMPATIBILITY_SUGGESTION_ONLY,
        REVIEW_COMPATIBILITY_UNSUPPORTED,
    }:
        return normalized
    if not normalized:
        return REVIEW_COMPATIBILITY_UNSUPPORTED
    return REVIEW_COMPATIBILITY_UNSUPPORTED


def _copy_config_snapshot(target_dir: Path) -> Path:
    from standardize import cli as standardize_cli

    source_dir = Path(standardize_cli.CONFIG_DIR).resolve()
    if target_dir.exists():
        shutil.rmtree(target_dir)
    shutil.copytree(source_dir, target_dir)
    return target_dir


def _run_patched_standardize_cli(
    *,
    settings: WebAppSettings,
    job: JobRecord,
    output_dir: Path,
    config_dir: Path,
    stdout_path: Path,
    stderr_path: Path,
    cancel_requested: Callable[[], bool] | None = None,
    timeout_seconds: int | None = None,
) -> dict[str, Any]:
    argv = [
        "--input-dir",
        job.input_path,
        "--template",
        job.template_path,
        "--output-dir",
        str(output_dir),
        "--output-run-subdir",
        "none",
        "--provider-priority",
        job.provider_priority,
    ]
    if job.source_image_dir and Path(job.source_image_dir).exists():
        argv.extend(["--source-image-dir", job.source_image_dir])
    argv.extend(list(settings.standardize_flags))

    logical_command = [settings.python_executable, "-m", "standardize.cli", *argv]
    payload = {"config_dir": str(config_dir), "argv": argv}
    runner_command = [
        settings.python_executable,
        "-c",
        PATCHED_STANDARDIZE_CLI_SCRIPT,
        json.dumps(payload, ensure_ascii=False),
    ]
    stdout_path.parent.mkdir(parents=True, exist_ok=True)
    stderr_path.parent.mkdir(parents=True, exist_ok=True)
    effective_timeout_seconds = max(int(timeout_seconds or job.timeout_seconds or settings.job_timeout_seconds), 1)
    try:
        with stdout_path.open("wb") as stdout_handle, stderr_path.open("wb") as stderr_handle:
            process = subprocess.Popen(
                runner_command,
                cwd=str(REPO_ROOT),
                stdout=stdout_handle,
                stderr=stderr_handle,
            )
            started_at = time.monotonic()
            cancelled = False
            while True:
                exit_code = process.poll()
                if exit_code is not None:
                    exit_code = int(exit_code)
                    break
                if cancel_requested is not None and cancel_requested():
                    cancelled = True
                    process.terminate()
                    try:
                        process.wait(timeout=10)
                    except subprocess.TimeoutExpired:
                        process.kill()
                        process.wait(timeout=5)
                    break
                if time.monotonic() - started_at > effective_timeout_seconds:
                    process.kill()
                    raise subprocess.TimeoutExpired(runner_command, timeout=effective_timeout_seconds)
                time.sleep(0.25)
        if cancelled:
            stderr_path.write_text("rerun cancelled by user request\n", encoding="utf-8")
            exit_code = -2
    except subprocess.TimeoutExpired:
        stderr_path.write_text(f"rerun timed out after {effective_timeout_seconds} seconds\n", encoding="utf-8")
        exit_code = -1
    return {
        "exit_code": exit_code,
        "logical_command": subprocess.list2cmdline([str(part) for part in logical_command]),
        "runner_command": subprocess.list2cmdline([str(part) for part in runner_command]),
        "cancelled": exit_code == -2,
    }


def _review_action_lookup_keys(row: dict[str, Any]) -> list[str]:
    review_item_id = str(row.get("review_item_id", "")).strip()
    review_id = str(row.get("review_id", "")).strip()
    action_type = str(row.get("action_type", "")).strip()
    action_value = str(row.get("action_value", "")).strip()
    source_ref = str(row.get("source_ref", "")).strip()
    keys = []
    if review_item_id:
        keys.append(f"review_item_id:{review_item_id}")
    if review_id or action_type or action_value:
        keys.append(f"review_action:{review_id}|{action_type}|{action_value}")
    if review_id or source_ref or action_type:
        keys.append(f"review_source_action:{review_id}|{source_ref}|{action_type}")
    return keys


def _merge_apply_result_row(row: dict[str, Any], original_rows_by_key: dict[str, dict[str, Any]], *, status: str) -> dict[str, Any]:
    original: dict[str, Any] = {}
    for key in _review_action_lookup_keys(row):
        original = original_rows_by_key.get(key, {})
        if original:
            break
    merged = dict(original)
    merged.update({str(key): value for key, value in row.items()})
    merged["applied_status"] = status
    if "backend_review_id" not in merged:
        merged["backend_review_id"] = merged.get("review_id", "")
    return merged


def _write_generic_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in fieldnames:
                fieldnames.append(str(key))
    if not fieldnames:
        fieldnames = ["status"]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: _serialize_value(row.get(key, "")) for key in fieldnames})


def _timestamp_slug() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def count_csv_rows(path: Path) -> int:
    if not path.exists():
        return 0
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        try:
            next(reader)
        except StopIteration:
            return 0
        return sum(1 for _ in reader)


def _load_review_queue_items(path: Path) -> list[ReviewItemRecord]:
    rows: list[ReviewItemRecord] = []
    for row in _read_csv_rows(path):
        reason_codes = _parse_json_list(row.get("reason_codes", ""))
        meta = _parse_json_object(row.get("meta_json", ""))
        mapping_code, mapping_name = _parse_mapping_candidate_text(row.get("mapping_candidates", ""))
        evidence_path, evidence_label = _first_evidence_path(row)
        backend_review_id = row.get("review_id", "").strip()
        stable_source_ref = str(meta.get("source_cell_ref", "")).strip() or row.get("source_file", "").strip()
        review_item_id = backend_review_id or _stable_review_item_id(
            "review_queue",
            row.get("doc_id", ""),
            row.get("page_no", ""),
            row.get("statement_type", ""),
            row.get("row_label_std", ""),
            row.get("period_key", ""),
            stable_source_ref,
        )
        rows.append(
            ReviewItemRecord(
                review_item_id=review_item_id,
                review_id=backend_review_id,
                source_file=row.get("source_file", "").strip(),
                source_type="review_queue",
                priority_score=_to_float(row.get("priority_score"), 0.0),
                reason_code=reason_codes[0] if reason_codes else "",
                reason_codes=reason_codes,
                reason_label_zh=_reason_label_zh(reason_codes[0] if reason_codes else ""),
                doc_id=row.get("doc_id", "").strip(),
                page_no=_to_int(row.get("page_no")),
                statement_type=row.get("statement_type", "").strip(),
                row_label_raw=row.get("row_label_raw", "").strip(),
                row_label_std=row.get("row_label_std", "").strip(),
                mapping_code=mapping_code,
                mapping_name=mapping_name,
                period_key=row.get("period_key", "").strip(),
                value_raw=row.get("value_raw", "").strip(),
                value_num=_to_float_or_none(row.get("value_num")),
                provider=row.get("provider", "").strip(),
                source_cell_ref=str(meta.get("source_cell_ref", "")).strip(),
                evidence_path=evidence_path,
                evidence_label=evidence_label,
                current_status=REVIEW_STATUS_UNRESOLVED,
                source_ref=stable_source_ref or review_item_id,
                related_conflict_ids=_parse_json_list(row.get("related_conflict_ids", "")),
                related_validation_ids=_parse_json_list(row.get("related_validation_ids", "")),
                related_fact_ids=_parse_json_list(row.get("related_fact_ids", "")),
                meta={
                    "evidence_cell_path": row.get("evidence_cell_path", "").strip(),
                    "evidence_row_path": row.get("evidence_row_path", "").strip(),
                    "evidence_table_path": row.get("evidence_table_path", "").strip(),
                    **meta,
                },
            )
        )
    return rows


def _load_issue_items(path: Path) -> list[ReviewItemRecord]:
    rows: list[ReviewItemRecord] = []
    for row in _read_csv_rows(path):
        issue_type = row.get("issue_type", "").strip()
        item_id = _stable_review_item_id(
            "issue",
            row.get("doc_id", ""),
            row.get("page_no", ""),
            row.get("source_cell_ref", ""),
            issue_type,
            row.get("message", ""),
        )
        meta = _parse_json_object(row.get("meta_json", ""))
        reason_code = f"issue:{issue_type}" if issue_type else "issue"
        raw_label = row.get("text_clean", "").strip() or row.get("text_raw", "").strip() or row.get("message", "").strip()
        rows.append(
            ReviewItemRecord(
                review_item_id=item_id,
                review_id="",
                source_file=row.get("source_file", "").strip(),
                source_type="issue",
                priority_score=_priority_from_issue(row.get("severity", "").strip(), issue_type),
                reason_code=reason_code,
                reason_codes=[reason_code],
                reason_label_zh=_reason_label_zh(reason_code),
                doc_id=row.get("doc_id", "").strip(),
                page_no=_to_int(row.get("page_no")),
                row_label_raw=raw_label,
                provider=row.get("provider", "").strip(),
                source_cell_ref=row.get("source_cell_ref", "").strip(),
                current_status=REVIEW_STATUS_UNRESOLVED,
                source_ref=row.get("source_cell_ref", "").strip() or item_id,
                meta=meta,
            )
        )
    return rows


def _load_validation_items(path: Path) -> list[ReviewItemRecord]:
    rows: list[ReviewItemRecord] = []
    for row in _read_csv_rows(path):
        status = row.get("status", "").strip()
        if status not in {"fail", "review"}:
            continue
        refs = _parse_json_list(row.get("evidence_fact_refs", ""))
        first_ref = refs[0] if refs else ""
        ref_meta = _parse_source_cell_ref(first_ref)
        rule_name = row.get("rule_name", "").strip()
        reason_code = f"validation:{rule_name}:{status}" if rule_name else f"validation:{status}"
        rows.append(
            ReviewItemRecord(
                review_item_id=row.get("validation_id", "").strip() or _stable_review_item_id("validation", rule_name, first_ref),
                review_id="",
                source_file="",
                source_type="validation",
                priority_score=5.0 if status == "fail" else 4.0,
                reason_code=reason_code,
                reason_codes=[reason_code],
                reason_label_zh=_reason_label_zh(reason_code),
                doc_id=row.get("doc_id", "").strip() or ref_meta.get("doc_id", ""),
                page_no=ref_meta.get("page_no"),
                statement_type=row.get("statement_type", "").strip(),
                row_label_raw=row.get("message", "").strip(),
                row_label_std=rule_name,
                period_key=row.get("period_key", "").strip(),
                provider=ref_meta.get("provider", ""),
                source_cell_ref=first_ref,
                current_status=REVIEW_STATUS_UNRESOLVED,
                source_ref=first_ref or row.get("validation_id", "").strip(),
                related_validation_ids=[row.get("validation_id", "").strip()],
                meta=_parse_json_object(row.get("meta_json", "")),
            )
        )
    return rows


def _load_conflict_items(path: Path) -> list[ReviewItemRecord]:
    rows: list[ReviewItemRecord] = []
    for row in _read_csv_rows(path):
        decision = row.get("decision", "").strip()
        needs_review = str(row.get("needs_review", "")).strip().lower() in {"1", "true", "yes"}
        if decision not in {"review_required", "unresolved"} and not needs_review:
            continue
        providers = row.get("providers", "").strip()
        candidate_fact_id = _first_fact_id_from_json(row.get("candidate_values_json", "")) or row.get("accepted_fact_id", "").strip()
        reason_code = f"conflict:{decision or 'review_required'}"
        rows.append(
            ReviewItemRecord(
                review_item_id=row.get("conflict_id", "").strip() or _stable_review_item_id("conflict", providers, row.get("row_label_std", "")),
                review_id="",
                source_file="",
                source_type="conflict",
                priority_score=5.0 if decision == "review_required" or needs_review else 4.0,
                reason_code=reason_code,
                reason_codes=[reason_code],
                reason_label_zh=_reason_label_zh(reason_code),
                doc_id=row.get("doc_id", "").strip(),
                page_no=_to_int(row.get("page_no")),
                statement_type=row.get("statement_type", "").strip(),
                row_label_std=row.get("row_label_std", "").strip(),
                period_key=row.get("period_key", "").strip(),
                provider=providers or row.get("accepted_provider", "").strip(),
                value_raw=row.get("validation_delta", "").strip(),
                current_status=REVIEW_STATUS_UNRESOLVED,
                source_ref=row.get("conflict_id", "").strip() or providers,
                related_conflict_ids=[row.get("conflict_id", "").strip()] if row.get("conflict_id", "").strip() else [],
                candidate_conflict_fact_id=candidate_fact_id,
                meta=_parse_json_object(row.get("meta_json", "")),
            )
        )
    return rows


def _load_unplaced_fact_items(path: Path) -> list[ReviewItemRecord]:
    rows: list[ReviewItemRecord] = []
    for row in _read_csv_rows(path):
        reason = row.get("unplaced_reason", "").strip() or "unplaced"
        reason_code = f"unplaced:{reason}"
        rows.append(
            ReviewItemRecord(
                review_item_id=row.get("fact_id", "").strip() or _stable_review_item_id("unplaced", row.get("source_cell_ref", "")),
                review_id="",
                source_file="",
                source_type="unplaced_fact",
                priority_score=3.5 if reason == "unmapped" else 2.5,
                reason_code=reason_code,
                reason_codes=[reason_code],
                reason_label_zh=_reason_label_zh(reason_code),
                doc_id=row.get("doc_id", "").strip(),
                page_no=_to_int(row.get("page_no")),
                statement_type=row.get("statement_type", "").strip(),
                row_label_raw=row.get("row_label_raw", "").strip(),
                row_label_std=row.get("row_label_std", "").strip(),
                mapping_code=row.get("mapping_code", "").strip(),
                mapping_name=row.get("mapping_name", "").strip(),
                period_key=row.get("period_key", "").strip(),
                value_raw=row.get("value_raw", "").strip(),
                value_num=_to_float_or_none(row.get("value_num")),
                provider=row.get("provider", "").strip(),
                source_cell_ref=row.get("source_cell_ref", "").strip(),
                current_status=REVIEW_STATUS_UNRESOLVED,
                source_ref=row.get("source_cell_ref", "").strip() or row.get("fact_id", "").strip(),
                related_fact_ids=[row.get("fact_id", "").strip()] if row.get("fact_id", "").strip() else [],
            )
        )
    return rows


def _load_mapping_candidate_items(path: Path) -> list[ReviewItemRecord]:
    grouped: dict[str, dict[str, str]] = {}
    for row in _read_csv_rows(path):
        group_key = "|".join(
            [
                row.get("doc_id", "").strip(),
                row.get("page_no", "").strip(),
                row.get("source_cell_ref", "").strip(),
                row.get("row_label_std", "").strip() or row.get("row_label_raw", "").strip(),
            ]
        )
        current = grouped.get(group_key)
        if current is None or _candidate_sort_tuple(row) < _candidate_sort_tuple(current):
            grouped[group_key] = row

    rows: list[ReviewItemRecord] = []
    for row in grouped.values():
        review_required = str(row.get("review_required", "")).strip().lower() in {"1", "true", "yes"}
        reason_code = "mapping:candidate_review" if review_required else "mapping:candidate"
        candidate_code = row.get("candidate_code", "").strip()
        candidate_name = row.get("candidate_name", "").strip()
        rows.append(
            ReviewItemRecord(
                review_item_id=_stable_review_item_id("mapping", row.get("source_cell_ref", ""), candidate_code, candidate_name),
                review_id="",
                source_file="",
                source_type="mapping_candidate",
                priority_score=round(2.0 + _to_float(row.get("candidate_score"), 0.0), 3),
                reason_code=reason_code,
                reason_codes=[reason_code],
                reason_label_zh=_reason_label_zh(reason_code),
                doc_id=row.get("doc_id", "").strip(),
                page_no=_to_int(row.get("page_no")),
                statement_type=row.get("statement_type", "").strip(),
                row_label_raw=row.get("row_label_raw", "").strip(),
                row_label_std=row.get("row_label_std", "").strip(),
                mapping_code=candidate_code,
                mapping_name=candidate_name,
                provider=row.get("provider", "").strip(),
                source_cell_ref=row.get("source_cell_ref", "").strip(),
                current_status=REVIEW_STATUS_UNRESOLVED,
                source_ref=row.get("source_cell_ref", "").strip() or candidate_code,
                meta=_parse_json_object(row.get("meta_json", "")),
            )
        )
    return rows


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [{str(key).strip(): ("" if value is None else str(value).strip()) for key, value in row.items()} for row in csv.DictReader(handle)]


def _apply_saved_actions(items: Iterable[ReviewItemRecord], actions: Iterable[ReviewActionRecord]) -> None:
    action_map = {item.review_item_id: item for item in actions}
    for item in items:
        action = action_map.get(item.review_item_id)
        if action is None:
            continue
        item.action_type = action.action_type
        item.action_value = action.action_value
        item.reviewer_note = action.reviewer_note
        item.reviewer_name = action.reviewer_name
        item.current_status = action.review_status or _derive_review_status(action.action_type)


def _normalize_review_item_identity(items: Iterable[ReviewItemRecord]) -> None:
    for item in items:
        if not item.review_item_id:
            item.review_item_id = _stable_review_item_id(
                item.source_type or "review",
                item.doc_id,
                item.page_no,
                item.row_label_std,
                item.row_label_raw,
                item.period_key,
                item.source_cell_ref,
                item.provider,
            )
        if not item.source_ref:
            item.source_ref = item.source_cell_ref or item.review_id or item.review_item_id


def _derive_apply_target_type(action_type: str) -> str:
    if action_type in {"accept_mapping_candidate", "accept_mapping_alias"}:
        return "mapping_alias"
    if action_type == "set_mapping_override":
        return "local_fact_mapping"
    if action_type == "set_conflict_winner":
        return "conflict"
    if action_type == "mark_not_financial_fact":
        return "fact_suppression"
    if action_type == "suppress_false_positive":
        return "ocr_false_positive_suppression"
    if action_type in {"ignore", "defer", "request_reocr"}:
        return "review_item"
    return "review_item"


def _build_export_row(item: ReviewItemRecord, action: ReviewActionRecord) -> tuple[dict[str, Any], dict[str, Any]]:
    export_action_type = action.action_type
    action_type_mapped = False
    action_value = action.action_value
    compatibility_notes: list[str] = []
    used_reocr_fallback = False
    apply_target_type = _derive_apply_target_type(export_action_type)

    if export_action_type == "accept_mapping_candidate":
        export_action_type = "accept_mapping_alias"
        action_type_mapped = True
        action_value = action_value or item.mapping_code
        compatibility_notes.append("accept_mapping_candidate 已映射为 accept_mapping_alias 以兼容现有 backend。")

    if export_action_type == "request_reocr" and not (action_value or item.suggested_reocr_task_id):
        action_value = f"web_reocr_request:{item.review_item_id}"
        used_reocr_fallback = True
        compatibility_notes.append("request_reocr 未提供后端任务 ID，已回退为 Web 侧占位 task_id。")

    if export_action_type == "set_mapping_override":
        action_value = action_value or item.mapping_code
    if export_action_type == "set_conflict_winner":
        action_value = action_value or item.candidate_conflict_fact_id

    row = {header: "" for header in EXPORT_HEADERS}
    row.update(
        {
            "review_id": item.review_id,
            "backend_review_id": item.review_id,
            "priority_score": item.priority_score,
            "source_type": item.source_type,
            "fact_id": item.related_fact_ids[0] if item.related_fact_ids else "",
            "source_cell_ref": item.source_cell_ref,
            "related_conflict_ids": ",".join(item.related_conflict_ids),
            "doc_id": item.doc_id,
            "page_no": item.page_no if item.page_no is not None else "",
            "statement_type": item.statement_type,
            "row_label_raw": item.row_label_raw,
            "row_label_std": item.row_label_std,
            "period_key": item.period_key,
            "value_raw": item.value_raw,
            "value_num": item.value_num if item.value_num is not None else "",
            "provider": item.provider,
            "candidate_mapping_code": item.mapping_code,
            "candidate_mapping_name": item.mapping_name,
            "candidate_conflict_fact_id": item.candidate_conflict_fact_id,
            "candidate_period_override": item.candidate_period_override,
            "suggested_reocr_task_id": item.suggested_reocr_task_id or (action_value if export_action_type == "request_reocr" else ""),
            "action_type": export_action_type,
            "action_value": action_value,
            "reviewer_note": action.reviewer_note,
            "reviewer_name": action.reviewer_name,
            "review_status": action.review_status,
            "applied_status": "",
            "apply_message": "",
            "review_item_id": item.review_item_id,
            "source_ref": action.source_ref or item.source_ref,
            "apply_target_type": apply_target_type,
            "created_at": action.created_at,
            "original_action_type": action.action_type,
        }
    )

    compatibility_status = REVIEW_COMPATIBILITY_BACKEND_READY
    if not item.review_id:
        compatibility_status = REVIEW_COMPATIBILITY_PARTIAL
        compatibility_notes.append("该条目缺少 backend_review_id，直接用于现有 apply-review-actions 时可能被后端拒绝。")
    if export_action_type not in BACKEND_SUPPORTED_ACTIONS:
        compatibility_status = REVIEW_COMPATIBILITY_UNSUPPORTED
        compatibility_notes.append("导出 action_type 仍未被现有 backend 支持。")
    if not item.review_id and export_action_type in {"accept_mapping_alias", "set_mapping_override"}:
        compatibility_status = REVIEW_COMPATIBILITY_SUGGESTION_ONLY
        compatibility_notes.append("映射类动作缺少 backend_review_id，当前仅适合作为建议导出。")
    if export_action_type in {"accept_mapping_alias", "set_mapping_override"} and not row["candidate_mapping_code"] and not row["action_value"]:
        compatibility_status = REVIEW_COMPATIBILITY_PARTIAL
        compatibility_notes.append("映射动作缺少 candidate_mapping_code/action_value。")
    if export_action_type == "set_conflict_winner" and not row["candidate_conflict_fact_id"] and not row["action_value"]:
        compatibility_status = REVIEW_COMPATIBILITY_PARTIAL
        compatibility_notes.append("冲突动作缺少 candidate_conflict_fact_id/action_value。")

    incompatibility_reason = "；".join(dict.fromkeys(note for note in compatibility_notes if note))
    row["apply_compatibility_status"] = compatibility_status
    row["apply_incompatibility_reason"] = incompatibility_reason
    row["compatibility_status"] = compatibility_status
    row["compatibility_note"] = incompatibility_reason
    return row, {
        "compatibility_status": compatibility_status,
        "compatibility_note": incompatibility_reason,
        "used_reocr_fallback": used_reocr_fallback,
        "action_type_mapped": action_type_mapped,
    }


def _write_export_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=EXPORT_HEADERS)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: _serialize_value(row.get(header, "")) for header in EXPORT_HEADERS})


def _write_export_workbook(path: Path, rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Actions"
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(EXPORT_HEADERS))}1"
    worksheet.append(EXPORT_HEADERS)
    for row in rows:
        worksheet.append([_serialize_value(row.get(header, "")) for header in EXPORT_HEADERS])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _serialize_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"), sort_keys=True)
    return value if value is not None else ""


def _sort_key(sort_by: str):
    if sort_by == "page_asc":
        return lambda item: (item.page_no if item.page_no is not None else 10**9, -(item.priority_score or 0.0), item.source_type, item.review_item_id)
    if sort_by == "source_type":
        return lambda item: (item.source_type, -(item.priority_score or 0.0), item.page_no if item.page_no is not None else 10**9, item.review_item_id)
    if sort_by == "compatibility":
        return lambda item: (_compatibility_sort_rank(item.apply_compatibility_status), -(item.priority_score or 0.0), item.review_item_id)
    return _default_sort_key


def _default_sort_key(item: ReviewItemRecord):
    return (-(item.priority_score or 0.0), item.page_no if item.page_no is not None else 10**9, item.source_type, item.review_item_id)


def _derive_review_status(action_type: str) -> str:
    if action_type == "ignore":
        return REVIEW_STATUS_IGNORED
    if action_type == "defer":
        return REVIEW_STATUS_DEFERRED
    if action_type == "request_reocr":
        return REVIEW_STATUS_REOCR_REQUESTED
    if action_type:
        return REVIEW_STATUS_RESOLVED
    return REVIEW_STATUS_UNRESOLVED


def _stable_review_item_id(prefix: str, *parts: object) -> str:
    payload = "|".join(str(part or "").strip() for part in parts)
    digest = hashlib.sha1(payload.encode("utf-8")).hexdigest()[:12]
    return f"{prefix}_{digest}"


def _parse_json_list(raw: str) -> list[str]:
    if not raw:
        return []
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError:
        parsed = [value.strip() for value in raw.split(",") if value.strip()]
    if isinstance(parsed, list):
        return [str(value).strip() for value in parsed if str(value).strip()]
    return [str(parsed).strip()] if str(parsed).strip() else []


def _parse_json_object(raw: str) -> dict[str, Any]:
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _parse_source_cell_ref(value: str) -> dict[str, Any]:
    parts = [part.strip() for part in value.split(":")]
    if len(parts) < 3:
        return {"doc_id": "", "page_no": None, "provider": ""}
    page_no = _to_int(parts[1])
    return {"doc_id": parts[0], "page_no": page_no, "provider": parts[2] if len(parts) > 2 else ""}


def _parse_mapping_candidate_text(raw: str) -> tuple[str, str]:
    text = raw.strip()
    if not text:
        return "", ""
    first = text.split(";", 1)[0].strip()
    if " " not in first:
        return first, ""
    code, remainder = first.split(" ", 1)
    name = remainder.rsplit("(", 1)[0].strip()
    return code.strip(), name


def _first_evidence_path(row: dict[str, str]) -> tuple[str, str]:
    for key, label in (
        ("evidence_cell_path", "单元格证据"),
        ("evidence_row_path", "行证据"),
        ("evidence_table_path", "表格证据"),
    ):
        value = row.get(key, "").strip()
        if value:
            return value, label
    return "", ""


def _priority_from_issue(severity: str, issue_type: str) -> float:
    base = {"error": 5.0, "warning": 3.5, "info": 2.0}.get(severity, 2.5)
    if issue_type == "suspicious_value":
        base += 0.5
    return base


def _reason_label_zh(reason_code: str) -> str:
    return reason_code_label_zh(reason_code)


def _candidate_sort_tuple(row: dict[str, str]) -> tuple[int, float]:
    return (_to_int(row.get("candidate_rank"), 999), -_to_float(row.get("candidate_score"), 0.0))


def _first_fact_id_from_json(raw: str) -> str:
    if not raw:
        return ""
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError:
        return ""
    if isinstance(parsed, dict):
        values = parsed.values()
    elif isinstance(parsed, list):
        values = [parsed]
    else:
        return ""
    for entries in values:
        if not isinstance(entries, list):
            continue
        for entry in entries:
            if isinstance(entry, dict) and str(entry.get("fact_id", "")).strip():
                return str(entry["fact_id"]).strip()
    return ""


def _to_int(value: object, default: int | None = None) -> int | None:
    text = str(value or "").strip()
    if not text:
        return default
    try:
        return int(float(text))
    except ValueError:
        return default


def _to_float(value: object, default: float) -> float:
    text = str(value or "").strip()
    if not text:
        return default
    try:
        return float(text)
    except ValueError:
        return default


def _to_float_or_none(value: object) -> float | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None
