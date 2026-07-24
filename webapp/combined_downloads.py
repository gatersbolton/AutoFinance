from __future__ import annotations

import csv
import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from project_paths import REPO_ROOT, STANDARD_METRICS_GENERATED_ROOT, WEB_GENERATED_ROOT
from .review_quality import display_period_role, has_temporal_key, is_invalid_metric_name


COMBINED_WORKBOOK_DOWNLOAD_NAME = "数据表.xlsx"
PRIMARY_DOWNLOAD_LABEL = "下载数据表"
NUMBER_FORMAT = "#,##0.00"
DATE_FORMAT = "yyyy-mm-dd"

DATA_TOTAL_COLUMNS = [
    "填表日期",
    "当前条目日期",
    "期间类型",
    "公司名",
    "原始指标名",
    "标准指标编码",
    "标准指标名称",
    "指标数值",
    "映射状态",
    "映射方法",
    "口径说明",
    "是否需要人工校对",
]

STANDARD_COLUMNS = [
    "填表日期",
    "当前条目日期",
    "期间类型",
    "公司名",
    "原始指标名",
    "标准指标编码",
    "标准指标名称",
    "指标数值",
    "映射方法",
    "映射状态",
    "口径说明",
    "是否需要人工校对",
]

RAW_COLUMNS = [
    "填表日期",
    "当前条目日期",
    "期间类型",
    "公司名",
    "指标名",
    "指标数值",
]

REVIEW_COLUMNS = [
    "原始指标名",
    "当前映射",
    "映射状态",
    "是否需要人工校对",
    "口径说明",
]

EXPLANATION_LINES = [
    "本工作簿包含原始识别数据和标准化映射结果。",
    "原始数据：未做标准术语映射。",
    "标准化数据：已根据系统规则映射为标准指标。",
    "期间类型用于区分同一填表日期下的期初、期末、本期、上期等口径，入库时建议与填表日期一起作为期间维度。",
    "“是否需要人工校对”为“是”的行建议人工检查。",
    "数值仅做展示格式化，内部数据不应被随意改写。",
]

PERIOD_ROLE_LABELS_ZH = {
    "beginning": "期初数",
    "ending": "期末数",
    "previous_ending": "上期期末",
    "current_point": "当前时点",
    "current_period": "本期",
    "current_year": "本年",
    "previous_period": "上期",
    "previous_year": "上年",
    "amount": "金额",
    "explicit_date": "明确日期",
}


def build_combined_metrics_workbook(
    raw_metrics_path: str | Path | None,
    standardized_metrics_path: str | Path | None,
    mapping_review_items_path: str | Path | None,
    output_path: str | Path,
    metadata: dict[str, Any] | None = None,
    unified_review_actions_path: str | Path | None = None,
) -> dict[str, Any]:
    metadata = metadata or {}
    output = _resolve_path(output_path)
    raw_path = _resolve_optional_path(raw_metrics_path)
    standard_path = _resolve_optional_path(standardized_metrics_path)
    review_path = _resolve_optional_path(mapping_review_items_path)
    warnings: list[str] = []

    raw_rows = _read_table(raw_path, preferred_sheet="raw_metrics") if raw_path and raw_path.exists() else []
    if not raw_rows:
        warnings.append("未找到可用的原始数据。")
    raw_detail_path = raw_path.parent / "raw_metrics_detailed.csv" if raw_path else None
    raw_detail_rows = _read_table(raw_detail_path, preferred_sheet="raw_metrics_detailed") if raw_detail_path and raw_detail_path.exists() else []
    raw_rows = _with_raw_metric_ids(raw_rows, raw_detail_rows)
    raw_rows = _filter_valid_metric_rows(raw_rows)

    standard_rows = (
        _read_table(standard_path, preferred_sheet="standardized_metrics") if standard_path and standard_path.exists() else []
    )
    if not standard_rows:
        warnings.append("未找到可用的标准化数据，工作簿将只包含已生成的数据。")
    standard_detail_path = standard_path.parent / "standardized_metrics_detailed.csv" if standard_path else None
    standard_detail_rows = (
        _read_table(standard_detail_path, preferred_sheet="standardized_metrics_detailed")
        if standard_detail_path and standard_detail_path.exists()
        else []
    )
    standard_rows = _with_raw_metric_ids(standard_rows, standard_detail_rows)
    standard_rows = _filter_valid_metric_rows(standard_rows)

    if (not review_path or not review_path.exists()) and standard_path and standard_path.exists():
        sibling_review = standard_path.parent / "mapping_review_items.csv"
        if sibling_review.exists():
            review_path = sibling_review
    review_rows = _read_table(review_path, preferred_sheet="mapping_review_items") if review_path and review_path.exists() else []
    review_rows = _filter_valid_metric_rows(review_rows)
    action_path = _resolve_optional_path(unified_review_actions_path)
    action_rows = _load_unified_review_actions(action_path)
    overlay_summary = _apply_unified_review_overlays(
        raw_rows,
        standard_rows,
        review_rows,
        action_rows,
        mapping_not_before_timestamp=standard_path.stat().st_mtime if standard_path and standard_path.exists() else None,
    )

    combined_rows = _build_combined_rows(raw_rows, standard_rows)
    normalized_standard_rows = [_normalize_standard_row(row) for row in standard_rows]
    normalized_raw_rows = [_normalize_raw_row(row) for row in raw_rows]
    normalized_review_rows = _build_review_rows(review_rows, standard_rows)

    workbook = Workbook()
    workbook.remove(workbook.active)

    sheets_created: list[str] = []
    counters = {"numeric": 0, "date": 0}

    _write_sheet(workbook, "数据总表", DATA_TOTAL_COLUMNS, combined_rows, counters)
    sheets_created.append("数据总表")

    if normalized_standard_rows:
        _write_sheet(workbook, "标准化数据", STANDARD_COLUMNS, normalized_standard_rows, counters)
        sheets_created.append("标准化数据")

    if normalized_raw_rows:
        _write_sheet(workbook, "原始数据", RAW_COLUMNS, normalized_raw_rows, counters)
        sheets_created.append("原始数据")

    if normalized_review_rows:
        _write_sheet(workbook, "术语映射校对", REVIEW_COLUMNS, normalized_review_rows, counters)
        sheets_created.append("术语映射校对")

    _write_explanation_sheet(workbook, metadata, warnings)
    sheets_created.append("说明")

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)

    summary = {
        "pass": output.exists(),
        "workbook_path": str(output),
        "output_path": str(output),
        "sheets_created": sheets_created,
        "raw_metrics_source": str(raw_path) if raw_path else "",
        "standardized_metrics_source": str(standard_path) if standard_path else "",
        "mapping_review_items_source": str(review_path) if review_path else "",
        "raw_rows_total": len(raw_rows),
        "standardized_rows_total": len(standard_rows),
        "combined_rows_total": len(combined_rows),
        "numeric_cells_formatted_total": counters["numeric"],
        "date_cells_formatted_total": counters["date"],
        "numeric_formatting_pass": counters["numeric"] > 0 or not combined_rows,
        "date_formatting_pass": counters["date"] > 0 or not combined_rows,
        "primary_download_label": PRIMARY_DOWNLOAD_LABEL,
        "advanced_downloads_available": bool(raw_path or standard_path),
        "warnings": warnings,
        "path_hygiene_pass": _path_hygiene_pass(output, metadata),
        "unified_review_actions_source": str(action_path) if action_path else "",
        **overlay_summary,
    }
    for key in ("doc_id", "job_id", "stage"):
        if key in metadata:
            summary[key] = metadata[key]

    summary_path = _resolve_optional_path(metadata.get("summary_path"))
    if summary_path:
        _write_json(summary_path, summary)
    return summary


def build_workbook_preview(workbook_path: str | Path, *, max_rows_per_sheet: int = 300, max_cols_per_sheet: int = 40) -> dict[str, Any]:
    path = _resolve_path(workbook_path)
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheets: list[dict[str, Any]] = []
        for worksheet in workbook.worksheets:
            rows: list[list[str]] = []
            for row in worksheet.iter_rows(max_row=max_rows_per_sheet, max_col=max_cols_per_sheet):
                rows.append([_format_preview_cell(cell.value, cell.number_format) for cell in row])
            while rows and all(value == "" for value in rows[-1]):
                rows.pop()
            sheets.append(
                {
                    "name": worksheet.title,
                    "rows": rows,
                    "rows_rendered": len(rows),
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "truncated": worksheet.max_row > max_rows_per_sheet or worksheet.max_column > max_cols_per_sheet,
                }
            )
    finally:
        workbook.close()
    return {
        "workbook_path": str(path),
        "filename": path.name,
        "sheets": sheets,
        "sheets_total": len(sheets),
    }


def _format_preview_cell(value: Any, number_format: str = "") -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        format_text = str(number_format or "")
        if ".00" in format_text or "0.00" in format_text:
            return f"{float(value):,.2f}"
        return f"{value:,}" if isinstance(value, int) else str(value)
    return str(value)


def _with_raw_metric_ids(rows: list[dict[str, Any]], detail_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for index, row in enumerate(rows):
        item = dict(row)
        detail = detail_rows[index] if index < len(detail_rows) else {}
        raw_metric_id = _first(item, "raw_metric_id") or _first(detail, "raw_metric_id", "source_cell_ref")
        if raw_metric_id:
            item["_raw_metric_id"] = str(raw_metric_id)
        if not _first(item, "期间类型", "period_role"):
            period_role = _period_role_from_row(detail)
            if period_role:
                item["期间类型"] = period_role
        result.append(item)
    return result


def _filter_valid_metric_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    # Missing dates/period roles are retained and marked for review later.
    # Numeric-only metric labels remain extraction noise and are filtered.
    return [row for row in rows if not is_invalid_metric_name(_row_metric_name(row))]


def _row_metric_name(row: dict[str, Any]) -> Any:
    return _first(row, "原始指标名", "指标名", "metric_name", "original_metric_name", "raw_metric_name")


def _row_has_temporal_key(row: dict[str, Any]) -> bool:
    return has_temporal_key(
        _first(row, "当前条目日期", "item_date"),
        _first(row, "period_role_norm", "期间类型", "period_role"),
        _first(row, "period_role_raw") or _period_role_from_row(row),
    )


def _temporal_review_required(row: dict[str, Any]) -> bool:
    return not _row_has_temporal_key(row)


def _append_review_note(value: Any, note: str) -> str:
    text = str(value or "").strip()
    if not text:
        return note
    return text if note in text else f"{text}；{note}"


def _load_unified_review_actions(path: Path | None) -> list[dict[str, Any]]:
    if path is None or not path.exists():
        return []
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return []
    return [item for item in payload if isinstance(item, dict)] if isinstance(payload, list) else []


def _apply_unified_review_overlays(
    raw_rows: list[dict[str, Any]],
    standard_rows: list[dict[str, Any]],
    review_rows: list[dict[str, Any]],
    action_rows: list[dict[str, Any]],
    *,
    mapping_not_before_timestamp: float | None = None,
) -> dict[str, Any]:
    value_overrides, mapping_overrides = _latest_unified_review_overrides(
        action_rows,
        mapping_not_before_timestamp=mapping_not_before_timestamp,
    )
    for rows in (raw_rows, standard_rows, review_rows):
        for row in rows:
            raw_metric_id = _row_raw_metric_id(row)
            if not raw_metric_id:
                continue
            value_action = value_overrides.get(raw_metric_id)
            if value_action:
                _apply_value_override(row, value_action)
            mapping_action = mapping_overrides.get(raw_metric_id)
            if mapping_action:
                _apply_mapping_override(row, mapping_action)
    return {
        "unified_review_actions_total": len(action_rows),
        "unified_review_value_overrides_total": len(value_overrides),
        "unified_review_mapping_overrides_total": len(mapping_overrides),
    }


def _latest_unified_review_overrides(
    action_rows: list[dict[str, Any]], *, mapping_not_before_timestamp: float | None = None
) -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]]]:
    value_overrides: dict[str, dict[str, Any]] = {}
    mapping_overrides: dict[str, dict[str, Any]] = {}
    for action in action_rows:
        raw_metric_id = str(action.get("raw_metric_id", "") or "")
        if not raw_metric_id:
            continue
        edit_type = str(action.get("edit_type", "") or "")
        if edit_type in {"value_change", "reset_value"}:
            value_overrides[raw_metric_id] = action
        elif edit_type in {"mapping_change", "reset_mapping"}:
            if mapping_not_before_timestamp is not None:
                action_timestamp = _action_created_timestamp(action)
                if action_timestamp is not None and action_timestamp < mapping_not_before_timestamp:
                    continue
            mapping_overrides[raw_metric_id] = action
    return value_overrides, mapping_overrides


def _row_raw_metric_id(row: dict[str, Any]) -> str:
    return str(row.get("_raw_metric_id") or row.get("raw_metric_id") or row.get("source_cell_ref") or "")


def _apply_value_override(row: dict[str, Any], action: dict[str, Any]) -> None:
    value = str(action.get("new_value", "") or "")
    for key in ("指标数值", "metric_value", "value_raw"):
        if key in row:
            row[key] = value


def _apply_mapping_override(row: dict[str, Any], action: dict[str, Any]) -> None:
    code = str(action.get("new_code", "") or "")
    name = str(action.get("new_name", "") or "")
    for key in ("标准指标编码", "standard_code", "candidate_code", "current_code", "final_code", "selected_code"):
        if key in row:
            row[key] = code
    for key in ("标准指标名称", "standard_name", "candidate_name", "current_name", "final_name", "selected_name"):
        if key in row:
            row[key] = name
    mapped = bool(code or name)
    for key in ("映射状态", "mapping_status"):
        if key in row:
            row[key] = "mapped" if mapped else "unmapped"
    for key in ("映射方法", "mapping_method", "candidate_method"):
        if key in row:
            row[key] = "manual_once" if mapped else "none"
    for key in ("是否需要人工校对", "review_required"):
        if key in row:
            row[key] = "否" if mapped else "是"
    if "mapping_decision" in row:
        row["mapping_decision"] = "accept_once" if mapped else "reject"


def _action_created_timestamp(action: dict[str, Any]) -> float | None:
    created_at = str(action.get("created_at", "") or "").strip()
    if not created_at:
        return None
    try:
        return datetime.fromisoformat(created_at.replace("Z", "+00:00")).timestamp()
    except ValueError:
        return None


def _build_combined_rows(raw_rows: list[dict[str, Any]], standard_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if standard_rows:
        raw_lookup = {_row_raw_metric_id(row): row for row in raw_rows if _row_raw_metric_id(row)}
        return [_normalize_combined_from_standard(row, raw_lookup.get(_row_raw_metric_id(row), {})) for row in standard_rows]
    return [_normalize_combined_from_raw(row) for row in raw_rows]


def _normalize_combined_from_standard(row: dict[str, Any], raw_row: dict[str, Any] | None = None) -> dict[str, Any]:
    normalized = _normalize_standard_row(row)
    if not normalized.get("期间类型") and raw_row:
        normalized["期间类型"] = _normalize_raw_row(raw_row).get("期间类型", "")
    return {column: normalized.get(column, "") for column in DATA_TOTAL_COLUMNS}


def _normalize_combined_from_raw(row: dict[str, Any]) -> dict[str, Any]:
    raw = _normalize_raw_row(row)
    temporal_review = _temporal_review_required(row)
    return {
        "填表日期": raw.get("填表日期", ""),
        "当前条目日期": raw.get("当前条目日期", ""),
        "期间类型": raw.get("期间类型", ""),
        "公司名": raw.get("公司名", ""),
        "原始指标名": raw.get("指标名", ""),
        "标准指标编码": "",
        "标准指标名称": "",
        "指标数值": raw.get("指标数值", ""),
        "映射状态": "未标准化",
        "映射方法": "",
        "口径说明": "日期或期间缺失，需人工校对。" if temporal_review else "",
        "是否需要人工校对": "是" if temporal_review else "",
    }


def _normalize_standard_row(row: dict[str, Any]) -> dict[str, Any]:
    review_required = row.get("是否需要人工校对")
    if review_required in (None, ""):
        review_required = _review_required_from_status(_first(row, "映射状态", "mapping_status"))
    temporal_review = _temporal_review_required(row)
    if temporal_review:
        review_required = "是"
    notes = _first(row, "口径说明", "notes", "issue_reason")
    if temporal_review:
        notes = _append_review_note(notes, "日期或期间缺失，需人工校对。")
    return {
        "填表日期": _first(row, "填表日期", "fill_date"),
        "当前条目日期": _first(row, "当前条目日期", "item_date"),
        "期间类型": _first(row, "期间类型", "period_role") or _period_role_from_row(row),
        "公司名": _first(row, "公司名", "company_name"),
        "原始指标名": _first(row, "原始指标名", "指标名", "metric_name", "original_metric_name"),
        "标准指标编码": _first(row, "标准指标编码", "standard_code", "candidate_code", "current_code"),
        "标准指标名称": _first(row, "标准指标名称", "standard_name", "candidate_name", "current_name"),
        "指标数值": _first(row, "指标数值", "metric_value", "value_raw"),
        "映射方法": _first(row, "映射方法", "mapping_method", "candidate_method"),
        "映射状态": _first(row, "映射状态", "mapping_status"),
        "口径说明": notes,
        "是否需要人工校对": _normalize_yes_no(review_required),
    }


def _normalize_raw_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "填表日期": _first(row, "填表日期", "fill_date"),
        "当前条目日期": _first(row, "当前条目日期", "item_date"),
        "期间类型": _first(row, "期间类型", "period_role") or _period_role_from_row(row),
        "公司名": _first(row, "公司名", "company_name"),
        "指标名": _first(row, "指标名", "metric_name", "原始指标名"),
        "指标数值": _first(row, "指标数值", "metric_value", "value_raw"),
    }


def _build_review_rows(review_rows: list[dict[str, Any]], standard_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    source_rows = review_rows if review_rows else standard_rows
    result: list[dict[str, Any]] = []
    for row in source_rows:
        code = _first(row, "candidate_code", "标准指标编码", "current_code", "standard_code")
        name = _first(row, "candidate_name", "标准指标名称", "current_name", "standard_name")
        status = _first(row, "mapping_status", "映射状态")
        review_required = _first(row, "是否需要人工校对")
        if review_required == "":
            review_required = _review_required_from_status(status)
        temporal_review = _temporal_review_required(row)
        if temporal_review:
            review_required = "是"
        notes = _first(row, "口径说明", "issue_reason", "notes")
        if temporal_review:
            notes = _append_review_note(notes, "日期或期间缺失，需人工校对。")
        result.append(
            {
                "原始指标名": _first(row, "原始指标名", "original_metric_name", "指标名", "metric_name"),
                "当前映射": f"{code} {name}".strip(),
                "映射状态": status,
                "是否需要人工校对": _normalize_yes_no(review_required),
                "口径说明": notes,
            }
        )
    return result


def _write_sheet(
    workbook: Workbook,
    title: str,
    columns: list[str],
    rows: list[dict[str, Any]],
    counters: dict[str, int],
) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column, "") for column in columns])
    _format_sheet(sheet, columns, counters)


def _write_explanation_sheet(workbook: Workbook, metadata: dict[str, Any], warnings: Iterable[str]) -> None:
    sheet = workbook.create_sheet("说明")
    sheet["A1"] = "说明"
    sheet["A1"].font = Font(bold=True)
    row_index = 2
    for line in EXPLANATION_LINES:
        sheet.cell(row=row_index, column=1, value=line)
        row_index += 1
    if metadata.get("doc_id") or metadata.get("job_id"):
        row_index += 1
        sheet.cell(row=row_index, column=1, value=f"文档编号：{metadata.get('doc_id', '') or metadata.get('job_id', '')}")
        row_index += 1
    warning_list = list(warnings)
    if warning_list:
        row_index += 1
        sheet.cell(row=row_index, column=1, value="生成提示")
        sheet.cell(row=row_index, column=1).font = Font(bold=True)
        for warning in warning_list:
            row_index += 1
            sheet.cell(row=row_index, column=1, value=warning)
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 72


def _format_sheet(sheet, columns: list[str], counters: dict[str, int]) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="EAF2EF")
    border = Border(bottom=Side(style="thin", color="C9D4CE"))
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    date_column_indexes = {index for index, column in enumerate(columns, start=1) if column in {"填表日期", "当前条目日期"}}
    number_column_indexes = {index for index, column in enumerate(columns, start=1) if column == "指标数值"}

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if cell.column in number_column_indexes:
                parsed_number = _parse_number(cell.value)
                if parsed_number is not None:
                    cell.value = parsed_number
                    cell.number_format = NUMBER_FORMAT
                    cell.alignment = Alignment(horizontal="right")
                    counters["numeric"] += 1
                continue
            if cell.column in date_column_indexes:
                parsed_date = _parse_date(cell.value)
                if parsed_date is not None:
                    cell.value = parsed_date
                    cell.number_format = DATE_FORMAT
                    counters["date"] += 1

    for index, column in enumerate(columns, start=1):
        letter = get_column_letter(index)
        width = _column_width(sheet, index, column)
        sheet.column_dimensions[letter].width = width


def _column_width(sheet, index: int, column: str) -> float:
    preferred = {
        "填表日期": 13,
        "当前条目日期": 13,
        "期间类型": 12,
        "公司名": 24,
        "指标名": 26,
        "原始指标名": 28,
        "标准指标编码": 16,
        "标准指标名称": 28,
        "指标数值": 16,
        "映射状态": 14,
        "映射方法": 14,
        "是否需要人工校对": 16,
        "口径说明": 30,
        "当前映射": 32,
    }.get(column, 14)
    max_width = preferred
    for cell in sheet.iter_rows(min_col=index, max_col=index, values_only=True):
        value = "" if cell[0] is None else str(cell[0])
        max_width = max(max_width, min(_display_width(value) + 2, 42))
    return min(max_width, 42)


def _read_table(path: Path, *, preferred_sheet: str = "") -> list[dict[str, Any]]:
    if path.suffix.lower() == ".xlsx":
        return _read_xlsx(path, preferred_sheet=preferred_sheet)
    return _read_csv(path)


def _read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def _read_xlsx(path: Path, *, preferred_sheet: str = "") -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[preferred_sheet] if preferred_sheet and preferred_sheet in workbook.sheetnames else workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = ["" if value is None else str(value).strip() for value in rows[0]]
    result: list[dict[str, Any]] = []
    for values in rows[1:]:
        row = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers)) if headers[index]}
        if any(value not in (None, "") for value in row.values()):
            result.append(row)
    return result


def _parse_number(value: Any) -> float | int | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if not text:
        return None
    if "%" in text:
        return None
    negative = text.startswith("(") and text.endswith(")")
    if negative:
        text = text[1:-1].strip()
    text = text.replace(",", "").replace("，", "").replace(" ", "").replace("\u00a0", "")
    text = text.replace("￥", "").replace("¥", "").replace("元", "")
    if not re.fullmatch(r"[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?", text):
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    if negative:
        number = -number
    if number.is_integer() and "." not in text and "e" not in text.lower():
        return int(number)
    return number


def _parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if "T" in text:
        try:
            return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
        except ValueError:
            pass
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    match = re.fullmatch(r"(\d{4})年(\d{1,2})月(\d{1,2})日", text)
    if match:
        try:
            return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        except ValueError:
            return None
    return None


def _first(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        value = row.get(key)
        if value not in (None, ""):
            return value
    return ""


def _period_role_from_row(row: dict[str, Any]) -> str:
    norm = str(_first(row, "period_role_norm", "期间类型", "period_role") or "").strip()
    raw = str(_first(row, "period_role_raw", "header_path") or "").strip()
    return display_period_role(norm, raw)


def _normalize_yes_no(value: Any) -> str:
    if isinstance(value, bool):
        return "是" if value else "否"
    text = str(value or "").strip()
    if not text:
        return ""
    lowered = text.lower()
    if lowered in {"true", "1", "yes", "y", "review_required", "unmapped"} or text == "是":
        return "是"
    if lowered in {"false", "0", "no", "n", "mapped"} or text == "否":
        return "否"
    return text


def _review_required_from_status(status: Any) -> str:
    text = str(status or "").strip().lower()
    if not text:
        return ""
    return "否" if text == "mapped" else "是"


def _display_width(value: str) -> int:
    return sum(2 if ord(char) > 127 else 1 for char in value)


def _resolve_optional_path(value: Any) -> Path | None:
    if value in (None, ""):
        return None
    return _resolve_path(value)


def _resolve_path(value: str | Path) -> Path:
    path = Path(value)
    if not path.is_absolute():
        path = REPO_ROOT / path
    return path.resolve()


def _path_hygiene_pass(output: Path, metadata: dict[str, Any]) -> bool:
    raw_roots = metadata.get("path_hygiene_roots") or [WEB_GENERATED_ROOT, STANDARD_METRICS_GENERATED_ROOT]
    roots = [_resolve_path(root) for root in raw_roots]
    return any(_is_within(output, root) for root in roots)


def _is_within(path: Path, root: Path) -> bool:
    try:
        path.resolve().relative_to(root.resolve())
        return True
    except ValueError:
        return False


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
