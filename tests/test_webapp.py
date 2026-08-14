from __future__ import annotations

import csv
import json
import os
import shutil
import subprocess
import tempfile
import time
import unittest
from datetime import datetime, timedelta, timezone
from pathlib import Path
from unittest import mock

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from project_paths import REPO_ROOT
from scripts.deployment_check import main as deployment_check_main
from standard_map.store import LocalMappingStore
from standard_map.search import search_standard_terms
from webapp.accounting_export import build_accounting_workbook
from webapp.config import WebAppSettings
from webapp.deployment import run_deployment_preflight
from webapp.document_library import (
    build_delete_plan,
    document_to_job,
    load_document,
    metadata_path_for,
    update_document_status,
    write_document,
)
from webapp.document_models import STATUS_COMPLETED, STATUS_QUEUED
from webapp.db import (
    claim_next_queued_job,
    get_job,
    get_review_operation,
    init_db,
    list_review_actions,
    recover_abandoned_running_jobs,
    update_job,
    update_review_operation,
)
from webapp.jobs import build_job_stage_flow, discover_output_files, job_stage_label_zh
from webapp.main import create_app
from webapp.models import (
    DOCUMENT_PIPELINE_STAGE_RAW_METRICS,
    DOCUMENT_PIPELINE_STAGE_STANDARD_METRICS,
    JOB_MODE_DOCUMENT_PIPELINE,
    JobRecord,
)
from webapp.review import export_review_actions, get_review_dir, load_review_items, filter_review_items
from webapp.runner import run_worker_once
from webapp.simple_flow import (
    _label_matches,
    _normalize_ocr_label,
    _resolve_mapping_term_bbox_json,
    _resolve_source_table_cell_bbox_json,
    combined_download_summary_path,
    load_mapping_review_items,
    load_raw_review_items,
    load_simple_flow_state,
    mapping_review_dir,
    raw_review_dir,
    raw_step_summary_path,
    standard_step_summary_path,
    source_preview_rotation_degrees,
)
from webapp.unified_review import (
    format_metric_number,
    load_unified_review_items,
    parse_metric_number_input,
    unified_review_dir,
)


class WebAppTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory()
        self.temp_path = Path(self.tempdir.name)
        self.corpus_root = self.temp_path / "corpus"
        self.runtime_root = self.temp_path / "generated" / "web"
        self.raw_metrics_root = self.temp_path / "generated" / "raw_metrics"
        self.standard_metrics_root = self.temp_path / "generated" / "standard_metrics"
        self.template_path = REPO_ROOT / "data" / "templates" / "会计报表.xlsx"
        self.secret_path = self.temp_path / "secret"
        self.settings = self.make_settings()
        self.settings.ensure_directories()
        init_db(self.settings)
        self.sample_input_dir = self._create_minimal_ocr_input()
        self.client_cm = TestClient(create_app(self.settings))
        self.client = self.client_cm.__enter__()

    def tearDown(self) -> None:
        self.client_cm.__exit__(None, None, None)
        self.tempdir.cleanup()

    def make_settings(self, **overrides) -> WebAppSettings:
        defaults = dict(
            env_mode="dev",
            runtime_root=self.runtime_root,
            uploads_root=self.runtime_root / "uploads",
            jobs_root=self.runtime_root / "jobs",
            results_root=self.runtime_root / "results",
            logs_root=self.runtime_root / "logs",
            deletions_root=self.runtime_root / "deletions",
            raw_metrics_root=self.raw_metrics_root,
            standard_metrics_root=self.standard_metrics_root,
            db_path=self.runtime_root / "webapp.sqlite3",
            corpus_root=self.corpus_root,
            library_root=self.corpus_root / "library",
            template_path=self.template_path,
            secret_path=self.secret_path,
            deepseek_env_path=self.temp_path / "deepseek.env",
            enable_local_worker=False,
            auto_run_upload_ocr=False,
            worker_poll_seconds=1,
            job_timeout_seconds=120,
            auth_required=False,
            admin_password="",
        )
        defaults.update(overrides)
        return WebAppSettings(**defaults)

    def _write_secret_file(self, *, aliyun: bool = True, tencent: bool = False, secret_value: str = "demo-secret") -> None:
        lines: list[str] = []
        if aliyun:
            lines.extend(
                [
                    "aliyun:",
                    "  AccessKeyId: demo-id",
                    f"  AccessKeySecret: {secret_value}",
                ]
            )
        if tencent:
            lines.extend(
                [
                    "",
                    "tencent:",
                    "  SecretId: demo-tencent-id",
                    f"  SecretKey: {secret_value}",
                ]
            )
        self.secret_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    def _create_minimal_ocr_input(self) -> Path:
        target_doc_root = self.corpus_root / "CASE1"
        input_pdf_dir = target_doc_root / "input"
        input_pdf_dir.mkdir(parents=True, exist_ok=True)
        (input_pdf_dir / "sample.pdf").write_bytes(b"%PDF-1.4\n%mock\n")

        ocr_dir = target_doc_root / "ocr_outputs"
        doc_dir = ocr_dir / "aliyun_table" / "demo_doc"
        raw_dir = doc_dir / "raw"
        raw_dir.mkdir(parents=True, exist_ok=True)
        (raw_dir / "page_0001.json").write_text("{}", encoding="utf-8")
        (doc_dir / "result.json").write_text(
            json.dumps(
                {
                    "provider": "aliyun_table",
                    "pages": [{"page_number": 1, "text": "mock", "raw_file": "raw/page_0001.json"}],
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        return ocr_dir

    def _create_job(self, display_name: str = "mock job") -> str:
        response = self.client.post(
            "/jobs",
            data={
                "mode": "existing_ocr_outputs",
                "display_name": display_name,
                "existing_ocr_path": str(self.sample_input_dir),
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 303)
        return response.headers["location"].rstrip("/").split("/")[-1]

    def _write_fake_workbook(self, path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        worksheet["A1"] = "mock"
        workbook.create_sheet("_meta_summary")
        workbook.save(path)

    def _write_json(self, path: Path, payload: dict[str, object]) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    def _write_csv(self, path: Path, rows: list[dict[str, object]], fieldnames: list[str]) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            for row in rows:
                writer.writerow(row)

    def _column_index(self, worksheet, header: str) -> int:
        for cell in worksheet[1]:
            if cell.value == header:
                return cell.column
        raise AssertionError(f"missing header: {header}")

    def _create_raw_metrics_fixture(
        self,
        *,
        metric_name: str = "往来款",
        metric_value: str = "100",
        evidence_path: Path | None = None,
        source_file: Path | str | None = None,
        bbox_json: str = '[{"x":1,"y":2},{"x":3,"y":4}]',
        confidence: str = "",
        text_confidence: str = "",
        value_confidence: str = "",
        period_role_norm: str = "ending",
        period_role_raw: str = "期末数",
        unit_raw: str = "元",
        value_raw: str = "",
        statement_type: str = "balance_sheet",
    ) -> Path:
        self.raw_metrics_root.mkdir(parents=True, exist_ok=True)
        tempdir = tempfile.TemporaryDirectory(dir=self.raw_metrics_root)
        self.addCleanup(tempdir.cleanup)
        run_dir = Path(tempdir.name) / "RUN_WEB_TEST"
        run_dir.mkdir(parents=True, exist_ok=True)
        raw_path = run_dir / "raw_metrics.csv"
        self._write_csv(
            raw_path,
            [
                {
                    "填表日期": "2022-12-31",
                    "当前条目日期": "2022-12-31",
                    "公司名": "AAA有限公司",
                    "指标名": metric_name,
                    "指标数值": metric_value,
                }
            ],
            ["填表日期", "当前条目日期", "公司名", "指标名", "指标数值"],
        )
        source_pdf = evidence_path or (self.corpus_root / "CASE1" / "input" / "sample.pdf")
        detailed_source_file = source_file if source_file is not None else "fixture.json"
        self._write_csv(
            run_dir / "raw_metrics_detailed.csv",
            [
                {
                    "source_cell_ref": "CASE1:1:aliyun_table:0:1-1:2-2",
                    "page_no": "1",
                    "bbox_json": bbox_json,
                    "text_confidence": text_confidence,
                    "value_confidence": value_confidence or confidence,
                    "evidence_path": str(source_pdf),
                    "source_file": str(detailed_source_file),
                    "provider": "aliyun_table",
                    "doc_id": "CASE1",
                    "table_id": "0",
                    "logical_subtable_id": "0_sub1",
                    "row_index": "1",
                    "col_index": "2",
                    "row_label_clean": metric_name,
                    "period_role_raw": period_role_raw,
                    "period_role_norm": period_role_norm,
                    "statement_type": statement_type,
                    "statement_name_raw": "利润表" if statement_type == "income_statement" else "资产负债表",
                    "value_type": "amount",
                    "value_raw": value_raw or metric_value,
                    "unit_raw": unit_raw,
                    "unit_multiplier": {"元": "1", "千元": "1000", "万元": "10000", "亿元": "100000000"}.get(unit_raw, "1"),
                    "confidence": confidence,
                }
            ],
            [
                "source_cell_ref",
                "page_no",
                "bbox_json",
                "text_confidence",
                "value_confidence",
                "evidence_path",
                "source_file",
                "provider",
                "doc_id",
                "table_id",
                "logical_subtable_id",
                "row_index",
                "col_index",
                "row_label_clean",
                "period_role_raw",
                "period_role_norm",
                "statement_type",
                "statement_name_raw",
                "value_type",
                "value_raw",
                "unit_raw",
                "unit_multiplier",
                "confidence",
            ],
        )
        self.addCleanup(lambda: shutil.rmtree(self.standard_metrics_root / Path(tempdir.name).name, ignore_errors=True))
        return raw_path

    def _create_two_table_raw_metrics_fixture(self) -> Path:
        self.raw_metrics_root.mkdir(parents=True, exist_ok=True)
        tempdir = tempfile.TemporaryDirectory(dir=self.raw_metrics_root)
        self.addCleanup(tempdir.cleanup)
        run_dir = Path(tempdir.name) / "RUN_WEB_TEST"
        run_dir.mkdir(parents=True, exist_ok=True)
        raw_path = run_dir / "raw_metrics.csv"
        rows = [
            {
                "填表日期": "2022-12-31",
                "当前条目日期": "2022-12-31",
                "公司名": "AAA有限公司",
                "指标名": "货币资金",
                "指标数值": "100",
            },
            {
                "填表日期": "2022-12-31",
                "当前条目日期": "2022-12-31",
                "公司名": "AAA有限公司",
                "指标名": "短期借款",
                "指标数值": "200",
            },
        ]
        self._write_csv(raw_path, rows, ["填表日期", "当前条目日期", "公司名", "指标名", "指标数值"])
        source_pdf = self.corpus_root / "CASE1" / "input" / "sample.pdf"
        detailed_rows = [
            {
                "source_cell_ref": "CASE1:1:aliyun_table:1:1-1:2-2",
                "page_no": "1",
                "bbox_json": '[{"x":1,"y":2},{"x":3,"y":4}]',
                "evidence_path": str(source_pdf),
                "source_file": "fixture.json",
                "provider": "aliyun_table",
                "doc_id": "CASE1",
                "table_id": "1",
                "logical_subtable_id": "1_sub1",
                "row_index": "1",
                "col_index": "2",
                "row_label_clean": "货币资金",
                "period_role_raw": "期末数",
                "period_role_norm": "ending",
                "statement_type": "balance_sheet",
                "statement_name_raw": "资产负债表",
            },
            {
                "source_cell_ref": "CASE1:1:aliyun_table:2:1-1:2-2",
                "page_no": "1",
                "bbox_json": '[{"x":5,"y":6},{"x":7,"y":8}]',
                "evidence_path": str(source_pdf),
                "source_file": "fixture.json",
                "provider": "aliyun_table",
                "doc_id": "CASE1",
                "table_id": "2",
                "logical_subtable_id": "2_sub1",
                "row_index": "1",
                "col_index": "2",
                "row_label_clean": "短期借款",
                "period_role_raw": "期末数",
                "period_role_norm": "ending",
                "statement_type": "balance_sheet",
                "statement_name_raw": "资产负债表",
            },
        ]
        self._write_csv(
            run_dir / "raw_metrics_detailed.csv",
            detailed_rows,
            [
                "source_cell_ref",
                "page_no",
                "bbox_json",
                "evidence_path",
                "source_file",
                "provider",
                "doc_id",
                "table_id",
                "logical_subtable_id",
                "row_index",
                "col_index",
                "row_label_clean",
                "period_role_raw",
                "period_role_norm",
                "statement_type",
                "statement_name_raw",
            ],
        )
        self.addCleanup(lambda: shutil.rmtree(self.standard_metrics_root / Path(tempdir.name).name, ignore_errors=True))
        return raw_path

    def _upload_library_pdf(self, filename: str = "A公司财务报表.pdf") -> str:
        response = self.client.post(
            "/documents/upload",
            files=[("uploaded_files", (filename, b"%PDF-1.4\n%mock\n%%EOF\n", "application/pdf"))],
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 303)
        doc_dirs = sorted((self.corpus_root / "library").iterdir(), key=lambda path: path.name)
        self.assertTrue(doc_dirs)
        return doc_dirs[-1].name

    def _write_tiny_library_ocr(self, doc_id: str) -> None:
        document = load_document(self.settings, doc_id, refresh=False)
        provider_doc_dir = Path(document.ocr_output_dir) / "aliyun_table" / "demo_doc"
        raw_dir = provider_doc_dir / "raw"
        raw_dir.mkdir(parents=True, exist_ok=True)
        cells = [
            (1, 0, 0, 0, 0, "项目"),
            (2, 0, 0, 1, 1, "行次"),
            (3, 0, 0, 2, 2, "期初数"),
            (4, 0, 0, 3, 3, "期末数"),
            (5, 1, 1, 0, 0, "货币资金"),
            (6, 1, 1, 1, 1, "1"),
            (7, 1, 1, 2, 2, "100"),
            (8, 1, 1, 3, 3, "200"),
        ]
        raw_payload = {
            "Data": {
                "content": "资产负债表\n编制单位：AAA有限公司\n2022年12月31日\n单位：元",
                "tableHeadTail": [
                    {
                        "head": ["资产负债表", "编制单位：AAA有限公司", "2022年12月31日", "单位：元"],
                        "tail": [],
                    }
                ],
                "prism_tablesInfo": [
                    {
                        "tableId": "1",
                        "xCellSize": 4,
                        "yCellSize": 2,
                        "cellInfos": [
                            {
                                "tableCellId": cell_id,
                                "ysc": row_start,
                                "yec": row_end,
                                "xsc": col_start,
                                "xec": col_end,
                                "word": text,
                                "pos": [
                                    {"x": 10 + col_start * 80, "y": 20 + row_start * 30},
                                    {"x": 70 + col_end * 80, "y": 20 + row_start * 30},
                                    {"x": 70 + col_end * 80, "y": 45 + row_end * 30},
                                    {"x": 10 + col_start * 80, "y": 45 + row_end * 30},
                                ],
                            }
                            for cell_id, row_start, row_end, col_start, col_end, text in cells
                        ],
                    }
                ],
            }
        }
        (raw_dir / "page_0001.json").write_text(json.dumps(raw_payload, ensure_ascii=False, indent=2), encoding="utf-8")
        (provider_doc_dir / "result.json").write_text(
            json.dumps(
                {
                    "provider": "aliyun_table",
                    "pages": [
                        {
                            "page_number": 1,
                            "text": "资产负债表\n编制单位：AAA有限公司\n2022年12月31日\n单位：元",
                            "raw_file": "raw/page_0001.json",
                        }
                    ],
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        update_document_status(self.settings, doc_id, ocr_status=STATUS_COMPLETED)

    def _attach_raw_summary_to_job(self, job_id: str, raw_path: Path) -> None:
        job = get_job(self.settings, job_id)
        self.assertIsNotNone(job)
        payload = {
            "pass": True,
            "run_id": raw_path.parent.name,
            "doc_id": raw_path.parent.parent.name,
            "output_dir": str(raw_path.parent),
            "raw_metrics_csv": str(raw_path),
            "raw_metrics_xlsx": str(raw_path.parent / "raw_metrics.xlsx"),
            "raw_metrics_detailed_csv": str(raw_path.parent / "raw_metrics_detailed.csv"),
            "output_files": [str(raw_path), str(raw_path.parent / "raw_metrics_detailed.csv")],
        }
        raw_step_summary_path(job).parent.mkdir(parents=True, exist_ok=True)
        raw_step_summary_path(job).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    def _attach_raw_summary_to_document(self, doc_id: str, raw_path: Path) -> None:
        document = load_document(self.settings, doc_id, refresh=False)
        job = document_to_job(self.settings, document)
        payload = {
            "pass": True,
            "run_id": raw_path.parent.name,
            "doc_id": raw_path.parent.parent.name,
            "output_dir": str(raw_path.parent),
            "raw_metrics_csv": str(raw_path),
            "raw_metrics_xlsx": str(raw_path.parent / "raw_metrics.xlsx"),
            "raw_metrics_detailed_csv": str(raw_path.parent / "raw_metrics_detailed.csv"),
            "output_files": [str(raw_path), str(raw_path.parent / "raw_metrics_detailed.csv")],
        }
        raw_step_summary_path(job).parent.mkdir(parents=True, exist_ok=True)
        raw_step_summary_path(job).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    def _run_next_worker_item(self) -> None:
        run_worker_once(self.settings)

    def _prepare_review_job(self, *, include_optional: bool = True, outside_evidence: bool = False) -> str:
        job_id = self._create_job("review job")
        job = get_job(self.settings, job_id)
        self.assertIsNotNone(job)
        update_job(self.settings, job_id, status="needs_review", current_stage="completed", progress_summary="ready for review")
        output_dir = Path(job.output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        self._write_fake_workbook(output_dir / "会计报表_填充结果.xlsx")
        self._write_json(
            output_dir / "run_summary.json",
            {
                "run_id": "RUN_REVIEW_JOB_001",
                "review_total": 2,
                "validation_fail_total": 1,
                "mapped_facts_ratio": 0.4,
                "exportable_facts_total": 3,
                "integrity_fail_total": 0,
            },
        )
        self._write_json(output_dir / "pipeline_completion_summary.json", {"status": "success", "last_successful_stage": "export"})
        self._write_json(output_dir / "artifact_integrity.json", {"integrity_fail_total": 0, "integrity_review_total": 0})
        self._write_json(output_dir / "review_summary.json", {"review_total": 2})
        self._write_json(output_dir / "validation_summary.json", {"validation_fail_total": 1})
        self._write_json(Path(job.result_dir) / "job_summary.json", {"job_id": job_id})
        self._write_json(Path(job.result_dir) / "job_quality_summary.json", {"final_job_status": "needs_review"})
        self._write_json(Path(job.result_dir) / "job_log_bundle.json", {"log_files": []})

        review_pack_dir = output_dir / "review_pack"
        review_pack_dir.mkdir(parents=True, exist_ok=True)
        inside_evidence = review_pack_dir / "REV_case_1_cell.png"
        inside_evidence.write_bytes(
            b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x02\x00\x00\x00\x90wS\xde"
            b"\x00\x00\x00\x0cIDAT\x08\xd7c\xf8\xcf\xc0\x00\x00\x03\x01\x01\x00\xc9\xfe\x92\xef\x00\x00\x00\x00IEND\xaeB`\x82"
        )
        outside_path = self.temp_path / "outside_evidence.png"
        outside_path.write_bytes(b"outside")
        evidence_path = outside_path if outside_evidence else inside_evidence

        self._write_csv(
            output_dir / "review_queue.csv",
            [
                {
                    "review_id": "REV_case_1",
                    "priority_score": "6.0",
                    "reason_codes": json.dumps(["mapping:unmapped"], ensure_ascii=False),
                    "doc_id": "CASE1",
                    "page_no": "1",
                    "statement_type": "balance_sheet",
                    "row_label_raw": "货币资金",
                    "row_label_std": "货币资金",
                    "period_key": "2022-12-31__期末数",
                    "value_raw": "100",
                    "value_num": "100.0",
                    "provider": "aliyun_table",
                    "source_file": str(output_dir / "raw" / "page_0001.json"),
                    "bbox": "",
                    "related_fact_ids": json.dumps(["F_001"], ensure_ascii=False),
                    "related_conflict_ids": json.dumps(["CON_001"], ensure_ascii=False),
                    "related_validation_ids": json.dumps(["VAL_001"], ensure_ascii=False),
                    "mapping_candidates": "ZT_001 货币资金 (manual,0.99)",
                    "evidence_cell_path": str(evidence_path),
                    "evidence_row_path": "",
                    "evidence_table_path": "",
                    "meta_json": json.dumps({"source_cell_ref": "CASE1:1:aliyun_table:0:1-1:1-1"}, ensure_ascii=False),
                }
            ],
            [
                "review_id",
                "priority_score",
                "reason_codes",
                "doc_id",
                "page_no",
                "statement_type",
                "row_label_raw",
                "row_label_std",
                "period_key",
                "value_raw",
                "value_num",
                "provider",
                "source_file",
                "bbox",
                "related_fact_ids",
                "related_conflict_ids",
                "related_validation_ids",
                "mapping_candidates",
                "evidence_cell_path",
                "evidence_row_path",
                "evidence_table_path",
                "meta_json",
            ],
        )
        self._write_fake_workbook(output_dir / "review_workbook.xlsx")

        if include_optional:
            self._write_csv(
                output_dir / "issues.csv",
                [
                    {
                        "doc_id": "CASE1",
                        "page_no": "2",
                        "provider": "aliyun_table",
                        "source_file": str(output_dir / "raw" / "page_0002.json"),
                        "table_id": "0",
                        "logical_subtable_id": "0_sub1",
                        "source_cell_ref": "CASE1:2:aliyun_table:0:2-2:2-2",
                        "issue_type": "suspicious_value",
                        "severity": "warning",
                        "message": "expected_numeric_but_unparseable",
                        "text_raw": "-",
                        "text_clean": "-",
                        "status": "open",
                        "meta_json": "{}",
                    }
                ],
                [
                    "doc_id",
                    "page_no",
                    "provider",
                    "source_file",
                    "table_id",
                    "logical_subtable_id",
                    "source_cell_ref",
                    "issue_type",
                    "severity",
                    "message",
                    "text_raw",
                    "text_clean",
                    "status",
                    "meta_json",
                ],
            )
            self._write_csv(
                output_dir / "validation_results.csv",
                [
                    {
                        "validation_id": "VAL_001",
                        "doc_id": "CASE1",
                        "statement_type": "balance_sheet",
                        "period_key": "2022-12-31__期末数",
                        "rule_name": "subtotal_check",
                        "rule_type": "equation",
                        "lhs_value": "1",
                        "rhs_value": "2",
                        "diff_value": "1",
                        "tolerance": "0.01",
                        "status": "fail",
                        "evidence_fact_refs": json.dumps(["CASE1:3:aliyun_table:0:3-3:3-3"], ensure_ascii=False),
                        "message": "subtotal mismatch",
                        "meta_json": "{}",
                    }
                ],
                [
                    "validation_id",
                    "doc_id",
                    "statement_type",
                    "period_key",
                    "rule_name",
                    "rule_type",
                    "lhs_value",
                    "rhs_value",
                    "diff_value",
                    "tolerance",
                    "status",
                    "evidence_fact_refs",
                    "message",
                    "meta_json",
                ],
            )
            self._write_csv(
                output_dir / "mapping_candidates.csv",
                [
                    {
                        "doc_id": "CASE1",
                        "page_no": "1",
                        "provider": "aliyun_table",
                        "statement_type": "balance_sheet",
                        "row_label_raw": "货币资金",
                        "row_label_std": "货币资金",
                        "normalized_label": "货币资金",
                        "candidate_code": "ZT_001",
                        "candidate_name": "货币资金",
                        "candidate_rank": "1",
                        "candidate_score": "0.99",
                        "candidate_method": "manual",
                        "relation_type": "",
                        "review_required": "True",
                        "source_cell_ref": "CASE1:1:aliyun_table:0:1-1:1-1",
                        "meta_json": "{}",
                    }
                ],
                [
                    "doc_id",
                    "page_no",
                    "provider",
                    "statement_type",
                    "row_label_raw",
                    "row_label_std",
                    "normalized_label",
                    "candidate_code",
                    "candidate_name",
                    "candidate_rank",
                    "candidate_score",
                    "candidate_method",
                    "relation_type",
                    "review_required",
                    "source_cell_ref",
                    "meta_json",
                ],
            )
        return job_id

    def _fake_review_rerun(self, profile: str = "improved"):
        def _runner(
            *,
            settings,
            job,
            output_dir: Path,
            config_dir: Path,
            stdout_path: Path,
            stderr_path: Path,
            cancel_requested=None,
            timeout_seconds=None,
        ):
            stdout_path.parent.mkdir(parents=True, exist_ok=True)
            stderr_path.parent.mkdir(parents=True, exist_ok=True)
            stdout_path.write_text(f"fake rerun for {profile}\n", encoding="utf-8")
            stderr_path.write_text("", encoding="utf-8")
            output_dir.mkdir(parents=True, exist_ok=True)
            self._write_fake_workbook(output_dir / "会计报表_填充结果.xlsx")
            review_total = 1 if profile == "improved" else 2
            validation_fail_total = 0 if profile == "improved" else 1
            mapped_ratio = 0.7 if profile == "improved" else 0.4
            exportable_total = 5 if profile == "improved" else 3
            self._write_json(
                output_dir / "run_summary.json",
                {
                    "run_id": "RUN_RERUN_001",
                    "review_total": review_total,
                    "validation_fail_total": validation_fail_total,
                    "mapped_facts_ratio": mapped_ratio,
                    "exportable_facts_total": exportable_total,
                    "integrity_fail_total": 0,
                },
            )
            self._write_json(output_dir / "artifact_integrity.json", {"run_id": "RUN_RERUN_001", "integrity_fail_total": 0, "integrity_review_total": 0})
            self._write_json(output_dir / "review_summary.json", {"run_id": "RUN_RERUN_001", "review_total": review_total})
            self._write_json(output_dir / "validation_summary.json", {"run_id": "RUN_RERUN_001", "validation_fail_total": validation_fail_total})
            self._write_json(output_dir / "pipeline_completion_summary.json", {"run_id": "RUN_RERUN_001", "status": "success", "last_successful_stage": "export"})
            return {
                "exit_code": 0,
                "logical_command": "python -m standardize.cli --output-dir ...",
                "runner_command": "python -c ...",
            }

        return _runner

    def _fake_failed_review_rerun(self):
        def _runner(
            *,
            settings,
            job,
            output_dir: Path,
            config_dir: Path,
            stdout_path: Path,
            stderr_path: Path,
            cancel_requested=None,
            timeout_seconds=None,
        ):
            stdout_path.parent.mkdir(parents=True, exist_ok=True)
            stderr_path.parent.mkdir(parents=True, exist_ok=True)
            stdout_path.write_text("fake rerun failed\n", encoding="utf-8")
            stderr_path.write_text("rerun failed\n", encoding="utf-8")
            return {
                "exit_code": -1,
                "logical_command": "python -m standardize.cli --output-dir ...",
                "runner_command": "python -c ...",
                "cancelled": False,
            }

        return _runner

    def _fake_subprocess(self, profile: str = "clean_success"):
        def _runner(*, command, stdout_path: Path, stderr_path: Path, timeout_seconds: int):
            stdout_path.parent.mkdir(parents=True, exist_ok=True)
            stderr_path.parent.mkdir(parents=True, exist_ok=True)
            stdout_path.write_text(f"fake run for {profile}\n", encoding="utf-8")
            stderr_path.write_text("", encoding="utf-8")

            if "-m" in command and "standardize.cli" in command:
                output_dir = Path(command[command.index("--output-dir") + 1])
                self._write_fake_standardize_outputs(output_dir, profile)
                return mock.Mock(returncode=0)

            return mock.Mock(returncode=0)

        return _runner

    def _fake_failed_subprocess(self, raw_error: str):
        def _runner(*, command, stdout_path: Path, stderr_path: Path, timeout_seconds: int):
            stdout_path.parent.mkdir(parents=True, exist_ok=True)
            stderr_path.parent.mkdir(parents=True, exist_ok=True)
            stdout_path.write_text("", encoding="utf-8")
            stderr_path.write_text(raw_error, encoding="utf-8")
            return mock.Mock(returncode=1)

        return _runner

    def _write_fake_standardize_outputs(self, output_dir: Path, profile: str) -> None:
        output_dir.mkdir(parents=True, exist_ok=True)
        self._write_fake_workbook(output_dir / "会计报表_填充结果.xlsx")
        self._write_json(
            output_dir / "run_summary.json",
            {
                "run_id": "RUN_TEST_001",
                "integrity_fail_total": 0,
                "review_total": 0,
                "validation_fail_total": 0,
            },
        )
        self._write_json(
            output_dir / "artifact_integrity.json",
            {
                "run_id": "RUN_TEST_001",
                "integrity_fail_total": 0,
                "integrity_review_total": 0,
                "checks_total": 3,
            },
        )
        self._write_json(
            output_dir / "validation_summary.json",
            {"run_id": "RUN_TEST_001", "validation_fail_total": 0},
        )
        self._write_json(
            output_dir / "review_summary.json",
            {"run_id": "RUN_TEST_001", "review_total": 0},
        )
        self._write_json(
            output_dir / "pipeline_completion_summary.json",
            {"run_id": "RUN_TEST_001", "status": "success", "current_stage": "", "last_successful_stage": "export"},
        )
        self._write_json(
            output_dir / "full_run_contract_summary.json",
            {"run_id": "RUN_TEST_001", "contract_fail_total": 0},
        )
        (output_dir / "issues.csv").write_text("issue\n", encoding="utf-8")
        (output_dir / "validation_results.csv").write_text("result\n", encoding="utf-8")

        if profile == "warning":
            self._write_json(
                output_dir / "artifact_integrity.json",
                {
                    "run_id": "RUN_TEST_001",
                    "integrity_fail_total": 0,
                    "integrity_review_total": 2,
                    "checks_total": 3,
                },
            )
        elif profile == "needs_review":
            self._write_json(
                output_dir / "review_summary.json",
                {"run_id": "RUN_TEST_001", "review_total": 4},
            )
            self._write_json(
                output_dir / "validation_summary.json",
                {"run_id": "RUN_TEST_001", "validation_fail_total": 2},
            )
            (output_dir / "review_queue.csv").write_text("review\n", encoding="utf-8")
            self._write_fake_workbook(output_dir / "review_workbook.xlsx")
        elif profile == "missing_workbook":
            (output_dir / "会计报表_填充结果.xlsx").unlink(missing_ok=True)

    def test_app_starts_and_home_page_returns_200(self):
        response = self.client.get("/")
        self.assertEqual(response.status_code, 200)
        self.assertIn("财务报表数据提取", response.text)

    def test_simplified_home_page_hides_old_technical_labels(self):
        response = self.client.get("/")
        self.assertEqual(response.status_code, 200)
        self.assertIn("上传新的 PDF", response.text)
        self.assertIn("已保存 PDF", response.text)
        for text in ("新建任务", "待复核项目", "operation", "provider conflict", "raw JSON", "task id", "queue", "debug"):
            self.assertNotIn(text, response.text)

    def test_new_job_page_returns_200(self):
        response = self.client.get("/jobs/new")
        self.assertEqual(response.status_code, 200)
        self.assertIn("新建任务", response.text)

    def test_document_library_home_returns_200(self):
        response = self.client.get("/")
        self.assertEqual(response.status_code, 200)
        self.assertIn("已保存 PDF", response.text)

    def test_document_upload_saves_pdf_and_metadata_under_library(self):
        doc_id = self._upload_library_pdf("B公司审计报告.pdf")
        document = load_document(self.settings, doc_id, refresh=False)
        self.assertTrue(Path(document.pdf_path).exists())
        self.assertTrue(str(Path(document.pdf_path).resolve()).startswith(str((self.corpus_root / "library" / doc_id / "input").resolve())))
        self.assertEqual(document.original_filename, "B公司审计报告.pdf")
        self.assertTrue(Path(document.metadata_path).exists())

    def test_document_batch_uploads_each_pdf_separately_and_queues_individual_jobs(self):
        self._write_secret_file(aliyun=True)
        created = self.client.post(
            "/api/document-batches",
            data={"expected_files": "2"},
        )
        self.assertEqual(created.status_code, 201)
        batch = created.json()
        for index, filename in enumerate(("甲公司年报.pdf", "乙公司年报.pdf")):
            uploaded = self.client.post(
                batch["upload_url"],
                data={"upload_index": str(index)},
                files={
                    "uploaded_file": (
                        filename,
                        b"%PDF-1.4\n%mock\n%%EOF\n",
                        "application/pdf",
                    )
                },
            )
            self.assertEqual(uploaded.status_code, 201)

        queued = self.client.post(batch["queue_url"])
        self.assertEqual(queued.status_code, 200)
        summary = queued.json()
        self.assertEqual(summary["counts"]["uploaded"], 2)
        self.assertEqual(summary["counts"]["queued"], 2)
        self.assertEqual([item["upload_index"] for item in summary["items"]], [0, 1])
        doc_ids = [item["doc_id"] for item in summary["items"]]
        self.assertEqual(len(set(doc_ids)), 2)
        for item in summary["items"]:
            job = get_job(self.settings, item["job_id"])
            self.assertIsNotNone(job)
            self.assertEqual(job.mode, JOB_MODE_DOCUMENT_PIPELINE)
            self.assertEqual(job.status, "queued")
            document = load_document(self.settings, item["doc_id"], refresh=False)
            self.assertEqual(document.ocr_status, STATUS_QUEUED)
            self.assertTrue(Path(document.pdf_path).exists())

        detail = self.client.get(batch["detail_url"])
        self.assertEqual(detail.status_code, 200)
        self.assertIn("低配服务器一次只处理一份文件", detail.text)
        self.assertIn("甲公司年报.pdf", detail.text)
        self.assertIn("乙公司年报.pdf", detail.text)

    def test_document_batch_upload_index_is_idempotent(self):
        created = self.client.post(
            "/api/document-batches",
            data={"expected_files": "1"},
        ).json()
        request_kwargs = {
            "data": {"upload_index": "0"},
            "files": {
                "uploaded_file": (
                    "同一文件.pdf",
                    b"%PDF-1.4\n%mock\n%%EOF\n",
                    "application/pdf",
                )
            },
        }
        first = self.client.post(created["upload_url"], **request_kwargs)
        second = self.client.post(created["upload_url"], **request_kwargs)
        self.assertEqual(first.status_code, 201)
        self.assertEqual(second.status_code, 200)
        self.assertEqual(first.json()["doc_id"], second.json()["doc_id"])
        self.assertFalse(second.json()["created"])
        self.assertEqual(
            len(list((self.corpus_root / "library").glob("doc_*"))),
            1,
        )

    def test_incomplete_document_batch_can_be_resumed_after_page_reload(self):
        created = self.client.post(
            "/api/document-batches",
            data={"expected_files": "2"},
        ).json()
        uploaded = self.client.post(
            created["upload_url"],
            data={"upload_index": "0"},
            files={
                "uploaded_file": (
                    "resume-first.pdf",
                    b"%PDF-1.4\n%mock\n%%EOF\n",
                    "application/pdf",
                )
            },
        )
        self.assertEqual(uploaded.status_code, 201)

        detail = self.client.get(created["detail_url"])
        self.assertEqual(detail.status_code, 200)
        self.assertIn("继续上传", detail.text)
        self.assertIn(
            f"/documents/upload?resume_batch_id={created['batch_id']}",
            detail.text,
        )

        resumed = self.client.get(
            "/documents/upload",
            params={"resume_batch_id": created["batch_id"]},
        )
        self.assertEqual(resumed.status_code, 200)
        self.assertIn("已恢复未完成批次", resumed.text)
        self.assertIn('data-resume-expected-files="2"', resumed.text)
        self.assertIn(
            f'data-resume-batch-id="{created["batch_id"]}"',
            resumed.text,
        )
        self.assertIn("重新选择原批次的全部文件并保持原顺序", resumed.text)

    def test_document_batch_limits_and_incomplete_queue_are_rejected(self):
        too_many = self.client.post(
            "/api/document-batches",
            data={"expected_files": str(self.settings.max_upload_batch_files + 1)},
        )
        self.assertEqual(too_many.status_code, 400)
        created = self.client.post(
            "/api/document-batches",
            data={"expected_files": "2"},
        ).json()
        uploaded = self.client.post(
            created["upload_url"],
            data={"upload_index": "0"},
            files={
                "uploaded_file": (
                    "only-one.pdf",
                    b"%PDF-1.4\n%mock\n%%EOF\n",
                    "application/pdf",
                )
            },
        )
        self.assertEqual(uploaded.status_code, 201)
        queued = self.client.post(created["queue_url"])
        self.assertEqual(queued.status_code, 400)
        self.assertIn("当前仅收到 1 个", queued.json()["detail"])

    def test_document_batch_queue_is_atomic_when_one_file_is_running(self):
        self._write_secret_file(aliyun=True)
        created = self.client.post(
            "/api/document-batches",
            data={"expected_files": "2"},
        ).json()
        uploaded_items = []
        for index in range(2):
            uploaded_items.append(
                self.client.post(
                    created["upload_url"],
                    data={"upload_index": str(index)},
                    files={
                        "uploaded_file": (
                            f"atomic-{index}.pdf",
                            b"%PDF-1.4\n%mock\n%%EOF\n",
                            "application/pdf",
                        )
                    },
                ).json()
            )

        first_doc_id = uploaded_items[0]["doc_id"]
        second_doc_id = uploaded_items[1]["doc_id"]
        self.client.post(
            f"/documents/{first_doc_id}/start-ocr",
            follow_redirects=False,
        )
        running = claim_next_queued_job(self.settings)
        self.assertEqual(running.job_id, first_doc_id)

        response = self.client.post(created["queue_url"])

        self.assertEqual(response.status_code, 400)
        self.assertIn("正在处理中", response.json()["detail"])
        self.assertIsNone(get_job(self.settings, second_doc_id))
        batch = self.client.get(created["status_url"]).json()
        self.assertEqual(batch["status"], "uploading")

    def test_document_upload_rejects_pdf_over_page_limit(self):
        import fitz

        pdf = fitz.open()
        pdf.new_page()
        pdf.new_page()
        payload = pdf.tobytes()
        pdf.close()
        settings = self.make_settings(max_pdf_pages=1)
        settings.ensure_directories()
        init_db(settings)
        with TestClient(create_app(settings)) as client:
            response = client.post(
                "/documents/upload",
                files={
                    "uploaded_files": (
                        "two-pages.pdf",
                        payload,
                        "application/pdf",
                    )
                },
            )
        self.assertEqual(response.status_code, 400)
        self.assertIn("超过 1 页限制", response.text)

    def test_document_start_route_only_enqueues_work(self):
        self._write_secret_file(aliyun=True)
        doc_id = self._upload_library_pdf()
        with mock.patch("webapp.document_library.run_document_ocr") as run_ocr:
            response = self.client.post(
                f"/documents/{doc_id}/start-ocr",
                follow_redirects=False,
            )
        self.assertEqual(response.status_code, 303)
        run_ocr.assert_not_called()
        document = load_document(self.settings, doc_id, refresh=False)
        self.assertEqual(document.ocr_status, STATUS_QUEUED)
        job = get_job(self.settings, doc_id)
        self.assertEqual(job.status, "queued")
        home = self.client.get("/")
        self.assertIn("处理中", home.text)

    def test_document_cannot_be_deleted_while_background_job_is_active(self):
        self._write_secret_file(aliyun=True)
        doc_id = self._upload_library_pdf()
        self.client.post(
            f"/documents/{doc_id}/start-ocr",
            follow_redirects=False,
        )

        response = self.client.post(
            f"/documents/{doc_id}/delete",
            follow_redirects=False,
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("正在后台处理中", response.json()["detail"])
        self.assertTrue(Path(load_document(self.settings, doc_id).pdf_path).exists())

    def test_document_with_existing_results_stays_queued_when_reenqueued(self):
        self._write_secret_file(aliyun=True)
        doc_id = self._upload_library_pdf()
        self._write_tiny_library_ocr(doc_id)
        update_document_status(
            self.settings,
            doc_id,
            ocr_status=STATUS_COMPLETED,
            raw_metrics_status=STATUS_COMPLETED,
            standard_metrics_status=STATUS_COMPLETED,
        )

        response = self.client.post(
            f"/documents/{doc_id}/rerun-ocr",
            follow_redirects=False,
        )

        self.assertEqual(response.status_code, 303)
        document = load_document(self.settings, doc_id)
        self.assertEqual(document.ocr_status, STATUS_QUEUED)
        self.assertEqual(document.raw_metrics_status, STATUS_QUEUED)
        self.assertEqual(document.standard_metrics_status, STATUS_QUEUED)
        job = get_job(self.settings, doc_id)
        self.assertEqual(job.status, "queued")
        stage_flow = build_job_stage_flow(job)
        self.assertEqual(stage_flow[0]["label_zh"], "上传")
        self.assertEqual(stage_flow[0]["state"], "current")
        self.assertEqual(stage_flow[1]["state"], "pending")

    def test_sqlite_queue_allows_only_one_running_document(self):
        self._write_secret_file(aliyun=True)
        created = self.client.post(
            "/api/document-batches",
            data={"expected_files": "2"},
        ).json()
        for index in range(2):
            self.client.post(
                created["upload_url"],
                data={"upload_index": str(index)},
                files={
                    "uploaded_file": (
                        f"serial-{index}.pdf",
                        b"%PDF-1.4\n%mock\n%%EOF\n",
                        "application/pdf",
                    )
                },
            )
        self.client.post(created["queue_url"])
        first = claim_next_queued_job(self.settings)
        second = claim_next_queued_job(self.settings)
        self.assertIsNotNone(first)
        self.assertEqual(first.status, "running")
        self.assertIsNone(second)

    def test_stale_running_job_is_failed_before_queue_continues(self):
        self._write_secret_file(aliyun=True)
        created = self.client.post(
            "/api/document-batches",
            data={"expected_files": "2"},
        ).json()
        for index in range(2):
            self.client.post(
                created["upload_url"],
                data={"upload_index": str(index)},
                files={
                    "uploaded_file": (
                        f"recover-{index}.pdf",
                        b"%PDF-1.4\n%mock\n%%EOF\n",
                        "application/pdf",
                    )
                },
            )
        self.client.post(created["queue_url"])
        first = claim_next_queued_job(self.settings)
        future = datetime.now(timezone.utc) + timedelta(
            seconds=first.timeout_seconds
            + self.settings.worker_stale_job_grace_seconds
            + 5
        )
        recovered = recover_abandoned_running_jobs(self.settings, now=future)
        self.assertEqual([job.job_id for job in recovered], [first.job_id])
        self.assertEqual(get_job(self.settings, first.job_id).status, "failed")
        second = claim_next_queued_job(self.settings)
        self.assertIsNotNone(second)
        self.assertNotEqual(second.job_id, first.job_id)
        self.assertEqual(second.status, "running")

    def test_worker_runs_document_pipeline_to_structured_outputs(self):
        self._write_secret_file(aliyun=True)
        created = self.client.post(
            "/api/document-batches",
            data={"expected_files": "1"},
        ).json()
        uploaded = self.client.post(
            created["upload_url"],
            data={"upload_index": "0"},
            files={
                "uploaded_file": (
                    "pipeline.pdf",
                    b"%PDF-1.4\n%mock\n%%EOF\n",
                    "application/pdf",
                )
            },
        ).json()
        self.client.post(created["queue_url"])
        doc_id = uploaded["doc_id"]

        def fake_ocr(settings, target_doc_id, *, rerun):
            self._write_tiny_library_ocr(target_doc_id)
            update_document_status(settings, target_doc_id, ocr_status=STATUS_COMPLETED)
            return {
                "pass": True,
                "command_executed": "mock-document-ocr",
                "error_message": "",
            }

        with mock.patch(
            "webapp.document_library.run_document_ocr",
            side_effect=fake_ocr,
        ):
            result = run_worker_once(self.settings)
        self.assertIsNotNone(result)
        self.assertIn(result.status, {"succeeded", "needs_review"})
        document = load_document(self.settings, doc_id)
        self.assertEqual(document.ocr_status, STATUS_COMPLETED)
        self.assertEqual(document.raw_metrics_status, STATUS_COMPLETED)
        self.assertEqual(document.standard_metrics_status, STATUS_COMPLETED)
        state = load_simple_flow_state(document_to_job(self.settings, document), self.settings)
        self.assertTrue(state["raw_ready"])
        self.assertTrue(state["standard_ready"])
        self.assertTrue(state["combined_ready"])
        pipeline_summary = self.settings.results_root / doc_id / "document_pipeline_summary.json"
        self.assertTrue(pipeline_summary.exists())
        self.assertTrue(json.loads(pipeline_summary.read_text(encoding="utf-8"))["pass"])

    def test_document_home_no_ocr_button_state(self):
        doc_id = self._upload_library_pdf()
        response = self.client.get("/")
        self.assertEqual(response.status_code, 200)
        self.assertIn("开始识别", response.text)
        self.assertIn("删除", response.text)
        row_start = response.text.index("A公司财务报表.pdf")
        row_text = response.text[row_start : row_start + 2000]
        self.assertNotIn("继续处理", row_text)

    def test_document_home_ocr_completed_button_state(self):
        doc_id = self._upload_library_pdf()
        update_document_status(self.settings, doc_id, ocr_status=STATUS_COMPLETED)
        response = self.client.get("/")
        self.assertEqual(response.status_code, 200)
        self.assertIn("重新OCR", response.text)
        self.assertIn("继续处理", response.text)

    def test_document_continue_blocks_when_ocr_missing(self):
        doc_id = self._upload_library_pdf()
        response = self.client.get(f"/documents/{doc_id}/continue", follow_redirects=False)
        self.assertEqual(response.status_code, 303)
        self.assertIn("%E8%AF%B7%E5%85%88%E7%82%B9%E5%87%BB", response.headers["location"])

    def test_document_step1_raw_extraction_from_fixture_ocr(self):
        doc_id = self._upload_library_pdf()
        self._write_tiny_library_ocr(doc_id)
        with mock.patch("webapp.simple_flow.run_raw_metrics_step") as raw_step:
            response = self.client.post(
                f"/documents/{doc_id}/raw-metrics/run",
                follow_redirects=False,
            )
        self.assertEqual(response.status_code, 303)
        raw_step.assert_not_called()
        document = load_document(self.settings, doc_id)
        self.assertEqual(document.raw_metrics_status, STATUS_QUEUED)
        self.assertEqual(document.standard_metrics_status, STATUS_QUEUED)
        queued_job = get_job(self.settings, doc_id)
        self.assertEqual(
            queued_job.requested_stage,
            DOCUMENT_PIPELINE_STAGE_RAW_METRICS,
        )
        self.assertEqual(
            job_stage_label_zh(queued_job),
            "等待提取原始数据（复用已有 OCR）",
        )
        stage_flow = build_job_stage_flow(queued_job)
        self.assertEqual(stage_flow[0]["state"], "done")
        self.assertEqual(stage_flow[1]["state"], "done")
        self.assertEqual(stage_flow[1]["status_label_zh"], "复用已有 OCR")
        self.assertEqual(stage_flow[2]["state"], "current")
        self.assertEqual(stage_flow[2]["status_label_zh"], "排队等待提取原始数据")

        result = run_worker_once(self.settings)
        self.assertIsNotNone(result)
        document = load_document(self.settings, doc_id)
        self.assertEqual(document.raw_metrics_status, STATUS_COMPLETED)
        self.assertEqual(document.standard_metrics_status, STATUS_COMPLETED)
        state = load_simple_flow_state(document_to_job(self.settings, document))
        self.assertTrue(state["raw_ready"])
        self.assertTrue(state["standard_ready"])
        self.assertTrue(Path(state["raw_metrics_csv"]).exists())
        self.assertTrue(str(Path(state["raw_metrics_csv"]).resolve()).startswith(str((self.raw_metrics_root / doc_id).resolve())))
        self.assertTrue(state["combined_ready"])
        workbook = load_workbook(Path(state["combined_metrics_xlsx"]))
        self.assertIn("数据总表", workbook.sheetnames)
        self.assertIn("原始数据", workbook.sheetnames)
        self.assertIn("标准化数据", workbook.sheetnames)
        self.assertIn("说明", workbook.sheetnames)

    def test_document_metadata_copied_from_windows_uses_current_library_paths(self):
        doc_id = self._upload_library_pdf()
        self._write_tiny_library_ocr(doc_id)
        metadata_path = metadata_path_for(self.settings, doc_id)
        payload = json.loads(metadata_path.read_text(encoding="utf-8"))
        payload["pdf_path"] = rf"C:\Users\gater\Desktop\Python\finance\AutoFinance\data\corpus\library\{doc_id}\input\A公司财务报表.pdf"
        payload["input_dir"] = rf"C:\Users\gater\Desktop\Python\finance\AutoFinance\data\corpus\library\{doc_id}\input"
        payload["ocr_output_dir"] = rf"C:\Users\gater\Desktop\Python\finance\AutoFinance\data\corpus\library\{doc_id}\ocr_outputs"
        metadata_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

        document = load_document(self.settings, doc_id, refresh=False)
        self.assertEqual(Path(document.input_dir).resolve(), (self.corpus_root / "library" / doc_id / "input").resolve())
        self.assertEqual(Path(document.ocr_output_dir).resolve(), (self.corpus_root / "library" / doc_id / "ocr_outputs").resolve())

        response = self.client.post(f"/documents/{doc_id}/raw-metrics/run", follow_redirects=False)
        self.assertEqual(response.status_code, 303)
        run_worker_once(self.settings)
        state = load_simple_flow_state(document_to_job(self.settings, load_document(self.settings, doc_id)))
        self.assertTrue(state["raw_ready"])

    def test_document_standard_mapping_queue_requires_raw_metrics(self):
        doc_id = self._upload_library_pdf()
        self._write_tiny_library_ocr(doc_id)

        response = self.client.post(
            f"/documents/{doc_id}/standard-metrics/run",
            follow_redirects=False,
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("请先生成原始数据", response.json()["detail"])
        self.assertIsNone(get_job(self.settings, doc_id))

    def test_document_step2_standard_mapping_after_raw_metrics_exists(self):
        doc_id = self._upload_library_pdf()
        self._write_tiny_library_ocr(doc_id)
        self.client.post(f"/documents/{doc_id}/raw-metrics/run", follow_redirects=False)
        run_worker_once(self.settings)
        with mock.patch("webapp.simple_flow.run_standard_metrics_step") as standard_step:
            response = self.client.post(
                f"/documents/{doc_id}/standard-metrics/run",
                follow_redirects=False,
            )
        self.assertEqual(response.status_code, 303)
        standard_step.assert_not_called()
        queued_document = load_document(self.settings, doc_id)
        self.assertEqual(queued_document.standard_metrics_status, STATUS_QUEUED)
        self.assertEqual(
            get_job(self.settings, doc_id).requested_stage,
            DOCUMENT_PIPELINE_STAGE_STANDARD_METRICS,
        )
        queued_job = get_job(self.settings, doc_id)
        self.assertEqual(
            job_stage_label_zh(queued_job),
            "等待重新映射标准科目（复用已有原始数据）",
        )
        stage_flow = build_job_stage_flow(queued_job)
        self.assertEqual(stage_flow[0]["state"], "done")
        self.assertEqual(stage_flow[1]["state"], "done")
        self.assertEqual(stage_flow[1]["status_label_zh"], "复用已有 OCR")
        self.assertEqual(stage_flow[2]["state"], "current")
        self.assertEqual(
            stage_flow[2]["status_label_zh"],
            "排队等待重新映射（复用原始数据）",
        )
        run_worker_once(self.settings)
        document = load_document(self.settings, doc_id)
        self.assertEqual(document.standard_metrics_status, STATUS_COMPLETED)
        state = load_simple_flow_state(document_to_job(self.settings, document))
        self.assertTrue(state["standard_ready"])
        self.assertTrue(Path(state["standardized_metrics_csv"]).exists())
        continue_response = self.client.get(f"/documents/{doc_id}/continue")
        self.assertEqual(continue_response.status_code, 200)
        self.assertIn("document-action-panel", continue_response.text)
        self.assertNotIn("stage-flow", continue_response.text)
        self.assertNotIn("stage-card", continue_response.text)
        self.assertIn(f'href="/documents/{doc_id}/download/combined_metrics_xlsx"', continue_response.text)
        self.assertIn(f'href="/documents/{doc_id}/download-preview/combined_metrics_xlsx"', continue_response.text)
        self.assertIn(">下载数据表</a>", continue_response.text)
        self.assertIn(">预览下载版</a>", continue_response.text)
        self.assertIn("高级下载", continue_response.text)
        self.assertIn(f'href="/documents/{doc_id}/download/raw_metrics_csv"', continue_response.text)
        self.assertIn(f'href="/documents/{doc_id}/download/standardized_metrics_csv"', continue_response.text)
        self.assertNotIn(">下载原始数据</a>", continue_response.text)
        self.assertNotIn(">下载标准化数据</a>", continue_response.text)
        self.assertIn(f'href="/documents/{doc_id}/proofread"', continue_response.text)
        self.assertIn(f'action="/documents/{doc_id}/standard-metrics/run"', continue_response.text)
        self.assertIn("重新生成标准指标 / 标准映射", continue_response.text)

        preview_response = self.client.get(f"/documents/{doc_id}/download-preview/combined_metrics_xlsx")
        self.assertEqual(preview_response.status_code, 200)
        self.assertIn("实际下载版", preview_response.text)
        self.assertIn("数据总表", preview_response.text)
        self.assertIn("标准化数据", preview_response.text)
        self.assertIn(f'href="/documents/{doc_id}/download/combined_metrics_xlsx"', preview_response.text)

        browser_submit = self.client.post(
            f"/documents/{doc_id}/mapping/decision",
            data={
                "review_item_id": "maprev_000001",
                "selected_code": "ZT_001",
                "selected_name": "货币资金",
                "decision": "accept_once",
            },
            headers={"accept": "text/html,application/xhtml+xml"},
            follow_redirects=False,
        )
        self.assertEqual(browser_submit.status_code, 303)
        self.assertEqual(browser_submit.headers["location"], f"/documents/{doc_id}/proofread")

    def test_document_delete_confirmation_lists_associated_files(self):
        doc_id = self._upload_library_pdf()
        self._write_tiny_library_ocr(doc_id)
        (self.raw_metrics_root / doc_id / "RUN_TEST").mkdir(parents=True, exist_ok=True)
        (self.standard_metrics_root / doc_id / "RUN_TEST").mkdir(parents=True, exist_ok=True)
        response = self.client.get(f"/documents/{doc_id}/delete-confirm")
        self.assertEqual(response.status_code, 200)
        for text in ("原始 PDF", "OCR 输出", "原始数据结果", "标准化数据结果", "相关网页任务文件", "确认删除"):
            self.assertIn(text, response.text)

    def test_document_delete_action_removes_allowed_paths_and_writes_summary(self):
        doc_id = self._upload_library_pdf()
        self._write_tiny_library_ocr(doc_id)
        raw_dir = self.raw_metrics_root / doc_id / "RUN_TEST"
        standard_dir = self.standard_metrics_root / doc_id / "RUN_TEST"
        raw_dir.mkdir(parents=True, exist_ok=True)
        standard_dir.mkdir(parents=True, exist_ok=True)
        (raw_dir / "raw_metrics.csv").write_text("x\n", encoding="utf-8")
        (standard_dir / "standardized_metrics.csv").write_text("x\n", encoding="utf-8")
        response = self.client.post(f"/documents/{doc_id}/delete", follow_redirects=False)
        self.assertEqual(response.status_code, 303)
        self.assertFalse((self.corpus_root / "library" / doc_id).exists())
        self.assertFalse((self.raw_metrics_root / doc_id).exists())
        self.assertFalse((self.standard_metrics_root / doc_id).exists())
        summaries = sorted(self.settings.deletions_root.glob(f"{doc_id}_*_delete_summary.json"))
        self.assertTrue(summaries)
        summary = json.loads(summaries[-1].read_text(encoding="utf-8"))
        self.assertEqual(summary["status"], "deleted")

    def test_document_delete_rejects_unsafe_paths(self):
        doc_id = self._upload_library_pdf()
        document = load_document(self.settings, doc_id, refresh=False)
        outside_pdf = self.temp_path / "outside.pdf"
        outside_pdf.write_bytes(b"%PDF-1.4\n")
        document.pdf_path = str(outside_pdf)
        write_document(self.settings, document)
        plan = build_delete_plan(self.settings, load_document(self.settings, doc_id, refresh=False))
        self.assertTrue(plan["unsafe_paths"])
        response = self.client.post(f"/documents/{doc_id}/delete", follow_redirects=False)
        self.assertEqual(response.status_code, 400)
        self.assertTrue((self.corpus_root / "library" / doc_id).exists())

    def test_document_home_advanced_links_are_not_prominent(self):
        response = self.client.get("/")
        self.assertEqual(response.status_code, 200)
        self.assertIn('class="advanced-links"', response.text)
        self.assertIn("管理员工具", response.text)
        self.assertNotIn("Provider Priority", response.text)

    def test_step2_route_runs_standard_map_on_raw_metrics_fixture(self):
        job_id = self._create_job("step2 route")
        raw_path = self._create_raw_metrics_fixture(metric_name="往来款")
        response = self.client.post(
            f"/jobs/{job_id}/standard-metrics/run",
            data={"raw_metrics_path": str(raw_path)},
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 303)
        job = get_job(self.settings, job_id)
        state = load_simple_flow_state(job)
        self.assertTrue(state["standard_ready"])
        self.assertTrue(Path(state["standardized_metrics_csv"]).exists())
        self.assertTrue(str(Path(state["standardized_metrics_csv"]).resolve()).startswith(str(self.standard_metrics_root.resolve())))
        self.assertTrue(state["combined_ready"])
        self.assertTrue(Path(state["combined_metrics_xlsx"]).exists())
        self.assertTrue(str(Path(state["combined_metrics_xlsx"]).resolve()).startswith(str(self.settings.results_root.resolve())))
        self.assertTrue(combined_download_summary_path(job).exists())
        self.assertTrue((self.settings.runtime_root / "combined_download_summary.json").exists())

    def test_combined_download_workbook_formatting_and_route(self):
        job_id = self._create_job("combined download")
        raw_path = self._create_raw_metrics_fixture(metric_name="货币资金", metric_value="12345.67")
        self.client.post(
            f"/jobs/{job_id}/standard-metrics/run",
            data={"raw_metrics_path": str(raw_path)},
            follow_redirects=False,
        )
        job = get_job(self.settings, job_id)
        state = load_simple_flow_state(job)
        workbook_path = Path(state["combined_metrics_xlsx"])
        self.assertTrue(workbook_path.exists())

        workbook = load_workbook(workbook_path)
        self.assertEqual(workbook.sheetnames[:3], ["数据总表", "标准化数据", "原始数据"])
        for sheet_name in ("数据总表", "标准化数据", "原始数据", "术语映射校对", "说明"):
            self.assertIn(sheet_name, workbook.sheetnames)

        total_sheet = workbook["数据总表"]
        self.assertEqual(total_sheet.freeze_panes, "A2")
        self.assertTrue(total_sheet.auto_filter.ref)
        self.assertGreater(total_sheet.column_dimensions["D"].width, 20)
        value_col = self._column_index(total_sheet, "指标数值")
        fill_date_col = self._column_index(total_sheet, "填表日期")
        period_role_col = self._column_index(total_sheet, "期间类型")
        value_cell = total_sheet.cell(row=2, column=value_col)
        fill_date_cell = total_sheet.cell(row=2, column=fill_date_col)
        self.assertIsInstance(value_cell.value, (int, float))
        self.assertEqual(value_cell.value, 12345.67)
        self.assertEqual(total_sheet.cell(row=2, column=period_role_col).value, "期末数")
        self.assertEqual(value_cell.number_format, "#,##0.00")
        self.assertEqual(fill_date_cell.number_format, "yyyy-mm-dd")

        summary = json.loads(combined_download_summary_path(job).read_text(encoding="utf-8"))
        self.assertTrue(summary["pass"])
        self.assertEqual(summary["primary_download_label"], "下载数据表")
        self.assertTrue(summary["advanced_downloads_available"])
        self.assertTrue(summary["path_hygiene_pass"])
        self.assertGreater(summary["numeric_cells_formatted_total"], 0)
        self.assertGreater(summary["date_cells_formatted_total"], 0)

        download = self.client.get(f"/jobs/{job_id}/download/combined_metrics_xlsx")
        self.assertEqual(download.status_code, 200)
        self.assertTrue(download.content.startswith(b"PK"))
        csv_download = self.client.get(f"/jobs/{job_id}/download/combined_metrics_csv")
        self.assertEqual(csv_download.status_code, 200)
        self.assertTrue(csv_download.content.startswith(b"\xef\xbb\xbf"))

        preview = self.client.get(f"/jobs/{job_id}/download-preview/combined_metrics_xlsx")
        self.assertEqual(preview.status_code, 200)
        self.assertIn("实际下载版", preview.text)
        self.assertIn("数据总表", preview.text)
        self.assertIn("标准化数据", preview.text)
        self.assertIn("12,345.67", preview.text)
        self.assertIn("期末数", preview.text)
        self.assertIn(f'href="/jobs/{job_id}/download/combined_metrics_xlsx"', preview.text)

        detail = self.client.get(f"/jobs/{job_id}")
        self.assertEqual(detail.status_code, 200)
        self.assertIn("下载数据表", detail.text)
        self.assertIn("预览下载版", detail.text)
        self.assertIn("高级下载", detail.text)
        self.assertIn("原始数据 CSV", detail.text)
        self.assertIn("标准化数据 CSV", detail.text)
        self.assertNotIn("下载原始数据表 Excel", detail.text)
        self.assertNotIn("下载标准化数据表 Excel", detail.text)

    def test_manual_accounting_workbook_generation_uses_current_corrected_csv(self):
        job_id = self._create_job("accounting workbook")
        raw_path = self._create_raw_metrics_fixture(metric_name="货币资金", metric_value="12345.67")
        self.client.post(
            f"/jobs/{job_id}/standard-metrics/run",
            data={"raw_metrics_path": str(raw_path)},
            follow_redirects=False,
        )

        response = self.client.post(f"/jobs/{job_id}/generate-accounting-workbook")

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.content.startswith(b"PK"))
        job = get_job(self.settings, job_id)
        state = load_simple_flow_state(job)
        self.assertTrue(state["accounting_workbook_ready"])
        workbook = load_workbook(Path(state["accounting_workbook"]), data_only=True)
        self.assertIn("生成说明", workbook.sheetnames)
        sheet = workbook[workbook.sheetnames[0]]
        headers = [sheet.cell(row=3, column=column).value for column in range(1, sheet.max_column + 1)]
        self.assertEqual(headers, ["科目名称", "2022-12-31__期末数"])
        self.assertEqual(sheet["B4"].value, 12345.67)
        self.assertNotIn("金额", headers)
        workbook.close()

        summary = json.loads(Path(state["accounting_export_summary"]).read_text(encoding="utf-8"))
        self.assertEqual(summary["written_cells_total"], 1)
        self.assertEqual(summary["conflicted_cells_total"], 0)
        self.assertEqual(summary["skipped_reason_counts"], {})
        download = self.client.get(f"/jobs/{job_id}/download/accounting_workbook")
        self.assertEqual(download.status_code, 200)

        save = self.client.post(
            f"/jobs/{job_id}/proofread/save",
            json={
                "edits": [
                    {
                        "item_id": "unirev_000001",
                        "raw_metric_id": "CASE1:1:aliyun_table:0:1-1:2-2",
                        "raw_metric_ids": ["CASE1:1:aliyun_table:0:1-1:2-2"],
                        "edit_type": "date_change",
                        "previous_date": "2022-12-31",
                        "new_date": "2023-12-31",
                    }
                ]
            },
        )
        self.assertEqual(save.status_code, 200)
        stale_download = self.client.get(f"/jobs/{job_id}/download/accounting_workbook")
        self.assertEqual(stale_download.status_code, 409)
        self.assertEqual(stale_download.json()["detail"], "校对结果已变化，请重新生成会计报表。")

    def test_accounting_export_migrates_legacy_code_by_matching_subject_name(self):
        current_csv = self.temp_path / "legacy-code-current.csv"
        self._write_csv(
            current_csv,
            [
                {
                    "填表日期": "2022-12-31",
                    "当前条目日期": "2022-12-31",
                    "期间类型": "期末数",
                    "报表类型": "",
                    "标准指标编码": "ZT_002",
                    "标准指标名称": "短期借款",
                    "指标数值": "136000000",
                    "是否需要人工校对": "否",
                },
                {
                    "填表日期": "2022-12-31",
                    "当前条目日期": "2022-12-31",
                    "期间类型": "期末数",
                    "报表类型": "",
                    "标准指标编码": "ZT_002",
                    "标准指标名称": "无法识别科目",
                    "指标数值": "99",
                    "是否需要人工校对": "否",
                },
            ],
            [
                "填表日期",
                "当前条目日期",
                "期间类型",
                "报表类型",
                "标准指标编码",
                "标准指标名称",
                "指标数值",
                "是否需要人工校对",
            ],
        )
        output = self.temp_path / "legacy-code-accounting.xlsx"
        summary = build_accounting_workbook(
            template_path=self.template_path,
            current_data_csv=current_csv,
            output_path=output,
        )

        self.assertEqual(summary["written_cells_total"], 1)
        self.assertEqual(
            summary["skipped_reason_counts"],
            {"标准指标编码与名称不一致，请重新映射": 1},
        )
        workbook = load_workbook(output, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        self.assertEqual(sheet["B71"].value, 136000000)
        self.assertIsNone(sheet["B5"].value)
        explanation = workbook["生成说明"]
        explanation_values = [cell.value for row in explanation.iter_rows() for cell in row]
        self.assertIn("未写入：标准指标编码与名称不一致，请重新映射", explanation_values)
        workbook.close()

        blank_code_csv = self.temp_path / "blank-code-current.csv"
        blank_code_row = {
            "填表日期": "2022-12-31",
            "当前条目日期": "2022-12-31",
            "期间类型": "期末数",
            "报表类型": "",
            "标准指标编码": "",
            "标准指标名称": "短期借款",
            "指标数值": "1",
            "是否需要人工校对": "否",
        }
        self._write_csv(blank_code_csv, [blank_code_row], list(blank_code_row))
        blank_code_summary = build_accounting_workbook(
            template_path=self.template_path,
            current_data_csv=blank_code_csv,
            output_path=self.temp_path / "blank-code-accounting.xlsx",
        )
        self.assertEqual(blank_code_summary["written_cells_total"], 0)
        self.assertEqual(
            blank_code_summary["skipped_reason_counts"],
            {"未映射到当前会计报表模板": 1},
        )

    def test_raw_review_page_loads_and_raw_actions_are_saved(self):
        job_id = self._create_job("raw review")
        raw_path = self._create_raw_metrics_fixture(metric_name="货币资金")
        self._attach_raw_summary_to_job(job_id, raw_path)
        response = self.client.get(f"/jobs/{job_id}/raw-review")
        self.assertEqual(response.status_code, 200)
        self.assertIn("原始数据校对", response.text)
        self.assertIn("data-sheet-tabs", response.text)
        self.assertIn(">保存</button>", response.text)
        self.assertIn(">下一页</button>", response.text)
        self.assertNotIn("通过选中单元格", response.text)
        self.assertNotIn("跳过选中单元格", response.text)
        self.assertNotIn("保存表格修改", response.text)
        action_response = self.client.post(
            f"/jobs/{job_id}/raw-review/actions",
            data={
                "review_item_id": "rawrev_000001",
                "action": "edit",
                "edits_json": json.dumps(
                    [
                        {
                            "review_item_id": "rawrev_000001",
                            "raw_metric_id": "CASE1:1:aliyun_table:0:1-1:2-2",
                            "row_index": "1",
                            "col_index": "2",
                            "metric_name": "货币资金",
                            "value": "999.88",
                        }
                    ],
                    ensure_ascii=False,
                ),
                "reviewer_note": "ok",
            },
            follow_redirects=False,
        )
        self.assertEqual(action_response.status_code, 303)
        saved_response = self.client.get(f"/jobs/{job_id}/raw-review/items/rawrev_000001")
        self.assertEqual(saved_response.status_code, 200)
        self.assertIn("999.88", saved_response.text)
        job = get_job(self.settings, job_id)
        self.assertTrue((raw_review_dir(job) / "raw_review_actions.csv").exists())
        self.assertTrue((raw_review_dir(job) / "raw_review_actions.json").exists())

    def test_raw_review_next_table_saves_current_sheet_and_redirects(self):
        job_id = self._create_job("raw review next")
        raw_path = self._create_two_table_raw_metrics_fixture()
        self._attach_raw_summary_to_job(job_id, raw_path)
        response = self.client.post(
            f"/jobs/{job_id}/raw-review/actions",
            data={
                "review_item_id": "rawrev_000001",
                "action": "next_table",
                "edits_json": json.dumps(
                    [
                        {
                            "review_item_id": "rawrev_000001",
                            "raw_metric_id": "CASE1:1:aliyun_table:1:1-1:2-2",
                            "row_index": "1",
                            "col_index": "2",
                            "metric_name": "货币资金",
                            "value": "321.00",
                        }
                    ],
                    ensure_ascii=False,
                ),
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 303)
        self.assertTrue(response.headers["location"].endswith("/jobs/%s/raw-review/items/rawrev_000002" % job_id))
        saved_response = self.client.get(f"/jobs/{job_id}/raw-review/items/rawrev_000001")
        self.assertEqual(saved_response.status_code, 200)
        self.assertIn("321.00", saved_response.text)

    def test_mapping_review_page_loads_and_mapping_actions_are_saved(self):
        job_id = self._create_job("mapping review")
        raw_path = self._create_raw_metrics_fixture(metric_name="货币资金")
        self.client.post(
            f"/jobs/{job_id}/standard-metrics/run",
            data={"raw_metrics_path": str(raw_path)},
            follow_redirects=False,
        )
        response = self.client.get(f"/jobs/{job_id}/mapping-review")
        self.assertEqual(response.status_code, 200)
        self.assertIn("术语映射校对", response.text)
        self.assertIn('data-mapping-review-workbench', response.text)
        self.assertIn('class="spreadsheet-table mapping-table"', response.text)
        self.assertIn("原始术语", response.text)
        self.assertIn("标准术语", response.text)
        self.assertIn("搜索标准术语，如：短期、借款、68、dqjk", response.text)
        self.assertIn("data-standard-term-input", response.text)
        self.assertIn("data-mapping-cell", response.text)
        self.assertIn("data-bbox=", response.text)
        self.assertIn("data-page-no=", response.text)
        self.assertIn("data-page-image-key=", response.text)
        self.assertIn("精确匹配，无需决策", response.text)
        self.assertNotIn("不采纳", response.text)
        self.assertNotIn("仅本次采用", response.text)
        self.assertNotIn("采用并记住", response.text)
        self.assertNotIn("<th>指标数值</th>", response.text)
        approve = self.client.post(
            f"/jobs/{job_id}/mapping-review/actions",
            data={"review_item_id": "maprev_000001", "action": "approve_mapping", "reviewer_note": "ok"},
            follow_redirects=False,
        )
        self.assertEqual(approve.status_code, 303)
        skip = self.client.post(
            f"/jobs/{job_id}/mapping-review/actions",
            data={"review_item_id": "maprev_000001", "action": "skip_mapping", "reviewer_note": "skip"},
            follow_redirects=False,
        )
        self.assertEqual(skip.status_code, 303)
        change = self.client.post(
            f"/jobs/{job_id}/mapping-review/actions",
            data={
                "review_item_id": "maprev_000001",
                "action": "change_mapping",
                "selected_code": "ZT_068",
                "selected_name": "短期借款",
                "reviewer_note": "change",
            },
            follow_redirects=False,
        )
        self.assertEqual(change.status_code, 303)
        job = get_job(self.settings, job_id)
        actions_path = mapping_review_dir(job) / "mapping_review_actions.json"
        self.assertTrue(actions_path.exists())
        self.assertTrue((mapping_review_dir(job) / "mapping_review_actions.csv").exists())
        actions = json.loads(actions_path.read_text(encoding="utf-8"))
        self.assertEqual([item["action"] for item in actions], ["approve_mapping", "skip_mapping", "change_mapping"])
        self.assertEqual(actions[-1]["original_metric_name"], "货币资金")
        self.assertEqual(actions[-1]["previous_code"], "ZT_001")
        self.assertEqual(actions[-1]["previous_name"], "货币资金")
        self.assertEqual(actions[-1]["selected_code"], "ZT_068")
        self.assertEqual(actions[-1]["selected_name"], "短期借款")
        self.assertTrue(str(actions_path.resolve()).startswith(str(self.runtime_root.resolve())))

    def test_stage15_mapping_decision_routes_store_preview_and_config_hygiene(self):
        job_id = self._create_job("stage15 decisions")
        raw_path = self._create_raw_metrics_fixture(metric_name="阶段十五待映射")
        config_path = REPO_ROOT / "config" / "standard_term_aliases.yml"
        config_before = config_path.read_text(encoding="utf-8")
        self.client.post(
            f"/jobs/{job_id}/standard-metrics/run",
            data={"raw_metrics_path": str(raw_path)},
            follow_redirects=False,
        )
        job = get_job(self.settings, job_id)
        state = load_simple_flow_state(job)
        output_dir = Path(state["standardized_metrics_csv"]).parent

        invalid_code = self.client.post(
            f"/jobs/{job_id}/mapping/accept-once",
            data={
                "review_item_id": "maprev_000001",
                "selected_code": "ZT_999",
                "selected_name": "伪造科目",
            },
        )
        self.assertEqual(invalid_code.status_code, 400)

        mismatched_name = self.client.post(
            f"/jobs/{job_id}/mapping/accept-once",
            data={
                "review_item_id": "maprev_000001",
                "selected_code": "ZT_001",
                "selected_name": "短期借款",
            },
        )
        self.assertEqual(mismatched_name.status_code, 400)

        accept_once = self.client.post(
            f"/jobs/{job_id}/mapping/accept-once",
            data={
                "review_item_id": "maprev_000001",
                "selected_code": "ZT_001",
                "selected_name": "货币资金",
                "note": "once",
            },
        )
        self.assertEqual(accept_once.status_code, 200)
        self.assertEqual(accept_once.json()["decision"], "accept_once")
        with (output_dir / "standardized_metrics_detailed.csv").open("r", encoding="utf-8-sig", newline="") as handle:
            row = next(csv.DictReader(handle))
        self.assertEqual(row["标准指标编码"], "ZT_001")
        self.assertEqual(row["映射方法"], "manual_once")
        store = LocalMappingStore(self.settings.mapping_store_path)
        self.assertEqual(store.count("term_aliases", where="enabled = 1 AND COALESCE(source, '') != 'base'"), 0)

        remember = self.client.post(
            f"/jobs/{job_id}/mapping/accept-and-remember",
            data={
                "review_item_id": "maprev_000001",
                "selected_code": "ZT_068",
                "selected_name": "短期借款",
                "note": "remember",
            },
        )
        self.assertEqual(remember.status_code, 200)
        self.assertEqual(remember.json()["decision"], "accept_and_remember")
        aliases = store.alias_rows(include_base=False)
        remembered_alias = next(
            row
            for row in aliases
            if row["alias"] == "阶段十五待映射" and row["standard_code"] == "ZT_068"
        )
        self.assertEqual(remembered_alias["scope_company"], "AAA有限公司")
        self.assertEqual(remembered_alias["scope_statement_type"], "balance_sheet")
        self.assertTrue((self.settings.mapping_store_root / "local_aliases_export.yml").exists())
        self.assertTrue((self.settings.mapping_store_root / "mapping_decisions_audit.csv").exists())

        reject = self.client.post(
            f"/jobs/{job_id}/mapping/reject",
            data={"review_item_id": "maprev_000001", "note": "reject"},
        )
        self.assertEqual(reject.status_code, 200)
        self.assertEqual(reject.json()["decision"], "reject")

        decisions_csv = mapping_review_dir(job) / "mapping_decisions.csv"
        decisions_json = mapping_review_dir(job) / "mapping_decisions.json"
        decisions_summary = mapping_review_dir(job) / "mapping_decision_summary.json"
        self.assertTrue(decisions_csv.exists())
        self.assertTrue(decisions_json.exists())
        self.assertTrue(decisions_summary.exists())
        with decisions_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            decisions = list(csv.DictReader(handle))
        self.assertEqual([row["decision"] for row in decisions], ["accept_once", "accept_and_remember", "reject"])

        candidates = self.client.get(
            "/api/mapping/candidates",
            params={
                "raw_metric_name": "阶段十五待映射",
                "company_name": "AAA有限公司",
                "statement_type": "balance_sheet",
            },
        )
        self.assertEqual(candidates.status_code, 200)
        self.assertEqual(candidates.json()["mapping_method"], "local_alias")
        self.assertEqual(candidates.json()["standard_code"], "ZT_068")

        before_aliases = store.count("term_aliases", where="enabled = 1 AND COALESCE(source, '') != 'base'")
        preview = self.client.get(f"/jobs/{job_id}/mapping/bulk-confidence-preview?threshold=0.9")
        self.assertEqual(preview.status_code, 200)
        self.assertEqual(preview.json()["future_decision"], "accept_once")
        self.assertFalse(preview.json()["mutated_mappings"])
        self.assertEqual(before_aliases, store.count("term_aliases", where="enabled = 1 AND COALESCE(source, '') != 'base'"))
        self.assertTrue((output_dir / "confidence_bulk_accept_preview.json").exists())
        self.assertEqual(config_path.read_text(encoding="utf-8"), config_before)

    def test_stage15_2_llm_ai_suggestions_and_bulk_apply_routes(self):
        job_id = self._create_job("stage15.2 llm")
        raw_path = self._create_raw_metrics_fixture(metric_name="总收入")
        with mock.patch.dict(os.environ, {"LLM_MAPPING_MOCK": "true", "DEEPSEEK_API_KEY": ""}, clear=False):
            run_response = self.client.post(
                f"/jobs/{job_id}/standard-metrics/run",
                data={"raw_metrics_path": str(raw_path)},
                follow_redirects=False,
            )
        self.assertEqual(run_response.status_code, 303)
        job = get_job(self.settings, job_id)
        state = load_simple_flow_state(job)
        output_dir = Path(state["standardized_metrics_csv"]).parent

        response = self.client.get(f"/jobs/{job_id}/mapping-review")
        self.assertEqual(response.status_code, 200)
        self.assertIn("AI建议", response.text)
        self.assertIn("status--mapping-llm_suggested", response.text)
        self.assertIn("93%", response.text)
        self.assertIn("本操作只对当前文件生效，不会写入本地映射库。", response.text)
        self.assertIn("data-bulk-confidence-preview-form", response.text)
        self.assertIn('method="get" data-bulk-confidence-preview-form', response.text)
        self.assertIn('value="90"', response.text)
        self.assertIn("确认本次采纳", response.text)
        self.assertIn("以后遇到相同术语将自动映射。", response.text)

        preview = self.client.get(f"/jobs/{job_id}/mapping/bulk-confidence-preview?threshold=90")
        self.assertEqual(preview.status_code, 200)
        self.assertEqual(preview.json()["threshold"], 0.9)
        self.assertEqual(preview.json()["eligible_total"], 1)
        self.assertEqual(preview.json()["would_apply_decision"], "accept_once")

        with mock.patch.dict(os.environ, {"LLM_MAPPING_MOCK": "true", "DEEPSEEK_API_KEY": ""}, clear=False):
            rerun_response = self.client.post(
                f"/jobs/{job_id}/standard-metrics/run",
                data={"raw_metrics_path": str(raw_path)},
                follow_redirects=False,
            )
        self.assertEqual(rerun_response.status_code, 303)
        rerun_state = load_simple_flow_state(job)
        rerun_output_dir = Path(rerun_state["standardized_metrics_csv"]).parent
        rerun_llm_summary = json.loads((rerun_output_dir / "llm_mapping_summary.json").read_text(encoding="utf-8"))
        self.assertEqual(rerun_llm_summary["cached_suggestions_total"], 1)
        output_dir = rerun_output_dir

        store = LocalMappingStore(self.settings.mapping_store_path)
        before_aliases = store.count("term_aliases", where="enabled = 1 AND COALESCE(source, '') != 'base'")
        apply_response = self.client.post(f"/jobs/{job_id}/mapping/bulk-accept-confidence", data={"threshold": "90"})
        self.assertEqual(apply_response.status_code, 200)
        self.assertEqual(apply_response.json()["applied_total"], 1)
        self.assertFalse(apply_response.json()["mutated_local_alias_store"])
        self.assertEqual(before_aliases, store.count("term_aliases", where="enabled = 1 AND COALESCE(source, '') != 'base'"))
        self.assertTrue((output_dir / "confidence_bulk_accept_apply_summary.json").exists())
        with (output_dir / "mapping_decisions.csv").open("r", encoding="utf-8-sig", newline="") as handle:
            decisions = list(csv.DictReader(handle))
        self.assertEqual(decisions[-1]["decision"], "accept_once")

        browser_submit = self.client.post(
            f"/jobs/{job_id}/mapping/decision",
            data={
                "review_item_id": "maprev_000001",
                "selected_code": "ST_001",
                "selected_name": "总营业额",
                "decision": "accept_once",
            },
            headers={"accept": "text/html,application/xhtml+xml"},
            follow_redirects=False,
        )
        self.assertEqual(browser_submit.status_code, 303)
        self.assertEqual(browser_submit.headers["location"], f"/jobs/{job_id}/proofread")

        proofread = self.client.get(f"/jobs/{job_id}/proofread")
        self.assertEqual(proofread.status_code, 200)
        self.assertIn("AI建议", proofread.text)
        self.assertIn("本操作只对当前文件生效，不会写入本地映射库。", proofread.text)
        self.assertIn("已本次采用", proofread.text)
        self.assertNotIn("不采纳", proofread.text)
        self.assertNotIn("采用并记住", proofread.text)

    def test_unified_proofread_page_loads_combined_table_and_saves_edits(self):
        job_id = self._create_job("unified review")
        raw_path = self._create_raw_metrics_fixture(metric_name="货币资金", metric_value="12345.67", confidence="0.92", text_confidence="0.91")
        self.client.post(
            f"/jobs/{job_id}/standard-metrics/run",
            data={"raw_metrics_path": str(raw_path)},
            follow_redirects=False,
        )
        response = self.client.get(f"/jobs/{job_id}/proofread")
        self.assertEqual(response.status_code, 200)
        self.assertIn("data-unified-proofread-workbench", response.text)
        self.assertIn("/static/app.js?v=current-review-20260813-1", response.text)
        self.assertIn("/static/style.css?v=current-review-20260813-1", response.text)
        self.assertIn("source-panel", response.text)
        self.assertIn("sheet-panel", response.text)
        self.assertIn("data-page-image-key=", response.text)
        for text in ("原始术语", "填表日期：", "单位：", "期间数值一", "期间数值二", "期末数", "标准术语", "状态", "映射决策"):
            self.assertIn(text, response.text)
        self.assertNotIn("<th>表格日期</th>", response.text)
        self.assertEqual(response.text.count("data-section-date-input"), 1)
        self.assertEqual(response.text.count("data-section-unit-select"), 1)
        self.assertEqual(response.text.count("section-metadata-page"), 1)
        self.assertNotIn("原单位", response.text)
        self.assertNotIn("→ 人民币元", response.text)
        self.assertNotIn("期间类型", response.text)
        self.assertIn("资产负债表", response.text)
        self.assertNotIn("来源表 0", response.text)
        self.assertNotIn("拆分区 1", response.text)
        self.assertIn("精确匹配，无需决策", response.text)
        self.assertNotIn("不采纳", response.text)
        self.assertNotIn("仅本次采用", response.text)
        self.assertNotIn("采用并记住", response.text)
        self.assertIn("12,345.67", response.text)
        self.assertIn('class="confidence-switch"', response.text)
        self.assertIn("显示OCR置信度", response.text)
        self.assertIn("data-ocr-confidence-text hidden", response.text)
        self.assertIn('class="todo-note unified-confidence-summary" data-ocr-confidence-text hidden', response.text)
        self.assertIn("data-unified-status-filter", response.text)
        self.assertIn("data-unified-confidence-filter", response.text)
        self.assertIn("data-unified-confidence-threshold", response.text)
        self.assertIn('value="90"', response.text)
        self.assertIn("data-unified-sort", response.text)
        self.assertIn("data-bulk-confidence-preview-form", response.text)
        self.assertIn("data-bulk-confidence-apply-button", response.text)
        self.assertIn("OCR识别文字 置信度91%", response.text)
        self.assertIn("OCR识别数字 置信度92%", response.text)
        self.assertIn("词语映射 置信度100%", response.text)
        self.assertIn("data-mapping-confidence-text", response.text)
        self.assertIn("预览下载版", response.text)
        self.assertIn("data-unified-download-preview", response.text)
        self.assertNotIn("来源页码", response.text)
        self.assertNotIn("unified-col-source-page", response.text)
        self.assertNotIn(">通过</button>", response.text)
        self.assertNotIn(">跳过</button>", response.text)
        self.assertNotIn(">修改映射</button>", response.text)
        self.assertNotIn("data-sheet-tabs", response.text)
        self.assertNotIn("没有找到标准术语", response.text)

        self.assertEqual(format_metric_number("396149420.6"), "396,149,420.60")
        self.assertEqual(format_metric_number("396149420.62"), "396,149,420.62")
        self.assertEqual(format_metric_number("396149420.625"), "396,149,420.63")
        self.assertEqual(parse_metric_number_input("396,149,420.6")["value"], "396149420.60")
        self.assertEqual(parse_metric_number_input("396,149,420.62")["value"], "396149420.62")
        self.assertEqual(parse_metric_number_input("396,149,420.625")["value"], "396149420.63")
        self.assertTrue(parse_metric_number_input("396,149,420.625")["precision_adjusted"])
        self.assertFalse(parse_metric_number_input("12.34%")["valid"])

        save_response = self.client.post(
            f"/jobs/{job_id}/proofread/save",
            json={
                "reviewer_name": "auditor",
                "edits": [
                    {
                        "item_id": "unirev_000001",
                        "edit_type": "value_change",
                        "previous_value": "12345.67",
                        "new_value": "12,346.675",
                    },
                    {
                        "item_id": "unirev_000001",
                        "edit_type": "mapping_change",
                        "previous_code": "ZT_001",
                        "previous_name": "货币资金",
                        "new_code": "ZT_068",
                        "new_name": "短期借款",
                    },
                ],
            },
        )
        self.assertEqual(save_response.status_code, 200)
        self.assertTrue(save_response.json()["combined_workbook_refreshed"])
        self.assertTrue(Path(save_response.json()["combined_metrics_xlsx"]).exists())
        self.assertEqual(save_response.json()["precision_warnings_total"], 1)
        job = get_job(self.settings, job_id)
        action_dir = unified_review_dir(job)
        actions_path = action_dir / "unified_review_actions.json"
        csv_path = action_dir / "unified_review_actions.csv"
        summary_path = action_dir / "unified_review_summary.json"
        self.assertTrue(actions_path.exists())
        self.assertTrue(csv_path.exists())
        self.assertTrue(summary_path.exists())
        actions = json.loads(actions_path.read_text(encoding="utf-8"))
        self.assertEqual([item["edit_type"] for item in actions[-2:]], ["value_change", "mapping_change"])
        self.assertEqual(actions[-2]["new_value"], "12346.68")
        self.assertEqual(actions[-1]["new_code"], "ZT_068")
        self.assertEqual(actions[-1]["new_name"], "短期借款")
        state = load_simple_flow_state(job)
        workbook = load_workbook(Path(state["combined_metrics_xlsx"]), data_only=True)
        total_sheet = workbook["数据总表"]
        value_col = self._column_index(total_sheet, "指标数值")
        code_col = self._column_index(total_sheet, "标准指标编码")
        name_col = self._column_index(total_sheet, "标准指标名称")
        self.assertEqual(total_sheet.cell(row=2, column=value_col).value, 12346.68)
        self.assertEqual(total_sheet.cell(row=2, column=code_col).value, "ZT_068")
        self.assertEqual(total_sheet.cell(row=2, column=name_col).value, "短期借款")
        workbook.close()
        preview_response = self.client.get(f"/jobs/{job_id}/download-preview/combined_metrics_xlsx")
        self.assertEqual(preview_response.status_code, 200)
        self.assertIn("12,346.68", preview_response.text)
        self.assertIn("短期借款", preview_response.text)
        saved_page = self.client.get(f"/jobs/{job_id}/proofread")
        self.assertEqual(saved_page.status_code, 200)
        self.assertIn("数值已修改", saved_page.text)
        self.assertIn("术语已修改", saved_page.text)
        self.assertNotIn(">已修改</span>", saved_page.text)
        self.assertIn('data-value-changed="true"', saved_page.text)
        self.assertIn('data-mapping-changed="true"', saved_page.text)
        self.assertIn('data-original-value="12345.67"', saved_page.text)
        self.assertIn('data-saved-value="12346.68"', saved_page.text)
        self.assertIn('data-original-code="ZT_001"', saved_page.text)
        self.assertIn('data-saved-code="ZT_068"', saved_page.text)
        self.assertTrue((raw_review_dir(job) / "raw_review_actions.json").exists())
        self.assertTrue((mapping_review_dir(job) / "mapping_review_actions.json").exists())
        self.assertTrue(str(actions_path.resolve()).startswith(str(self.runtime_root.resolve())))
        self.assertFalse((REPO_ROOT / "unified_review_actions.csv").exists())

    def test_unified_proofread_saves_date_and_source_unit_corrections(self):
        job_id = self._create_job("unified date unit review")
        raw_path = self._create_raw_metrics_fixture(
            metric_name="货币资金",
            metric_value="1000000.00",
            value_raw="100.00",
            unit_raw="万元",
        )
        self.client.post(
            f"/jobs/{job_id}/standard-metrics/run",
            data={"raw_metrics_path": str(raw_path)},
            follow_redirects=False,
        )

        page = self.client.get(f"/jobs/{job_id}/proofread")
        self.assertEqual(page.status_code, 200)
        self.assertIn("data-section-date-input", page.text)
        self.assertIn("data-section-unit-select", page.text)
        self.assertEqual(page.text.count("data-section-date-input"), 1)
        self.assertEqual(page.text.count("data-section-unit-select"), 1)
        self.assertIn("保存并生成会计报表", page.text)

        response = self.client.post(
            f"/jobs/{job_id}/proofread/save",
            json={
                "edits": [
                    {
                        "item_id": "unirev_000001",
                        "raw_metric_id": "CASE1:1:aliyun_table:0:1-1:2-2",
                        "raw_metric_ids": ["CASE1:1:aliyun_table:0:1-1:2-2"],
                        "edit_type": "date_change",
                        "previous_date": "2022-12-31",
                        "new_date": "2023-12-31",
                    },
                    {
                        "item_id": "unirev_000001",
                        "raw_metric_id": "CASE1:1:aliyun_table:0:1-1:2-2",
                        "edit_type": "unit_change",
                        "previous_unit": "万元",
                        "new_unit": "千元",
                        "previous_value": "1000000.00",
                        "new_value": "100000.00",
                    },
                ]
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["date_changes_total"], 1)
        self.assertEqual(response.json()["unit_changes_total"], 1)
        job = get_job(self.settings, job_id)
        state = load_simple_flow_state(job)
        self.assertTrue(state["combined_csv_ready"])
        with Path(state["combined_metrics_csv"]).open("r", encoding="utf-8-sig", newline="") as handle:
            row = next(csv.DictReader(handle))
        self.assertEqual(row["填表日期"], "2023-12-31")
        self.assertEqual(row["当前条目日期"], "2023-12-31")
        self.assertEqual(row["指标数值"], "100000.00")
        self.assertEqual(row["原始单位"], "千元")
        self.assertEqual(row["标准单位"], "人民币元")

        saved_page = self.client.get(f"/jobs/{job_id}/proofread")
        self.assertIn("日期已修改", saved_page.text)
        self.assertIn("单位已修改", saved_page.text)
        self.assertIn('data-date-changed="true"', saved_page.text)
        self.assertIn('data-unit-changed="true"', saved_page.text)
        self.assertNotIn("日期待校对", saved_page.text)

    def test_unified_proofread_section_metadata_updates_every_row_in_one_table(self):
        job_id = self._create_job("unified table metadata")
        raw_path = self._create_raw_metrics_fixture(metric_name="货币资金", metric_value="1000000.00")
        run_dir = raw_path.parent
        raw_ids = [
            "CASE1:1:aliyun_table:0:1-1:2-2",
            "CASE1:1:aliyun_table:0:2-2:2-2",
        ]
        rows = [
            {
                "填表日期": "2022-12-31",
                "当前条目日期": "2022-12-31",
                "公司名": "AAA有限公司",
                "指标名": metric_name,
                "指标数值": value,
            }
            for metric_name, value in (("货币资金", "1000000.00"), ("短期借款", "2000000.00"))
        ]
        self._write_csv(
            raw_path,
            rows,
            ["填表日期", "当前条目日期", "公司名", "指标名", "指标数值"],
        )
        detailed_path = run_dir / "raw_metrics_detailed.csv"
        with detailed_path.open("r", encoding="utf-8-sig", newline="") as handle:
            fieldnames = list(csv.DictReader(handle).fieldnames or [])
        detailed_rows = []
        for index, (raw_id, metric_name, value) in enumerate(
            zip(raw_ids, ("货币资金", "短期借款"), ("100.00", "200.00")),
            start=1,
        ):
            detailed_rows.append(
                {
                    "source_cell_ref": raw_id,
                    "page_no": "1",
                    "bbox_json": '[{"x":1,"y":2},{"x":3,"y":4}]',
                    "evidence_path": str(self.corpus_root / "CASE1" / "input" / "sample.pdf"),
                    "source_file": "fixture.json",
                    "provider": "aliyun_table",
                    "doc_id": "CASE1",
                    "table_id": "0",
                    "logical_subtable_id": f"0_sub{index}",
                    "row_index": str(index),
                    "col_index": "2",
                    "row_label_clean": metric_name,
                    "period_role_raw": "期末数",
                    "period_role_norm": "ending",
                    "statement_type": "balance_sheet",
                    "statement_name_raw": "资产负债表",
                    "value_type": "amount",
                    "value_raw": value,
                    "unit_raw": "万元",
                    "unit_multiplier": "10000",
                }
            )
        self._write_csv(detailed_path, detailed_rows, fieldnames)
        self.client.post(
            f"/jobs/{job_id}/standard-metrics/run",
            data={"raw_metrics_path": str(raw_path)},
            follow_redirects=False,
        )

        page = self.client.get(f"/jobs/{job_id}/proofread")
        self.assertEqual(page.status_code, 200)
        self.assertEqual(page.text.count("data-section-date-input"), 1)
        self.assertEqual(page.text.count("data-section-unit-select"), 1)
        self.assertEqual(page.text.count("section-metadata-page"), 1)
        self.assertIn("资产负债表", page.text)
        self.assertNotIn("资产负债表-左半部分", page.text)
        self.assertNotIn("资产负债表-右半部分", page.text)

        response = self.client.post(
            f"/jobs/{job_id}/proofread/save",
            json={
                "edits": [
                    {
                        "item_id": "unirev_000001",
                        "raw_metric_id": raw_ids[0],
                        "raw_metric_ids": raw_ids,
                        "edit_type": "date_change",
                        "previous_date": "2022-12-31",
                        "new_date": "2023-12-31",
                    },
                    {
                        "item_id": "unirev_000001",
                        "raw_metric_id": raw_ids[0],
                        "raw_metric_ids": raw_ids,
                        "edit_type": "unit_change",
                        "previous_unit": "万元",
                        "new_unit": "千元",
                    },
                ]
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["date_changes_total"], 2)
        self.assertEqual(response.json()["unit_changes_total"], 2)
        state = load_simple_flow_state(get_job(self.settings, job_id))
        with Path(state["combined_metrics_csv"]).open("r", encoding="utf-8-sig", newline="") as handle:
            saved_rows = list(csv.DictReader(handle))
        self.assertEqual([row["填表日期"] for row in saved_rows], ["2023-12-31", "2023-12-31"])
        self.assertEqual([row["原始单位"] for row in saved_rows], ["千元", "千元"])
        self.assertEqual([row["指标数值"] for row in saved_rows], ["100000.00", "200000.00"])

        reset_response = self.client.post(
            f"/jobs/{job_id}/proofread/save",
            json={
                "edits": [
                    {
                        "item_id": "unirev_000001",
                        "raw_metric_id": raw_ids[0],
                        "raw_metric_ids": raw_ids,
                        "edit_type": "reset_date",
                        "previous_date": "2023-12-31",
                    },
                    {
                        "item_id": "unirev_000001",
                        "raw_metric_id": raw_ids[0],
                        "raw_metric_ids": raw_ids,
                        "edit_type": "reset_unit",
                        "previous_unit": "千元",
                    },
                ]
            },
        )
        self.assertEqual(reset_response.status_code, 200)
        reset_state = load_simple_flow_state(get_job(self.settings, job_id))
        with Path(reset_state["combined_metrics_csv"]).open("r", encoding="utf-8-sig", newline="") as handle:
            reset_rows = list(csv.DictReader(handle))
        self.assertEqual([row["填表日期"] for row in reset_rows], ["2022-12-31", "2022-12-31"])
        self.assertEqual([row["原始单位"] for row in reset_rows], ["万元", "万元"])
        self.assertEqual([row["指标数值"] for row in reset_rows], ["1000000.00", "2000000.00"])

    def test_unified_proofread_merges_beginning_and_ending_values_into_one_row(self):
        job_id = self._create_job("unified merged periods")
        raw_path = self._create_raw_metrics_fixture(metric_name="货币资金", metric_value="100")
        run_dir = raw_path.parent
        self._write_csv(
            raw_path,
            [
                {
                    "填表日期": "2022-12-31",
                    "当前条目日期": "2022-01-01",
                    "期间类型": "期初数",
                    "公司名": "AAA有限公司",
                    "指标名": "应收帐款",
                    "指标数值": "74954432.97",
                },
                {
                    "填表日期": "2022-12-31",
                    "当前条目日期": "2022-12-31",
                    "期间类型": "期末数",
                    "公司名": "AAA有限公司",
                    "指标名": "应收帐款",
                    "指标数值": "80163547.04",
                },
            ],
            ["填表日期", "当前条目日期", "期间类型", "公司名", "指标名", "指标数值"],
        )
        self._write_csv(
            run_dir / "raw_metrics_detailed.csv",
            [
                {
                    "source_cell_ref": "CASE1:1:aliyun_table:0:1-1:2-2",
                    "page_no": "1",
                    "bbox_json": '[{"x":1,"y":2},{"x":3,"y":4}]',
                    "text_confidence": "0.96",
                    "value_confidence": "0.91",
                    "evidence_path": str(self.corpus_root / "CASE1" / "input" / "sample.pdf"),
                    "source_file": "fixture.json",
                    "provider": "aliyun_table",
                    "doc_id": "CASE1",
                    "table_id": "0",
                    "logical_subtable_id": "0_sub1",
                    "row_index": "1",
                    "col_index": "2",
                    "row_label_raw": "应收帐款",
                    "row_label_clean": "应收帐款",
                    "period_role_raw": "期初数",
                    "period_role_norm": "beginning",
                    "statement_type": "balance_sheet",
                    "statement_name_raw": "资产负债表",
                    "value_type": "amount",
                    "confidence": "0.91",
                },
                {
                    "source_cell_ref": "CASE1:1:aliyun_table:0:1-1:3-3",
                    "page_no": "1",
                    "bbox_json": '[{"x":5,"y":6},{"x":7,"y":8}]',
                    "text_confidence": "0.96",
                    "value_confidence": "0.93",
                    "evidence_path": str(self.corpus_root / "CASE1" / "input" / "sample.pdf"),
                    "source_file": "fixture.json",
                    "provider": "aliyun_table",
                    "doc_id": "CASE1",
                    "table_id": "0",
                    "logical_subtable_id": "0_sub1",
                    "row_index": "1",
                    "col_index": "3",
                    "row_label_raw": "应收帐款",
                    "row_label_clean": "应收帐款",
                    "period_role_raw": "期末数",
                    "period_role_norm": "ending",
                    "statement_type": "balance_sheet",
                    "statement_name_raw": "资产负债表",
                    "value_type": "amount",
                    "confidence": "0.93",
                },
            ],
            [
                "source_cell_ref",
                "page_no",
                "bbox_json",
                "text_confidence",
                "value_confidence",
                "evidence_path",
                "source_file",
                "provider",
                "doc_id",
                "table_id",
                "logical_subtable_id",
                "row_index",
                "col_index",
                "row_label_raw",
                "row_label_clean",
                "period_role_raw",
                "period_role_norm",
                "statement_type",
                "statement_name_raw",
                "value_type",
                "confidence",
            ],
        )
        self.client.post(
            f"/jobs/{job_id}/standard-metrics/run",
            data={"raw_metrics_path": str(raw_path)},
            follow_redirects=False,
        )
        job = get_job(self.settings, job_id)
        self.assertIsNotNone(job)
        items = load_unified_review_items(job)
        self.assertEqual(len(items), 1)
        self.assertEqual(items[0]["original_metric_name"], "应收帐款")
        self.assertEqual(items[0]["table_date"], "2022-12-31")
        self.assertEqual(items[0]["beginning_value"]["current_value"], "74954432.97")
        self.assertEqual(items[0]["ending_value"]["current_value"], "80163547.04")
        self.assertEqual(items[0]["value_confidence_score"], "0.910000")

        response = self.client.get(f"/jobs/{job_id}/proofread")
        self.assertEqual(response.status_code, 200)
        self.assertNotIn("期间类型", response.text)
        self.assertEqual(response.text.count('data-original-metric-name="应收帐款"'), 1)
        self.assertIn("74,954,432.97", response.text)
        self.assertIn("80,163,547.04", response.text)

    def test_unified_proofread_merges_current_and_previous_period_values_into_one_row(self):
        job_id = self._create_job("unified merged current previous periods")
        raw_path = self._create_raw_metrics_fixture(metric_name="主营业务收入", metric_value="100")
        run_dir = raw_path.parent
        self._write_csv(
            raw_path,
            [
                {
                    "填表日期": "2022-12-31",
                    "当前条目日期": "",
                    "期间类型": "本期",
                    "公司名": "AAA有限公司",
                    "指标名": "主营业务收入",
                    "指标数值": "251143230.20",
                },
                {
                    "填表日期": "2022-12-31",
                    "当前条目日期": "",
                    "期间类型": "上年累计",
                    "公司名": "AAA有限公司",
                    "指标名": "主营业务收入",
                    "指标数值": "227585011.97",
                },
            ],
            ["填表日期", "当前条目日期", "期间类型", "公司名", "指标名", "指标数值"],
        )
        self._write_csv(
            run_dir / "raw_metrics_detailed.csv",
            [
                {
                    "source_cell_ref": "CASE1:2:aliyun_table:0:1-1:2-2",
                    "page_no": "2",
                    "bbox_json": '[{"x":1,"y":2},{"x":3,"y":4}]',
                    "text_confidence": "",
                    "value_confidence": "0.91",
                    "evidence_path": str(self.corpus_root / "CASE1" / "input" / "sample.pdf"),
                    "source_file": "fixture.json",
                    "provider": "aliyun_table",
                    "doc_id": "CASE1",
                    "table_id": "0",
                    "logical_subtable_id": "0_sub1",
                    "row_index": "1",
                    "col_index": "2",
                    "row_label_raw": "主营业务收入",
                    "row_label_clean": "主营业务收入",
                    "header_path": "本年累计数",
                    "period_role_raw": "本期",
                    "period_role_norm": "current_period",
                    "statement_type": "income_statement",
                    "statement_name_raw": "利润表",
                    "value_type": "amount",
                    "confidence": "0.91",
                },
                {
                    "source_cell_ref": "CASE1:2:aliyun_table:0:1-1:3-3",
                    "page_no": "2",
                    "bbox_json": '[{"x":5,"y":6},{"x":7,"y":8}]',
                    "text_confidence": "",
                    "value_confidence": "0.92",
                    "evidence_path": str(self.corpus_root / "CASE1" / "input" / "sample.pdf"),
                    "source_file": "fixture.json",
                    "provider": "aliyun_table",
                    "doc_id": "CASE1",
                    "table_id": "0",
                    "logical_subtable_id": "0_sub1",
                    "row_index": "1",
                    "col_index": "3",
                    "row_label_raw": "主营业务收入",
                    "row_label_clean": "主营业务收入",
                    "header_path": "上年累计数",
                    "period_role_raw": "上年累计",
                    "period_role_norm": "previous_period",
                    "statement_type": "income_statement",
                    "statement_name_raw": "利润表",
                    "value_type": "amount",
                    "confidence": "0.92",
                },
            ],
            [
                "source_cell_ref",
                "page_no",
                "bbox_json",
                "text_confidence",
                "value_confidence",
                "evidence_path",
                "source_file",
                "provider",
                "doc_id",
                "table_id",
                "logical_subtable_id",
                "row_index",
                "col_index",
                "row_label_raw",
                "row_label_clean",
                "header_path",
                "period_role_raw",
                "period_role_norm",
                "statement_type",
                "statement_name_raw",
                "value_type",
                "confidence",
            ],
        )
        self.client.post(
            f"/jobs/{job_id}/standard-metrics/run",
            data={"raw_metrics_path": str(raw_path)},
            follow_redirects=False,
        )
        job = get_job(self.settings, job_id)
        self.assertIsNotNone(job)
        items = load_unified_review_items(job)
        self.assertEqual(len(items), 1)
        self.assertEqual(items[0]["original_metric_name"], "主营业务收入")
        self.assertEqual(items[0]["beginning_value"]["slot_label"], "本年累计数")
        self.assertEqual(items[0]["ending_value"]["slot_label"], "上年累计数")
        self.assertEqual(items[0]["beginning_value"]["current_value"], "251143230.20")
        self.assertEqual(items[0]["ending_value"]["current_value"], "227585011.97")

        response = self.client.get(f"/jobs/{job_id}/proofread")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.text.count('data-original-metric-name="主营业务收入"'), 1)
        self.assertIn("本年累计数", response.text)
        self.assertIn("上年累计数", response.text)
        self.assertIn("251,143,230.20", response.text)
        self.assertIn("227,585,011.97", response.text)

    def test_unified_proofread_single_amount_column_uses_generic_empty_second_header(self):
        job_id = self._create_job("unified single amount period")
        raw_path = self._create_raw_metrics_fixture(metric_name="支付给职工以及为职工支付的现金", metric_value="2698533.22")
        run_dir = raw_path.parent
        self._write_csv(
            raw_path,
            [
                {
                    "填表日期": "2022-12-31",
                    "当前条目日期": "2022-12-31",
                    "期间类型": "金额",
                    "公司名": "AAA有限公司",
                    "指标名": "支付给职工以及为职工支付的现金",
                    "指标数值": "2698533.22",
                }
            ],
            ["填表日期", "当前条目日期", "期间类型", "公司名", "指标名", "指标数值"],
        )
        self._write_csv(
            run_dir / "raw_metrics_detailed.csv",
            [
                {
                    "source_cell_ref": "CASE1:3:aliyun_table:0:64-64:2-2",
                    "page_no": "3",
                    "bbox_json": '[{"x":1,"y":2},{"x":3,"y":4}]',
                    "text_confidence": "",
                    "value_confidence": "0.91",
                    "evidence_path": str(self.corpus_root / "CASE1" / "input" / "sample.pdf"),
                    "source_file": "fixture.json",
                    "provider": "aliyun_table",
                    "doc_id": "CASE1",
                    "table_id": "0",
                    "logical_subtable_id": "0_sub1",
                    "row_index": "64",
                    "col_index": "2",
                    "row_label_raw": "支付给职工以及为职工支付的现金",
                    "row_label_clean": "支付给职工以及为职工支付的现金",
                    "header_path": "金额",
                    "period_role_raw": "金额",
                    "period_role_norm": "amount",
                    "statement_type": "cash_flow",
                    "statement_name_raw": "现金流量表",
                    "value_type": "amount",
                    "confidence": "0.91",
                }
            ],
            [
                "source_cell_ref",
                "page_no",
                "bbox_json",
                "text_confidence",
                "value_confidence",
                "evidence_path",
                "source_file",
                "provider",
                "doc_id",
                "table_id",
                "logical_subtable_id",
                "row_index",
                "col_index",
                "row_label_raw",
                "row_label_clean",
                "header_path",
                "period_role_raw",
                "period_role_norm",
                "statement_type",
                "statement_name_raw",
                "value_type",
                "confidence",
            ],
        )
        self.client.post(
            f"/jobs/{job_id}/standard-metrics/run",
            data={"raw_metrics_path": str(raw_path)},
            follow_redirects=False,
        )
        job = get_job(self.settings, job_id)
        self.assertIsNotNone(job)
        items = load_unified_review_items(job)
        self.assertEqual(len(items), 1)
        self.assertEqual(items[0]["period_columns"][0]["label"], "金额")
        self.assertEqual(items[0]["period_columns"][1]["label"], "期间数值二")

        response = self.client.get(f"/jobs/{job_id}/proofread")
        self.assertEqual(response.status_code, 200)
        self.assertIn("金额", response.text)
        self.assertIn("期间数值二", response.text)
        self.assertNotIn("现金流量表 期初数 期末数", response.text)

    def test_unified_proofread_section_label_infers_statement_name_and_split_position(self):
        job_id = self._create_job("unified readable section label")
        source_file = self.temp_path / "tencent_balance_page.json"
        source_file.write_text(
            json.dumps(
                {
                    "TableDetections": [
                        {
                            "TableCoordPoint": [
                                {"X": 100, "Y": 100},
                                {"X": 900, "Y": 100},
                                {"X": 900, "Y": 500},
                                {"X": 100, "Y": 500},
                            ],
                            "Cells": [
                                {"RowTl": 0, "RowBr": 1, "ColTl": 0, "ColBr": 1, "Text": "资产"},
                                {"RowTl": 0, "RowBr": 1, "ColTl": 1, "ColBr": 2, "Text": "行次"},
                                {"RowTl": 0, "RowBr": 1, "ColTl": 2, "ColBr": 3, "Text": "年初数"},
                                {"RowTl": 0, "RowBr": 1, "ColTl": 3, "ColBr": 4, "Text": "期末数"},
                                {"RowTl": 0, "RowBr": 1, "ColTl": 4, "ColBr": 5, "Text": "负债及所有者权益"},
                                {"RowTl": 0, "RowBr": 1, "ColTl": 5, "ColBr": 6, "Text": "行次"},
                                {"RowTl": 0, "RowBr": 1, "ColTl": 6, "ColBr": 7, "Text": "年初数"},
                                {"RowTl": 0, "RowBr": 1, "ColTl": 7, "ColBr": 8, "Text": "期末数"},
                                {"RowTl": 2, "RowBr": 3, "ColTl": 4, "ColBr": 5, "Text": "短期借款"},
                                {"RowTl": 2, "RowBr": 3, "ColTl": 6, "ColBr": 7, "Text": "37,550,000.00"},
                                {"RowTl": 2, "RowBr": 3, "ColTl": 7, "ColBr": 8, "Text": "136,000,000.00"},
                            ],
                        }
                    ],
                    "Angle": 0,
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        raw_path = self._create_raw_metrics_fixture(metric_name="短期借款", metric_value="136000000")
        run_dir = raw_path.parent
        self._write_csv(
            raw_path,
            [
                {
                    "填表日期": "2022-12-31",
                    "当前条目日期": "2022-12-31",
                    "期间类型": "期末数",
                    "公司名": "AAA有限公司",
                    "指标名": "短期借款",
                    "指标数值": "136000000",
                }
            ],
            ["填表日期", "当前条目日期", "期间类型", "公司名", "指标名", "指标数值"],
        )
        self._write_csv(
            run_dir / "raw_metrics_detailed.csv",
            [
                {
                    "source_cell_ref": "CASE1:1:tencent_table_v3:1:2-2:7-7",
                    "page_no": "1",
                    "bbox_json": "",
                    "text_confidence": "",
                    "value_confidence": "0.94",
                    "evidence_path": str(self.corpus_root / "CASE1" / "input" / "sample.pdf"),
                    "source_file": str(source_file),
                    "provider": "tencent_table_v3",
                    "doc_id": "CASE1",
                    "table_id": "1",
                    "logical_subtable_id": "1_sub2",
                    "row_index": "2",
                    "col_index": "7",
                    "row_label_raw": "短期借款",
                    "row_label_clean": "短期借款",
                    "period_role_raw": "期末数",
                    "period_role_norm": "ending",
                    "statement_type": "",
                    "statement_name_raw": "",
                    "value_type": "amount",
                    "confidence": "0.94",
                }
            ],
            [
                "source_cell_ref",
                "page_no",
                "bbox_json",
                "text_confidence",
                "value_confidence",
                "evidence_path",
                "source_file",
                "provider",
                "doc_id",
                "table_id",
                "logical_subtable_id",
                "row_index",
                "col_index",
                "row_label_raw",
                "row_label_clean",
                "period_role_raw",
                "period_role_norm",
                "statement_type",
                "statement_name_raw",
                "value_type",
                "confidence",
            ],
        )
        self.client.post(
            f"/jobs/{job_id}/standard-metrics/run",
            data={"raw_metrics_path": str(raw_path)},
            follow_redirects=False,
        )

        response = self.client.get(f"/jobs/{job_id}/proofread")
        self.assertEqual(response.status_code, 200)
        self.assertIn("资产负债表-右半部分", response.text)
        self.assertNotIn("来源表 1", response.text)
        self.assertNotIn("拆分区 2", response.text)

    def test_unified_proofread_rejects_invalid_numeric_edit_and_documents_link_to_it(self):
        job_id = self._create_job("unified invalid")
        raw_path = self._create_raw_metrics_fixture(metric_name="货币资金", metric_value="100")
        self.client.post(
            f"/jobs/{job_id}/standard-metrics/run",
            data={"raw_metrics_path": str(raw_path)},
            follow_redirects=False,
        )
        invalid = self.client.post(
            f"/jobs/{job_id}/proofread/save",
            json={"edits": [{"item_id": "unirev_000001", "edit_type": "value_change", "new_value": "abc"}]},
        )
        self.assertEqual(invalid.status_code, 400)
        self.assertIn("数值格式有误", invalid.text)

        doc_id = self._upload_library_pdf()
        update_document_status(self.settings, doc_id, ocr_status=STATUS_COMPLETED, raw_metrics_status=STATUS_COMPLETED)
        self._attach_raw_summary_to_document(doc_id, raw_path)
        continue_response = self.client.get(f"/documents/{doc_id}/continue")
        self.assertEqual(continue_response.status_code, 200)
        self.assertIn(f'href="/documents/{doc_id}/proofread"', continue_response.text)

    def test_unified_proofread_source_attrs_confidence_missing_and_autocomplete_bug_contract(self):
        job_id = self._create_job("unified attrs")
        raw_path = self._create_raw_metrics_fixture(metric_name="货币资金", metric_value="396149420.62", bbox_json="")
        self.client.post(
            f"/jobs/{job_id}/standard-metrics/run",
            data={"raw_metrics_path": str(raw_path)},
            follow_redirects=False,
        )
        response = self.client.get(f"/jobs/{job_id}/proofread")
        self.assertEqual(response.status_code, 200)
        self.assertIn("396,149,420.62", response.text)
        self.assertIn("未记录", response.text)
        self.assertIn("data-term-bbox=", response.text)
        self.assertIn("data-value-bbox=", response.text)
        self.assertIn("当前项目未记录位置", response.text)

        script = (REPO_ROOT / "webapp" / "static" / "app.js").read_text(encoding="utf-8")
        self.assertIn('if (!query.trim())', script)
        self.assertIn('event.key === "Escape"', script)
        self.assertIn('input.addEventListener("blur"', script)
        self.assertIn('!event.target.closest(".standard-term-picker")', script)
        self.assertIn("closeAutocompleteResults(root);", script)
        self.assertIn('statusBadge("value_changed", "数值已修改")', script)
        self.assertIn('statusBadge("term_changed", "术语已修改")', script)
        self.assertIn("valueInput.dataset.savedValue", script)
        self.assertIn("data-ocr-confidence-text", script)
        self.assertIn("openUnifiedDownloadPreview", script)
        self.assertIn("data-section-date-input", script)
        self.assertIn("data-section-unit-select", script)
        self.assertNotIn("data-row-page-chip", script)
        self.assertIn("sectionRowsToSort", script)
        self.assertIn("header.hidden = !hasVisibleRows", script)
        self.assertNotIn('setAttribute("data-original-value"', script)
        self.assertNotIn("search();\n    });\n    input.addEventListener(\"click\"", script)
        style = (REPO_ROOT / "webapp" / "static" / "style.css").read_text(encoding="utf-8")
        self.assertIn(".unified-review-table thead th", style)
        self.assertIn("position: static;", style)
        self.assertIn(".table-section-row th", style)
        self.assertIn("top: 0;", style)

    def test_unified_proofread_filters_numeric_metric_name_but_keeps_missing_period_for_review(self):
        job_id = self._create_job("unified filters bad source rows")
        raw_path = self._create_raw_metrics_fixture(metric_name="货币资金", metric_value="100")
        run_dir = raw_path.parent
        self._write_csv(
            raw_path,
            [
                {
                    "填表日期": "2022-12-31",
                    "当前条目日期": "2022-12-31",
                    "期间类型": "期末数",
                    "公司名": "AAA有限公司",
                    "指标名": "货币资金",
                    "指标数值": "100",
                },
                {
                    "填表日期": "2022-12-31",
                    "当前条目日期": "",
                    "期间类型": "42,940,481.00",
                    "公司名": "AAA有限公司",
                    "指标名": "26",
                    "指标数值": "7132218.77",
                },
                {
                    "填表日期": "2022-12-31",
                    "当前条目日期": "",
                    "期间类型": "",
                    "公司名": "AAA有限公司",
                    "指标名": "其他应付款",
                    "指标数值": "113866652.00",
                },
            ],
            ["填表日期", "当前条目日期", "期间类型", "公司名", "指标名", "指标数值"],
        )
        coarse_bbox = '[{"X":77,"Y":211},{"X":1159,"Y":211},{"X":1159,"Y":1513},{"X":77,"Y":1513}]'
        self._write_csv(
            run_dir / "raw_metrics_detailed.csv",
            [
                {
                    "source_cell_ref": "CASE1:1:aliyun_table:0:1-1:2-2",
                    "page_no": "1",
                    "bbox_json": '[{"x":1,"y":2},{"x":3,"y":4}]',
                    "text_confidence": "",
                    "value_confidence": "",
                    "evidence_path": str(self.corpus_root / "CASE1" / "input" / "sample.pdf"),
                    "source_file": "fixture.json",
                    "provider": "aliyun_table",
                    "doc_id": "CASE1",
                    "table_id": "0",
                    "logical_subtable_id": "0_sub1",
                    "row_index": "1",
                    "col_index": "2",
                    "row_label_raw": "货币资金",
                    "row_label_clean": "货币资金",
                    "period_role_raw": "期末数",
                    "period_role_norm": "ending",
                    "statement_type": "balance_sheet",
                    "statement_name_raw": "资产负债表",
                    "value_type": "amount",
                    "confidence": "",
                },
                {
                    "source_cell_ref": "CASE1:1:tencent_table_v3:5:6-8:1-3",
                    "page_no": "1",
                    "bbox_json": coarse_bbox,
                    "text_confidence": "",
                    "value_confidence": "",
                    "evidence_path": str(self.corpus_root / "CASE1" / "input" / "sample.pdf"),
                    "source_file": "fixture.json",
                    "provider": "tencent_table_v3",
                    "doc_id": "CASE1",
                    "table_id": "5",
                    "logical_subtable_id": "5_sub1",
                    "row_index": "6",
                    "col_index": "1",
                    "row_label_raw": "53,295,859.26",
                    "row_label_clean": "26",
                    "period_role_raw": "42,940,481.00",
                    "period_role_norm": "unknown",
                    "statement_type": "balance_sheet",
                    "statement_name_raw": "资产负债表",
                    "value_type": "amount",
                    "confidence": "",
                },
                {
                    "source_cell_ref": "CASE1:1:tencent_table_v3:5:8-10:3-5",
                    "page_no": "1",
                    "bbox_json": coarse_bbox,
                    "text_confidence": "",
                    "value_confidence": "",
                    "evidence_path": str(self.corpus_root / "CASE1" / "input" / "sample.pdf"),
                    "source_file": "fixture.json",
                    "provider": "tencent_table_v3",
                    "doc_id": "CASE1",
                    "table_id": "5",
                    "logical_subtable_id": "5_sub1",
                    "row_index": "8",
                    "col_index": "3",
                    "row_label_raw": "其他应付款",
                    "row_label_clean": "其他应付款",
                    "period_role_raw": "",
                    "period_role_norm": "unknown",
                    "statement_type": "balance_sheet",
                    "statement_name_raw": "资产负债表",
                    "value_type": "amount",
                    "confidence": "",
                },
            ],
            [
                "source_cell_ref",
                "page_no",
                "bbox_json",
                "text_confidence",
                "value_confidence",
                "evidence_path",
                "source_file",
                "provider",
                "doc_id",
                "table_id",
                "logical_subtable_id",
                "row_index",
                "col_index",
                "row_label_raw",
                "row_label_clean",
                "period_role_raw",
                "period_role_norm",
                "statement_type",
                "statement_name_raw",
                "value_type",
                "confidence",
            ],
        )
        self.client.post(
            f"/jobs/{job_id}/standard-metrics/run",
            data={"raw_metrics_path": str(raw_path)},
            follow_redirects=False,
        )
        job = get_job(self.settings, job_id)
        self.assertIsNotNone(job)

        unified_items = load_unified_review_items(job)
        mapping_items = load_mapping_review_items(job)
        self.assertEqual([item["original_metric_name"] for item in unified_items], ["货币资金", "其他应付款"])
        missing_period_item = next(item for item in unified_items if item["original_metric_name"] == "其他应付款")
        self.assertTrue(missing_period_item["temporal_review_required"])
        self.assertEqual(missing_period_item["base_status"], "review_required")
        self.assertFalse(any(item.get("original_metric_name") == "26" for item in mapping_items))
        self.assertTrue(any(item.get("original_metric_name") == "其他应付款" for item in mapping_items))

        response = self.client.get(f"/jobs/{job_id}/proofread")
        self.assertEqual(response.status_code, 200)
        self.assertIn("货币资金", response.text)
        self.assertIn("期末数", response.text)
        self.assertNotIn(">26<", response.text)
        self.assertIn("其他应付款", response.text)
        self.assertIn("日期待校对", response.text)
        self.assertNotIn("42,940,481.00", response.text)
        self.assertNotIn(coarse_bbox, response.text)

    def test_standard_term_search_supports_code_name_and_pinyin_initials(self):
        for query in ("68", "068", "ZT_068", "短期", "借款", "dqjk"):
            results = search_standard_terms(query, limit=5)
            self.assertTrue(results, query)
            self.assertEqual(results[0]["code"], "ZT_068")
            self.assertEqual(results[0]["name"], "短期借款")

        response = self.client.get("/api/standard-terms/search?q=dqjk")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["results"][0]["display_label"], "ZT_068 短期借款")

    def test_webapp_base_path_serves_pages_api_and_redirects(self):
        settings = self.make_settings(base_path="/AutoFinance")
        settings.ensure_directories()
        init_db(settings)
        with TestClient(create_app(settings)) as client:
            response = client.get("/AutoFinance/")
            self.assertEqual(response.status_code, 200)
            self.assertIn('href="/AutoFinance/documents/upload"', response.text)
            self.assertIn("/AutoFinance/static/style.css", response.text)
            self.assertIn("/AutoFinance/static/app.js", response.text)
            self.assertIn('window.__APP_BASE_PATH__ = "/AutoFinance"', response.text)
            self.assertNotIn('href="/documents/upload"', response.text)

            static_response = client.get("/AutoFinance/static/app.js")
            self.assertEqual(static_response.status_code, 200)

            api_response = client.get("/AutoFinance/api/standard-terms/search?q=68")
            self.assertEqual(api_response.status_code, 200)
            self.assertEqual(api_response.json()["results"][0]["display_label"], "ZT_068 短期借款")

            redirect_response = client.get("/AutoFinance/advanced", follow_redirects=False)
            self.assertEqual(redirect_response.status_code, 303)
            self.assertEqual(redirect_response.headers["location"], "/AutoFinance/jobs")

            health_response = client.get("/healthz")
            self.assertEqual(health_response.status_code, 200)

    def test_mapping_review_escapes_unsafe_bbox_and_keeps_source_attrs(self):
        job_id = self._create_job("mapping unsafe bbox")
        raw_path = self._create_raw_metrics_fixture(metric_name="货币资金", bbox_json='"><script>alert(1)</script>')
        self.client.post(
            f"/jobs/{job_id}/standard-metrics/run",
            data={"raw_metrics_path": str(raw_path)},
            follow_redirects=False,
        )
        response = self.client.get(f"/jobs/{job_id}/mapping-review")
        self.assertEqual(response.status_code, 200)
        self.assertIn("data-bbox=", response.text)
        self.assertIn("data-bbox-state=", response.text)
        self.assertNotIn("<script>alert(1)</script>", response.text)

    def test_mapping_label_match_allows_ocr_tail_noise(self):
        target = _normalize_ocr_label("负债和所有者权益(或股东权益)")
        candidate = _normalize_ocr_label("负债和所有者权益(或股东权益包")
        unrelated = _normalize_ocr_label("一年内到期的长期借款")
        self.assertTrue(_label_matches(candidate, target))
        self.assertFalse(_label_matches(unrelated, _normalize_ocr_label("长期借款")))

    def test_tencent_shared_table_polygon_gets_grid_cell_bbox_for_review_highlight(self):
        source_file = self.temp_path / "tencent_shared_polygon.json"
        table_polygon = [
            {"X": 100, "Y": 200},
            {"X": 500, "Y": 200},
            {"X": 500, "Y": 600},
            {"X": 100, "Y": 600},
        ]
        source_file.write_text(
            json.dumps(
                {
                    "Angle": 0,
                    "TableDetections": [
                        {
                            "Type": "table",
                            "TableCoordPoint": table_polygon,
                            "Cells": [
                                {"RowTl": 0, "RowBr": 1, "ColTl": 0, "ColBr": 1, "Text": "项目", "Polygon": table_polygon},
                                {"RowTl": 0, "RowBr": 1, "ColTl": 2, "ColBr": 3, "Text": "期末数", "Polygon": table_polygon},
                                {"RowTl": 2, "RowBr": 3, "ColTl": 0, "ColBr": 1, "Text": "货币资金", "Polygon": table_polygon},
                                {"RowTl": 2, "RowBr": 3, "ColTl": 2, "ColBr": 3, "Text": "100", "Polygon": table_polygon},
                                {"RowTl": 3, "RowBr": 4, "ColTl": 3, "ColBr": 4, "Text": "200", "Polygon": table_polygon},
                            ],
                        }
                    ],
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        detailed = {
            "source_file": str(source_file),
            "table_id": "1",
            "row_index": "2",
            "col_index": "2",
        }

        value_bbox = json.loads(_resolve_source_table_cell_bbox_json(detailed))
        self.assertEqual(value_bbox[0], {"x": 300.0, "y": 400.0})
        self.assertEqual(value_bbox[2], {"x": 400.0, "y": 500.0})

        term_bbox = json.loads(_resolve_mapping_term_bbox_json({}, detailed, "货币资金"))
        self.assertEqual(term_bbox[0], {"x": 100.0, "y": 400.0})
        self.assertEqual(term_bbox[2], {"x": 200.0, "y": 500.0})

    def test_rotated_source_page_is_rendered_in_ocr_orientation(self):
        source_file = self.temp_path / "rotated_page.json"
        source_file.write_text(
            json.dumps(
                {
                    "Data": {
                        "angle": 90,
                        "width": 1649,
                        "height": 1157,
                        "orgWidth": 1157,
                        "orgHeight": 1649,
                    }
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        self.assertEqual(source_preview_rotation_degrees({"source_file": str(source_file)}), -90)
        tencent_source_file = self.temp_path / "tencent_rotated_page.json"
        tencent_source_file.write_text(
            json.dumps({"Angle": 27000, "Data": ""}, ensure_ascii=False),
            encoding="utf-8",
        )
        self.assertEqual(source_preview_rotation_degrees({"source_file": str(tencent_source_file)}), -90)

    def test_unsafe_evidence_pdf_paths_are_rejected(self):
        outside_pdf = self.temp_path / "outside.pdf"
        outside_pdf.write_bytes(b"%PDF-1.4\n%outside\n")
        job_id = self._create_job("unsafe evidence")
        raw_path = self._create_raw_metrics_fixture(metric_name="货币资金", evidence_path=outside_pdf)
        self._attach_raw_summary_to_job(job_id, raw_path)
        response = self.client.get(f"/jobs/{job_id}/raw-review/evidence/rawrev_000001")
        self.assertEqual(response.status_code, 400)

    def test_stale_evidence_path_falls_back_to_job_source_pdf(self):
        job_id = self._create_job("stale evidence")
        stale_local_path = Path("Z:/old-local-machine/missing-source.pdf")
        raw_path = self._create_raw_metrics_fixture(metric_name="货币资金", evidence_path=stale_local_path)
        self._attach_raw_summary_to_job(job_id, raw_path)

        page = self.client.get(f"/jobs/{job_id}/raw-review")
        self.assertEqual(page.status_code, 200)
        self.assertIn("/raw-review/page-image/rawrev_000001", page.text)

        evidence = self.client.get(f"/jobs/{job_id}/raw-review/evidence/rawrev_000001")
        self.assertEqual(evidence.status_code, 200)
        self.assertIn("application/pdf", evidence.headers.get("content-type", ""))

    def test_stale_source_file_falls_back_for_rotation_and_term_bbox(self):
        source_json = self.sample_input_dir / "aliyun_table" / "demo_doc" / "raw" / "page_0001.json"
        source_json.write_text(
            json.dumps(
                {
                    "Data": {
                        "angle": 90,
                        "prism_tablesInfo": [
                            {
                                "tableId": "0",
                                "cellInfos": [
                                    {"ysc": 1, "yec": 1, "xsc": 0, "xec": 0, "word": "短期借款", "pos": [10, 20, 80, 40]},
                                    {"ysc": 1, "yec": 1, "xsc": 2, "xec": 2, "word": "100", "pos": [200, 20, 260, 40]},
                                ],
                            }
                        ],
                    }
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        job_id = self._create_job("stale source file")
        stale_source_file = Path("C:/Users/gater/Desktop/Python/finance/AutoFinance/data/corpus/library/CASE1/ocr_outputs/aliyun_table/demo_doc/raw/page_0001.json")
        raw_path = self._create_raw_metrics_fixture(metric_name="短期借款", source_file=stale_source_file)
        self._attach_raw_summary_to_job(job_id, raw_path)
        job = get_job(self.settings, job_id)

        raw_items = load_raw_review_items(job)
        self.assertEqual(Path(raw_items[0]["source_file"]).resolve(), source_json.resolve())
        self.assertEqual(source_preview_rotation_degrees(raw_items[0]), -90)

        standard_dir = self.standard_metrics_root / "WEB_TEST_STALE_SOURCE" / "RUN_WEB_TEST"
        self.addCleanup(lambda: shutil.rmtree(standard_dir.parent, ignore_errors=True))
        standard_dir.mkdir(parents=True, exist_ok=True)
        standard_csv = standard_dir / "standardized_metrics.csv"
        self._write_csv(standard_csv, [{"raw_metric_id": "CASE1:1:aliyun_table:0:1-1:2-2"}], ["raw_metric_id"])
        self._write_csv(
            standard_dir / "mapping_review_items.csv",
            [
                {
                    "review_item_id": "maprev_000001",
                    "raw_metric_id": "CASE1:1:aliyun_table:0:1-1:2-2",
                    "original_metric_name": "短期借款",
                    "candidate_code": "ZT_068",
                    "candidate_name": "短期借款",
                    "mapping_status": "review_required",
                    "mapping_method": "candidate",
                    "source_page_no": "1",
                }
            ],
            [
                "review_item_id",
                "raw_metric_id",
                "original_metric_name",
                "candidate_code",
                "candidate_name",
                "mapping_status",
                "mapping_method",
                "source_page_no",
            ],
        )
        standard_step_summary_path(job).parent.mkdir(parents=True, exist_ok=True)
        standard_step_summary_path(job).write_text(
            json.dumps({"standardized_metrics_csv": str(standard_csv)}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

        mapping_items = load_mapping_review_items(job)
        self.assertIn('"x":10.0', mapping_items[0]["source_term_bbox_json"])
        self.assertNotIn('"x":200.0', mapping_items[0]["source_term_bbox_json"])

        unified_items = load_unified_review_items(job)
        self.assertIn('"x":10.0', unified_items[0]["source_term_bbox_json"])
        self.assertNotIn('"x":200.0', unified_items[0]["source_term_bbox_json"])

    def test_simple_flow_output_files_stay_under_data_generated(self):
        job_id = self._create_job("path hygiene step2")
        raw_path = self._create_raw_metrics_fixture(metric_name="往来款")
        self.client.post(
            f"/jobs/{job_id}/standard-metrics/run",
            data={"raw_metrics_path": str(raw_path)},
            follow_redirects=False,
        )
        job = get_job(self.settings, job_id)
        state = load_simple_flow_state(job)
        output_dir = Path(state["standard_summary"]["output_dir"]).resolve()
        self.assertTrue(str(output_dir).startswith(str(self.standard_metrics_root.resolve())))
        self.assertTrue(str(Path(job.output_dir)).startswith(str(self.settings.jobs_root)))
        self.assertFalse((REPO_ROOT / "standardized_metrics.csv").exists())

    def test_create_standardize_only_job_from_existing_ocr_path(self):
        job_id = self._create_job("CASE1 smoke")
        job = get_job(self.settings, job_id)
        self.assertIsNotNone(job)
        self.assertEqual(job.status, "queued")

    def test_invalid_upload_extension_is_rejected(self):
        response = self.client.post(
            "/jobs",
            data={"mode": "upload_pdf", "display_name": "bad upload"},
            files={"uploaded_files": ("notes.txt", b"not a pdf", "text/plain")},
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("不支持的上传文件类型", response.text)

    def test_upload_pdf_missing_ocr_credentials_returns_chinese_error(self):
        response = self.client.post(
            "/jobs",
            data={
                "mode": "upload_pdf",
                "display_name": "missing secret",
                "upload_provider_mode": "aliyun_table",
            },
            files={"uploaded_files": ("demo.pdf", b"%PDF-1.4\n%mock\n", "application/pdf")},
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("当前未配置阿里云 OCR 密钥", response.text)

    def test_create_upload_pdf_job_with_selected_provider_and_mock_runtime(self):
        settings = self.make_settings(auto_run_upload_ocr=True)
        settings.ensure_directories()
        init_db(settings)
        with mock.patch.dict(
            os.environ,
            {
                "WEBAPP_UPLOAD_OCR_MOCK_MODE": "copy_fixture",
                "WEBAPP_UPLOAD_OCR_MOCK_SOURCE_DIR": str(self.sample_input_dir),
            },
            clear=False,
        ):
            with TestClient(create_app(settings)) as client:
                response = client.post(
                    "/jobs",
                    data={
                        "mode": "upload_pdf",
                        "display_name": "upload demo",
                        "upload_provider_mode": "tencent_table_v3",
                    },
                    files={"uploaded_files": ("demo.pdf", b"%PDF-1.4\n%mock\n", "application/pdf")},
                    follow_redirects=False,
                )
        self.assertEqual(response.status_code, 303)
        job_id = response.headers["location"].rstrip("/").split("/")[-1]
        job = get_job(settings, job_id)
        self.assertIsNotNone(job)
        self.assertEqual(job.provider_mode, "tencent_table_v3")
        self.assertEqual(job.status, "queued")
        self.assertTrue((settings.uploads_root / job_id / "demo.pdf").exists())

    def test_upload_pdf_worker_captures_ocr_and_standardize_logs_separately(self):
        settings = self.make_settings(auto_run_upload_ocr=True)
        settings.ensure_directories()
        init_db(settings)
        with mock.patch.dict(
            os.environ,
            {
                "WEBAPP_UPLOAD_OCR_MOCK_MODE": "copy_fixture",
                "WEBAPP_UPLOAD_OCR_MOCK_SOURCE_DIR": str(self.sample_input_dir),
            },
            clear=False,
        ):
            with TestClient(create_app(settings)) as client:
                response = client.post(
                    "/jobs",
                    data={
                        "mode": "upload_pdf",
                        "display_name": "upload run",
                        "upload_provider_mode": "cloud_first",
                    },
                    files={"uploaded_files": ("demo.pdf", b"%PDF-1.4\n%mock\n", "application/pdf")},
                    follow_redirects=False,
                )
                self.assertEqual(response.status_code, 303)
                job_id = response.headers["location"].rstrip("/").split("/")[-1]
                with mock.patch("webapp.runner._run_subprocess", side_effect=self._fake_subprocess(profile="clean_success")):
                    run_worker_once(settings)
        job = get_job(settings, job_id)
        self.assertIsNotNone(job)
        self.assertEqual(job.status, "succeeded")
        ocr_stdout = settings.logs_root / job_id / "ocr_stdout.txt"
        standardize_stdout = settings.logs_root / job_id / "standardize_stdout.txt"
        self.assertTrue(ocr_stdout.exists())
        self.assertTrue(standardize_stdout.exists())
        ocr_stage_summary = json.loads((settings.results_root / job_id / "ocr_stage_summary.json").read_text(encoding="utf-8"))
        self.assertTrue(ocr_stage_summary["used_mock"])
        self.assertFalse(ocr_stage_summary["cloud_ocr_executed"])
        log_bundle = json.loads((settings.results_root / job_id / "job_log_bundle.json").read_text(encoding="utf-8"))
        log_names = {item["name"] for item in log_bundle["log_files"]}
        self.assertIn("ocr_stdout.txt", log_names)
        self.assertIn("standardize_stdout.txt", log_names)

    def test_deployment_preflight_success_and_script_writes_summary(self):
        self._write_secret_file(aliyun=True)
        settings = self.make_settings(
            env_mode="prod",
            auth_required=True,
            admin_password="demo-pass",
            upload_ocr_method="aliyun_table",
        )
        settings.ensure_directories()
        summary = run_deployment_preflight(settings, deployment_profile="aliyun", min_free_bytes=1)
        self.assertTrue(summary["pass"])

        output_path = settings.runtime_root / "deployment_check_summary.json"
        with mock.patch.dict(
            os.environ,
            {
                "WEBAPP_ENV": "prod",
                "WEBAPP_AUTH_REQUIRED": "1",
                "WEBAPP_ADMIN_PASSWORD": "demo-pass",
                "WEBAPP_QUEUE_BACKEND": "local",
                "WEBAPP_UPLOAD_OCR_METHOD": "aliyun_table",
                "WEBAPP_RUNTIME_ROOT": str(settings.runtime_root),
                "WEBAPP_UPLOADS_ROOT": str(settings.uploads_root),
                "WEBAPP_JOBS_ROOT": str(settings.jobs_root),
                "WEBAPP_RESULTS_ROOT": str(settings.results_root),
                "WEBAPP_LOGS_ROOT": str(settings.logs_root),
                "WEBAPP_DB_PATH": str(settings.db_path),
                "WEBAPP_TEMPLATE_PATH": str(settings.template_path),
                "WEBAPP_SECRET_PATH": str(settings.secret_path),
            },
            clear=False,
        ):
            exit_code = deployment_check_main(["--profile", "aliyun", "--output", str(output_path), "--min-free-mb", "1"])
        self.assertEqual(exit_code, 0)
        self.assertTrue(output_path.exists())

    def test_deployment_preflight_failure_when_prod_missing_password(self):
        settings = self.make_settings(
            env_mode="prod",
            auth_required=True,
            admin_password="",
            upload_ocr_method="aliyun_table",
        )
        summary = run_deployment_preflight(settings, deployment_profile="aliyun", min_free_bytes=1)
        self.assertFalse(summary["pass"])
        error_text = "\n".join(summary["errors"])
        self.assertIn("WEBAPP_ADMIN_PASSWORD", error_text)

    def test_system_status_api_does_not_expose_secret_values(self):
        self._write_secret_file(aliyun=True, secret_value="TOP_SECRET_STAGE11")
        response = self.client.get("/api/system-status")
        self.assertEqual(response.status_code, 200)
        self.assertNotIn("TOP_SECRET_STAGE11", response.text)
        self.assertNotIn("demo-id", response.text)

    def test_docker_compose_config_validates_when_available(self):
        if shutil.which("docker") is None:
            self.skipTest("docker not installed")
        compose_version = subprocess.run(
            ["docker", "compose", "version"],
            cwd=str(REPO_ROOT),
            capture_output=True,
            text=True,
            check=False,
        )
        if compose_version.returncode != 0:
            self.skipTest("docker compose plugin unavailable")
        result = subprocess.run(
            ["docker", "compose", "-f", "docker-compose.yml", "-f", "docker-compose.aliyun.yml", "config"],
            cwd=str(REPO_ROOT),
            capture_output=True,
            text=True,
            check=False,
        )
        self.assertEqual(result.returncode, 0, msg=result.stderr or result.stdout)

    def test_succeeded_with_warnings_job_quality_classification(self):
        job_id = self._create_job("warning job")
        with mock.patch("webapp.runner._run_subprocess", side_effect=self._fake_subprocess(profile="warning")):
            run_worker_once(self.settings)
        job = get_job(self.settings, job_id)
        self.assertEqual(job.status, "succeeded_with_warnings")
        quality_summary_path = Path(job.result_dir) / "job_quality_summary.json"
        self.assertTrue(quality_summary_path.exists())
        quality_summary = json.loads(quality_summary_path.read_text(encoding="utf-8"))
        self.assertEqual(quality_summary["final_job_status"], "succeeded_with_warnings")
        self.assertEqual(quality_summary["artifact_integrity_review_total"], 2)
        self.assertEqual(quality_summary["command_exit_code"], 0)

    def test_failed_command_classification_and_error_translation(self):
        job_id = self._create_job("failed job")
        raw_error = "Template workbook does not exist: mock-template.xlsx"
        with mock.patch("webapp.runner._run_subprocess", side_effect=self._fake_failed_subprocess(raw_error)):
            run_worker_once(self.settings)
        job = get_job(self.settings, job_id)
        self.assertEqual(job.status, "failed")
        self.assertIn("模板文件不存在", job.user_friendly_error)
        quality_summary = json.loads((Path(job.result_dir) / "job_quality_summary.json").read_text(encoding="utf-8"))
        self.assertEqual(quality_summary["final_job_status"], "failed")
        self.assertIn("模板文件不存在", quality_summary["user_friendly_error"])

    def test_job_detail_page_contains_chinese_user_facing_summary(self):
        job_id = self._create_job("needs review job")
        with mock.patch("webapp.runner._run_subprocess", side_effect=self._fake_subprocess(profile="needs_review")):
            run_worker_once(self.settings)
        response = self.client.get(f"/jobs/{job_id}")
        self.assertEqual(response.status_code, 200)
        self.assertIn("处理状态", response.text)
        self.assertIn("是否成功生成会计报表", response.text)
        self.assertIn("下一步建议", response.text)
        self.assertIn("已生成结果，但建议复核", response.text)

    def test_output_discovery_handles_missing_optional_files(self):
        job = JobRecord(
            job_id="job_manual",
            display_name="manual",
            mode="existing_ocr_outputs",
            provider_mode="cloud_first",
            input_path=str(self.sample_input_dir),
            source_image_dir="",
            upload_dir="",
            ocr_output_dir="",
            template_path=str(self.template_path),
            output_dir=str(self.runtime_root / "jobs" / "job_manual" / "standardize"),
            result_dir=str(self.runtime_root / "results" / "job_manual"),
            log_dir=str(self.runtime_root / "logs" / "job_manual"),
            provider_priority="aliyun,tencent",
            status="succeeded",
            current_stage="completed",
            progress_summary="done",
            created_at="",
            updated_at="",
            started_at="",
            finished_at="",
            error_message="",
            raw_error_message="",
            user_friendly_error="",
            recommended_action="",
            run_id="RUN_TEST_001",
            command_executed="",
            exit_code=0,
            timeout_seconds=120,
        )
        output_dir = Path(job.output_dir)
        result_dir = Path(job.result_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        result_dir.mkdir(parents=True, exist_ok=True)
        self._write_fake_workbook(output_dir / "会计报表_填充结果.xlsx")
        (output_dir / "run_summary.json").write_text("{}", encoding="utf-8")
        artifacts = {item.slug: item for item in discover_output_files(job)}
        self.assertTrue(artifacts["filled_workbook"].exists)
        self.assertFalse(artifacts["review_workbook"].exists)
        self.assertFalse(artifacts["quality_summary"].exists)

    def test_no_generated_web_files_go_into_repo_root(self):
        job_id = self._create_job("path smoke")
        with mock.patch("webapp.runner._run_subprocess", side_effect=self._fake_subprocess(profile="warning")):
            run_worker_once(self.settings)
        job = get_job(self.settings, job_id)
        self.assertTrue(str(Path(job.output_dir)).startswith(str(self.settings.jobs_root)))
        self.assertTrue(str(Path(job.result_dir)).startswith(str(self.settings.results_root)))
        self.assertTrue(str(self.settings.db_path).startswith(str(self.settings.runtime_root)))
        self.assertFalse((REPO_ROOT / job.job_id).exists())

    def test_job_detail_status_endpoint_returns_payload(self):
        job_id = self._create_job("status smoke")
        with mock.patch("webapp.runner._run_subprocess", side_effect=self._fake_subprocess(profile="warning")):
            run_worker_once(self.settings)
        status_response = self.client.get(f"/jobs/{job_id}/status")
        self.assertEqual(status_response.status_code, 200)
        payload = status_response.json()
        self.assertEqual(payload["job"]["job_id"], job_id)
        self.assertIn("quality_summary", payload)
        self.assertIn("output_files", payload)

    def test_review_dashboard_route_returns_200_for_job_with_review_files(self):
        job_id = self._prepare_review_job(include_optional=True)
        response = self.client.get(f"/jobs/{job_id}/review")
        self.assertEqual(response.status_code, 200)
        self.assertIn("复核看板", response.text)
        self.assertIn("待复核项目", response.text)
        self.assertIn("已提交动作", response.text)
        self.assertIn("原始问题行", response.text)

    def test_review_dashboard_handles_missing_optional_files(self):
        job_id = self._prepare_review_job(include_optional=False)
        response = self.client.get(f"/jobs/{job_id}/review")
        self.assertEqual(response.status_code, 200)
        self.assertIn("不可用", response.text)

    def test_review_item_loader_parses_review_queue_csv(self):
        job_id = self._prepare_review_job(include_optional=False)
        job = get_job(self.settings, job_id)
        items, _ = load_review_items(self.settings, job)
        review_item = next(item for item in items if item.source_type == "review_queue")
        self.assertEqual(review_item.review_item_id, "REV_case_1")
        self.assertEqual(review_item.reason_code, "mapping:unmapped")
        self.assertEqual(review_item.mapping_code, "ZT_001")

    def test_review_item_loader_parses_issues_and_validation_results(self):
        job_id = self._prepare_review_job(include_optional=True)
        job = get_job(self.settings, job_id)
        items, _ = load_review_items(self.settings, job)
        source_types = {item.source_type for item in items}
        self.assertIn("issue", source_types)
        self.assertIn("validation", source_types)

    def test_review_item_filtering_supports_source_type_and_status(self):
        job_id = self._prepare_review_job(include_optional=True)
        response = self.client.post(
            f"/jobs/{job_id}/review/actions",
            data={
                "review_item_id": "REV_case_1",
                "action_type": "defer",
                "action_value": "",
                "reviewer_note": "later",
                "reviewer_name": "tester",
                "next_url": f"/jobs/{job_id}/review/items",
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 303)
        job = get_job(self.settings, job_id)
        items, _ = load_review_items(self.settings, job)
        deferred = filter_review_items(items, status="deferred")
        mapping_items = filter_review_items(items, source_type="review_queue")
        self.assertEqual(len(deferred), 1)
        self.assertEqual(deferred[0].review_item_id, "REV_case_1")
        self.assertEqual(len(mapping_items), 1)

    def test_submitting_review_action_stores_it(self):
        job_id = self._prepare_review_job(include_optional=True)
        response = self.client.post(
            f"/jobs/{job_id}/review/actions",
            data={
                "review_item_id": "REV_case_1",
                "action_type": "ignore",
                "action_value": "",
                "reviewer_note": "false alert",
                "reviewer_name": "auditor",
                "next_url": f"/jobs/{job_id}/review/items",
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 303)
        actions = list_review_actions(self.settings, job_id)
        self.assertEqual(len(actions), 1)
        self.assertEqual(actions[0].action_type, "ignore")
        self.assertEqual(actions[0].reviewer_name, "auditor")

    def test_exporting_review_actions_creates_csv_xlsx_json_under_job_review_dir(self):
        job_id = self._prepare_review_job(include_optional=True)
        self.client.post(
            f"/jobs/{job_id}/review/actions",
            data={
                "review_item_id": "REV_case_1",
                "action_type": "defer",
                "action_value": "",
                "reviewer_note": "queue later",
                "reviewer_name": "auditor",
                "next_url": f"/jobs/{job_id}/review/items",
            },
            follow_redirects=False,
        )
        job = get_job(self.settings, job_id)
        result = export_review_actions(self.settings, job)
        review_dir = get_review_dir(job)
        self.assertTrue((review_dir / "review_actions_filled.csv").exists())
        self.assertTrue((review_dir / "review_actions_filled.xlsx").exists())
        self.assertTrue((review_dir / "review_action_export_summary.json").exists())
        self.assertTrue(str(result["csv_path"]).startswith(str(self.settings.jobs_root)))

    def test_evidence_path_outside_allowed_directories_is_rejected(self):
        job_id = self._prepare_review_job(include_optional=False, outside_evidence=True)
        response = self.client.get(f"/jobs/{job_id}/review/evidence/REV_case_1/cell")
        self.assertEqual(response.status_code, 404)

    def test_review_exports_do_not_go_into_repo_root(self):
        job_id = self._prepare_review_job(include_optional=True)
        self.client.post(
            f"/jobs/{job_id}/review/actions",
            data={
                "review_item_id": "REV_case_1",
                "action_type": "request_reocr",
                "action_value": "",
                "reviewer_note": "need targeted reocr",
                "reviewer_name": "auditor",
                "next_url": f"/jobs/{job_id}/review/items",
            },
            follow_redirects=False,
        )
        job = get_job(self.settings, job_id)
        review_dir = get_review_dir(job)
        export_review_actions(self.settings, job)
        self.assertTrue(str(review_dir).startswith(str(self.settings.jobs_root)))
        self.assertFalse((REPO_ROOT / "review_actions_filled.csv").exists())
        self.assertFalse((REPO_ROOT / "review_actions_filled.xlsx").exists())

    def test_review_dashboard_count_semantics_summary_is_generated(self):
        job_id = self._prepare_review_job(include_optional=True)
        response = self.client.get(f"/jobs/{job_id}/review")
        self.assertEqual(response.status_code, 200)
        self.assertIn("待复核项目", response.text)
        self.assertIn("已提交动作", response.text)
        self.assertIn("可处理项目", response.text)
        self.assertIn("原始问题行", response.text)
        self.assertIn("高优先级", response.text)
        job = get_job(self.settings, job_id)
        summary_path = get_review_dir(job) / "review_dashboard_counts_summary.json"
        self.assertTrue(summary_path.exists())
        summary = json.loads(summary_path.read_text(encoding="utf-8"))
        self.assertEqual(summary["total_review_items"], 4)
        self.assertEqual(summary["source_artifact_rows_total"], 4)
        self.assertEqual(summary["actions_submitted_total"], 0)
        self.assertEqual(summary["actionable_items_total"], 4)

    def test_review_items_page_contains_chinese_action_labels(self):
        job_id = self._prepare_review_job(include_optional=True)
        response = self.client.get(f"/jobs/{job_id}/review/items")
        self.assertEqual(response.status_code, 200)
        for text in (
            "暂缓处理",
            "忽略此项",
            "标记为非财务事实",
            "请求重新 OCR",
            "接受科目建议",
            "指定标准科目",
            "选择冲突赢家",
            "标记为误报",
        ):
            self.assertIn(text, response.text)

    def test_non_review_queue_items_get_stable_surrogate_review_item_id_and_source_ref(self):
        job_id = self._prepare_review_job(include_optional=True)
        job = get_job(self.settings, job_id)
        items_first, _ = load_review_items(self.settings, job)
        items_second, _ = load_review_items(self.settings, job)
        issue_first = next(item for item in items_first if item.source_type == "issue")
        issue_second = next(item for item in items_second if item.source_type == "issue")
        self.assertEqual(issue_first.review_item_id, issue_second.review_item_id)
        self.assertEqual(issue_first.review_id, "")
        self.assertTrue(issue_first.review_item_id.startswith("issue_"))
        self.assertTrue(issue_first.source_ref)

    def test_review_action_compatibility_summary_is_generated(self):
        job_id = self._prepare_review_job(include_optional=True)
        job = get_job(self.settings, job_id)
        items, _ = load_review_items(self.settings, job)
        review_item = next(item for item in items if item.source_type == "review_queue")
        issue_item = next(item for item in items if item.source_type == "issue")
        self.client.post(
            f"/jobs/{job_id}/review/actions",
            data={
                "review_item_id": review_item.review_item_id,
                "action_type": "ignore",
                "action_value": "",
                "reviewer_note": "close queue item",
                "reviewer_name": "auditor",
                "next_url": f"/jobs/{job_id}/review/items",
            },
            follow_redirects=False,
        )
        self.client.post(
            f"/jobs/{job_id}/review/actions",
            data={
                "review_item_id": issue_item.review_item_id,
                "action_type": "ignore",
                "action_value": "",
                "reviewer_note": "close issue item",
                "reviewer_name": "auditor",
                "next_url": f"/jobs/{job_id}/review/items",
            },
            follow_redirects=False,
        )
        result = export_review_actions(self.settings, job)
        self.assertTrue(Path(result["compatibility_summary_path"]).exists())
        summary = json.loads(Path(result["compatibility_summary_path"]).read_text(encoding="utf-8"))
        self.assertEqual(summary["actions_total"], 2)
        self.assertEqual(summary["backend_ready_total"], 1)
        self.assertEqual(summary["backend_partial_total"], 1)

    def test_apply_review_actions_route_creates_apply_outputs(self):
        job_id = self._prepare_review_job(include_optional=True)
        self.client.post(
            f"/jobs/{job_id}/review/actions",
            data={
                "review_item_id": "REV_case_1",
                "action_type": "ignore",
                "action_value": "",
                "reviewer_note": "close item",
                "reviewer_name": "auditor",
                "next_url": f"/jobs/{job_id}/review/items",
            },
            follow_redirects=False,
        )
        self.client.post(f"/jobs/{job_id}/review/export-actions", follow_redirects=False)
        response = self.client.post(f"/jobs/{job_id}/review/apply", follow_redirects=False)
        self.assertEqual(response.status_code, 303)
        self._run_next_worker_item()
        apply_status = self.client.get(f"/jobs/{job_id}/review/apply-status")
        self.assertEqual(apply_status.status_code, 200)
        latest_apply = apply_status.json()["latest_apply_summary"]
        self.assertEqual(latest_apply["applied_actions_total"], 1)
        self.assertEqual(latest_apply["rejected_actions_total"], 0)
        apply_dir = self.settings.jobs_root / job_id / "review" / latest_apply["apply_id"]
        self.assertTrue((apply_dir / "applied_review_actions.csv").exists())
        self.assertTrue((apply_dir / "review_apply_summary.json").exists())

    def test_rejected_unsupported_or_incompatible_actions_are_recorded(self):
        job_id = self._prepare_review_job(include_optional=True)
        job = get_job(self.settings, job_id)
        items, _ = load_review_items(self.settings, job)
        issue_item = next(item for item in items if item.source_type == "issue")
        self.client.post(
            f"/jobs/{job_id}/review/actions",
            data={
                "review_item_id": issue_item.review_item_id,
                "action_type": "defer",
                "action_value": "",
                "reviewer_note": "cannot backend apply",
                "reviewer_name": "auditor",
                "next_url": f"/jobs/{job_id}/review/items",
            },
            follow_redirects=False,
        )
        self.client.post(f"/jobs/{job_id}/review/export-actions", follow_redirects=False)
        self.client.post(f"/jobs/{job_id}/review/apply", follow_redirects=False)
        self._run_next_worker_item()
        latest_apply = self.client.get(f"/jobs/{job_id}/review/apply-status").json()["latest_apply_summary"]
        apply_dir = self.settings.jobs_root / job_id / "review" / latest_apply["apply_id"]
        rejected_path = apply_dir / "rejected_review_actions.csv"
        self.assertTrue(rejected_path.exists())
        rejected_text = rejected_path.read_text(encoding="utf-8-sig")
        self.assertIn("review_id_not_found", rejected_text)
        self.assertIn(issue_item.review_item_id, rejected_text)

    def test_apply_and_rerun_creates_new_rerun_output_directory(self):
        job_id = self._prepare_review_job(include_optional=True)
        self.client.post(
            f"/jobs/{job_id}/review/actions",
            data={
                "review_item_id": "REV_case_1",
                "action_type": "ignore",
                "action_value": "",
                "reviewer_note": "close item",
                "reviewer_name": "auditor",
                "next_url": f"/jobs/{job_id}/review/items",
            },
            follow_redirects=False,
        )
        self.client.post(f"/jobs/{job_id}/review/export-actions", follow_redirects=False)
        with mock.patch("webapp.review._run_patched_standardize_cli", side_effect=self._fake_review_rerun()):
            response = self.client.post(f"/jobs/{job_id}/review/apply-and-rerun", follow_redirects=False)
            self._run_next_worker_item()
        self.assertEqual(response.status_code, 303)
        rerun_dir = self.settings.jobs_root / job_id / "reruns" / "rerun_001" / "standardize"
        self.assertTrue(rerun_dir.exists())
        self.assertTrue((rerun_dir / "会计报表_填充结果.xlsx").exists())

    def test_review_rerun_delta_json_is_generated(self):
        job_id = self._prepare_review_job(include_optional=True)
        self.client.post(
            f"/jobs/{job_id}/review/actions",
            data={
                "review_item_id": "REV_case_1",
                "action_type": "ignore",
                "action_value": "",
                "reviewer_note": "close item",
                "reviewer_name": "auditor",
                "next_url": f"/jobs/{job_id}/review/items",
            },
            follow_redirects=False,
        )
        self.client.post(f"/jobs/{job_id}/review/export-actions", follow_redirects=False)
        with mock.patch("webapp.review._run_patched_standardize_cli", side_effect=self._fake_review_rerun()):
            self.client.post(f"/jobs/{job_id}/review/apply-and-rerun", follow_redirects=False)
            self._run_next_worker_item()
        delta_path = self.settings.results_root / job_id / "reruns" / "rerun_001" / "review_rerun_delta.json"
        self.assertTrue(delta_path.exists())
        delta = json.loads(delta_path.read_text(encoding="utf-8"))
        metric_map = {row["metric"]: row for row in delta["metrics"]}
        self.assertEqual(metric_map["review_total"]["before"], 2)
        self.assertEqual(metric_map["review_total"]["after"], 1)
        self.assertEqual(metric_map["validation_fail_total"]["after"], 0)

    def test_job_detail_shows_original_and_rerun_outputs(self):
        job_id = self._prepare_review_job(include_optional=True)
        self.client.post(
            f"/jobs/{job_id}/review/actions",
            data={
                "review_item_id": "REV_case_1",
                "action_type": "ignore",
                "action_value": "",
                "reviewer_note": "close item",
                "reviewer_name": "auditor",
                "next_url": f"/jobs/{job_id}/review/items",
            },
            follow_redirects=False,
        )
        self.client.post(f"/jobs/{job_id}/review/export-actions", follow_redirects=False)
        with mock.patch("webapp.review._run_patched_standardize_cli", side_effect=self._fake_review_rerun()):
            self.client.post(f"/jobs/{job_id}/review/apply-and-rerun", follow_redirects=False)
            self._run_next_worker_item()
        response = self.client.get(f"/jobs/{job_id}")
        self.assertEqual(response.status_code, 200)
        self.assertIn("结果版本", response.text)
        self.assertIn("original", response.text)
        self.assertIn("rerun_001", response.text)
        self.assertIn("当前推荐结果", response.text)

    def test_evidence_preview_fixture_has_real_preview_and_summary(self):
        job_id = self._prepare_review_job(include_optional=False)
        response = self.client.get(f"/jobs/{job_id}/review/evidence/REV_case_1/cell")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content[:8], b"\x89PNG\r\n\x1a\n")
        self.client.get(f"/jobs/{job_id}/review")
        job = get_job(self.settings, job_id)
        summary_path = get_review_dir(job) / "review_evidence_preview_summary.json"
        summary = json.loads(summary_path.read_text(encoding="utf-8"))
        self.assertGreaterEqual(summary["evidence_preview_available_count"], 1)
        self.assertTrue(summary["pass"])

    def test_review_apply_and_rerun_outputs_stay_under_web_runtime_paths(self):
        job_id = self._prepare_review_job(include_optional=True)
        self.client.post(
            f"/jobs/{job_id}/review/actions",
            data={
                "review_item_id": "REV_case_1",
                "action_type": "ignore",
                "action_value": "",
                "reviewer_note": "close item",
                "reviewer_name": "auditor",
                "next_url": f"/jobs/{job_id}/review/items",
            },
            follow_redirects=False,
        )
        self.client.post(f"/jobs/{job_id}/review/export-actions", follow_redirects=False)
        with mock.patch("webapp.review._run_patched_standardize_cli", side_effect=self._fake_review_rerun()):
            self.client.post(f"/jobs/{job_id}/review/apply-and-rerun", follow_redirects=False)
            self._run_next_worker_item()
        apply_root = self.settings.jobs_root / job_id / "review"
        rerun_root = self.settings.jobs_root / job_id / "reruns" / "rerun_001"
        rerun_result_root = self.settings.results_root / job_id / "reruns" / "rerun_001"
        self.assertTrue(str(apply_root).startswith(str(self.settings.jobs_root)))
        self.assertTrue(str(rerun_root).startswith(str(self.settings.jobs_root)))
        self.assertTrue(str(rerun_result_root).startswith(str(self.settings.results_root)))
        self.assertFalse((REPO_ROOT / "rerun_001").exists())

    def test_review_workbench_summary_is_generated(self):
        job_id = self._prepare_review_job(include_optional=True)
        response = self.client.get(f"/jobs/{job_id}/review")
        self.assertEqual(response.status_code, 200)
        for text in ("待复核总数", "已处理", "可自动应用", "有证据图片"):
            self.assertIn(text, response.text)
        job = get_job(self.settings, job_id)
        summary_path = get_review_dir(job) / "review_workbench_summary.json"
        self.assertTrue(summary_path.exists())
        summary = json.loads(summary_path.read_text(encoding="utf-8"))
        self.assertEqual(summary["review_items_total"], 4)
        self.assertEqual(summary["backend_ready_total"], 1)
        self.assertEqual(summary["backend_partial_total"], 2)
        self.assertEqual(summary["backend_suggestion_only_total"], 1)
        self.assertEqual(summary["evidence_available_total"], 1)
        self.assertTrue(summary["pass"])

    def test_review_items_can_filter_by_apply_compatibility(self):
        job_id = self._prepare_review_job(include_optional=True)
        job = get_job(self.settings, job_id)
        items, _ = load_review_items(self.settings, job)
        review_item = next(item for item in items if item.source_type == "review_queue")
        issue_item = next(item for item in items if item.source_type == "issue")
        response = self.client.get(f"/jobs/{job_id}/review/items?apply_compatibility=backend_ready")
        self.assertEqual(response.status_code, 200)
        self.assertIn(review_item.review_item_id, response.text)
        self.assertNotIn(issue_item.review_item_id, response.text)

    def test_review_items_can_filter_by_evidence_available(self):
        job_id = self._prepare_review_job(include_optional=True)
        job = get_job(self.settings, job_id)
        items, _ = load_review_items(self.settings, job)
        review_item = next(item for item in items if item.source_type == "review_queue")
        issue_item = next(item for item in items if item.source_type == "issue")
        response = self.client.get(f"/jobs/{job_id}/review/items?evidence_available=yes")
        self.assertEqual(response.status_code, 200)
        self.assertIn(review_item.review_item_id, response.text)
        self.assertNotIn(issue_item.review_item_id, response.text)

    def test_bulk_defer_creates_actions_and_summary(self):
        job_id = self._prepare_review_job(include_optional=True)
        job = get_job(self.settings, job_id)
        items, _ = load_review_items(self.settings, job)
        selected_ids = [item.review_item_id for item in items[:3]]
        response = self.client.post(
            f"/jobs/{job_id}/review/bulk-action",
            data={
                "selected_review_item_ids": selected_ids,
                "action_type": "defer",
                "action_value": "",
                "reviewer_note": "bulk defer",
                "reviewer_name": "auditor",
                "next_url": f"/jobs/{job_id}/review/items",
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 303)
        actions = list_review_actions(self.settings, job_id)
        self.assertEqual(len(actions), 3)
        self.assertTrue(all(action.action_type == "defer" for action in actions))
        summary_path = get_review_dir(job) / "bulk_review_action_summary.json"
        summary = json.loads(summary_path.read_text(encoding="utf-8"))
        self.assertEqual(summary["requested_total"], 3)
        self.assertEqual(summary["applied_total"], 3)
        self.assertEqual(summary["rejected_total"], 0)

    def test_bulk_ignore_creates_actions(self):
        job_id = self._prepare_review_job(include_optional=True)
        job = get_job(self.settings, job_id)
        items, _ = load_review_items(self.settings, job)
        selected_ids = [item.review_item_id for item in items[:2]]
        response = self.client.post(
            f"/jobs/{job_id}/review/bulk-action",
            data={
                "selected_review_item_ids": selected_ids,
                "action_type": "ignore",
                "action_value": "",
                "reviewer_note": "bulk ignore",
                "reviewer_name": "auditor",
                "next_url": f"/jobs/{job_id}/review/items",
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 303)
        actions = list_review_actions(self.settings, job_id)
        self.assertEqual(len(actions), 2)
        self.assertTrue(all(action.action_type == "ignore" for action in actions))

    def test_unsupported_bulk_action_items_are_rejected_with_reason(self):
        job_id = self._prepare_review_job(include_optional=True)
        job = get_job(self.settings, job_id)
        items, _ = load_review_items(self.settings, job)
        review_item = next(item for item in items if item.source_type == "review_queue")
        issue_item = next(item for item in items if item.source_type == "issue")
        response = self.client.post(
            f"/jobs/{job_id}/review/bulk-action",
            data={
                "selected_review_item_ids": [review_item.review_item_id, issue_item.review_item_id],
                "action_type": "accept_mapping_candidate",
                "action_value": "",
                "reviewer_note": "bulk mapping",
                "reviewer_name": "auditor",
                "next_url": f"/jobs/{job_id}/review/items",
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 303)
        summary = json.loads((get_review_dir(job) / "bulk_review_action_summary.json").read_text(encoding="utf-8"))
        self.assertEqual(summary["requested_total"], 2)
        self.assertEqual(summary["applied_total"], 1)
        self.assertEqual(summary["rejected_total"], 1)
        self.assertIn("compatibility_not_backend_ready", summary["rejected_reasons"])

    def test_review_apply_preview_summary_is_generated(self):
        job_id = self._prepare_review_job(include_optional=True)
        job = get_job(self.settings, job_id)
        items, _ = load_review_items(self.settings, job)
        review_item = next(item for item in items if item.source_type == "review_queue")
        issue_item = next(item for item in items if item.source_type == "issue")
        for item in (review_item, issue_item):
            self.client.post(
                f"/jobs/{job_id}/review/actions",
                data={
                    "review_item_id": item.review_item_id,
                    "action_type": "ignore",
                    "action_value": "",
                    "reviewer_note": "preview",
                    "reviewer_name": "auditor",
                    "next_url": f"/jobs/{job_id}/review/items",
                },
                follow_redirects=False,
            )
        self.client.post(f"/jobs/{job_id}/review/export-actions", follow_redirects=False)
        response = self.client.get(f"/jobs/{job_id}/review/apply-preview", follow_redirects=False)
        self.assertEqual(response.status_code, 303)
        summary = json.loads((get_review_dir(job) / "review_apply_preview_summary.json").read_text(encoding="utf-8"))
        self.assertEqual(summary["actions_total"], 2)
        self.assertEqual(summary["backend_ready_total"], 1)
        self.assertEqual(summary["partial_total"], 1)
        self.assertEqual(summary["likely_applied_total"], 1)
        self.assertEqual(summary["likely_rejected_total"], 1)

    def test_review_rerun_delta_explained_is_generated(self):
        job_id = self._prepare_review_job(include_optional=True)
        self.client.post(
            f"/jobs/{job_id}/review/actions",
            data={
                "review_item_id": "REV_case_1",
                "action_type": "ignore",
                "action_value": "",
                "reviewer_note": "close item",
                "reviewer_name": "auditor",
                "next_url": f"/jobs/{job_id}/review/items",
            },
            follow_redirects=False,
        )
        self.client.post(f"/jobs/{job_id}/review/export-actions", follow_redirects=False)
        with mock.patch("webapp.review._run_patched_standardize_cli", side_effect=self._fake_review_rerun()):
            self.client.post(f"/jobs/{job_id}/review/apply-and-rerun", follow_redirects=False)
            self._run_next_worker_item()
        explained_path = self.settings.results_root / job_id / "reruns" / "rerun_001" / "review_rerun_delta_explained.json"
        self.assertTrue(explained_path.exists())
        explained = json.loads(explained_path.read_text(encoding="utf-8"))
        self.assertIn("headline_status_before", explained)
        self.assertIn("headline_status_after", explained)
        self.assertTrue(explained["user_friendly_summary_zh"])
        self.assertIn("recommended_next_action_zh", explained)

    def test_operation_summary_status_endpoint_returns_latest_operation(self):
        job_id = self._prepare_review_job(include_optional=True)
        self.client.post(
            f"/jobs/{job_id}/review/actions",
            data={
                "review_item_id": "REV_case_1",
                "action_type": "ignore",
                "action_value": "",
                "reviewer_note": "close item",
                "reviewer_name": "auditor",
                "next_url": f"/jobs/{job_id}/review/items",
            },
            follow_redirects=False,
        )
        self.client.post(f"/jobs/{job_id}/review/export-actions", follow_redirects=False)
        with mock.patch("webapp.review._run_patched_standardize_cli", side_effect=self._fake_review_rerun()):
            self.client.post(f"/jobs/{job_id}/review/apply-and-rerun", follow_redirects=False)
            self._run_next_worker_item()
        response = self.client.get(f"/jobs/{job_id}/review/operation-status")
        self.assertEqual(response.status_code, 200)
        latest_operation = response.json()["latest_operation_summary"]
        self.assertEqual(latest_operation["operation_type"], "apply_and_rerun")
        self.assertEqual(latest_operation["status"], "succeeded")
        self.assertTrue(latest_operation["log_paths"])
        self.assertTrue((self.settings.jobs_root / job_id / "review" / "review_operation_summary.json").exists())

    def test_review_items_page_contains_stage_10_2_ux_labels_and_help(self):
        job_id = self._prepare_review_job(include_optional=True)
        response = self.client.get(f"/jobs/{job_id}/review/items")
        self.assertEqual(response.status_code, 200)
        for text in ("接受科目建议", "当前兼容性", "可自动应用", "批量动作", "对选中条目执行批量动作"):
            self.assertIn(text, response.text)

    def test_review_workbench_artifacts_do_not_go_into_repo_root(self):
        job_id = self._prepare_review_job(include_optional=True)
        job = get_job(self.settings, job_id)
        self.client.get(f"/jobs/{job_id}/review")
        self.client.post(
            f"/jobs/{job_id}/review/actions",
            data={
                "review_item_id": "REV_case_1",
                "action_type": "ignore",
                "action_value": "",
                "reviewer_note": "preview apply",
                "reviewer_name": "auditor",
                "next_url": f"/jobs/{job_id}/review/items",
            },
            follow_redirects=False,
        )
        self.client.post(f"/jobs/{job_id}/review/export-actions", follow_redirects=False)
        self.client.get(f"/jobs/{job_id}/review/apply-preview", follow_redirects=False)
        self.client.post(
            f"/jobs/{job_id}/review/bulk-action",
            data={
                "selected_review_item_ids": ["REV_case_1"],
                "action_type": "defer",
                "action_value": "",
                "reviewer_note": "bulk",
                "reviewer_name": "auditor",
                "next_url": f"/jobs/{job_id}/review/items",
            },
            follow_redirects=False,
        )
        self.client.post(f"/jobs/{job_id}/review/apply", follow_redirects=False)
        self._run_next_worker_item()
        review_dir = get_review_dir(job)
        self.assertTrue((review_dir / "review_workbench_summary.json").exists())
        self.assertTrue((review_dir / "review_apply_preview_summary.json").exists())
        self.assertTrue((review_dir / "bulk_review_action_summary.json").exists())
        self.assertTrue((review_dir / "review_operation_summary.json").exists())
        self.assertFalse((REPO_ROOT / "review_workbench_summary.json").exists())
        self.assertFalse((REPO_ROOT / "review_apply_preview_summary.json").exists())
        self.assertFalse((REPO_ROOT / "bulk_review_action_summary.json").exists())
        self.assertFalse((REPO_ROOT / "review_operation_summary.json").exists())

    def test_operation_creation_returns_quickly_and_status_endpoint(self):
        job_id = self._prepare_review_job(include_optional=True)
        self.client.post(
            f"/jobs/{job_id}/review/actions",
            data={
                "review_item_id": "REV_case_1",
                "action_type": "ignore",
                "action_value": "",
                "reviewer_note": "queue quick",
                "reviewer_name": "auditor",
                "next_url": f"/jobs/{job_id}/review/items",
            },
            follow_redirects=False,
        )
        self.client.post(f"/jobs/{job_id}/review/export-actions", follow_redirects=False)
        started = time.perf_counter()
        response = self.client.post(f"/jobs/{job_id}/review/apply-and-rerun", follow_redirects=False)
        elapsed = time.perf_counter() - started
        self.assertEqual(response.status_code, 303)
        self.assertLess(elapsed, 1.5)
        latest_operation = self.client.get(f"/jobs/{job_id}/review/operation-status").json()["latest_operation_summary"]
        self.assertEqual(latest_operation["operation_type"], "apply_and_rerun")
        self.assertEqual(latest_operation["status"], "queued")
        detail = self.client.get(f"/jobs/{job_id}/operations/{latest_operation['operation_id']}")
        self.assertEqual(detail.status_code, 200)
        self.assertEqual(detail.json()["operation"]["status"], "queued")

    def test_operation_stage_timeline_is_generated_for_async_apply_and_rerun(self):
        job_id = self._prepare_review_job(include_optional=True)
        self.client.post(
            f"/jobs/{job_id}/review/actions",
            data={
                "review_item_id": "REV_case_1",
                "action_type": "ignore",
                "action_value": "",
                "reviewer_note": "timeline",
                "reviewer_name": "auditor",
                "next_url": f"/jobs/{job_id}/review/items",
            },
            follow_redirects=False,
        )
        self.client.post(f"/jobs/{job_id}/review/export-actions", follow_redirects=False)
        with mock.patch("webapp.review._run_patched_standardize_cli", side_effect=self._fake_review_rerun()):
            self.client.post(f"/jobs/{job_id}/review/apply-and-rerun", follow_redirects=False)
            self._run_next_worker_item()
        latest_operation = self.client.get(f"/jobs/{job_id}/review/operation-status").json()["latest_operation_summary"]
        timeline_path = self.settings.jobs_root / job_id / "review" / "operations" / latest_operation["operation_id"] / "operation_stage_timeline.json"
        self.assertTrue(timeline_path.exists())
        timeline = json.loads(timeline_path.read_text(encoding="utf-8"))
        stages = [event["stage"] for event in timeline["events"]]
        self.assertIn("created", stages)
        self.assertIn("queued", stages)
        self.assertIn("running", stages)
        self.assertIn("running_standardize", stages)
        self.assertTrue((self.settings.jobs_root / job_id / "review" / "operation_stage_timeline.json").exists())

    def test_operation_lock_prevents_duplicate_running_apply_and_rerun(self):
        job_id = self._prepare_review_job(include_optional=True)
        self.client.post(
            f"/jobs/{job_id}/review/actions",
            data={
                "review_item_id": "REV_case_1",
                "action_type": "ignore",
                "action_value": "",
                "reviewer_note": "duplicate lock",
                "reviewer_name": "auditor",
                "next_url": f"/jobs/{job_id}/review/items",
            },
            follow_redirects=False,
        )
        self.client.post(f"/jobs/{job_id}/review/export-actions", follow_redirects=False)
        first = self.client.post(f"/jobs/{job_id}/review/apply-and-rerun", follow_redirects=False)
        self.assertEqual(first.status_code, 303)
        second = self.client.post(
            f"/jobs/{job_id}/review/apply-and-rerun",
            follow_redirects=False,
            headers={"accept": "application/json"},
        )
        self.assertEqual(second.status_code, 409)
        self.assertEqual(second.json()["error"], "duplicate_operation_blocked")
        lock_summary = json.loads((self.settings.jobs_root / job_id / "review" / "operation_lock_summary.json").read_text(encoding="utf-8"))
        self.assertTrue(lock_summary["blocked"])
        self.assertTrue(lock_summary["blocked_by_operation_id"])

    def test_retry_failed_operation_creates_new_operation_and_retry_summary(self):
        job_id = self._prepare_review_job(include_optional=True)
        self.client.post(
            f"/jobs/{job_id}/review/actions",
            data={
                "review_item_id": "REV_case_1",
                "action_type": "ignore",
                "action_value": "",
                "reviewer_note": "retry failed",
                "reviewer_name": "auditor",
                "next_url": f"/jobs/{job_id}/review/items",
            },
            follow_redirects=False,
        )
        self.client.post(f"/jobs/{job_id}/review/export-actions", follow_redirects=False)
        with mock.patch("webapp.review._run_patched_standardize_cli", side_effect=self._fake_failed_review_rerun()):
            self.client.post(f"/jobs/{job_id}/review/apply-and-rerun", follow_redirects=False)
            self._run_next_worker_item()
        failed_operation = self.client.get(f"/jobs/{job_id}/review/operation-status").json()["latest_operation_summary"]
        self.assertEqual(failed_operation["status"], "failed")

        with mock.patch("webapp.review._run_patched_standardize_cli", side_effect=self._fake_review_rerun()):
            retry_response = self.client.post(
                f"/jobs/{job_id}/operations/{failed_operation['operation_id']}/retry",
                follow_redirects=False,
            )
            self.assertEqual(retry_response.status_code, 303)
            self._run_next_worker_item()
        latest_operation = self.client.get(f"/jobs/{job_id}/review/operation-status").json()["latest_operation_summary"]
        self.assertEqual(latest_operation["status"], "succeeded")
        self.assertNotEqual(latest_operation["operation_id"], failed_operation["operation_id"])
        retry_summary = json.loads((self.settings.jobs_root / job_id / "review" / "operation_retry_summary.json").read_text(encoding="utf-8"))
        self.assertEqual(retry_summary["source_operation_id"], failed_operation["operation_id"])
        self.assertEqual(retry_summary["new_operation_id"], latest_operation["operation_id"])

    def test_cancel_operation_route_marks_queued_operation_cancelled(self):
        job_id = self._prepare_review_job(include_optional=True)
        self.client.post(
            f"/jobs/{job_id}/review/actions",
            data={
                "review_item_id": "REV_case_1",
                "action_type": "ignore",
                "action_value": "",
                "reviewer_note": "cancel queued",
                "reviewer_name": "auditor",
                "next_url": f"/jobs/{job_id}/review/items",
            },
            follow_redirects=False,
        )
        self.client.post(f"/jobs/{job_id}/review/export-actions", follow_redirects=False)
        self.client.post(f"/jobs/{job_id}/review/apply-and-rerun", follow_redirects=False)
        latest_operation = self.client.get(f"/jobs/{job_id}/review/operation-status").json()["latest_operation_summary"]
        cancel_response = self.client.post(
            f"/jobs/{job_id}/operations/{latest_operation['operation_id']}/cancel",
            follow_redirects=False,
        )
        self.assertEqual(cancel_response.status_code, 303)
        detail = self.client.get(f"/jobs/{job_id}/operations/{latest_operation['operation_id']}")
        self.assertEqual(detail.status_code, 200)
        self.assertEqual(detail.json()["operation"]["status"], "cancelled")

    def test_filter_dropdowns_show_chinese_labels_not_raw_enums(self):
        job_id = self._prepare_review_job(include_optional=True)
        response = self.client.get(f"/jobs/{job_id}/review/items")
        self.assertEqual(response.status_code, 200)
        self.assertIn('value="unresolved"', response.text)
        self.assertIn(">未处理</option>", response.text)
        self.assertIn(">复核队列</option>", response.text)
        self.assertIn(">科目映射问题</option>", response.text)
        self.assertIn(">可自动应用</option>", response.text)
        self.assertNotIn(">unresolved</option>", response.text)
        self.assertNotIn(">review_queue</option>", response.text)
        self.assertNotIn(">backend_ready</option>", response.text)

    def test_operation_status_and_type_are_shown_in_chinese(self):
        job_id = self._prepare_review_job(include_optional=True)
        self.client.post(
            f"/jobs/{job_id}/review/actions",
            data={
                "review_item_id": "REV_case_1",
                "action_type": "ignore",
                "action_value": "",
                "reviewer_note": "zh labels",
                "reviewer_name": "auditor",
                "next_url": f"/jobs/{job_id}/review/items",
            },
            follow_redirects=False,
        )
        self.client.post(f"/jobs/{job_id}/review/export-actions", follow_redirects=False)
        self.client.post(f"/jobs/{job_id}/review/apply-and-rerun", follow_redirects=False)
        response = self.client.get(f"/jobs/{job_id}")
        self.assertEqual(response.status_code, 200)
        self.assertIn("应用复核并重新生成", response.text)
        self.assertIn("排队中", response.text)

    def test_operation_log_tail_route_is_restricted_to_allowed_job_dirs(self):
        job_id = self._prepare_review_job(include_optional=True)
        self.client.post(
            f"/jobs/{job_id}/review/actions",
            data={
                "review_item_id": "REV_case_1",
                "action_type": "ignore",
                "action_value": "",
                "reviewer_note": "log tail",
                "reviewer_name": "auditor",
                "next_url": f"/jobs/{job_id}/review/items",
            },
            follow_redirects=False,
        )
        self.client.post(f"/jobs/{job_id}/review/export-actions", follow_redirects=False)
        self.client.post(f"/jobs/{job_id}/review/apply", follow_redirects=False)
        latest_operation = self.client.get(f"/jobs/{job_id}/review/operation-status").json()["latest_operation_summary"]
        allowed_log_path = self.settings.jobs_root / job_id / "review" / "operations" / latest_operation["operation_id"] / "operation.log"
        allowed_log_path.parent.mkdir(parents=True, exist_ok=True)
        allowed_log_path.write_text("allowed log\n", encoding="utf-8")
        outside_log_path = self.temp_path / "outside_operation.log"
        outside_log_path.write_text("outside log\n", encoding="utf-8")
        update_review_operation(
            self.settings,
            latest_operation["operation_id"],
            log_paths=[str(allowed_log_path), str(outside_log_path)],
        )
        response = self.client.get(f"/jobs/{job_id}/operations/{latest_operation['operation_id']}/logs")
        self.assertEqual(response.status_code, 200)
        log_tails = response.json()["log_tails"]
        self.assertEqual(len(log_tails), 1)
        self.assertIn("allowed log", log_tails[0]["tail"])

    def test_operation_artifacts_stay_under_generated_web_paths(self):
        job_id = self._prepare_review_job(include_optional=True)
        self.client.post(
            f"/jobs/{job_id}/review/actions",
            data={
                "review_item_id": "REV_case_1",
                "action_type": "ignore",
                "action_value": "",
                "reviewer_note": "artifacts under runtime",
                "reviewer_name": "auditor",
                "next_url": f"/jobs/{job_id}/review/items",
            },
            follow_redirects=False,
        )
        self.client.post(f"/jobs/{job_id}/review/export-actions", follow_redirects=False)
        with mock.patch("webapp.review._run_patched_standardize_cli", side_effect=self._fake_review_rerun()):
            self.client.post(f"/jobs/{job_id}/review/apply-and-rerun", follow_redirects=False)
            self._run_next_worker_item()
        latest_operation = self.client.get(f"/jobs/{job_id}/review/operation-status").json()["latest_operation_summary"]
        for raw_path in [*latest_operation.get("log_paths", []), *latest_operation.get("result_paths", [])]:
            resolved = Path(raw_path)
            if not resolved.is_absolute():
                resolved = REPO_ROOT / resolved
            self.assertTrue(str(resolved).startswith(str(self.runtime_root)))
        self.assertTrue((self.settings.jobs_root / job_id / "review" / "operations").exists())

    def test_auth_disabled_in_dev_works(self):
        response = self.client.get("/")
        self.assertEqual(response.status_code, 200)

    def test_auth_required_in_prod_without_password_fails_startup(self):
        settings = self.make_settings(env_mode="prod", auth_required=True, admin_password="")
        with self.assertRaisesRegex(RuntimeError, "WEBAPP_ADMIN_PASSWORD"):
            with TestClient(create_app(settings)):
                pass

    def test_auth_enabled_requires_basic_auth(self):
        settings = self.make_settings(env_mode="prod", auth_required=True, admin_password="secret-pass")
        with TestClient(create_app(settings)) as client:
            unauthorized = client.get("/")
            self.assertEqual(unauthorized.status_code, 401)
            authorized = client.get("/", auth=("admin", "secret-pass"))
            self.assertEqual(authorized.status_code, 200)


if __name__ == "__main__":  # pragma: no cover
    unittest.main()
