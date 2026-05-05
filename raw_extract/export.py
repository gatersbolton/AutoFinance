from __future__ import annotations

import csv
import json
from dataclasses import fields
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence

from openpyxl import Workbook

from .models import (
    DETAILED_OUTPUT_FIELDS,
    MAIN_OUTPUT_COLUMNS,
    CompanyResolution,
    FillDateResolution,
    RawMetricCandidate,
    RawMetricIssue,
    dataclass_row,
    serialize_value,
)


CANDIDATE_FIELDS = [
    "candidate_id",
    "accepted",
    "selection_status",
    "provider_rank",
    "duplicate_key",
    *DETAILED_OUTPUT_FIELDS,
]


def export_raw_metrics_run(
    *,
    output_dir: Path,
    accepted: Sequence[RawMetricCandidate],
    candidates: Sequence[RawMetricCandidate],
    issues: Sequence[RawMetricIssue],
    date_audits: Sequence[FillDateResolution],
    company_audits: Sequence[CompanyResolution],
    summary: Dict[str, Any],
    manifest: Dict[str, Any],
) -> List[str]:
    output_dir.mkdir(parents=True, exist_ok=True)

    write_dict_csv(output_dir / "raw_metrics.csv", [row.main_row() for row in accepted], MAIN_OUTPUT_COLUMNS)
    write_jsonl(output_dir / "raw_metrics.jsonl", [row.main_row() for row in accepted])
    write_dict_csv(output_dir / "raw_metrics_detailed.csv", [row.detailed_row() for row in accepted], DETAILED_OUTPUT_FIELDS)
    write_dict_csv(output_dir / "raw_metric_candidates.csv", [row.candidate_row() for row in candidates], CANDIDATE_FIELDS)
    write_dataclass_csv(output_dir / "raw_metrics_issues.csv", issues, RawMetricIssue)
    write_dataclass_csv(output_dir / "date_resolution_audit.csv", date_audits, FillDateResolution)
    write_dataclass_csv(output_dir / "company_resolution_audit.csv", company_audits, CompanyResolution)

    write_xlsx(output_dir / "raw_metrics.xlsx", accepted, candidates, issues)
    write_json(output_dir / "raw_metrics_summary.json", summary)
    write_json(output_dir / "raw_metrics_smoke_summary.json", summary)
    write_json(output_dir / "extraction_run_manifest.json", manifest)

    return sorted(str(path) for path in output_dir.iterdir() if path.is_file())


def write_xlsx(
    path: Path,
    accepted: Sequence[RawMetricCandidate],
    candidates: Sequence[RawMetricCandidate],
    issues: Sequence[RawMetricIssue],
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "raw_metrics"
    append_rows(sheet, MAIN_OUTPUT_COLUMNS, [row.main_row() for row in accepted])

    detailed = workbook.create_sheet("raw_metrics_detailed")
    append_rows(detailed, DETAILED_OUTPUT_FIELDS, [row.detailed_row() for row in accepted])

    candidate_sheet = workbook.create_sheet("raw_metric_candidates")
    append_rows(candidate_sheet, CANDIDATE_FIELDS, [row.candidate_row() for row in candidates])

    issue_sheet = workbook.create_sheet("raw_metrics_issues")
    issue_fields = [field.name for field in fields(RawMetricIssue)]
    append_rows(issue_sheet, issue_fields, [dataclass_row(issue) for issue in issues])
    workbook.save(path)


def append_rows(sheet, fieldnames: Sequence[str], rows: Iterable[Dict[str, Any]]) -> None:
    sheet.append(list(fieldnames))
    for row in rows:
        sheet.append([serialize_value(row.get(field)) for field in fieldnames])


def write_dict_csv(path: Path, rows: List[Dict[str, Any]], fieldnames: Sequence[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(fieldnames))
        writer.writeheader()
        for row in rows:
            writer.writerow({field: serialize_value(row.get(field)) for field in fieldnames})


def write_dataclass_csv(path: Path, rows: Sequence[Any], model_cls: Any) -> None:
    fieldnames = [field.name for field in fields(model_cls)]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(dataclass_row(row))


def write_jsonl(path: Path, rows: Iterable[Dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False, separators=(",", ":"), sort_keys=True))
            handle.write("\n")


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
